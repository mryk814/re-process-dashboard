from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "data" / "source" / "material_workbench_tutorial_v1.xlsx"
PROCESS_SOURCE = ROOT / "data" / "source" / "material_workbench_process_v1.xlsx"


def build(output: Path, fixture: str) -> None:
    workbook = load_workbook(
        PROCESS_SOURCE if fixture in {"partial-evidence", "unresolved-heat-series"} else SOURCE
    )
    try:
        if fixture == "partial-evidence":
            hot_sheet = workbook["熱延組織"]
            hot_headers = [cell.value for cell in hot_sheet[1]]
            hot_sheet.delete_cols(hot_headers.index("画像リンク先") + 1)
            anneal_sheet = workbook["焼鈍板組織"]
            anneal_headers = [cell.value for cell in anneal_sheet[1]]
            anneal_sheet.cell(
                row=1,
                column=anneal_headers.index("画像リンク名") + 1,
                value="焼鈍顕微鏡ファイル",
            )
        elif fixture == "unresolved-heat-series":
            history = workbook["焼鈍履歴"]
            headers = [cell.value for cell in history[1]]
            parent_index = headers.index("焼鈍履歴_key**")
            temperature_index = headers.index("温度[℃]")
            first_parent = next(
                row[parent_index].value
                for row in history.iter_rows(min_row=2)
                if row[parent_index].value
            )
            for row in history.iter_rows(min_row=2):
                if row[parent_index].value == first_parent:
                    row[temperature_index].value = None
            annealing = workbook["焼鈍条件-3CGL"]
            annealing_headers = [cell.value for cell in annealing[1]]
            annealing_parent_index = annealing_headers.index("焼鈍条件-3CGL_key**")
            line_speed_index = annealing_headers.index("LS[mpm]")
            for row in annealing.iter_rows(min_row=2):
                if row[annealing_parent_index].value == first_parent:
                    row[line_speed_index].value = None
        else:
            sheet = workbook["熱延"]
            sheet.title = "熱延条件（設備B）"
            headers = [cell.value for cell in sheet[1]]
            sheet.cell(
                row=1,
                column=headers.index("均熱温度[℃]") + 1,
                value="加熱温度[℃]",
            )
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("output", type=Path)
    parser.add_argument(
        "--fixture",
        choices=("renamed-source", "partial-evidence", "unresolved-heat-series"),
        default="renamed-source",
    )
    arguments = parser.parse_args()
    build(arguments.output, arguments.fixture)
