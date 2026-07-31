from __future__ import annotations

import json
from dataclasses import replace
from pathlib import Path

import material_workbench.bootstrap.resources as resources_module
import numpy as np
import pytest
from fastapi.testclient import TestClient
from material_workbench.app import create_app
from material_workbench.application.dataset_registration import register_managed_dataset
from material_workbench.application.material_lineage_candidates import (
    candidate_from_lineage,
    lineage_candidate_options,
)
from material_workbench.contracts.schemas import CandidateInput
from material_workbench.contracts.task_contracts import TaskDefinition
from material_workbench.data.importer import (
    _derived_anneal_feature_row,
    detect_dataset_profile_path,
    load_workbook_data,
)
from material_workbench.data.profile_workbench import (
    create_source_binding_draft,
    validate_workbook_profile,
)
from material_workbench.data.profiles.canonicalization import canonicalize_workbook
from material_workbench.data.profiles.loading import (
    load_dataset_profile,
    load_task_definitions,
    materialize_dataset_profile_document,
)
from material_workbench.data.profiles.requirements import task_data_requirements
from material_workbench.data.profiles.schema import DatasetProfileError
from material_workbench.data.profiles.validation import preflight_workbook
from material_workbench.developer_experience.source_inspection import (
    inspect_source_against_profiles,
)
from material_workbench.modeling.feature_pipeline import (
    build_feature_bundle,
)
from material_workbench.modeling.feature_pipeline import (
    candidate_from_observation as anneal_candidate_from_observation,
)
from material_workbench.modeling.hot_rolling_feature_pipeline import (
    build_hot_rolling_features,
)
from material_workbench.task_composition.catalog import registered_task_modules
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
TUTORIAL_SOURCE = ROOT / "data" / "source" / "material_workbench_tutorial_v2.xlsx"
PROCESS_SOURCE = ROOT / "data" / "source" / "material_workbench_process_v1.xlsx"
SOURCE = TUTORIAL_SOURCE
PROFILE = ROOT / "backend" / "src" / "material_workbench" / "data" / "dataset-input-profile-tutorial-base.json"
PROCESS_PROFILE = ROOT / "backend" / "src" / "material_workbench" / "data" / "dataset-input-profile-process-v1.json"
FLANK_WEAR_PROFILE = ROOT / "backend" / "src" / "material_workbench" / "data" / "dataset-input-profile-flank-wear-v1.json"
FLANK_WEAR_SOURCE = ROOT / "data" / "source" / "cutting_tool_flank_wear_synthetic_dataset.xlsx"


def _replace_numeric_column_unit(
    workbook: Workbook,
    sheet_name: str,
    original_column: str,
    replacement_column: str,
    transform,
) -> None:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index(original_column) + 1
    sheet.cell(row=1, column=column, value=replacement_column)
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row, column=column).value
        if isinstance(value, (int, float)):
            sheet.cell(row=row, column=column, value=transform(float(value)))


def test_non_identity_units_are_applied_to_observations_and_heat_series(
    tmp_path: Path,
) -> None:
    profile = load_dataset_profile(PROFILE)
    baseline_workbook = load_workbook(TUTORIAL_SOURCE, data_only=True)
    baseline = canonicalize_workbook(baseline_workbook, profile)
    baseline_workbook.close()

    workbook = load_workbook(TUTORIAL_SOURCE)
    _replace_numeric_column_unit(
        workbook,
        "熱延引張",
        "TS[MPa]",
        "TS[kPa]",
        lambda value: value * 1000,
    )
    _replace_numeric_column_unit(
        workbook,
        "焼鈍履歴",
        "到達時間[s]",
        "到達時間[min]",
        lambda value: value / 60,
    )
    _replace_numeric_column_unit(
        workbook,
        "焼鈍履歴",
        "実績温度[℃]",
        "実績温度[K]",
        lambda value: value + 273.15,
    )
    document = materialize_dataset_profile_document(PROFILE)
    document["shared"]["column_aliases"] = {
        "hot_tensile": {"TS[MPa]": "TS[kPa]"},
        "anneal_history": {
            "到達時間[s]": "到達時間[min]",
            "実績温度[℃]": "実績温度[K]",
        },
    }
    series = next(
        mapping
        for mapping in document["tasks"]["annealed-properties-v1"]["mappings"]
        if mapping["kind"] == "ordered_heat_series"
    )
    series["series_columns"]["time_source_unit"] = "min"
    series["series_columns"]["value_source_unit"] = "K"
    profile_path = tmp_path / "converted-profile.json"
    profile_path.write_text(json.dumps(document, ensure_ascii=False), encoding="utf-8")
    converted = canonicalize_workbook(workbook, load_dataset_profile(profile_path))
    workbook.close()

    baseline_observation = next(
        item for item in baseline.observations
        if item.task_id == "hot-rolled-properties-v1" and item.targets
    )
    converted_observation = next(
        item for item in converted.observations
        if item.id == baseline_observation.id
        and item.task_id == baseline_observation.task_id
    )
    assert converted_observation.targets["TS"] == pytest.approx(
        baseline_observation.targets["TS"]
    )
    series_identity = next(iter(baseline.heat_series))
    baseline_points = baseline.heat_series[series_identity]
    converted_points = converted.heat_series[series_identity]
    assert len(converted_points) == len(baseline_points)
    for baseline_point, converted_point in zip(baseline_points, converted_points):
        assert converted_point["time_s"] == pytest.approx(baseline_point["time_s"])
        assert converted_point["temperature_c"] == pytest.approx(
            baseline_point["temperature_c"]
        )


def test_non_identity_units_are_applied_before_observation_scoped_aggregation(
    tmp_path: Path,
) -> None:
    profile = load_dataset_profile(FLANK_WEAR_PROFILE)
    baseline_workbook = load_workbook(FLANK_WEAR_SOURCE, data_only=True)
    baseline = canonicalize_workbook(baseline_workbook, profile)
    baseline_workbook.close()

    workbook = load_workbook(FLANK_WEAR_SOURCE)
    _replace_numeric_column_unit(
        workbook,
        "逃げ面摩耗履歴",
        "切削距離[m]",
        "切削距離[cm]",
        lambda value: value * 100,
    )
    document = materialize_dataset_profile_document(FLANK_WEAR_PROFILE)
    document["shared"]["column_aliases"] = {
        "wear_history": {"切削距離[m]": "切削距離[cm]"},
    }
    distance_mapping = next(
        mapping
        for mapping in document["tasks"]["flank-wear-v1"]["mappings"]
        if mapping["path"] == "process.cutting_distance_m"
    )
    distance_mapping["source_unit"] = "cm"
    profile_path = tmp_path / "flank-wear-cm-profile.json"
    profile_path.write_text(json.dumps(document, ensure_ascii=False), encoding="utf-8")
    converted = canonicalize_workbook(workbook, load_dataset_profile(profile_path))
    workbook.close()

    identity = next(
        identity for identity, entity in baseline.entities.items()
        if "process.cutting_distance_m" in entity.values.get("flank-wear-v1", {})
    )
    assert converted.entities[identity].values["flank-wear-v1"][
        "process.cutting_distance_m"
    ] == pytest.approx(
        baseline.entities[identity].values["flank-wear-v1"][
            "process.cutting_distance_m"
        ]
    )


def test_tutorial_profile_keeps_relations_repeats_and_partial_targets_explicit() -> None:
    assert detect_dataset_profile_path(TUTORIAL_SOURCE).name == "dataset-input-profile-tutorial.json"
    data = load_workbook_data(TUTORIAL_SOURCE)

    assert data.profile_id == "thin-sheet-tutorial-v1"
    assert len(data.composition) == 4
    assert len(data.hot_rolling_features) == 6
    assert len(data.anneal_features) == 6
    assert len(data.observations) == 26
    assert len([row for row in data.observations if row["parent_key"] == "AN-02"]) == 4

    tt03 = next(row for row in data.observations if row["id"] == "TT-03")
    tt07 = next(row for row in data.observations if row["id"] == "TT-07")
    ht01 = next(row for row in data.observations if row["id"] == "HT-01")
    an06 = data.anneal_features["AN-06"]
    assert "YS[MPa]" not in tt03["outputs"]
    assert "EL[%]" not in tt07["outputs"]
    assert {
        row["composition_key"]
        for row in data.observations
        if row["parent_key"] == "AN-02"
    } == {"ME-01", "ME-02"}
    assert ht01["composition_key"] == "ME-01"
    assert len(ht01["relation_context_ids"]) >= 2
    assert an06["ls_mpm"] is None
    assert an06["feature_eligible"] is True
    assert [
        (option.process_key, option.melt_key)
        for option in lineage_candidate_options(data, "HT-01")
    ] == [("HR-01", "ME-01")]
    shared_candidate = candidate_from_lineage(
        data, "AN-02", process_key="AN-02", melt_key="ME-01"
    )
    shared_route_ids = set(
        shared_candidate.provenance.source_ref.relation_context_ids
    )
    assert shared_route_ids
    assert all(
        route.members.get("melt") in {None, "ME-01"}
        for route in data.relation_routes
        if route.id in shared_route_ids
    )
    assert data.detected_quality == []


def test_profile_is_driven_by_production_task_definitions() -> None:
    profile = load_dataset_profile()
    for task_id, task in profile.tasks.items():
        definition = profile.task_definitions[task_id]
        ordered_paths = tuple(
            field.path
            for group in sorted(definition.input_groups, key=lambda item: item.order)
            for field in sorted(group.fields, key=lambda item: item.order)
        )
        assert tuple(mapping.path for mapping in task.mappings) == ordered_paths


def test_preflight_aggregates_duplicate_and_missing_headers() -> None:
    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    melt = workbook["溶製"]
    melt.cell(1, 2).value = melt.cell(1, 1).value
    hot = workbook["熱延"]
    target = next(cell for cell in hot[1] if cell.value == "均熱温度[℃]")
    target.value = "renamed"

    with pytest.raises(DatasetProfileError) as caught:
        preflight_workbook(workbook, load_dataset_profile())

    assert any("duplicate header" in error for error in caught.value.errors)
    assert any("missing required column '均熱温度[℃]'" in error for error in caught.value.errors)


def test_task_unused_relation_columns_do_not_block_registration(
    tmp_path: Path,
) -> None:
    profile = load_dataset_profile(PROCESS_PROFILE)
    requirements = task_data_requirements(profile)
    required_entity_types = {
        join.entity_type
        for join in profile.shared.relation.joins
        if requirements.requires_relation(join)
    }
    optional_joins = [
        join
        for join in profile.shared.relation.joins
        if not requirements.requires_relation(join)
    ]
    assert {join.entity_type for join in optional_joins} == {
        "anneal_microstructure",
        "hot_microstructure",
    }
    baseline_inspection = inspect_source_against_profiles(
        PROCESS_SOURCE,
        profile_path=PROCESS_PROFILE,
    )
    baseline_candidate = next(
        candidate
        for candidate in baseline_inspection.candidates
        if Path(candidate.profile_path) == PROCESS_PROFILE.resolve()
    )
    relation_sheet = profile.sheet_for_role(profile.shared.relation.role)
    assert not (
        {join.column for join in optional_joins}
        & set(baseline_candidate.extra_columns.get(relation_sheet, []))
    )

    workbook = load_workbook(PROCESS_SOURCE, read_only=False, data_only=True)
    baseline_source = tmp_path / "baseline.xlsx"
    workbook.save(baseline_source)
    relation = workbook[profile.sheet_for_role(profile.shared.relation.role)]
    headers = [cell.value for cell in relation[1]]
    for join in optional_joins:
        column = next(name for name in join.source_columns if name in headers)
        relation.cell(row=1, column=headers.index(column) + 1, value=f"unmapped-{join.entity_type}")
    source = tmp_path / "unused-relations-unmapped.xlsx"
    workbook.save(source)
    workbook.close()

    baseline_workbook = load_workbook(baseline_source, read_only=True, data_only=True)
    try:
        baseline = canonicalize_workbook(baseline_workbook, profile)
    finally:
        baseline_workbook.close()
    modified = load_workbook(source, read_only=True, data_only=True)
    try:
        canonical = canonicalize_workbook(modified, profile)
    finally:
        modified.close()
    assert canonical.entities == baseline.entities
    assert canonical.observations == baseline.observations
    assert canonical.heat_series == baseline.heat_series
    assert [
        {
            entity_type: identity
            for entity_type, identity in relation_row.items()
            if entity_type in required_entity_types
        }
        for relation_row in canonical.relations
        if required_entity_types & relation_row.keys()
    ] == [
        {
            entity_type: identity
            for entity_type, identity in relation_row.items()
            if entity_type in required_entity_types
        }
        for relation_row in baseline.relations
        if required_entity_types & relation_row.keys()
    ]
    assert all(
        join.entity_type not in relation_row
        for join in optional_joins
        for relation_row in canonical.relations
    )

    draft = create_source_binding_draft(source, PROCESS_PROFILE)
    assert draft is not None
    assert draft["complete"] is True
    optional_slots = {
        slot["canonical_name"]: slot
        for slot in draft["slots"]
        if slot["semantic_kind"] == "relation_join" and not slot["required"]
    }
    assert {join.column for join in optional_joins} <= set(optional_slots)
    assert all(
        optional_slots[join.column]["state"] != "confirmed"
        for join in optional_joins
    )

    inspection = inspect_source_against_profiles(
        source,
        profile_path=PROCESS_PROFILE,
    )
    selected = next(
        candidate
        for candidate in inspection.candidates
        if Path(candidate.profile_path) == PROCESS_PROFILE.resolve()
    )
    missing_relation_columns = set(selected.missing_columns.get(relation_sheet, []))
    assert not ({join.column for join in optional_joins} & missing_relation_columns)
    assert not (
        {join.column for join in optional_joins}
        & set(selected.extra_columns.get(relation_sheet, []))
    )
    assert selected.validation_error is None

    report = validate_workbook_profile(source, PROCESS_PROFILE)
    assert report["registration_ready"] is True
    registered = register_managed_dataset(
        database=tmp_path / "workspace.db",
        source=source,
        library_root=tmp_path / "library",
        profile_path=PROCESS_PROFILE,
        name="unused relation columns",
    )
    assert registered.profile_id == profile.profile_id


def test_task_required_relation_column_still_blocks_registration(
    tmp_path: Path,
) -> None:
    profile = load_dataset_profile(PROCESS_PROFILE)
    requirements = task_data_requirements(profile)
    required_join = next(
        join
        for join in profile.shared.relation.joins
        if requirements.requires_relation(join)
    )
    workbook = load_workbook(PROCESS_SOURCE, read_only=False, data_only=True)
    relation = workbook[profile.sheet_for_role(profile.shared.relation.role)]
    headers = [cell.value for cell in relation[1]]
    column = next(name for name in required_join.source_columns if name in headers)
    relation.cell(row=1, column=headers.index(column) + 1, value="required-relation-unmapped")
    source = tmp_path / "required-relation-unmapped.xlsx"
    workbook.save(source)

    with pytest.raises(DatasetProfileError) as caught:
        preflight_workbook(workbook, profile)
    workbook.close()

    assert any(required_join.path in error for error in caught.value.errors)
    draft = create_source_binding_draft(source, PROCESS_PROFILE)
    assert draft is not None
    required_slot = next(
        slot
        for slot in draft["slots"]
        if slot["semantic_kind"] == "relation_join"
        and slot["canonical_name"] == required_join.column
    )
    assert required_slot["required"] is True
    assert required_slot["state"] != "confirmed"
    assert draft["complete"] is False

    inspection = inspect_source_against_profiles(
        source,
        profile_path=PROCESS_PROFILE,
    )
    selected = next(
        candidate
        for candidate in inspection.candidates
        if Path(candidate.profile_path) == PROCESS_PROFILE.resolve()
    )
    relation_sheet = profile.sheet_for_role(profile.shared.relation.role)
    assert required_join.column in selected.missing_columns[relation_sheet]


def test_parent_consistency_uses_only_task_route_parents() -> None:
    profile = load_dataset_profile(PROCESS_PROFILE)
    annealing = next(
        join
        for join in profile.shared.relation.joins
        if join.entity_type == "annealing"
    )
    scoped_annealing = annealing.model_copy(
        update={
            "parent_entity_types": (*annealing.parent_entity_types, "hot_microstructure"),
            "edge_parent_entity_types": ("cold_rolling",),
            "parent_consistency": "exactly_one",
        }
    )
    relation = profile.shared.relation.model_copy(
        update={
            "joins": tuple(
                scoped_annealing if join.entity_type == "annealing" else join
                for join in profile.shared.relation.joins
            )
        }
    )
    scoped_profile = profile.model_copy(
        update={"shared": profile.shared.model_copy(update={"relation": relation})}
    )

    requirements = task_data_requirements(scoped_profile)
    assert "cold_rolling" in requirements.relation_entity_types
    assert "hot_microstructure" not in requirements.relation_entity_types
    workbook = load_workbook(PROCESS_SOURCE, read_only=True, data_only=True)
    try:
        preflight_workbook(workbook, scoped_profile)
    finally:
        workbook.close()


def test_reordered_columns_and_unmapped_metadata_do_not_change_canonical_values(tmp_path: Path) -> None:
    baseline = load_workbook_data(SOURCE)
    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    melt = workbook["溶製"]
    c_column = next(cell.column for cell in melt[1] if cell.value == "C[mass%]")
    si_column = next(cell.column for cell in melt[1] if cell.value == "Si[mass%]")
    for row in range(1, melt.max_row + 1):
        left, right = melt.cell(row, c_column).value, melt.cell(row, si_column).value
        melt.cell(row, c_column).value, melt.cell(row, si_column).value = right, left
    extra_column = melt.max_column + 1
    melt.cell(1, extra_column).value = "unused_metadata"
    for row in range(2, melt.max_row + 1):
        melt.cell(row, extra_column).value = float(10_000 + row)
    modified = tmp_path / "reordered.xlsx"
    workbook.save(modified)

    actual = load_workbook_data(modified)
    assert actual.composition == baseline.composition
    assert actual.hot_rolling_features == baseline.hot_rolling_features
    assert actual.anneal_features == baseline.anneal_features
    assert actual.observations == baseline.observations
    assert actual.medians == baseline.medians

    def representative_vector(data) -> np.ndarray:
        parent = "AN-01"
        process = data.anneal_features[parent]
        candidate = CandidateInput(
            name="dataset-profile-golden",
            inputs={
                "composition": data.composition["ME-01"],
                "process": {"ls_mpm": process["ls_mpm"]},
                "categorical": {},
                "heat_pattern": [
                    {"time_s": point["time_s"], "temperature_c": point["temperature_c"]}
                    for point in process["heat_pattern"]
                ],
            },
        )
        return build_feature_bundle(candidate).values

    expected = representative_vector(baseline)
    assert expected.shape == (41,)
    assert np.isfinite(expected).all()
    np.testing.assert_allclose(representative_vector(actual), expected, rtol=1e-12, atol=1e-12)

    def representative_hot_vector(data) -> np.ndarray:
        process = data.hot_rolling_features["HR-01"]
        candidate = CandidateInput(
            name="hot-dataset-profile-golden",
            inputs={
                "composition": data.composition["ME-01"],
                "process": {
                    key: process[key]
                    for key in (
                        "soaking_temperature_c", "finish_temperature_c", "entry_thickness_mm",
                        "exit_thickness_mm", "hold_temperature_c", "hold_time_min",
                    )
                },
                "categorical": {},
                "heat_pattern": None,
            },
        )
        return build_hot_rolling_features(candidate, data.medians).values

    hot_expected = representative_hot_vector(baseline)
    assert hot_expected.shape == (37,)
    assert np.isfinite(hot_expected).all()
    np.testing.assert_allclose(representative_hot_vector(actual), hot_expected, rtol=1e-12, atol=1e-12)


def test_unknown_unit_and_duplicate_mapping_are_rejected(tmp_path: Path) -> None:
    raw = json.loads(PROFILE.read_text(encoding="utf-8"))
    mappings = raw["tasks"]["hot-rolled-properties-v1"]["mappings"]
    mappings[14]["source_unit"] = "mystery-temperature"
    mappings.append(dict(mappings[0]))
    invalid = tmp_path / "invalid-profile.json"
    invalid.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(DatasetProfileError) as caught:
        load_dataset_profile(invalid)

    assert any("unknown or implicit unit conversion" in error for error in caught.value.errors)
    assert any("multiple source mappings" in error for error in caught.value.errors)


@pytest.mark.parametrize(
    "mutation",
    [
        "nested_unknown_field",
        "duplicate_auxiliary_key",
        "unknown_range_key",
        "missing_relation_join",
        "missing_required_technical",
        "duplicate_entity_role",
        "empty_policy_values",
        "nested_task_id",
    ],
)
def test_profile_rejects_ambiguous_nested_measurement_contracts(tmp_path: Path, mutation: str) -> None:
    raw = json.loads(PROFILE.read_text(encoding="utf-8"))
    hot = raw["tasks"]["hot-rolled-properties-v1"]["observations"][0]
    if mutation == "nested_unknown_field":
        hot["auxiliary"][0]["guess_from_header"] = True
    elif mutation == "duplicate_auxiliary_key":
        hot["auxiliary"][1]["key"] = hot["auxiliary"][0]["key"]
    elif mutation == "unknown_range_key":
        raw["shared"]["physical_ranges"]["hot-rolled-properties-v1"]["invented"] = [0, 1]
    elif mutation == "missing_relation_join":
        raw["shared"]["relation"]["joins"].pop(0)
    elif mutation == "missing_required_technical":
        raw["shared"]["technical"] = [
            item for item in raw["shared"]["technical"]
            if not (item["role"] == "hot_rolling" and item["name"] == "equipment")
        ]
    elif mutation == "duplicate_entity_role":
        raw["shared"]["entities"][1]["role"] = raw["shared"]["entities"][0]["role"]
    elif mutation == "empty_policy_values":
        raw["shared"]["eligibility"][0]["accepted_values"] = []
    else:
        raw["tasks"]["annealed-properties-v1"]["task_id"] = "hot-rolled-properties-v1"
    invalid = tmp_path / f"{mutation}.json"
    invalid.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(DatasetProfileError):
        load_dataset_profile(invalid)


@pytest.mark.parametrize("mutation", ["missing_task", "missing_output"])
def test_profile_requires_every_task_and_output_exactly_once(tmp_path: Path, mutation: str) -> None:
    raw = json.loads(PROFILE.read_text(encoding="utf-8"))
    if mutation == "missing_task":
        del raw["tasks"]["hot-rolled-properties-v1"]
    else:
        raw["tasks"]["annealed-properties-v1"]["observations"][1]["targets"] = []
    invalid = tmp_path / f"{mutation}.json"
    invalid.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(DatasetProfileError):
        load_dataset_profile(invalid)


def test_preflight_rejects_empty_required_heat_series_and_policy_without_signal(tmp_path: Path) -> None:
    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    profile = load_dataset_profile()
    heat_mapping = next(
        mapping for task in profile.tasks.values() for mapping in task.mappings
        if mapping.kind == "ordered_heat_series"
    )
    assert heat_mapping.series_columns is not None
    heat_sheet = workbook[profile.sheet_for_role(heat_mapping.role)]
    headers = [cell.value for cell in heat_sheet[1]]
    time_column = headers.index(heat_mapping.series_columns.time) + 1
    value_column = headers.index(heat_mapping.series_columns.value) + 1
    for row in range(2, heat_sheet.max_row + 1):
        heat_sheet.cell(row, time_column).value = None
        heat_sheet.cell(row, value_column).value = None

    raw = json.loads(PROFILE.read_text(encoding="utf-8"))
    raw["shared"]["eligibility"][0]["accepted_values"] = ["TYPO"]
    invalid_profile = tmp_path / "policy-without-signal.json"
    invalid_profile.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(DatasetProfileError) as caught:
        preflight_workbook(workbook, load_dataset_profile(invalid_profile))

    assert any("at least two numeric points" in error for error in caught.value.errors)
    assert any("has no accepted source values" in error for error in caught.value.errors)


def test_non_numeric_heat_series_order_excludes_only_its_parent() -> None:
    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    profile = load_dataset_profile()
    heat_mapping = next(
        mapping for task in profile.tasks.values() for mapping in task.mappings
        if mapping.kind == "ordered_heat_series"
    )
    assert heat_mapping.series_columns is not None
    sheet = workbook[profile.sheet_for_role(heat_mapping.role)]
    order_column = next(
        cell.column for cell in sheet[1] if cell.value == heat_mapping.series_columns.order
    )
    parent = str(sheet.cell(2, next(
        cell.column for cell in sheet[1] if cell.value == heat_mapping.series_columns.parent
    )).value)
    sheet.cell(2, order_column).value = "bad-order"

    canonical = canonicalize_workbook(workbook, profile)

    assert (heat_mapping.parent_entity_type or "annealing", parent) not in canonical.heat_series


def test_preflight_executes_declared_parent_consistency(tmp_path: Path) -> None:
    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    raw = json.loads(PROFILE.read_text(encoding="utf-8"))
    hot_join_document = next(
        join for join in raw["shared"]["relation"]["joins"] if join["entity_type"] == "hot_rolling"
    )
    hot_join_document["parent_consistency"] = "exactly_one"
    strict_profile_path = tmp_path / "strict-parent-profile.json"
    strict_profile_path.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")
    profile = load_dataset_profile(strict_profile_path)
    relation = workbook[profile.sheet_for_role(profile.shared.relation.role)]
    headers = [cell.value for cell in relation[1]]
    hot_join = next(join for join in profile.shared.relation.joins if join.entity_type == "hot_rolling")
    melt_join = next(join for join in profile.shared.relation.joins if join.entity_type == "melt")
    hot_column = headers.index(hot_join.column) + 1
    melt_column = headers.index(melt_join.column) + 1
    first_row_by_hot: dict[str, int] = {}
    changed = False
    for row in range(2, relation.max_row + 1):
        hot_key = relation.cell(row, hot_column).value
        if hot_key is None:
            continue
        first_row = first_row_by_hot.setdefault(str(hot_key), row)
        if first_row != row:
            original_melt = relation.cell(first_row, melt_column).value
            relation.cell(row, melt_column).value = f"{original_melt}-CONFLICT"
            changed = True
            break
    assert changed

    with pytest.raises(DatasetProfileError) as caught:
        preflight_workbook(workbook, profile)

    assert any("conflicting parents" in error for error in caught.value.errors)


def test_canonical_entities_separate_values_from_source_metadata() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    dataset = canonicalize_workbook(workbook, load_dataset_profile())
    entity = dataset.entities[("melt", "ME-01")]
    anneal_values = entity.values["annealed-properties-v1"]

    assert set(anneal_values) == {f"composition.{name}" for name in ("C", "Si", "Mn", "P", "S", "Al", "Cu", "Ni", "Cr", "Mo", "Ti", "B", "O", "N")}
    assert "プロジェクト名" in entity.source_metadata
    assert entity.source_locator == {"sheet": "溶製", "row": 2}
    anneal = dataset.entities[("annealing", "AN-01")]
    anneal_feature_row = next(
        row for row in dataset.rows("anneal_features")
        if dataset.technical_value(row, "anneal_features", "parent_key") == "AN-01"
    )
    assert dataset.value(anneal_feature_row, "annealed-properties-v1", "process.ls_mpm") > 0
    assert len(anneal.values["annealed-properties-v1"]["heat_pattern"]) == 6


def test_canonical_entity_identity_does_not_merge_equal_keys_across_types() -> None:
    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    workbook["溶製"].cell(2, 1).value = "SAME-KEY"
    hot_key_column = next(cell.column for cell in workbook["熱延"][1] if cell.value == "熱延_key")
    workbook["熱延"].cell(2, hot_key_column).value = "SAME-KEY"
    relation = workbook["relation"]
    melt_relation_column = next(cell.column for cell in relation[1] if cell.value == "溶製_key")
    hot_relation_column = next(cell.column for cell in relation[1] if cell.value == "熱延_key")
    relation.cell(2, melt_relation_column).value = "SAME-KEY"
    relation.cell(2, hot_relation_column).value = "SAME-KEY"

    dataset = canonicalize_workbook(workbook, load_dataset_profile())

    assert dataset.entities[("melt", "SAME-KEY")].identity == ("melt", "SAME-KEY")
    assert dataset.entities[("hot_rolling", "SAME-KEY")].identity == ("hot_rolling", "SAME-KEY")
    assert dataset.relations[0]["melt"] == ("melt", "SAME-KEY")
    assert dataset.relations[0]["hot_rolling"] == ("hot_rolling", "SAME-KEY")


def test_lineage_candidate_uses_profile_key_mapping_not_external_header_names(tmp_path: Path) -> None:
    raw_profile = json.loads(PROFILE.read_text(encoding="utf-8"))
    melt_entity = next(item for item in raw_profile["shared"]["entities"] if item["role"] == "melt")
    melt_join = next(item for item in raw_profile["shared"]["relation"]["joins"] if item["entity_type"] == "melt")
    old_key = melt_entity["key"]
    melt_entity["key"] = "melt_id"
    melt_join["column"] = "melt_id"
    profile_path = tmp_path / "renamed-key-profile.json"
    profile_path.write_text(json.dumps(raw_profile, ensure_ascii=False), encoding="utf-8")

    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    for sheet_name in (raw_profile["shared"]["sheets"]["melt"], raw_profile["shared"]["sheets"]["relation"]):
        header = next(cell for cell in workbook[sheet_name][1] if cell.value == old_key)
        header.value = "melt_id"
    source = tmp_path / "renamed-key.xlsx"
    workbook.save(source)
    workbook.close()

    data = load_workbook_data(source, profile_path)
    candidate = candidate_from_lineage(data, "AN-01")

    assert len(data.composition) == 4
    assert candidate.inputs.composition == data.composition["ME-01"]


def test_imported_relations_keep_every_composition_that_shares_one_process(tmp_path: Path) -> None:
    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    melt_sheet = workbook["溶製"]
    melt_key_column = next(cell.column for cell in melt_sheet[1] if cell.value == "溶製_key")
    copied_melt = [cell.value for cell in melt_sheet[2]]
    copied_melt[melt_key_column - 1] = "ME-SHARED-PROCESS"
    melt_sheet.append(copied_melt)

    relation_sheet = workbook["relation"]
    relation_headers = {cell.value: cell.column for cell in relation_sheet[1]}
    source_relation = next(
        row
        for row in relation_sheet.iter_rows(min_row=2)
        if row[relation_headers["溶製_key"] - 1].value == "ME-01"
        and row[relation_headers["焼鈍_key"] - 1].value == "AN-01"
    )
    copied_relation = [cell.value for cell in source_relation]
    copied_relation[relation_headers["溶製_key"] - 1] = "ME-SHARED-PROCESS"
    relation_sheet.append(copied_relation)
    source = tmp_path / "shared-process.xlsx"
    workbook.save(source)
    workbook.close()

    data = load_workbook_data(source)
    process_options = [
        option
        for option in lineage_candidate_options(data, "ME-01")
        if option.process_key == "AN-01"
    ]

    assert {option.melt_key for option in process_options} == {"ME-01"}
    ambiguous = next(row for row in data.observations if row["id"] == "TT-01")
    assert not ambiguous["eligible"]
    assert "試験に対応する上流成分が複数あります" in ambiguous["eligibility_reasons"]
    with pytest.raises(ValueError, match="組み合わせを候補化できません"):
        candidate_from_lineage(
            data,
            "ME-01",
            process_key="AN-01",
            melt_key="ME-SHARED-PROCESS",
        )
    candidate = candidate_from_lineage(
        data,
        "ME-SHARED-PROCESS",
        process_key="AN-01",
        melt_key="ME-SHARED-PROCESS",
    )
    assert candidate.name == "実績 AN-01 / ME-SHARED-PROCESS"
    assert candidate.inputs.composition == data.composition["ME-SHARED-PROCESS"]


def test_training_observation_uses_composition_from_its_relation_route() -> None:
    data = load_workbook_data(PROCESS_SOURCE)

    first = next(row for row in data.observations if row["id"] == "HT-00006")
    second = next(row for row in data.observations if row["id"] == "HT-00007")

    assert first["parent_key"] == second["parent_key"] == "HR-00004"
    assert first["composition_key"] == second["composition_key"] == "ME-00003"
    assert first["eligible"] and second["eligible"]
    assert first["condition_context_id"] == "ME-00003::HR-00004"
    assert first["relation_context_ids"] == ["relationEx:15"]


def test_explicit_anneal_history_is_trainable_without_line_speed() -> None:
    data = load_workbook_data(PROCESS_SOURCE)
    points = [
        {"time_s": 0.0, "temperature_c": 25.0, "stage_name": "入口"},
        {"time_s": 45.0, "temperature_c": 780.0, "stage_name": "均熱"},
        {"time_s": 110.0, "temperature_c": 420.0, "stage_name": "冷却"},
    ]
    features = _derived_anneal_feature_row(
        "AN-NO-LS",
        {"ls_mpm": None},
        points,
    )
    observation = {
        "task_id": "annealed-properties-v1",
        "parent_key": "AN-NO-LS",
        "features": features,
        "composition": data.composition["ME-00001"],
    }
    candidate = anneal_candidate_from_observation(observation)

    assert features["ls_mpm"] is None
    assert features["feature_eligible"]
    assert candidate is not None
    assert candidate.inputs.process == {}
    assert candidate.inputs.heat_time_basis == "elapsed_time"
    assert np.isfinite(build_feature_bundle(candidate, data.medians).values).all()


def test_importer_accepts_task_and_profile_composition_addition_without_code_change(tmp_path: Path) -> None:
    definitions = load_task_definitions()
    task_id = "annealed-properties-v1"
    task_document = definitions[task_id].model_dump(mode="json")
    composition_group = next(group for group in task_document["input_groups"] if group["key"] == "composition")
    vanadium = dict(composition_group["fields"][-1])
    vanadium.update({"path": "composition.V", "order": len(composition_group["fields"]), "label": "V"})
    composition_group["fields"].append(vanadium)
    task_document["display_decimals"]["composition.V"] = 5
    definitions[task_id] = TaskDefinition.model_validate(task_document)

    raw_profile = json.loads(PROFILE.read_text(encoding="utf-8"))
    mappings = raw_profile["tasks"][task_id]["mappings"]
    insertion_index = next(
        index for index, mapping in enumerate(mappings) if not mapping["path"].startswith("composition.")
    )
    mappings.insert(insertion_index, {
        "path": "composition.V",
        "role": "melt",
        "column": "V[mass%]",
        "kind": "entity_scalar",
        "source_unit": "mass%",
        "canonical_unit": "mass%",
    })
    profile_path = tmp_path / "with-vanadium-profile.json"
    profile_path.write_text(json.dumps(raw_profile, ensure_ascii=False), encoding="utf-8")

    workbook = load_workbook(SOURCE, read_only=False, data_only=True)
    melt = workbook[raw_profile["shared"]["sheets"]["melt"]]
    column = melt.max_column + 1
    melt.cell(1, column).value = "V[mass%]"
    for row in range(2, melt.max_row + 1):
        melt.cell(row, column).value = 0.0123
    source = tmp_path / "with-vanadium.xlsx"
    workbook.save(source)
    workbook.close()

    data = load_workbook_data(source, profile_path, definitions)

    assert data.composition["ME-01"]["V"] == pytest.approx(0.0123)


def test_process_source_maps_prediction_fields_and_tolerates_optional_context() -> None:
    profile_document = json.loads(
        (
            ROOT / "backend" / "src" / "material_workbench" / "data"
            / "dataset-input-profile-process-v1.json"
        ).read_text(encoding="utf-8")
    )
    assert "extends" not in profile_document
    assert detect_dataset_profile_path(PROCESS_SOURCE).name == "dataset-input-profile-process-v1.json"
    data = load_workbook_data(PROCESS_SOURCE)

    assert data.profile_id == "material-workbench-process-v1"
    assert len(data.composition) == 120
    assert len(data.anneal_features) == 196
    assert data.hot_rolling_features["HR-00001"]["equipment"] == ""
    assert data.anneal_features["AN-00001"]["heat_pattern"][0]["time_s"] == 0.0

    annealed = [row for row in data.observations if row["task_id"] == "annealed-properties-v1"]
    tensile = [row for row in annealed if row["source"] == "焼鈍引張実績"]
    holes = [row for row in annealed if row["source"] == "焼鈍穴拡げ実績"]
    assert len(tensile) == 292
    assert len(holes) == 518
    assert all({"TS[MPa]", "YS[MPa]", "EL[%]"} <= set(row["outputs"]) for row in tensile)
    assert all(row["date"] is None for row in holes)
    history_only = [
        row for row in annealed
        if row["parent_key"] in {f"AN-{index:05d}" for index in range(191, 197)}
    ]
    assert len(history_only) == 23
    assert not any(row["eligible"] for row in history_only)
    assert all(
        row["eligibility_reasons"] == ["焼鈍履歴を特徴量化できません"]
        for row in history_only
    )

    hot = [row for row in data.observations if row["task_id"] == "hot-rolled-properties-v1"]
    assert sum("TS[MPa]" in row["outputs"] for row in hot) == 348
    assert {row["test_direction"] for row in hot} == {None}


def test_process_derives_heat_pattern_from_measurement_master_when_history_is_absent(
    tmp_path: Path,
) -> None:
    workbook = load_workbook(PROCESS_SOURCE, read_only=False, data_only=True)
    workbook.remove(workbook["焼鈍履歴"])
    profile = load_dataset_profile(
        ROOT / "backend" / "src" / "material_workbench" / "data" / "dataset-input-profile-process-v1.json"
    )

    canonical = canonicalize_workbook(workbook, profile)
    points = canonical.entities[("annealing", "AN-00001")].values[
        "annealed-properties-v1"
    ]["heat_pattern"]

    assert points[0] == {
        "time_s": 0.0,
        "temperature_c": 34.3,
        "stage_name": "開始",
        "stage_category": "ENTRY",
        "mapping_status": "測定点マスタ補完",
    }
    phf = next(point for point in points if point["stage_name"] == "PHF")
    assert phf["time_s"] == pytest.approx(60 * 15 / 119.742)
    assert phf["temperature_c"] == pytest.approx(164.1)
    assert all(
        right["time_s"] > left["time_s"]
        for left, right in zip(points, points[1:])
    )
    source = tmp_path / "v7-without-heat-history.xlsx"
    workbook.save(source)
    workbook.close()
    data = load_workbook_data(source, profile=profile)
    candidate = candidate_from_lineage(data, "AN-00001")

    assert candidate.inputs.heat_time_basis == "line_speed"
    assert candidate.inputs.heat_pattern is not None
    assert candidate.inputs.heat_pattern[1].mapping_status == "測定点マスタ補完"


def test_process_allows_master_stage_unused_by_current_condition_data() -> None:
    workbook = load_workbook(PROCESS_SOURCE, read_only=False, data_only=True)
    workbook.remove(workbook["焼鈍履歴"])
    annealing = workbook["焼鈍条件-3CGL"]
    phf_column = next(
        cell.column for cell in annealing[1] if cell.value == "PHF[℃]"
    )
    annealing.cell(1, phf_column).value = "未使用工程のため条件列なし"
    profile = load_dataset_profile(
        ROOT / "backend" / "src" / "material_workbench" / "data"
        / "dataset-input-profile-process-v1.json"
    )

    canonical = canonicalize_workbook(workbook, profile)
    points = canonical.heat_series[("annealing", "AN-00001")]

    assert len(points) >= 2
    assert "PHF" not in {point["stage_name"] for point in points}
    assert {point["mapping_status"] for point in points} == {"測定点マスタ補完"}


def test_process_accepts_parent_without_history_or_derivable_measurement_series() -> None:
    workbook = load_workbook(PROCESS_SOURCE, read_only=False, data_only=True)
    workbook.remove(workbook["焼鈍履歴"])
    annealing = workbook["焼鈍条件-3CGL"]
    key_column = next(
        cell.column for cell in annealing[1]
        if cell.value == "焼鈍条件-3CGL_key**"
    )
    entry_column = next(cell.column for cell in annealing[1] if cell.value == "開始[℃]")
    row = next(
        cells[0].row
        for cells in annealing.iter_rows(min_col=key_column, max_col=key_column)
        if cells[0].value == "AN-00001"
    )
    annealing.cell(row, entry_column).value = None
    profile = load_dataset_profile(
        ROOT / "backend" / "src" / "material_workbench" / "data" / "dataset-input-profile-process-v1.json"
    )

    canonical = canonicalize_workbook(workbook, profile)

    assert ("annealing", "AN-00001") in canonical.entities
    assert ("annealing", "AN-00001") not in canonical.heat_series


def test_invalid_workbook_disables_affected_tasks_before_their_runtime_initialization(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook = Workbook()
    source = tmp_path / "invalid.xlsx"
    workbook.save(source)
    database = tmp_path / "must-not-exist.db"
    runtime_called = False

    def forbidden_runtime(*args, **kwargs):
        nonlocal runtime_called
        runtime_called = True
        raise AssertionError("runtime must not initialize before dataset preflight")

    guarded_modules = {
        task_id: replace(module, runtime_factory=forbidden_runtime)
        if module.source_kind == "primary"
        else module
        for task_id, module in registered_task_modules().items()
    }
    monkeypatch.setattr(
        resources_module,
        "registered_task_modules",
        lambda: guarded_modules,
    )
    app = create_app(source, database)
    with TestClient(app) as client:
        health = client.get("/api/health").json()
        assert health["degraded"] is True
        assert health["tasks"]["annealed-properties-v1"]["availability"]["stage"] == "source"
        assert health["tasks"]["hot-rolled-properties-v1"]["availability"]["stage"] == "source"
        assert health["tasks"]["flank-wear-v1"]["availability"]["status"] == "available"

    assert runtime_called is False
    assert database.exists()
