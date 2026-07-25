from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

from material_workbench.data.observation_profile import (
    build_observation_training_dataset,
    inspect_observation_training_view,
    load_observation_profile,
    materialize_observation_training_dataset,
)


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "data" / "source" / "welding_consumable_multistage_synthetic_dataset.xlsx"
PROFILE = (
    ROOT
    / "backend"
    / "src"
    / "material_workbench"
    / "data"
    / "observation-profile-welding-consumable-stage-c-v1.json"
)


def _dataset(source: Path = SOURCE):
    return build_observation_training_dataset(source, load_observation_profile(PROFILE))


def test_stage_c_profile_builds_observation_family_views_without_collapsing_row_context() -> None:
    dataset = _dataset()

    assert {
        family: (view.summary.source_rows, view.summary.usable_input_rows, view.summary.split_groups)
        for family, view in dataset.views.items()
    } == {
        "tensile": (600, 600, 300),
        "charpy": (2700, 2700, 300),
        "corrosion": (103, 103, 103),
    }
    assert [item.target for item in dataset.views["tensile"].summary.targets] == ["TS", "YS", "EL", "RA"]
    assert [item.target for item in dataset.views["charpy"].summary.targets] == [
        "CHARPY_ENERGY",
        "BRITTLE_FRACTURE",
    ]
    assert [item.target for item in dataset.views["corrosion"].summary.targets] == ["CORROSION_RATE"]

    charpy = dataset.views["charpy"]
    assert "process.test_temperature_c" in charpy.feature_names
    assert "process.test_temperature_c" not in dataset.views["tensile"].feature_names
    assert "categorical.test_solution" not in charpy.feature_names
    assert "categorical.test_solution" in dataset.views["corrosion"].feature_names
    first_group = charpy.rows[0].split_group_key
    group_rows = [row for row in charpy.rows if row.split_group_key == first_group]
    assert len(group_rows) == 9
    assert len({row.inputs["process.test_temperature_c"] for row in group_rows}) == 3
    assert all(row.provenance.entity_keys["weld_run"] == first_group for row in group_rows)
    assert all(row.provenance.entity_keys["weld_metal"] for row in group_rows)
    assert all(row.metadata["specimen_position"] for row in group_rows)


def test_inspection_contract_reports_target_usage_exclusions_and_weld_run_groups() -> None:
    dataset = _dataset()

    page = inspect_observation_training_view(
        dataset,
        family="charpy",
        target="CHARPY_ENERGY",
        offset=25,
        limit=10,
    )

    assert page.source_rows == 2700
    assert page.usable_rows == 2700
    assert page.split_groups == 300
    assert page.exclusion_reasons == {}
    assert page.offset == 25
    assert len(page.rows) == 10
    assert all(row.family == "charpy" for row in page.rows)


def test_relation_rows_are_indexes_and_never_become_duplicate_training_rows(tmp_path: Path) -> None:
    changed = tmp_path / SOURCE.name
    shutil.copyfile(SOURCE, changed)
    workbook = load_workbook(changed)
    relation = workbook["relationEx"]
    relation.append([cell.value for cell in relation[2]])
    workbook.save(changed)
    workbook.close()

    dataset = _dataset(changed)
    tensile = dataset.views["tensile"]

    assert len(tensile.rows) == 600
    assert tensile.summary.source_rows == 600
    assert tensile.summary.usable_input_rows == 599
    assert tensile.summary.exclusion_reasons == {"relationが複数行に対応": 1}
    assert tensile.summary.targets[0].exclusion_reasons == {"relationが複数行に対応": 1}


def test_target_status_preserves_common_input_and_target_local_missing_reasons(
    tmp_path: Path,
) -> None:
    changed = tmp_path / SOURCE.name
    shutil.copyfile(SOURCE, changed)
    workbook = load_workbook(changed)
    tensile = workbook["引張試験"]
    tensile.cell(row=2, column=3).value = None
    weld_metal = workbook["溶着金属成分"]
    weld_metal.cell(row=2, column=4).value = None
    workbook.save(changed)
    workbook.close()

    dataset = _dataset(changed)
    view = dataset.views["tensile"]

    assert view.summary.usable_input_rows == 598
    assert view.summary.targets[0].target == "TS"
    assert view.summary.targets[0].usable_rows == 598
    assert view.summary.targets[0].exclusion_reasons == {
        "入力値なし: composition.C": 2,
        "値なし": 1,
    }
    assert view.summary.targets[1].usable_rows == 598
    assert view.summary.targets[1].exclusion_reasons == {
        "入力値なし: composition.C": 2,
    }
    assert view.rows[0].target_status["TS"].reasons == (
        "入力値なし: composition.C",
        "値なし",
    )
    assert view.rows[0].target_status["YS"].reasons == ("入力値なし: composition.C",)
    page = inspect_observation_training_view(dataset, family="tensile", target="TS")
    assert page.exclusion_reasons == {
        "入力値なし: composition.C": 2,
        "値なし": 1,
    }


def test_materialized_views_are_reproducible_and_keep_inspector_summary(tmp_path: Path) -> None:
    dataset = _dataset()
    destination = materialize_observation_training_dataset(dataset, tmp_path / "stage-c")

    summary = json.loads((destination / "summary.json").read_text(encoding="utf-8"))
    assert summary["profile_digest"] == dataset.profile_digest
    assert summary["source_sha256"] == dataset.source_sha256
    assert summary["families"]["charpy"]["source_rows"] == 2700
    assert sum(1 for _ in (destination / "tensile.jsonl").open(encoding="utf-8")) == 600
    assert sum(1 for _ in (destination / "charpy.jsonl").open(encoding="utf-8")) == 2700
    assert sum(1 for _ in (destination / "corrosion.jsonl").open(encoding="utf-8")) == 103
