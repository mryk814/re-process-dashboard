from copy import deepcopy
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
import numpy as np
import pytest

from material_workbench.modeling.feature_pipeline import build_feature_bundle
from material_workbench.contracts.schemas import CandidateInput, ScreeningRunResponse
from material_workbench.domain.services import _candidate_xlsx_names, candidate_from_lineage, import_candidates_xlsx


PROFILE_ROOT = Path(__file__).parents[1] / "src" / "material_workbench" / "data"


def _screening_body(candidate: dict) -> dict:
    return {
        "base_candidate_id": candidate["id"],
        "base_inputs": candidate["inputs"],
        "samples": 48,
        "seed": 20260719,
        "target": "TS",
        "target_goal": {"direction": "at_least", "lower": 500},
        "variables": {
            "composition.C": {"mode": "range", "min": 0.04, "max": 0.12},
            "process.ls_mpm": {"mode": "range", "min": 80, "max": 130},
        },
    }


def test_project_metadata_persists(client) -> None:
    payload = {"name": "DP検討", "description": "焼鈍条件の比較", "purpose": "TS 500 MPa", "task_id": "annealed-properties-v1", "target_values": {"TS": 500.0, "EL": 40.0}, "notes": "初回"}
    updated = client.put("/api/projects/default", json=payload)
    assert updated.status_code == 200
    assert client.get("/api/projects/default").json()["target_values"] == {"TS": 500.0, "EL": 40.0}


def test_latin_hypercube_is_deterministic_bounded_and_convertible(client) -> None:
    candidate = client.get("/api/projects/default/candidates").json()[0]
    first = client.post("/api/screening", json=_screening_body(candidate)).json()
    second = client.post("/api/screening", json=_screening_body(candidate)).json()
    assert len(first["points"]) == 48
    assert [point["inputs"] for point in first["points"]] == [point["inputs"] for point in second["points"]]
    assert all(0.04 <= point["inputs"]["composition.C"] <= 0.12 for point in first["points"])
    assert all(80 <= point["inputs"]["process.ls_mpm"] <= 130 for point in first["points"])
    created = client.post(f"/api/screening/{first['id']}/candidates", json={"point_indices": [0]})
    assert created.status_code == 201
    assert created.json()["candidates"][0]["name"].startswith("Screen")
    assert first["base_candidate_id"] == candidate["id"]
    assert first["base_inputs"] == candidate["inputs"]
    assert first["model_provenance"]["model"]["version"]
    assert first["design_space"]["schema_version"] == "design-space-definition/v1"
    assert first["design_space"]["task_id"] == "annealed-properties-v1"
    assert first["design_space_digest"].startswith("sha256:")
    assert first["proposal_strategy"] == {
        "id": "latin_hypercube_v1",
        "version": "1.0.0",
        "seed": 20260719,
        "requested_count": 48,
        "pool_multiplier": 4,
    }
    diagnostics = first["proposal_diagnostics"]
    assert diagnostics["generated_count"] == 192
    assert diagnostics["valid_count"] + diagnostics["rejected_count"] == 192
    assert diagnostics["evaluated_count"] == 48
    assert sum(diagnostics["rejected_by_reason"].values()) == diagnostics["rejected_count"]
    assert diagnostics["rejection_rate"] == pytest.approx(diagnostics["rejected_count"] / 192)
    assert first["score_contract"] == {
        "version": "screening-score/v3",
        "preference": "lower_is_better",
        "direction": "at_least",
        "target_value": 500.0,
        "lower": 500.0,
        "upper": None,
        "probability_available": True,
        "probability_semantics": "probability_of_achieving_goal",
        "ranking_policy": "support_tier_then_secondary_goals_then_score",
        "fallback": "directional_shortfall",
        "display_label": "目標以上ほど有望",
    }
    assert all(point["prediction"]["goal_direction"] == "at_least" for point in first["points"])
    support_rank = {"supported": 0, "caution": 1, "extrapolated": 2}
    representative_rank = [
        (support_rank[point["support"]["status"]], point["score"])
        for point in first["representative_points"]
    ]
    assert representative_rank == sorted(representative_rank)
    assert client.get(f"/api/screening/{first['id']}").json()["base_canonical_input"] == first["base_canonical_input"]
    assert any(run["id"] == first["id"] for run in client.get("/api/screening").json())

    legacy = deepcopy(first)
    legacy["schema_version"] = "screening-run/v3"
    legacy["target_value"] = 500
    legacy["secondary_targets"] = {}
    legacy.pop("target_goal")
    legacy.pop("secondary_goals")
    legacy["score_contract"] = {
        **legacy["score_contract"],
        "version": "screening-score/v2",
    }
    restored_legacy = ScreeningRunResponse.model_validate(legacy)
    assert restored_legacy.schema_version == "screening-run/v3"
    assert restored_legacy.__dict__["target_value"] == 500


def test_screening_request_rejects_removed_scalar_goal_fields(client) -> None:
    candidate = client.get("/api/projects/default/candidates").json()[0]
    payload = _screening_body(candidate)
    payload.pop("target_goal")
    payload["target_value"] = 500

    response = client.post("/api/screening", json=payload)

    assert response.status_code == 422


def test_screening_seed_is_reproducible_and_can_draw_another_sample(client) -> None:
    candidate = client.get("/api/projects/default/candidates").json()[0]
    first_payload = _screening_body(candidate)
    repeated_payload = deepcopy(first_payload)
    another_payload = deepcopy(first_payload)
    another_payload["seed"] = first_payload["seed"] + 1

    first = client.post("/api/screening", json=first_payload).json()
    repeated = client.post("/api/screening", json=repeated_payload).json()
    another = client.post("/api/screening", json=another_payload).json()

    assert [point["inputs"] for point in first["points"]] == [
        point["inputs"] for point in repeated["points"]
    ]
    assert [point["inputs"] for point in first["points"]] != [
        point["inputs"] for point in another["points"]
    ]
    assert first["proposal_strategy"]["seed"] == first_payload["seed"]
    assert another["proposal_strategy"]["seed"] == another_payload["seed"]


def test_screening_uses_unsaved_base_inputs_without_updating_the_candidate(client) -> None:
    candidate = client.get("/api/projects/default/candidates").json()[0]
    payload = _screening_body(candidate)
    payload["variables"] = {"composition.C": {"mode": "range", "min": 0.04, "max": 0.12}}
    payload["base_inputs"] = deepcopy(candidate["inputs"])
    payload["base_inputs"]["process"]["ls_mpm"] = 117.25
    payload["base_inputs"]["heat_pattern"][0]["stage_name"] = "探索用の工程名"

    response = client.post("/api/screening", json=payload)

    assert response.status_code == 201, response.text
    points = response.json()["points"]
    assert all(point["candidate"]["inputs"]["process"]["ls_mpm"] == 117.25 for point in points)
    assert all(point["candidate"]["inputs"]["heat_pattern"][0]["stage_name"] == "探索用の工程名" for point in points)
    persisted = client.get(f"/api/projects/default/candidates/{candidate['id']}").json()
    assert persisted["inputs"] == candidate["inputs"]


def test_screening_rejects_invalid_field_values_and_empty_candidate_set(client) -> None:
    candidate = client.get("/api/projects/default/candidates").json()[0]
    invalid = _screening_body(candidate)
    invalid["variables"]["process.removed_field"] = {"mode": "fixed", "value": 999}
    assert client.post("/api/screening", json=invalid).status_code == 422
    for item in client.get("/api/projects/default/candidates").json():
        assert client.delete(f"/api/projects/default/candidates/{item['id']}?expected_revision={item['revision']}").status_code == 204
    no_base = _screening_body(candidate)
    no_base["base_candidate_id"] = None
    assert client.post("/api/screening", json=no_base).status_code == 422


def test_screening_without_target_uses_support_distance_contract(client) -> None:
    candidate = client.get("/api/projects/default/candidates").json()[0]
    payload = _screening_body(candidate)
    payload["target_goal"] = None

    response = client.post("/api/screening", json=payload)

    assert response.status_code == 201
    result = response.json()
    assert result["score_contract"]["fallback"] == "support_distance"
    assert all(point["score"] == point["support"]["distance"] for point in result["points"])


def test_screening_between_goal_persists_rule_and_uses_inclusive_boundaries(client) -> None:
    candidate = client.get("/api/projects/default/candidates").json()[0]
    payload = _screening_body(candidate)
    payload["target_goal"] = {"direction": "between", "lower": 450, "upper": 550}
    payload["secondary_goals"] = {
        "YS": {"direction": "at_least", "lower": 300},
    }

    response = client.post("/api/screening", json=payload)

    assert response.status_code == 201, response.text
    run = response.json()
    assert run["schema_version"] == "screening-run/v4"
    assert run["target_goal"] == {"direction": "between", "lower": 450.0, "upper": 550.0}
    assert run["secondary_goals"] == {
        "YS": {"direction": "at_least", "lower": 300.0, "upper": None},
    }
    assert run["target_value"] is None
    assert run["secondary_targets"] == {}
    assert run["score_contract"]["direction"] == "between"
    assert run["score_contract"]["fallback"] == "range_shortfall"
    restored = client.get(f"/api/screening/{run['id']}").json()
    assert restored["target_goal"] == run["target_goal"]
    assert restored["secondary_goals"] == run["secondary_goals"]


def test_screening_opposite_direction_keeps_rule_but_does_not_invert_probability(client) -> None:
    candidate = client.get("/api/projects/default/candidates").json()[0]
    payload = _screening_body(candidate)
    payload["target_goal"] = {"direction": "at_most", "upper": 650}

    response = client.post("/api/screening", json=payload)

    assert response.status_code == 201, response.text
    run = response.json()
    assert run["score_contract"]["direction"] == "at_most"
    assert run["score_contract"]["probability_available"] is False
    assert all(point["prediction"]["goal_direction"] == "at_most" for point in run["points"])
    assert all(point["prediction"]["goal_value"] == 650 for point in run["points"])
    assert all(point["prediction"]["goal_probability"] is None for point in run["points"])
    assert all(point["goal_evaluation"]["method"] == "directional_shortfall" for point in run["points"])


def test_lineage_candidate_actuals_and_snapshot_restore(client) -> None:
    lineage_candidate = client.post("/api/projects/default/lineage/AN-01/candidate")
    assert lineage_candidate.status_code == 201
    candidate = lineage_candidate.json()
    assert len(candidate["inputs"]["heat_pattern"]) >= 2
    quality = client.get("/api/projects/default/quality").json()
    assert candidate["provenance"]["source_ref"]["data_source_digest"] == quality["dataset"]["source_sha256"]
    actual = client.post(f"/api/projects/default/candidates/{candidate['id']}/actuals", params={"expected_revision": candidate["revision"]}, json={"property": "TS", "mean": 505.2, "std": 4.2, "replicates": 3, "unit": "MPa", "experiment_no": "EXP-01", "measured_at": "2026-07-20", "note": "確認用"})
    assert actual.status_code == 201
    assert actual.json()["snapshot_id"]
    changed = {key: value for key, value in candidate.items() if key in {"name", "inputs", "provenance"}}
    changed["expected_revision"] = candidate["revision"]
    changed["inputs"]["process"]["ls_mpm"] = candidate["inputs"]["process"]["ls_mpm"] + 30
    assert client.put(f"/api/projects/default/candidates/{candidate['id']}", json=changed).status_code == 200
    comparison = client.get(f"/api/projects/default/candidates/{candidate['id']}/prediction-vs-actual").json()
    assert comparison["actuals"][0]["mean"] == 505.2
    assert comparison["comparisons"][0]["snapshot_id"] == actual.json()["snapshot_id"]
    assert comparison["comparisons"][0]["prediction"]["canonical_input"]["process"]["ls_mpm"] != changed["inputs"]["process"]["ls_mpm"]
    assert comparison["comparisons"][0]["provenance"]["training_data"]["source_sha256"]
    assert client.post(f"/api/projects/default/candidates/{candidate['id']}/actuals", params={"expected_revision": candidate["revision"] + 1}, json={"property": "TS", "mean": 500, "unit": "%"}).status_code == 422
    snapshot = client.post(f"/api/projects/default/candidates/{candidate['id']}/snapshots").json()
    restored = client.post(f"/api/projects/default/snapshots/{snapshot['id']}/restore")
    assert restored.status_code == 201
    assert restored.json()["id"] != candidate["id"]


def test_lineage_candidate_preserves_stage_order_and_boundaries(client) -> None:
    expected = candidate_from_lineage(client.app.state.data, "AN-03")
    response = client.post("/api/projects/default/lineage/AN-03/candidate")
    assert response.status_code == 201
    payload = {key: value for key, value in response.json().items() if key not in {"id", "project_id", "created_at", "updated_at"}}
    actual = CandidateInput.model_validate(payload)
    assert actual.inputs.heat_pattern == expected.inputs.heat_pattern
    assert actual.inputs.heat_pattern is not None
    assert all(not point.segment_start for point in actual.inputs.heat_pattern)
    assert all(right.time_s > left.time_s for left, right in zip(actual.inputs.heat_pattern, actual.inputs.heat_pattern[1:]))


def test_lineage_candidate_without_line_speed_predicts_and_rejects_ls_curve(client) -> None:
    response = client.post("/api/projects/default/lineage/AN-06/candidate")
    assert response.status_code == 201
    candidate = response.json()
    assert candidate["inputs"]["process"] == {}
    assert candidate["inputs"]["heat_time_basis"] == "elapsed_time"

    preview = client.post(
        f"/api/projects/default/candidates/{candidate['id']}/preview",
        params={"expected_revision": candidate["revision"]},
    )
    assert preview.status_code == 200
    assert set(preview.json()["predictions"]) == {"TS", "YS", "EL", "lambda"}

    curve = client.get(
        f"/api/projects/default/candidates/{candidate['id']}/response-curve",
        params={
            "expected_revision": candidate["revision"],
            "target": "TS",
            "variable": "process.ls_mpm",
            "points": 9,
        },
    )
    assert curve.status_code == 422
    assert "未設定" in curve.json()["message"]


def test_hot_lineage_candidate_uses_hot_rolling_inputs(client) -> None:
    expected = candidate_from_lineage(client.app.state.data, "HR-01")
    assert expected.inputs.heat_pattern is None
    assert set(expected.inputs.process) == {
        "soaking_temperature_c",
        "finish_temperature_c",
        "entry_thickness_mm",
        "exit_thickness_mm",
        "hold_temperature_c",
        "hold_time_min",
    }
    assert expected.provenance.source_ref.entity_type == "hot_rolling"


def test_candidate_excel_import_and_exports(client) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "C", "Si", "Mn", "P", "S", "Al", "Cu", "Ni", "Cr", "Mo", "Ti", "B", "O", "N", "ls_mpm", "heat_time_basis", "time_s_1", "temperature_c_1", "segment_start_1", "stage_name_1", "time_s_2", "temperature_c_2", "segment_start_2", "stage_name_2", "time_s_3", "temperature_c_3", "segment_start_3", "stage_name_3"])
    sheet.append(["Excel候補", 0.08, 0.3, 1.5, 0.01, 0.005, 0.04, 0.0, 0.02, 0.01, 0.0, 0.01, 0.0003, 0.002, 0.004, 100, "経過時間を直接指定", 0, 25, False, "加熱", 300, 810, True, "均熱", 650, 120, False, "冷却"])
    buffer = BytesIO()
    workbook.save(buffer)
    response = client.post("/api/projects/default/candidates/import", files={"file": ("candidates.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert response.status_code == 200
    assert response.json()["created"] == 1
    exported = client.get("/api/projects/default/candidates/export.xlsx")
    assert exported.status_code == 200 and exported.content[:2] == b"PK"
    exported_workbook = load_workbook(BytesIO(exported.content), read_only=True, data_only=True)
    exported_headers = [cell.value for cell in next(exported_workbook["候補"].iter_rows())]
    assert exported_headers[:4] == ["形式バージョン", "候補ID", "候補名", "C[mass%]"]
    assert "ライン速度[m/min]" in exported_headers
    assert "時間基準" in exported_headers
    assert "到達時間[s]_1" in exported_headers
    assert "実績温度[℃]_1" in exported_headers
    assert "TS[MPa]" in exported_headers
    assert "学習範囲判定" in exported_headers
    assert not {"ls_mpm", "time_s_1", "temperature_c_1", "support_status"} & set(exported_headers)
    exported_workbook.close()
    round_tripped, errors = import_candidates_xlsx(exported.content)
    assert not errors
    source = CandidateInput.model_validate({key: value for key, value in response.json()["candidates"][0].items() if key not in {"id", "project_id", "created_at", "updated_at"}})
    restored = next(candidate for candidate in round_tripped if candidate.name == source.name)
    assert restored.model_dump() == source.model_dump()
    defaults = client.app.state.task_registry.runtime_for("annealed-properties-v1").composition_defaults
    assert np.allclose(build_feature_bundle(restored, defaults).values, build_feature_bundle(source, defaults).values)
    quality = client.get("/api/projects/default/quality/export.csv")
    quality_csv = quality.content.decode("utf-8-sig")
    assert quality.status_code == 200 and "issue_id" in quality_csv
    assert "focus_entity_key" in quality_csv and "suggested_view" in quality_csv
    assert quality.headers["content-type"].startswith("text/csv")


def test_candidate_excel_import_rejects_unknown_heat_time_basis() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "name", "C", "Si", "Mn", "P", "S", "Al", "Cu", "Ni", "Cr", "Mo", "Ti", "B", "O", "N",
        "ls_mpm", "heat_time_basis", "time_s_1", "temperature_c_1", "time_s_2", "temperature_c_2",
    ])
    sheet.append([
        "誤記候補", 0.08, 0.3, 1.5, 0.01, 0.005, 0.04, 0.0, 0.02, 0.01, 0.0, 0.01, 0.0003,
        0.002, 0.004, 100, "LSにたぶん連動", 0, 25, 300, 810,
    ])
    buffer = BytesIO()
    workbook.save(buffer)

    imported, errors = import_candidates_xlsx(buffer.getvalue())

    assert imported == []
    assert errors == [{
        "row": 2,
        "message": "時間基準は「ライン速度連動」または「経過時間を直接指定」で入力してください",
    }]


def test_candidate_excel_template_explains_and_round_trips_the_project_contract(client) -> None:
    response = client.get("/api/projects/default/candidates/template.xlsx")

    assert response.status_code == 200
    assert response.content[:2] == b"PK"
    workbook = load_workbook(BytesIO(response.content), read_only=False, data_only=True)
    assert workbook.sheetnames == ["候補", "記入例", "入力ルール"]
    headers = [cell.value for cell in workbook["候補"][1]]
    assert headers[:3] == ["候補名", "C[mass%]", "Si[mass%]"]
    assert "ライン速度[m/min]" in headers
    assert "時間基準" in headers
    assert "経過時間[s]_1" in headers
    assert "温度[℃]_3" in headers
    assert not any(
        header.startswith(("工程境界_", "標準工程名_", "標準工程カテゴリ_"))
        for header in headers
    )
    assert workbook["候補"]["A2"].value is None
    assert workbook["記入例"]["A2"].value == "記入例（候補シートへコピーして変更）"
    guide_text = "\n".join(str(cell.value) for row in workbook["入力ルール"].iter_rows() for cell in row if cell.value is not None)
    assert "1行＝1候補" in guide_text
    assert "最低2点" in guide_text
    assert "工程が変わっても0へ戻さず昇順" in guide_text
    assert "学習データ範囲（参照）" in guide_text
    assert "mpm" not in guide_text
    for column, cell in enumerate(workbook["記入例"][2], start=1):
        workbook["候補"].cell(2, column).value = cell.value
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    untouched, untouched_errors = import_candidates_xlsx(response.content)
    assert not untouched_errors
    assert untouched == []
    imported, errors = import_candidates_xlsx(buffer.getvalue())
    assert not errors
    assert len(imported) == 1
    assert imported[0].name == "記入例（候補シートへコピーして変更）"
    assert imported[0].inputs.heat_time_basis == "line_speed"
    assert all(
        not point.segment_start and point.stage_name is None and point.stage_category is None
        for point in imported[0].inputs.heat_pattern or []
    )


def test_candidate_excel_import_rejects_rows_outside_the_project_contract(client) -> None:
    template = client.get("/api/projects/default/candidates/template.xlsx").content

    def invalid_workbook(header: str, value: object) -> bytes:
        workbook = load_workbook(BytesIO(template), read_only=False, data_only=True)
        candidate_sheet = workbook["候補"]
        for column, cell in enumerate(workbook["記入例"][2], start=1):
            candidate_sheet.cell(2, column).value = cell.value
        headers = {cell.value: cell.column for cell in candidate_sheet[1]}
        candidate_sheet.cell(2, headers[header]).value = value
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    invalid_values = [
        ("C[mass%]", None),
        ("C[mass%]", 1000),
        ("経過時間[s]_2", 0),
    ]
    for header, value in invalid_values:
        response = client.post(
            "/api/projects/default/candidates/import",
            files={"file": ("invalid.xlsx", invalid_workbook(header, value), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["created"] == 0
        assert response.json()["errors"][0]["row"] == 2


def test_candidate_xlsx_names_follow_source_profile() -> None:
    names = _candidate_xlsx_names(
        "annealed-properties-v1",
        str(PROFILE_ROOT / "dataset-input-profile-process-v1.json"),
    )
    assert names["composition.C"] == "C[%]"
    assert names["process.ls_mpm"] == "LS[mpm]"
    assert names["time_s"] == "total時間[秒]"
    assert names["temperature_c"] == "温度[℃]"
    assert names["stage_name"] == "工程"
    assert names["TS"] == "引張強さ[MPa]"


def test_candidate_xlsx_import_rejects_duplicate_headers() -> None:
    workbook = Workbook()
    workbook.active.append(["候補名", "C[mass%]", "C[mass%]"])
    buffer = BytesIO()
    workbook.save(buffer)

    candidates, errors = import_candidates_xlsx(buffer.getvalue())

    assert candidates == []
    assert errors == [{"row": 1, "message": "列名が重複しています: C[mass%]"}]


def test_non_heat_candidate_xlsx_import_does_not_require_heat_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "name", "C", "Si", "Mn", "P", "S", "Al", "Cu", "Ni", "Cr", "Mo", "Ti", "B", "O", "N",
        "soaking_temperature_c", "finish_temperature_c", "entry_thickness_mm", "exit_thickness_mm",
        "hold_temperature_c", "hold_time_min",
    ])
    sheet.append([
        "熱延候補", 0.08, 0.3, 1.5, 0.01, 0.005, 0.04, 0.0, 0.02, 0.01, 0.0, 0.01, 0.0003, 0.002, 0.004,
        1200, 900, 30, 3, 1180, 20,
    ])
    buffer = BytesIO()
    workbook.save(buffer)

    candidates, errors = import_candidates_xlsx(buffer.getvalue(), task_id="hot-rolled-properties-v1")

    assert not errors
    assert candidates[0].name == "熱延候補"
    assert candidates[0].inputs.heat_pattern is None


def test_candidate_delete_preserves_actuals_and_snapshots(client) -> None:
    source = client.get("/api/projects/default/candidates").json()[0]
    payload = {key: value for key, value in source.items() if key not in {"id", "project_id", "created_at", "updated_at"}}
    payload["name"] = "削除検証"
    candidate = client.post("/api/projects/default/candidates", json=payload).json()
    actual = client.post(f"/api/projects/default/candidates/{candidate['id']}/actuals", params={"expected_revision": candidate["revision"]}, json={"property": "TS", "mean": 500, "unit": "MPa"}).json()
    assert client.delete(f"/api/projects/default/candidates/{candidate['id']}?expected_revision={candidate['revision']}").status_code == 204
    assert client.get(f"/api/projects/default/candidates/{candidate['id']}").status_code == 404
    assert client.get(f"/api/projects/default/candidates/{candidate['id']}?include_archived=true").json()["archived_at"] is not None
    assert client.get(f"/api/projects/default/candidates/{candidate['id']}/actuals").status_code == 200
    assert client.app.state.store.get_snapshot(actual["snapshot_id"]) is not None
    assert len(client.app.state.store.list_actuals(candidate["id"])) == 1
