from __future__ import annotations

from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "data" / "source" / "material_workbench_tutorial_v2.xlsx"
PROFILE_SOURCE_NAME = "dataset-input-profile-tutorial"
MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _file_payload(contents: bytes, name: str = "new-source.xlsx") -> dict[str, tuple[str, BytesIO, str]]:
    return {"file": (name, BytesIO(contents), MEDIA_TYPE)}


def _workbook_copy_with_new_digest() -> bytes:
    workbook = load_workbook(SOURCE)
    workbook.properties.title = "Profile Workbench API test"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _renamed_tutorial_workbook(*, ambiguous_ts: bool = False) -> bytes:
    workbook = load_workbook(SOURCE)
    melt = workbook["溶製"]
    melt.title = "溶製_更新"
    headers = [cell.value for cell in melt[1]]
    melt.cell(row=1, column=headers.index("C[mass%]") + 1, value="炭素量[mass%]")
    hot_rolling = workbook["熱延"]
    hot_headers = [cell.value for cell in hot_rolling[1]]
    soaking_column = hot_headers.index("均熱温度[℃]") + 1
    hot_rolling.cell(row=1, column=soaking_column, value="均熱温度[K]")
    for row in range(2, hot_rolling.max_row + 1):
        value = hot_rolling.cell(row=row, column=soaking_column).value
        if isinstance(value, (int, float)):
            hot_rolling.cell(row=row, column=soaking_column, value=value + 273.15)
    if ambiguous_ts:
        tensile = workbook["熱延引張"]
        tensile_headers = [cell.value for cell in tensile[1]]
        ts_column = tensile_headers.index("TS[MPa]") + 1
        tensile.cell(row=1, column=ts_column, value="TS-A")
        tensile.insert_cols(ts_column + 1)
        tensile.cell(row=1, column=ts_column + 1, value="TS-B")
        for row in range(2, tensile.max_row + 1):
            tensile.cell(row=row, column=ts_column + 1, value=tensile.cell(row=row, column=ts_column).value)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _unknown_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "新規測定"
    sheet.append(["試料番号", "TS-A", "TS-B", "温度"])
    sheet.append(["A-1", 500, 505, 800])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_profile_options_only_expose_allowlisted_profiles(client: TestClient) -> None:
    response = client.get("/api/profile-workbench/profiles")

    assert response.status_code == 200
    profiles = response.json()
    assert PROFILE_SOURCE_NAME in {item["source_name"] for item in profiles}
    assert all(item["profile_digest"] and item["task_ids"] for item in profiles)


def test_inspect_auto_detects_profile_without_exposing_server_paths(client: TestClient) -> None:
    response = client.post(
        "/api/profile-workbench/inspect",
        files=_file_payload(SOURCE.read_bytes(), SOURCE.name),
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["source_filename"] == SOURCE.name
    profile = next(item for item in client.get("/api/profile-workbench/profiles").json() if item["source_name"] == PROFILE_SOURCE_NAME)
    assert body["selected_profile_digest"] == profile["profile_digest"]
    assert body["auto_detected"] is True
    assert body["validation"]["registration_ready"] is True
    assert body["validation"]["observations"] > 0
    assert body["sheets"]
    assert "source" not in body
    assert "profile" not in body


def test_inspect_rejects_arbitrary_profile_key(client: TestClient) -> None:
    response = client.post(
        "/api/profile-workbench/inspect",
        data={"profile_digest": "../../private.json"},
        files=_file_payload(SOURCE.read_bytes()),
    )

    assert response.status_code == 422
    assert "許可されたDataset Profile" in response.json()["message"]


def test_inspect_reports_invalid_workbook_as_validation_error(client: TestClient) -> None:
    response = client.post(
        "/api/profile-workbench/inspect",
        files=_file_payload(b"not an xlsx archive"),
    )

    assert response.status_code == 422
    assert "Excelファイルを読み取れません" in response.json()["message"]


def test_inspect_keeps_inventory_when_selected_profile_does_not_match(client: TestClient) -> None:
    profiles = client.get("/api/profile-workbench/profiles").json()
    flank_wear = next(item for item in profiles if item["source_name"] == "dataset-input-profile-flank-wear-v1")

    response = client.post(
        "/api/profile-workbench/inspect",
        data={"profile_digest": flank_wear["profile_digest"]},
        files=_file_payload(SOURCE.read_bytes(), SOURCE.name),
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["sheets"]
    assert body["selected_profile_digest"] == flank_wear["profile_digest"]
    assert body["validation"] is None
    assert body["profile_error"]


def test_inspect_unknown_workbook_returns_unconfirmed_binding_draft(client: TestClient) -> None:
    contents = _unknown_workbook()

    response = client.post(
        "/api/profile-workbench/inspect",
        files=_file_payload(contents, "unknown.xlsx"),
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["selected_profile_digest"] is None
    assert body["auto_detected"] is False
    assert body["binding_draft"] is not None
    assert body["binding_draft"]["source_sha256"] == sha256(contents).hexdigest()
    assert body["binding_draft"]["complete"] is False
    assert any(
        slot["required"] and slot["state"] == "unresolved"
        for slot in body["binding_draft"]["slots"]
    )


def test_register_adds_managed_dataset_and_reuses_duplicate(client: TestClient) -> None:
    contents = _workbook_copy_with_new_digest()
    profile = next(item for item in client.get("/api/profile-workbench/profiles").json() if item["source_name"] == PROFILE_SOURCE_NAME)
    before = client.get("/api/data-library/datasets").json()

    first = client.post(
        "/api/profile-workbench/register",
        data={"profile_digest": profile["profile_digest"], "expected_source_sha256": sha256(contents).hexdigest(), "name": "設備B 再評価"},
        files=_file_payload(contents),
    )

    assert first.status_code == 200, first.text
    first_body = first.json()
    assert first_body["reused_existing"] is False
    assert first_body["profile_id"] == "thin-sheet-tutorial-v1"
    after_first = client.get("/api/data-library/datasets").json()
    assert len(after_first) == len(before) + 1
    registered = next(
        item for item in after_first
        if item["dataset_revision"]["id"] == first_body["dataset_revision_id"]
    )
    assert registered["data_asset"]["locator_kind"] == "managed"
    assert Path(registered["data_asset"]["locator"]).is_file()
    assert Path(registered["data_asset"]["locator"]).is_relative_to(client.app.state.data_library_root)

    second = client.post(
        "/api/profile-workbench/register",
        data={"profile_digest": profile["profile_digest"], "expected_source_sha256": sha256(contents).hexdigest(), "name": "設備B 再評価"},
        files=_file_payload(contents),
    )

    assert second.status_code == 200, second.text
    assert second.json()["reused_existing"] is True
    assert second.json()["dataset_revision_id"] == first_body["dataset_revision_id"]
    assert len(client.get("/api/data-library/datasets").json()) == len(after_first)


def test_stage_b_profile_inspection_and_managed_registration(client: TestClient) -> None:
    source = (
        ROOT
        / "data/source/welding_consumable_multistage_synthetic_dataset.xlsx"
    )
    profile = next(
        item
        for item in client.get("/api/profile-workbench/profiles").json()
        if item["source_name"] == "welding-stage-b-profile-v1"
    )
    contents = source.read_bytes()
    inspection = client.post(
        "/api/profile-workbench/inspect",
        data={"profile_digest": profile["profile_digest"]},
        files=_file_payload(contents, source.name),
    )
    assert inspection.status_code == 200, inspection.text
    assert inspection.json()["validation"]["observations"] == 300
    assert inspection.json()["validation"]["observations_by_task"] == {
        "welding-consumable-stage-b-v1": 300
    }

    registration = client.post(
        "/api/profile-workbench/register",
        data={
            "profile_digest": profile["profile_digest"],
            "expected_source_sha256": sha256(contents).hexdigest(),
        },
        files=_file_payload(contents, source.name),
    )
    assert registration.status_code == 200, registration.text
    assert registration.json()["profile_id"] == "welding-consumable-stage-b-v1"
    dataset = next(
        item
        for item in client.get("/api/data-library/datasets").json()
        if item["dataset_revision"]["id"]
        == registration.json()["dataset_revision_id"]
    )
    assert dataset["data_asset"]["locator_kind"] == "managed"


def test_register_rejects_file_changed_after_inspection(client: TestClient) -> None:
    contents = _workbook_copy_with_new_digest()
    profile = next(item for item in client.get("/api/profile-workbench/profiles").json() if item["source_name"] == PROFILE_SOURCE_NAME)
    before = len(client.get("/api/data-library/datasets").json())

    response = client.post(
        "/api/profile-workbench/register",
        data={"profile_digest": profile["profile_digest"], "expected_source_sha256": "0" * 64},
        files=_file_payload(contents),
    )

    assert response.status_code == 409
    assert "内容が変わりました" in response.json()["message"]
    assert len(client.get("/api/data-library/datasets").json()) == before


def test_register_does_not_silently_reactivate_archived_dataset(client: TestClient) -> None:
    contents = _workbook_copy_with_new_digest()
    profile = next(item for item in client.get("/api/profile-workbench/profiles").json() if item["source_name"] == PROFILE_SOURCE_NAME)
    payload = {
        "profile_digest": profile["profile_digest"],
        "expected_source_sha256": sha256(contents).hexdigest(),
        "name": "archive test",
    }
    first = client.post("/api/profile-workbench/register", data=payload, files=_file_payload(contents))
    assert first.status_code == 200, first.text
    client.app.state.workspace_catalog.archive_dataset_revision(first.json()["dataset_revision_id"])

    second = client.post("/api/profile-workbench/register", data=payload, files=_file_payload(contents))

    assert second.status_code == 409
    assert "archive" in second.json()["message"]


def test_binding_draft_saves_standalone_profile_and_registers_in_same_session(
    client: TestClient,
    monkeypatch,
    tmp_path: Path,
) -> None:
    store = tmp_path / "personal-profiles"
    monkeypatch.setenv("WORKBENCH_PROFILE_STORE_PATH", str(store))
    contents = _renamed_tutorial_workbook()
    base = next(
        item
        for item in client.get("/api/profile-workbench/profiles").json()
        if item["source_name"] == PROFILE_SOURCE_NAME
    )
    inspection = client.post(
        "/api/profile-workbench/inspect",
        data={"profile_digest": base["profile_digest"]},
        files=_file_payload(contents, "updated-source.xlsx"),
    )

    assert inspection.status_code == 200, inspection.text
    draft = inspection.json()["binding_draft"]
    assert draft["source_sha256"] == sha256(contents).hexdigest()
    slots = {item["slot_id"]: item for item in draft["slots"]}
    assert slots["sheet:melt"]["state"] == "suggested"
    assert slots["column:melt:C[mass%]"]["state"] in {"suggested", "unresolved"}
    assert "K" in slots["column:hot_rolling:均熱温度[℃]"]["source_unit_candidates"]
    bindings = [
        {
            "slot_id": "sheet:melt",
            "state": "confirmed",
            "source_name": "溶製_更新",
        },
        {
            "slot_id": "column:melt:C[mass%]",
            "state": "confirmed",
            "source_name": "炭素量[mass%]",
            "source_unit": "mass%",
        },
        {
            "slot_id": "column:hot_rolling:均熱温度[℃]",
            "state": "confirmed",
            "source_name": "均熱温度[K]",
            "source_unit": "K",
        },
    ]
    missing_unit = client.post(
        "/api/profile-workbench/profiles/drafts",
        data={
            "base_profile_digest": base["profile_digest"],
            "expected_source_sha256": draft["source_sha256"],
            "bindings_json": json.dumps([
                {
                    key: value
                    for key, value in binding.items()
                    if not (
                        binding["slot_id"] == "column:hot_rolling:均熱温度[℃]"
                        and key == "source_unit"
                    )
                }
                for binding in bindings
            ], ensure_ascii=False),
        },
        files=_file_payload(contents, "updated-source.xlsx"),
    )
    assert missing_unit.status_code == 422
    assert "source unit must be explicitly confirmed" in missing_unit.json()["message"]

    mismatched_unit_bindings = [
        {
            **binding,
            **(
                {"source_unit": "degC"}
                if binding["slot_id"] == "column:hot_rolling:均熱温度[℃]"
                else {}
            ),
        }
        for binding in bindings
    ]
    mismatched_unit = client.post(
        "/api/profile-workbench/profiles/drafts",
        data={
            "base_profile_digest": base["profile_digest"],
            "expected_source_sha256": draft["source_sha256"],
            "bindings_json": json.dumps(mismatched_unit_bindings, ensure_ascii=False),
        },
        files=_file_payload(contents, "updated-source.xlsx"),
    )
    assert mismatched_unit.status_code == 422
    assert "does not match the workbook header unit" in mismatched_unit.json()["message"]

    saved = client.post(
        "/api/profile-workbench/profiles/drafts",
        data={
            "base_profile_digest": base["profile_digest"],
            "expected_source_sha256": draft["source_sha256"],
            "bindings_json": json.dumps(bindings, ensure_ascii=False),
        },
        files=_file_payload(contents, "updated-source.xlsx"),
    )

    assert saved.status_code == 200, saved.text
    body = saved.json()
    assert body["validation"]["registration_ready"] is True
    locator = Path(body["profile_locator"])
    assert locator.is_file()
    assert locator.is_relative_to(store)
    effective = json.loads(locator.read_text(encoding="utf-8"))
    assert "extends" not in effective
    assert effective["shared"]["sheets"]["melt"] == "溶製_更新"
    assert effective["shared"]["column_aliases"]["melt"]["C[mass%]"] == "炭素量[mass%]"
    assert effective["shared"]["column_aliases"]["hot_rolling"]["均熱温度[℃]"] == "均熱温度[K]"
    hot_mapping = next(
        mapping
        for mapping in effective["tasks"]["hot-rolled-properties-v1"]["mappings"]
        if mapping["path"] == "process.soaking_temperature_c"
    )
    assert hot_mapping["source_unit"] == "K"
    assert next(
        item
        for item in client.get("/api/profile-workbench/profiles").json()
        if item["profile_digest"] == body["profile_digest"]
    )["personal"] is True

    exported = client.get(f"/api/profile-workbench/profiles/{body['profile_digest']}/export")
    assert exported.status_code == 200
    assert exported.json() == effective
    assert saved.json()["profile_digest"] == body["profile_digest"]

    registered = client.post(
        "/api/profile-workbench/register",
        data={
            "profile_digest": body["profile_digest"],
            "expected_source_sha256": draft["source_sha256"],
            "name": "更新版データ",
        },
        files=_file_payload(contents, "updated-source.xlsx"),
    )
    assert registered.status_code == 200, registered.text
    assert registered.json()["profile_id"] == "thin-sheet-tutorial-v1"
    registered_dataset = next(
        item
        for item in client.get("/api/data-library/datasets").json()
        if item["dataset_revision"]["id"] == registered.json()["dataset_revision_id"]
    )
    assert Path(registered_dataset["profile_locator"]) == locator


def test_binding_draft_keeps_ambiguous_column_unresolved(
    client: TestClient,
    monkeypatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setenv("WORKBENCH_PROFILE_STORE_PATH", str(tmp_path / "profiles"))
    contents = _renamed_tutorial_workbook(ambiguous_ts=True)
    base = next(
        item
        for item in client.get("/api/profile-workbench/profiles").json()
        if item["source_name"] == PROFILE_SOURCE_NAME
    )
    inspection = client.post(
        "/api/profile-workbench/inspect",
        data={"profile_digest": base["profile_digest"]},
        files=_file_payload(contents),
    )

    assert inspection.status_code == 200, inspection.text
    slots = inspection.json()["binding_draft"]["slots"]
    ts_slot = next(
        item
        for item in slots
        if item["slot_id"] == "column:hot_tensile:TS[MPa]"
    )
    assert ts_slot["state"] == "unresolved"
    assert {"TS-A", "TS-B"}.issubset({
        item["source_name"] for item in ts_slot["candidates"]
    })

    save = client.post(
        "/api/profile-workbench/profiles/drafts",
        data={
            "base_profile_digest": base["profile_digest"],
            "expected_source_sha256": sha256(contents).hexdigest(),
            "bindings_json": "[]",
        },
        files=_file_payload(contents),
    )
    assert save.status_code == 422
    assert "remain unresolved" in save.json()["message"]


def test_profile_draft_rejects_source_changed_after_inspection(
    client: TestClient,
    monkeypatch,
    tmp_path: Path,
) -> None:
    monkeypatch.setenv("WORKBENCH_PROFILE_STORE_PATH", str(tmp_path / "profiles"))
    base = next(
        item
        for item in client.get("/api/profile-workbench/profiles").json()
        if item["source_name"] == PROFILE_SOURCE_NAME
    )
    contents = SOURCE.read_bytes()
    response = client.post(
        "/api/profile-workbench/profiles/drafts",
        data={
            "base_profile_digest": base["profile_digest"],
            "expected_source_sha256": "0" * 64,
            "bindings_json": "[]",
        },
        files=_file_payload(contents),
    )

    assert response.status_code == 409
    assert "内容が変わりました" in response.json()["message"]


def test_profile_draft_rejects_store_inside_repository(
    client: TestClient,
    monkeypatch,
) -> None:
    base = next(
        item
        for item in client.get("/api/profile-workbench/profiles").json()
        if item["source_name"] == PROFILE_SOURCE_NAME
    )
    monkeypatch.setenv("WORKBENCH_PROFILE_STORE_PATH", str(ROOT / ".unsafe-profiles"))
    contents = SOURCE.read_bytes()

    response = client.post(
        "/api/profile-workbench/profiles/drafts",
        data={
            "base_profile_digest": base["profile_digest"],
            "expected_source_sha256": sha256(contents).hexdigest(),
            "bindings_json": "[]",
        },
        files=_file_payload(contents),
    )

    assert response.status_code == 422
    assert "outside the repository" in response.json()["message"]
