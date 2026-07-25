from __future__ import annotations

from dataclasses import replace
import json
from pathlib import Path
import shutil

from openpyxl import load_workbook
import pytest
from pydantic import ValidationError
from material_workbench.data.stage_b_training import (
    StageBWorkbookProfile,
    build_stage_b_training_data,
    load_stage_b_profile,
)
from material_workbench.data.dataset_registration import register_managed_dataset
from material_workbench.data.profile_workbench import validate_workbook_profile
from material_workbench.modeling.tabular_model_builder import (
    build_tabular_package_from_data,
)
from material_workbench.modeling.model_packages import ModelPackageLoader
from material_workbench.persistence.workspace_catalog import WorkspaceCatalog
from material_workbench.task_modules import _load_welding_stage_b


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "data/source/welding_consumable_multistage_synthetic_dataset.xlsx"
PROFILE = (
    ROOT
    / "backend/src/material_workbench/data/welding-stage-b-profile-v1.json"
)


def test_stage_b_uses_weld_metal_observations_not_relation_rows() -> None:
    profile = load_stage_b_profile(PROFILE)
    result = build_stage_b_training_data(SOURCE, profile)

    assert result.data.row_count == 300
    assert len(result.data.observations) == 300
    assert len({row["id"] for row in result.data.observations}) == 300
    assert len({row["parent_key"] for row in result.data.observations}) == 300
    assert all(row["eligible"] for row in result.data.observations)
    assert all(len(row["composition"]) == 31 for row in result.data.observations)
    assert all(len(row["features"]) == 4 for row in result.data.observations)
    assert all(len(row["categorical"]) == 2 for row in result.data.observations)
    assert all(len(row["outputs"]) == 16 for row in result.data.observations)


def test_stage_b_cohorts_and_folds_are_target_specific_and_group_safe() -> None:
    profile = load_stage_b_profile(PROFILE)
    result = build_stage_b_training_data(SOURCE, profile)

    assert set(result.cohort_digests) == set(profile.weld_output_columns)
    assert set(result.fold_digests) == set(profile.weld_output_columns)
    assert set(result.fold_assignments) == set(profile.weld_output_columns)
    assert result.folds == 5
    assert all(
        sorted(set(assignment.values())) == [0, 1, 2, 3, 4]
        for assignment in result.fold_assignments.values()
    )
    assert result.missing_by_target == {
        target: 0 for target in profile.weld_output_columns
    }
    assert all(value.startswith("sha256:") for value in result.cohort_digests.values())
    assert all(value.startswith("sha256:") for value in result.fold_digests.values())
    assert result.profile_digest == profile.profile_digest
    assert result.transform_digest == profile.transform_digest


def test_stage_b_selected_profile_is_propagated_to_runtime_compilation() -> None:
    profile = load_stage_b_profile(PROFILE)
    fields = list(profile.welding_context)
    original_column = fields[0].column
    replacement_column = "電流[A]"
    fields[0] = fields[0].model_copy(update={"column": replacement_column})
    selected = profile.model_copy(update={
        "id": "welding-consumable-stage-b-selected-profile",
        "welding_context": tuple(fields),
    })

    baseline = _load_welding_stage_b(SOURCE)
    changed = _load_welding_stage_b(SOURCE, selected)  # type: ignore[arg-type]

    assert baseline.observations[0]["features"]["heat_input_kj_per_mm"] != (
        changed.observations[0]["features"]["heat_input_kj_per_mm"]
    )
    assert changed.lifecycle_profile == selected
    assert changed.profile_path == f"catalog:{selected.id}"
    assert original_column != replacement_column


def test_stage_b_resolver_uses_the_pinned_application_profile(
    client,
    monkeypatch,
) -> None:
    project = client.app.state.store.get_project("welding-stage-b-default")
    assert project is not None
    profile = load_stage_b_profile(PROFILE)
    fields = list(profile.welding_context)
    fields[0] = fields[0].model_copy(update={"column": "電流[A]"})
    selected = profile.model_copy(update={
        "id": "welding-consumable-stage-b-application-profile",
        "welding_context": tuple(fields),
    })
    resolver = client.app.state.project_runtime_resolver
    original = resolver._dataset_resources
    calls = 0

    def selected_application_profile(*args, **kwargs):
        nonlocal calls
        path, loaded_profile, source_sha, profile_digest = original(*args, **kwargs)
        calls += 1
        if calls == 1:
            return path, selected, source_sha, selected.profile_digest
        return path, loaded_profile, source_sha, profile_digest

    monkeypatch.setattr(
        resolver,
        "_dataset_resources",
        selected_application_profile,
    )
    resolver._cache.clear()
    resolved = resolver.resolve(project)

    assert resolved.runtime.data.lifecycle_profile == profile
    assert resolved.context_runtime.data.lifecycle_profile == selected
    assert resolved.runtime.data.observations[0]["features"][
        "heat_input_kj_per_mm"
    ] != resolved.context_runtime.data.observations[0]["features"][
        "heat_input_kj_per_mm"
    ]


def test_stage_b_profile_rejects_axis_or_basis_drift() -> None:
    raw = load_stage_b_profile(PROFILE).model_dump(mode="json")
    invalid_profiles = []
    missing_input = json.loads(json.dumps(raw))
    missing_input["raw_component_columns"].pop("other")
    invalid_profiles.append(missing_input)
    reordered_output = json.loads(json.dumps(raw))
    reordered_output["output_axes"] = list(reversed(reordered_output["output_axes"]))
    invalid_profiles.append(reordered_output)
    changed_basis = json.loads(json.dumps(raw))
    changed_basis["input_basis"] = "mass% flux core"
    invalid_profiles.append(changed_basis)

    for invalid in invalid_profiles:
        with pytest.raises(ValidationError):
            StageBWorkbookProfile.model_validate(invalid)


def test_stage_b_missing_targets_keep_inputs_and_create_target_cohorts(
    tmp_path: Path,
) -> None:
    changed = tmp_path / "stage-b-missing-targets.xlsx"
    shutil.copyfile(SOURCE, changed)
    workbook = load_workbook(changed)
    sheet = workbook["溶着金属成分"]
    headers = {
        cell.value: cell.column for cell in sheet[1]
    }
    sheet.cell(2, headers["C[%]"]).value = None
    sheet.cell(3, headers["O[%]"]).value = None
    workbook.save(changed)
    workbook.close()

    profile = load_stage_b_profile(PROFILE)
    result = build_stage_b_training_data(changed, profile)

    assert result.data.row_count == 300
    assert all(row["eligible"] for row in result.data.observations)
    assert result.missing_by_target["C"] == 1
    assert result.missing_by_target["O"] == 1
    assert result.missing_by_target["Mn"] == 0
    assert result.cohort_digests["C"] != result.cohort_digests["Mn"]
    assert result.fold_digests["C"] != result.fold_digests["Mn"]


def test_stage_b_incomplete_upstream_group_is_not_eligible(tmp_path: Path) -> None:
    changed = tmp_path / "stage-b-empty-weld-run.xlsx"
    shutil.copyfile(SOURCE, changed)
    workbook = load_workbook(changed)
    sheet = workbook["relationEx"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    weld_key = sheet.cell(2, headers["溶着金属成分_key**"]).value
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, headers["溶着金属成分_key**"]).value == weld_key:
            sheet.cell(row, headers["溶接施工_key**"]).value = None
    workbook.save(changed)
    workbook.close()

    result = build_stage_b_training_data(changed, load_stage_b_profile(PROFILE))
    observation = next(row for row in result.data.observations if row["id"] == weld_key)

    assert not observation["eligible"]
    assert not observation["parent_key"]
    assert "incomplete upstream relation" in observation["exclusion_reasons"]


def test_stage_b_three_fold_assignment_is_used_by_package(tmp_path: Path) -> None:
    profile = load_stage_b_profile(PROFILE).model_copy(update={"folds": 3})
    training = build_stage_b_training_data(SOURCE, profile)
    destination = tmp_path / "stage-b-three-fold"
    contract = {
        "profile_digest": training.profile_digest,
        "transform_digest": training.transform_digest,
        "cohort_digests": training.cohort_digests,
        "fold_digests": training.fold_digests,
        "fold_assignments": training.fold_assignments,
        "folds": training.folds,
        "missing_by_target": training.missing_by_target,
    }
    build_tabular_package_from_data(
        training.data,
        PROFILE,
        destination,
        training_contract=contract,
    )

    report = json.loads(
        (destination / "reports/quality-report.json").read_text(encoding="utf-8")
    )
    manifest = json.loads(
        (destination / "manifest.json").read_text(encoding="utf-8")
    )
    assert report["folds"] == 3
    assert all(
        predictor["config"]["validation"].startswith("3-fold")
        for predictor in manifest["predictors"]
    )
    assert all(
        sorted(set(assignment.values())) == [0, 1, 2]
        for assignment in training.fold_assignments.values()
    )


def test_stage_b_package_rejects_empty_group(tmp_path: Path) -> None:
    training = build_stage_b_training_data(SOURCE, load_stage_b_profile(PROFILE))
    observations = [dict(row) for row in training.data.observations]
    observations[0] = {**observations[0], "parent_key": ""}
    invalid = replace(training.data, observations=observations)

    with pytest.raises(ValueError, match="non-empty parent group"):
        build_tabular_package_from_data(invalid, PROFILE, tmp_path / "invalid")


def test_stage_b_managed_registration_runs_compiler_preflight(tmp_path: Path) -> None:
    report = validate_workbook_profile(SOURCE, PROFILE)
    assert report["registration_ready"]
    assert report["observations"] == 300
    assert report["observations_by_task"] == {
        "welding-consumable-stage-b-v1": 300
    }

    result = register_managed_dataset(
        database=tmp_path / "catalog.db",
        source=SOURCE,
        library_root=tmp_path / "library",
        profile_path=PROFILE,
    )
    catalog = WorkspaceCatalog(tmp_path / "catalog.db")
    revision = catalog.get_dataset_revision(result.dataset_revision_id)
    assert revision is not None
    assert result.profile_id == "welding-consumable-stage-b-v1"
    assert Path(result.locator).is_file()


def test_stage_b_package_records_profile_cohort_fold_and_smoke_contracts() -> None:
    package = ModelPackageLoader().load(
        ROOT / "models/packages/welding-consumable-stage-b-ridge-v1"
    )
    assert package.manifest.task_id == "welding-consumable-stage-b-v1"
    assert len(package.manifest.predictors) == 16
    stats = package.artifact_path(
        "reference/training_stats.json"
    ).read_text(encoding="utf-8")
    for marker in (
        "profile_digest",
        "transform_digest",
        "cohort_digests",
        "fold_digests",
        "missing_by_target",
    ):
        assert marker in stats


def test_stage_b_task_predicts_and_compares_actual_measurement(client) -> None:
    tasks = client.get("/api/task-definitions").json()
    task = next(
        item for item in tasks
        if item["definition"]["task_definition"]["id"]
        == "welding-consumable-stage-b-v1"
    )
    assert task["definition"]["availability"]["status"] == "available"
    projects = client.get("/api/projects").json()
    project = next(
        item for item in projects if item["id"] == "welding-stage-b-default"
    )
    package = client.get(f"/api/projects/{project['id']}/model-package")
    assert package.status_code == 200, package.text
    assert package.json()["id"] == "welding-consumable-stage-b-ridge-v1"
    candidates = client.get(
        f"/api/projects/{project['id']}/candidates"
    ).json()
    candidate = candidates[0]
    preview = client.post(
        f"/api/projects/{project['id']}/candidates/{candidate['id']}/preview",
        params={"expected_revision": candidate["revision"]},
    )
    assert preview.status_code == 200, preview.text
    prediction = preview.json()["predictions"]["C"]
    assert prediction["unit"] == "mass% deposited metal"
    assert preview.json()["model_meta"]["training_data"]["profile_digest"].startswith(
        "sha256:"
    )

    actual = client.post(
        f"/api/projects/{project['id']}/candidates/{candidate['id']}/actuals",
        params={"expected_revision": candidate["revision"]},
        json={
            "property": "C",
            "mean": prediction["value"],
            "std": 0.001,
            "replicates": 1,
            "unit": "mass% deposited metal",
            "experiment_no": "STAGE-B-TEST",
            "note": "Stage B unit integration",
        },
    )
    assert actual.status_code == 201
    comparison = client.get(
        f"/api/projects/{project['id']}/candidates/{candidate['id']}/prediction-vs-actual"
    )
    assert comparison.status_code == 200
    assert comparison.json()["comparisons"][0]["actual"]["property"] == "C"


def test_stage_b_is_available_in_developer_training_inspector(client) -> None:
    profiles = client.get("/api/developer/observation-training-profiles")
    assert profiles.status_code == 200
    profile = next(
        item for item in profiles.json()
        if item["profile_id"] == "welding-consumable-stage-b-v1"
    )
    assert profile["families"][0]["source_rows"] == 300
    page = client.get(
        "/api/developer/observation-training-data",
        params={
            "profile_id": "welding-consumable-stage-b-v1",
            "family": "stage_b",
            "target": "C",
            "limit": 10,
        },
    )
    assert page.status_code == 200
    assert page.json()["source_rows"] == 300
    assert page.json()["usable_rows"] == 300
