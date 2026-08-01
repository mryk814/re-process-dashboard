from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from decision_workbench.developer_experience.task_scaffolding import (
    ScaffoldField,
    create_task_scaffold,
)


def _workbook(path: Path, *, sheet_name: str = "測定") -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["sample_id", "temperature", "route", "strength"])
    sheet.append(["001", 700, "A", 300])
    sheet.append(["002", 750, "B", 360])
    sheet.append(["003", 800, "A", 420])
    hidden = workbook.create_sheet("計算用")
    hidden.sheet_state = "hidden"
    hidden.append(["internal"])
    workbook.save(path)
    workbook.close()
    return path


def _fields() -> list[ScaffoldField]:
    return [
        ScaffoldField("sample_id", "categorical", "sample_id", "試料", None),
        ScaffoldField(
            "temperature",
            "process",
            "temperature_c",
            "温度",
            "°C",
            allowed_range=(0, 1500),
            default_range=(600, 900),
            training_range=(700, 800),
        ),
        ScaffoldField("route", "categorical", "route", "工程", None),
        ScaffoldField(
            "strength",
            "output",
            "strength_mpa",
            "強度",
            "MPa",
            plausible_range=(0, 2000),
            display_range=(250, 500),
        ),
    ]


def test_xlsx_inspection_requires_explicit_visible_sheet_and_preserves_stored_types(
    client: TestClient,
    tmp_path: Path,
) -> None:
    source = _workbook(tmp_path / "typed.xlsx")

    with source.open("rb") as stream:
        inventory = client.post(
            "/api/data-library/csv-onboarding/inspect",
            files={"file": (source.name, stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert inventory.status_code == 200, inventory.text
    payload = inventory.json()
    assert payload["source_kind"] == "xlsx"
    assert payload["requires_sheet_selection"] is True
    assert payload["columns"] == []
    assert payload["worksheets"] == [
        {"name": "測定", "state": "visible"},
        {"name": "計算用", "state": "hidden"},
    ]

    with source.open("rb") as stream:
        selected = client.post(
            "/api/data-library/csv-onboarding/inspect",
            files={"file": (source.name, stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_name": "測定"},
        )
    assert selected.status_code == 200, selected.text
    selected_payload = selected.json()
    assert selected_payload["selected_sheet"] == "測定"
    assert selected_payload["requires_sheet_selection"] is False
    assert selected_payload["rows"] == 3
    assert selected_payload["reader_policy"] == "xlsx-stored-values-no-formulas/v1"
    sample_id = selected_payload["columns"][0]
    assert sample_id["kind"] == "categorical"
    assert sample_id["choices"] == ["001", "002", "003"]


@pytest.mark.parametrize("case", ["merged", "formula", "hidden"])
def test_xlsx_inspection_fails_closed_for_report_workbooks(
    client: TestClient,
    tmp_path: Path,
    case: str,
) -> None:
    source = _workbook(tmp_path / f"{case}.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "測定"
    sheet.append(["input", "output"])
    sheet.append([1, 2])
    sheet.append([2, 4])
    sheet.append([3, 6])
    if case == "merged":
        sheet.merge_cells("A2:B2")
    elif case == "formula":
        sheet["B2"] = "=A2*2"
    else:
        sheet.sheet_state = "hidden"
        workbook.create_sheet("visible")
    workbook.save(source)
    workbook.close()

    with source.open("rb") as stream:
        response = client.post(
            "/api/data-library/csv-onboarding/inspect",
            files={"file": (source.name, stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"sheet_name": "測定"},
        )
    assert response.status_code == 422
    assert "Profile Workbench" in response.json()["next_action"]
    assert {
        "merged": "merged cell",
        "formula": "formula cell",
        "hidden": "hidden sheet",
    }[case] in response.json()["message"]


def test_csv_and_xlsx_create_the_same_canonical_snapshot_and_profile(
    tmp_path: Path,
) -> None:
    xlsx = _workbook(tmp_path / "source.xlsx")
    csv_source = tmp_path / "source.csv"
    with csv_source.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerows([
            ["sample_id", "temperature", "route", "strength"],
            ["001", 700, "A", 300],
            ["002", 750, "B", 360],
            ["003", 800, "A", 420],
        ])

    csv_result = create_task_scaffold(
        source=csv_source,
        task_id="equivalent-tabular-v1",
        label="等価表",
        fields=_fields(),
        grain_confirmation="one-row-one-observation",
        relation_confirmation="no-relations",
        store=tmp_path / "csv-store",
    )
    xlsx_result = create_task_scaffold(
        source=xlsx,
        sheet="測定",
        task_id="equivalent-tabular-v1",
        label="等価表",
        fields=_fields(),
        grain_confirmation="one-row-one-observation",
        relation_confirmation="no-relations",
        store=tmp_path / "xlsx-store",
    )

    assert csv_result.source_path.read_bytes() == xlsx_result.source_path.read_bytes()
    assert csv_result.profile_path is not None
    assert xlsx_result.profile_path is not None
    assert json.loads(csv_result.profile_path.read_text(encoding="utf-8")) == json.loads(
        xlsx_result.profile_path.read_text(encoding="utf-8")
    )
    xlsx_scaffold = json.loads(
        (xlsx_result.root / "scaffold.json").read_text(encoding="utf-8")
    )
    assert xlsx_scaffold["source"]["selected_sheet"] == "測定"
    assert xlsx_scaffold["source"]["reader_policy"] == "xlsx-stored-values-no-formulas/v1"
