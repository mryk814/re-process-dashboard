from __future__ import annotations

from pathlib import Path
import sqlite3

from openpyxl import load_workbook
import pytest

from decision_workbench.contracts.dataset_disposition_contracts import (
    DATASET_DISPOSITION_SCHEMA_VERSION,
    DatasetDisposition,
    DatasetOperationEligibility,
    DatasetTaskDisposition,
    build_dataset_disposition,
    build_count_disposition,
    compare_dispositions,
    disposition_digest,
)
from decision_workbench.data.profiles.canonicalization import (
    CanonicalDataset,
    CanonicalEntity,
    CanonicalObservation,
)
from decision_workbench.data.profiles.schema import DatasetProfileError
from decision_workbench.data.profiles.loading import load_dataset_profile
from decision_workbench.data.file_integrity import file_sha256
from decision_workbench.data.importer import load_workbook_data
from decision_workbench.data.profile_workbench import validate_source_profile
from decision_workbench.application.material_lineage_candidates import (
    lineage_candidate_options,
)
from decision_workbench.modeling.feature_pipeline import candidate_from_observation
from decision_workbench.application.dataset_registration import register_dataset_records
from decision_workbench.persistence.dataset_disposition_migration import (
    migrate_dataset_disposition_storage,
)
from decision_workbench.persistence.workspace_catalog import WorkspaceCatalog
from decision_workbench.persistence.workspace_catalog_migration import (
    migrate_workspace_catalog,
)


ROOT = Path(__file__).resolve().parents[2]
PROFILE_PATH = ROOT / "backend" / "src" / "decision_workbench" / "data" / "dataset-input-profile-process-v1.json"


def test_disposition_keeps_parent_observation_and_excludes_only_series_operations() -> None:
    profile = load_dataset_profile(PROFILE_PATH)
    task_id = "annealed-properties-v1"
    missing = ("annealing", "parent-missing-series")
    resolved = ("annealing", "parent-with-series")
    canonical = CanonicalDataset(
        profile=profile,
        source_rows={},
        entities={
            missing: CanonicalEntity(missing, {}, {}, {}),
            resolved: CanonicalEntity(resolved, {}, {}, {}),
        },
        relations=(),
        observations=(
            CanonicalObservation(
                task_id=task_id,
                source_role="tensile_test",
                id="obs-missing",
                parent_identity=missing,
                targets={"yield_strength_mpa": 1.0},
                canonical_measurements={"yield_strength_mpa": 1.0},
                metadata={},
                source_locator={"sheet": "ignored", "row": 2},
                policy_results={},
            ),
            CanonicalObservation(
                task_id=task_id,
                source_role="tensile_test",
                id="obs-resolved",
                parent_identity=resolved,
                targets={"yield_strength_mpa": 2.0},
                canonical_measurements={"yield_strength_mpa": 2.0},
                metadata={},
                source_locator={"sheet": "ignored", "row": 3},
                policy_results={},
            ),
        ),
        heat_series={resolved: [{"time_s": 0.0, "temperature_c": 20.0}, {"time_s": 1.0, "temperature_c": 800.0}]},
        heat_series_reasons={missing: ("insufficient_points",)},
    )

    disposition = build_dataset_disposition(
        canonical,
        source_sha256="a" * 64,
        profile_digest="sha256:" + "b" * 64,
        canonicalization_contract_digest="sha256:" + "c" * 64,
    )

    task = disposition.task_dispositions[task_id]
    assert disposition.schema_version == DATASET_DISPOSITION_SCHEMA_VERSION
    assert task.entity_count == 2
    assert task.observation_count == 2
    assert task.usable_observation_count == 1
    assert task.heat_series_parent_count == 1
    assert task.unresolved_heat_series_parent_count == 1
    assert task.operation_eligibility.lineage == "retained"
    assert task.operation_eligibility.observation_browse == "retained"
    assert task.operation_eligibility.training == "excluded_without_required_series"
    assert task.operation_eligibility.candidate_reference == "excluded_without_required_series"
    assert task.operation_eligibility.similarity == "excluded_without_required_series"
    assert task.operation_eligibility.prediction_input == "requires_user_supplied_series"
    assert task.reason_counts == {"insufficient_points": 1}
    assert disposition_digest(disposition).startswith("sha256:")
    serialized = disposition.model_dump(mode="json")
    assert "source_locator" not in str(serialized)
    assert "parent-missing-series" not in str(serialized)


def test_existing_dataset_rows_reload_as_unknown_legacy_without_inference(tmp_path: Path) -> None:
    database = tmp_path / "legacy.db"
    migrate_workspace_catalog(database)
    with sqlite3.connect(database) as connection:
        connection.execute(
            "INSERT INTO data_assets(id,sha256,original_filename,media_type,locator_kind,locator,created_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (
                "asset-legacy",
                "d" * 64,
                "legacy.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "bundled",
                "C:/private/legacy.xlsx",
                "2026-01-01T00:00:00+00:00",
            ),
        )
        connection.execute(
            "INSERT INTO dataset_profile_revisions(id,profile_id,revision,name,profile_digest,canonical_contract_digest,effective_profile_json,created_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (
                "profile-legacy",
                "profile",
                1,
                "profile",
                "sha256:" + "e" * 64,
                "sha256:" + "f" * 64,
                "{}",
                "2026-01-01T00:00:00+00:00",
            ),
        )
        connection.execute(
            "INSERT INTO dataset_revisions(id,data_asset_id,profile_revision_id,canonicalization_contract_digest,dataset_digest,created_at) "
            "VALUES (?,?,?,?,?,?)",
            (
                "dataset-legacy",
                "asset-legacy",
                "profile-legacy",
                "sha256:" + "f" * 64,
                "legacy-dataset-digest",
                "2026-01-01T00:00:00+00:00",
            ),
        )
        connection.commit()

    migrate_dataset_disposition_storage(database)
    dataset = WorkspaceCatalog(database).get_dataset_revision(
        "dataset-legacy", include_archived=True
    )

    assert dataset is not None
    assert dataset.disposition_status == "unknown_legacy"
    assert dataset.disposition_digest is None
    assert dataset.disposition_json is None


def test_registered_disposition_persists_and_reloads(tmp_path: Path) -> None:
    source = ROOT / "data" / "source" / "material_workbench_process_v1.xlsx"
    catalog = WorkspaceCatalog(tmp_path / "workspace.db")
    result = register_dataset_records(
        catalog=catalog,
        source_path=source,
        source_sha256=file_sha256(source),
        profile_path=PROFILE_PATH,
        locator_kind="bundled",
        locator=source,
        name="process disposition",
    )

    stored = WorkspaceCatalog(tmp_path / "workspace.db").get_dataset_revision(
        result.dataset_revision_id
    )
    assert stored is not None
    assert stored.disposition_status == "recorded"
    assert stored.disposition_digest == result.disposition_digest
    assert stored.disposition_json is not None
    assert stored.disposition_json.schema_version == DATASET_DISPOSITION_SCHEMA_VERSION


def test_actual_candidate_and_lineage_operations_share_required_series_boundary() -> None:
    source = ROOT / "data" / "source" / "material_workbench_process_v1.xlsx"
    data = load_workbook_data(source)
    row = next(
        item
        for item in data.observations
        if item["task_id"] == "annealed-properties-v1"
    )
    parent_key = str(row["parent_key"])
    data.anneal_features[parent_key]["heat_pattern"] = []
    row["features"]["heat_pattern"] = []

    assert candidate_from_observation(row) is None
    assert lineage_candidate_options(data, parent_key) == []

    resolved = next(
        item
        for item in data.observations
        if item["task_id"] == "annealed-properties-v1"
        and len(item["features"].get("heat_pattern", [])) >= 2
    )
    assert candidate_from_observation(resolved) is not None


def test_all_required_series_unusable_is_refused_before_registration(tmp_path: Path) -> None:
    source = ROOT / "data" / "source" / "material_workbench_process_v1.xlsx"
    output = tmp_path / "all-unusable.xlsx"
    workbook = load_workbook(source)
    try:
        history = workbook["焼鈍履歴"]
        history_headers = [cell.value for cell in history[1]]
        history_temperature = history_headers.index("温度[℃]") + 1
        for row in history.iter_rows(min_row=2):
            row[history_temperature - 1].value = None
        annealing = workbook["焼鈍条件-3CGL"]
        annealing_headers = [cell.value for cell in annealing[1]]
        line_speed = annealing_headers.index("LS[mpm]") + 1
        for row in annealing.iter_rows(min_row=2):
            row[line_speed - 1].value = None
        hot_tensile = workbook["熱延引張実績"]
        for row in hot_tensile.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        workbook.save(output)
    finally:
        workbook.close()

    with pytest.raises(DatasetProfileError, match="no parent with at least two numeric points"):
        validate_source_profile(output, PROFILE_PATH)


def test_disposition_diff_exposes_identity_and_bounded_task_changes() -> None:
    operation = DatasetOperationEligibility(
        lineage="retained",
        observation_browse="retained",
        training="eligible",
        candidate_reference="eligible",
        similarity="eligible",
        prediction_input="eligible",
    )

    def task(*, entities: int, observations: int) -> DatasetTaskDisposition:
        return DatasetTaskDisposition(
            entity_count=entities,
            observation_count=observations,
            usable_observation_count=observations,
            heat_series_parent_count=0,
            unresolved_heat_series_parent_count=0,
            operation_eligibility=operation,
        )

    previous = DatasetDisposition(
        source_sha256="a" * 64,
        profile_digest="profile-before",
        canonicalization_contract_digest="contract",
        canonical_dataset_digest="dataset-before",
        task_dispositions={
            "common": task(entities=2, observations=2),
            "removed": task(entities=1, observations=1),
        },
    )
    current = DatasetDisposition(
        source_sha256="b" * 64,
        profile_digest="profile-after",
        canonicalization_contract_digest="contract",
        canonical_dataset_digest="dataset-after",
        task_dispositions={
            "common": task(entities=3, observations=4),
            "added": task(entities=1, observations=1),
        },
    )

    diff = compare_dispositions(previous, current)

    assert diff.comparable is False
    assert diff.reason == "profile_changed"
    assert diff.changed is True
    assert diff.source_changed is True
    assert diff.profile_changed is True
    assert diff.previous_source_sha256 == "a" * 64
    assert diff.current_source_sha256 == "b" * 64
    assert diff.previous_profile_digest == "profile-before"
    assert diff.current_profile_digest == "profile-after"
    assert diff.added_task_ids == ["added"]
    assert diff.removed_task_ids == ["removed"]
    assert diff.changed_task_ids == ["added", "common", "removed"]
    assert diff.task_diffs["common"].count_deltas == {
        "entity_count": 1,
        "observation_count": 2,
        "usable_observation_count": 2,
        "heat_series_parent_count": 0,
        "unresolved_heat_series_parent_count": 0,
    }


def test_count_disposition_requires_explicit_usable_count_when_provided() -> None:
    disposition = build_count_disposition(
        source_sha256="a" * 64,
        profile_digest="profile",
        canonicalization_contract_digest="contract",
        task_ids=["task"],
        entities=3,
        observations_by_task={"task": 3},
        usable_observations_by_task={"task": 1},
    )
    assert disposition.task_dispositions["task"].usable_observation_count == 1

    with pytest.raises(ValueError, match="usable_observations_by_task"):
        build_count_disposition(
            source_sha256="a" * 64,
            profile_digest="profile",
            canonicalization_contract_digest="contract",
            task_ids=["task"],
            entities=3,
            observations_by_task={"task": 3},
            usable_observations_by_task={"task": 4},
        )
