from __future__ import annotations

import hashlib
import json
from pathlib import Path
import shutil
import sys

import pytest
from openpyxl import load_workbook

from material_workbench.adapters.builtin_deterministic_linear import (
    DeterministicLinearArtifact,
    ScientificTransformResult,
)
from material_workbench.contracts.blend_contracts import (
    BlendItem,
    CommercialMaterialCatalog,
    RevisionRef,
    SparseBlend,
)
from material_workbench.contracts.stage_a_contracts import ScientificBlendInput
from material_workbench.modeling.model_package_verify import (
    verify_deterministic_transform_package,
)
from material_workbench.modeling.model_packages import ModelPackageLoader, PackageContractError


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend/scripts"))

import build_welding_stage_a_package as package_builder  # noqa: E402


PACKAGE = ROOT / "models/packages/welding-stage-a-deterministic-v1"
CATALOG = ROOT / "models/catalogs/welding-stage-a-commercial-v2.json"
DESIGN_SPACE = ROOT / "models/design-spaces/welding-stage-a-v2.json"
SOURCE = ROOT / "data/source/welding_consumable_multistage_synthetic_dataset.xlsx"
ORACLE_COMPONENTS = (
    "Fe", "C", "Si", "Mn", "Cr", "Ni", "Mo", "Ti", "B", "Al",
    "Mg", "Nb", "V", "Cu", "Zr", "Ca", "N", "O", "S", "P",
    "CaF2", "TiO2", "SiO2", "Al2O3", "MgO", "ZrO2", "K2O", "Na2O",
    "CaCO3", "Fe2O3", "other",
)
ORACLE_SOURCE_COLUMNS = tuple(
    f"{'その他' if component == 'other' else component}[%]"
    for component in ORACLE_COMPONENTS
)


def _sheet_records(workbook: object, sheet_name: str) -> list[dict[str, object]]:
    rows = workbook[sheet_name].iter_rows(values_only=True)
    headers = tuple(str(value) for value in next(rows))
    return [dict(zip(headers, row, strict=True)) for row in rows]


def _full_blend(
    scientific: ScientificBlendInput,
    catalog: CommercialMaterialCatalog,
) -> SparseBlend:
    return SparseBlend(
        items=scientific.items,
        hoop_id=scientific.hoop_id,
        fill_ratio=scientific.fill_ratio,
        balance_material_id=max(scientific.items, key=lambda item: item.ratio).material_id,
        scientific_master=scientific.scientific_master,
        commercial_catalog=catalog.ref,
        design_space=RevisionRef(
            resource_id="test-stage-a-space",
            revision=1,
            digest="sha256:" + "1" * 64,
        ),
    )


def _loaded() -> tuple[object, object, CommercialMaterialCatalog]:
    package = ModelPackageLoader().load(PACKAGE)
    transform = package.load_transform("material-composition")
    catalog = CommercialMaterialCatalog.model_validate_json(
        CATALOG.read_text(encoding="utf-8")
    )
    return package, transform, catalog


def _rewrite_json_artifact(root: Path, relative_path: str, mutate: object) -> None:
    artifact_path = root / relative_path
    payload = json.loads(artifact_path.read_text(encoding="utf-8"))
    mutate(payload)  # type: ignore[operator]
    artifact_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    manifest_path = root / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    artifact_entry = next(
        item for item in manifest["artifacts"] if item["path"] == relative_path
    )
    artifact_entry["sha256"] = hashlib.sha256(artifact_path.read_bytes()).hexdigest()
    artifact_entry["bytes"] = artifact_path.stat().st_size
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _rewrite_artifact(root: Path, mutate: object) -> None:
    _rewrite_json_artifact(root, "transform/stage-a.json", mutate)


def test_stage_a_builder_is_reproducible_and_keeps_source_read_only(
    tmp_path: Path,
) -> None:
    source_before = hashlib.sha256(SOURCE.read_bytes()).hexdigest()
    package = tmp_path / "package"
    catalog = tmp_path / "catalog.json"

    package_builder.build_package(SOURCE, package, catalog)

    assert hashlib.sha256(SOURCE.read_bytes()).hexdigest() == source_before
    for expected in PACKAGE.rglob("*"):
        if not expected.is_file():
            continue
        relative = expected.relative_to(PACKAGE)
        assert (package / relative).read_bytes() == expected.read_bytes()
    assert catalog.read_bytes() == CATALOG.read_bytes()
    assert json.loads(
        catalog.with_name("catalog-design-space.json").read_text(encoding="utf-8")
    ) == json.loads(DESIGN_SPACE.read_text(encoding="utf-8"))


def test_price_only_rebuild_changes_catalog_but_not_scientific_package(
    tmp_path: Path,
) -> None:
    repriced_source = tmp_path / "repriced.xlsx"
    shutil.copyfile(SOURCE, repriced_source)
    workbook = load_workbook(repriced_source)
    sheet = workbook["原料マスタ"]
    headers = [cell.value for cell in sheet[1]]
    price_column = headers.index("単価[円/kg]") + 1
    sheet.cell(row=2, column=price_column).value += 123.0
    workbook.save(repriced_source)
    workbook.close()

    original_package = tmp_path / "original-package"
    original_catalog = tmp_path / "original-catalog.json"
    repriced_package = tmp_path / "repriced-package"
    repriced_catalog = tmp_path / "repriced-catalog.json"
    package_builder.build_package(SOURCE, original_package, original_catalog)
    package_builder.build_package(
        repriced_source,
        repriced_package,
        repriced_catalog,
    )

    original_files = {
        path.relative_to(original_package): path.read_bytes()
        for path in original_package.rglob("*")
        if path.is_file()
    }
    repriced_files = {
        path.relative_to(repriced_package): path.read_bytes()
        for path in repriced_package.rglob("*")
        if path.is_file()
    }
    assert repriced_files == original_files
    assert repriced_catalog.read_bytes() != original_catalog.read_bytes()
    assert (
        repriced_catalog.with_name("repriced-catalog-design-space.json").read_bytes()
        != original_catalog.with_name("original-catalog-design-space.json").read_bytes()
    )
    assert (
        ModelPackageLoader().load(original_package).manifest_sha256
        == ModelPackageLoader().load(repriced_package).manifest_sha256
    )


def test_builder_rejects_source_component_rename_instead_of_inventing_an_axis(
    tmp_path: Path,
) -> None:
    renamed_source = tmp_path / "renamed-axis.xlsx"
    shutil.copyfile(SOURCE, renamed_source)
    workbook = load_workbook(renamed_source)
    sheet = workbook["原料成分"]
    headers = [cell.value for cell in sheet[1]]
    other_column = headers.index("その他[%]") + 1
    sheet.cell(row=1, column=other_column).value = "other[%]"
    workbook.save(renamed_source)
    workbook.close()

    with pytest.raises(ValueError, match="fixed Stage A source-column contract"):
        package_builder.build_package(
            renamed_source,
            tmp_path / "package",
            tmp_path / "catalog.json",
        )


def test_stage_a_package_smoke_and_all_120_blends_match_golden() -> None:
    report = verify_deterministic_transform_package(PACKAGE)
    package, transform, _ = _loaded()
    golden = json.loads(
        package.artifact_path("reference/stage-a-golden-120.json").read_text(
            encoding="utf-8"
        )
    )

    assert report.runtime_type == "builtin.deterministic_linear.v1"
    assert (
        golden["scientific_source_digest"]
        == package.manifest.provenance.training_data_id
    )
    assert len(golden["rows"]) == 120
    for row in golden["rows"]:
        blend = ScientificBlendInput.model_validate(row["blend"])
        expected = ScientificTransformResult.model_validate(row["expected"])
        assert transform.transform(blend) == expected
        assert sum(expected.material_composition.values()) == pytest.approx(
            100.0, abs=2e-5
        )


@pytest.mark.parametrize("field", ["training_data_id", "feature_dataset_id"])
def test_verifier_rejects_scientific_provenance_tampering(
    tmp_path: Path,
    field: str,
) -> None:
    tampered = tmp_path / field
    shutil.copytree(PACKAGE, tampered)
    manifest_path = tampered / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["provenance"][field] = "sha256:" + "f" * 64
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(PackageContractError, match="provenance.*scientific master"):
        verify_deterministic_transform_package(tampered)


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (
            lambda payload: payload.update(
                {"scientific_source_digest": "sha256:" + "f" * 64}
            ),
            "golden scientific digest",
        ),
        (
            lambda payload: payload["rows"][1].update(
                {"blend_id": payload["rows"][0]["blend_id"]}
            ),
            "golden blend ids",
        ),
        (
            lambda payload: payload["rows"][0]["expected"][
                "material_composition"
            ].pop("other"),
            "fixed output axis",
        ),
        (
            lambda payload: payload.update(
                {"schema_version": "stage-a-golden/v0"}
            ),
            "invalid deterministic verification artifact",
        ),
        (
            lambda payload: payload["rows"].pop(),
            "golden row count",
        ),
    ],
)
def test_verifier_rejects_golden_contract_tampering(
    tmp_path: Path,
    mutation: object,
    message: str,
) -> None:
    tampered = tmp_path / message.replace(" ", "-")
    shutil.copytree(PACKAGE, tampered)
    _rewrite_json_artifact(
        tampered,
        "reference/stage-a-golden-120.json",
        mutation,
    )

    with pytest.raises(PackageContractError, match=message):
        verify_deterministic_transform_package(tampered)


def test_all_120_golden_rows_match_the_workbook_generation_formula() -> None:
    package = ModelPackageLoader().load(PACKAGE)
    golden = json.loads(
        package.artifact_path("reference/stage-a-golden-120.json").read_text(
            encoding="utf-8"
        )
    )
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    try:
        material_rows = _sheet_records(workbook, "原料成分")
        hoop_rows = _sheet_records(workbook, "フープマスタ")
        blend_heads_rows = _sheet_records(workbook, "配合")
        blend_item_rows = _sheet_records(workbook, "配合明細")
    finally:
        workbook.close()
    material_compositions = {
        row["原料_key**"]: {
            component: float(row[column] or 0)
            for component, column in zip(
                ORACLE_COMPONENTS,
                ORACLE_SOURCE_COLUMNS,
                strict=True,
            )
        }
        for row in material_rows
    }
    hoop_compositions = {
        row["フープ_key**"]: {
            component: float(row.get(column) or 0)
            for component, column in zip(
                ORACLE_COMPONENTS,
                ORACLE_SOURCE_COLUMNS,
                strict=True,
            )
        }
        for row in hoop_rows
    }
    blend_heads = {
        row["配合_key**"]: row for row in blend_heads_rows
    }
    blend_lines: dict[str, list[dict[str, object]]] = {}
    for row in blend_item_rows:
        blend_lines.setdefault(row["配合_key**"], []).append(row)

    for golden_row in golden["rows"]:
        blend_id = golden_row["blend_id"]
        head = blend_heads[blend_id]
        fill = float(head["充填率[%]"]) / 100.0
        hoop = hoop_compositions[head["フープ_key**"]]
        direct = {
            component: (1.0 - fill) * hoop[component]
            for component in ORACLE_COMPONENTS
        }
        for line in blend_lines[blend_id]:
            material_id = line["原料_key**"]
            whole_wire_fraction = fill * float(line["配合比[%]"]) / 100.0
            for component, value in material_compositions[material_id].items():
                direct[component] += whole_wire_fraction * value
        assert golden_row["expected"]["material_composition"] == pytest.approx(
            direct, abs=1e-10
        )


def test_sparse_order_zero_line_and_cost_breakdown_are_stable() -> None:
    _, transform, catalog = _loaded()
    scientific = ScientificBlendInput.model_validate_json(
        (PACKAGE / "smoke/input.json").read_text(encoding="utf-8")
    )
    baseline = transform.transform(scientific)
    used = {item.material_id for item in scientific.items}
    unused = next(item.material_id for item in catalog.materials if item.material_id not in used)
    reordered = scientific.model_copy(
        update={
            "items": tuple(reversed(scientific.items))
            + (BlendItem(material_id=unused, ratio=0),)
        }
    )

    assert transform.transform(reordered) == baseline
    blend = _full_blend(scientific, catalog)
    evaluated = transform.execute(blend, catalog)
    assert evaluated.powder_blend_cost_yen_per_kg_core == pytest.approx(
        sum(item.contribution_yen_per_kg_core for item in evaluated.material_cost_contributions)
    )
    assert sum(
        item.share_of_blend_cost for item in evaluated.material_cost_contributions
    ) == pytest.approx(1.0)
    assert set(evaluated.auxiliary_features) == {"alloy_powder_d50_um"}


def test_stage_a_rejects_unknown_material_unit_mismatch_and_artifact_tamper(
    tmp_path: Path,
) -> None:
    _, transform, _ = _loaded()
    blend = ScientificBlendInput.model_validate_json(
        (PACKAGE / "smoke/input.json").read_text(encoding="utf-8")
    )
    unknown = blend.model_copy(
        update={
            "items": blend.items
            + (BlendItem(material_id="RM-NOT-IN-PACKAGE", ratio=0),)
        }
    )
    with pytest.raises(PackageContractError, match="unknown Stage A material"):
        transform.transform(unknown)

    wrong_unit = tmp_path / "wrong-unit"
    shutil.copytree(PACKAGE, wrong_unit)
    _rewrite_artifact(
        wrong_unit,
        lambda payload: payload["compiler"].update(
            {"core_ratio_unit": "fraction_core"}
        ),
    )
    with pytest.raises(PackageContractError, match="invalid deterministic-linear artifact"):
        ModelPackageLoader().load(wrong_unit).load_transform("material-composition")

    renamed_axis = tmp_path / "renamed-axis"
    shutil.copytree(PACKAGE, renamed_axis)
    _rewrite_artifact(
        renamed_axis,
        lambda payload: payload["scientific_master"]["components"].__setitem__(
            -1, "その他"
        ),
    )
    with pytest.raises(PackageContractError, match="canonical Stage A axis"):
        ModelPackageLoader().load(renamed_axis).load_transform("material-composition")

    tampered = tmp_path / "tampered"
    shutil.copytree(PACKAGE, tampered)
    with (tampered / "transform/stage-a.json").open("ab") as stream:
        stream.write(b" ")
    with pytest.raises(PackageContractError, match="artifact (size|hash) mismatch"):
        ModelPackageLoader().load(tampered)


def test_old_package_remains_reproducible_after_new_scientific_master_revision(
    tmp_path: Path,
) -> None:
    _, old_transform, _ = _loaded()
    old_blend = ScientificBlendInput.model_validate_json(
        (PACKAGE / "smoke/input.json").read_text(encoding="utf-8")
    )
    old_result = old_transform.transform(old_blend)

    new_package = tmp_path / "revision-2"
    shutil.copytree(PACKAGE, new_package)

    def update_master(payload: dict[str, object]) -> None:
        master = payload["scientific_master"]
        master["revision"] = 2
        master["materials"][0]["d50_um"] += 1

    _rewrite_artifact(new_package, update_master)
    artifact = DeterministicLinearArtifact.model_validate_json(
        (new_package / "transform/stage-a.json").read_text(encoding="utf-8")
    )
    manifest_path = new_package / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["package_version"] = "2.0.0"
    manifest["deterministic_transforms"][0][
        "scientific_master_digest"
    ] = artifact.scientific_master.ref.digest
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    new_transform = ModelPackageLoader().load(new_package).load_transform(
        "material-composition"
    )
    with pytest.raises(PackageContractError, match="scientific master"):
        new_transform.transform(old_blend)
    assert old_transform.transform(old_blend) == old_result
