"""Author the next tutorial workbook revision by adding microstructure evidence.

The shipped tutorial workbook declares 熱延組織 / 焼鈍組織 sheets but has no rows,
so the minimal teaching dataset cannot demonstrate image evidence at all.

This script never edits the existing workbook. It reads a source revision, adds
microstructure rows plus their relation links, and writes a **new** revision file
next to it. Source workbooks stay immutable; Package provenance keeps pointing at
whichever revision it was trained on.

Generated micrographs are synthetic and deterministic: the same revision always
produces byte-identical PNGs, so dataset digests stay reproducible.
"""
from __future__ import annotations

import argparse
import hashlib
import struct
import sys
import zlib
from datetime import date
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

IMAGE_WIDTH = 256
IMAGE_HEIGHT = 192
MAGNIFICATIONS = ("100x", "500x")
ETCHANT = "ナイタール"
SECTION = "L断面"
POSITION = "板厚中央"
# One row deliberately has no file so the UI must show "image not available"
# instead of pretending a micrograph exists.
MISSING_IMAGE_STATE = "未取込"
AVAILABLE_IMAGE_STATE = "利用可"
REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_ROOT = (REPO_ROOT / "data" / "source").resolve()
DEFAULT_GENERATED_ROOT = REPO_ROOT / "artifacts" / "tutorial-v2"


def _inside_source_root(path: Path) -> bool:
    resolved = path.resolve()
    return resolved == SOURCE_ROOT or SOURCE_ROOT in resolved.parents


def _validate_destinations(destination: Path, images_root: Path) -> None:
    protected = [
        str(path)
        for path in (destination, images_root)
        if _inside_source_root(path)
    ]
    if protected:
        raise ValueError(
            "data/source is a read-only source of truth; generate under artifacts/ "
            f"and promote the reviewed revision separately: {', '.join(protected)}"
        )


def _png_grayscale(width: int, height: int, pixels: bytes) -> bytes:
    """Minimal deterministic 8-bit grayscale PNG encoder (no image dependency)."""

    raw = b"".join(
        b"\x00" + pixels[row * width : (row + 1) * width] for row in range(height)
    )

    def chunk(tag: bytes, payload: bytes) -> bytes:
        body = tag + payload
        return (
            struct.pack(">I", len(payload))
            + body
            + struct.pack(">I", zlib.crc32(body) & 0xFFFFFFFF)
        )

    return (
        b"\x89PNG\r\n\x1a\n"
        + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 0, 0, 0, 0))
        + chunk(b"IDAT", zlib.compress(raw, 9))
        + chunk(b"IEND", b"")
    )


def _micrograph(seed_key: str, magnification: str) -> bytes:
    """A grain-like texture. Synthetic demo evidence, not a real observation."""

    seed = int.from_bytes(hashlib.sha256(seed_key.encode("utf-8")).digest()[:8], "big")
    rng = np.random.default_rng(seed)
    grain_count = 40 if magnification == "100x" else 12
    centres = rng.random((grain_count, 2)) * (IMAGE_WIDTH, IMAGE_HEIGHT)
    shades = rng.integers(70, 210, size=grain_count)

    ys, xs = np.mgrid[0:IMAGE_HEIGHT, 0:IMAGE_WIDTH]
    distance = np.sqrt(
        (xs[..., None] - centres[:, 0]) ** 2 + (ys[..., None] - centres[:, 1]) ** 2
    )
    nearest = np.argmin(distance, axis=-1)
    image = shades[nearest].astype(np.float64)

    # Grain boundaries: darken where the two nearest centres are nearly equidistant.
    ordered = np.sort(distance, axis=-1)
    boundary = (ordered[..., 1] - ordered[..., 0]) < 1.6
    image[boundary] = 35.0
    image += rng.normal(0.0, 6.0, image.shape)
    return np.clip(image, 0, 255).astype(np.uint8).tobytes()


def _rows(sheet) -> tuple[list[str], list[list]]:
    values = list(sheet.iter_rows(values_only=True))
    header = [str(item) if item is not None else "" for item in values[0]]
    body = [list(row) for row in values[1:] if any(item is not None for item in row)]
    return header, body


def _entity_keys(book, sheet_name: str) -> list[str]:
    _header, body = _rows(book[sheet_name])
    return [str(row[0]) for row in body if row and row[0] is not None]


def _upstream_context(
    header: list[str], body: list[list], condition_column: str, carried: tuple[str, ...]
) -> dict[str, dict[str, object]]:
    """Reuse the chain the existing test rows already declare for each condition.

    A 組織 row has to sit on the same chain as the 引張 / 穴広げ rows of the same
    condition, otherwise the graph shows it only next to its own condition and
    hides it from the material and process it belongs to.
    """

    positions = {name: index for index, name in enumerate(header)}
    condition_at = positions[condition_column]
    context: dict[str, dict[str, object]] = {}
    for row in body:
        condition = row[condition_at]
        if condition is None:
            continue
        carried_values = {
            name: row[positions[name]] for name in carried if row[positions[name]] is not None
        }
        known = context.get(str(condition))
        # Rows differ in how much of the chain they fill in; keep the fullest one.
        if known is None or len(carried_values) > len(known):
            context[str(condition)] = carried_values
    return context


def build(source: Path, destination: Path, images_root: Path) -> None:
    _validate_destinations(destination, images_root)
    book = load_workbook(source)
    plan = (
        ("熱延", "熱延組織", "熱延組織_key", "HM", "hot", ("溶製_key", "熱延_key")),
        (
            "焼鈍",
            "焼鈍組織",
            "焼鈍組織_key",
            "AM",
            "anneal",
            ("溶製_key", "熱延_key", "冷延_key", "焼鈍_key"),
        ),
    )
    relation = book["relation"]
    relation_header, relation_body = _rows(relation)
    written_images: list[Path] = []
    added_relations = 0

    for entity_sheet, micro_sheet, relation_column, prefix, folder, carried in plan:
        conditions = _entity_keys(book, entity_sheet)
        context = _upstream_context(
            relation_header, relation_body, entity_sheet + "_key", carried
        )
        sheet = book[micro_sheet]
        header, existing = _rows(sheet)
        if existing:
            raise SystemExit(f"{micro_sheet} already has rows; refusing to rewrite")
        index = 0
        for condition in conditions:
            for magnification in MAGNIFICATIONS:
                index += 1
                key = f"{prefix}-{index:02d}"
                relative = f"images/{folder}/{condition}_{magnification}.png"
                # 最後の1件だけ画像を用意しない。取り込み漏れを画面が正直に扱えるか確かめる。
                available = not (
                    condition == conditions[-1] and magnification == MAGNIFICATIONS[-1]
                )
                if available:
                    target = images_root / folder / f"{condition}_{magnification}.png"
                    target.parent.mkdir(parents=True, exist_ok=True)
                    target.write_bytes(
                        _png_grayscale(
                            IMAGE_WIDTH,
                            IMAGE_HEIGHT,
                            _micrograph(f"{condition}:{magnification}", magnification),
                        )
                    )
                    written_images.append(target)
                record = {
                    f"{micro_sheet}_key": key,
                    "画像path": relative,
                    "倍率": magnification,
                    "腐食液": ETCHANT,
                    "観察位置": POSITION,
                    "断面方向": SECTION,
                    "画像状態": AVAILABLE_IMAGE_STATE if available else MISSING_IMAGE_STATE,
                    "撮影日": date(2026, 3, 1 + (index % 27)),
                    "撮影者": "demo",
                    "備考": None if available else "画像ファイル未取込",
                }
                sheet.append([record.get(column) for column in header])
                link = dict(context.get(condition, {}))
                link[entity_sheet + "_key"] = condition
                link[relation_column] = key
                relation.append([link.get(column) for column in relation_header])
                added_relations += 1

    destination.parent.mkdir(parents=True, exist_ok=True)
    book.save(destination)
    book.close()
    print(
        f"wrote {destination} (+{added_relations} relation rows, "
        f"{len(written_images)} images under {images_root})"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source", type=Path, default=Path("data/source/material_workbench_tutorial_v1.xlsx")
    )
    parser.add_argument(
        "--output", type=Path, default=DEFAULT_GENERATED_ROOT / "material_workbench_tutorial_v2.xlsx"
    )
    parser.add_argument("--images-root", type=Path, default=DEFAULT_GENERATED_ROOT / "images")
    arguments = parser.parse_args()
    try:
        build(arguments.source, arguments.output, arguments.images_root)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc
