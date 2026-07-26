"""Build the compact CALCE CS2 cycle-capacity dataset used by the app.

The raw CALCE archives remain outside the repository. This script accepts the
official ``CS2_33.zip`` ... ``CS2_36.zip`` files and deterministically derives
one row per complete discharge cycle.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import statistics
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE_ROOT = (ROOT / "data/source").resolve()
CELLS = {
    "CS2_33": 0.5,
    "CS2_34": 0.5,
    "CS2_35": 1.0,
    "CS2_36": 1.0,
}
MIN_COMPLETE_CAPACITY_AH = 0.5
MAX_PLAUSIBLE_CAPACITY_AH = 1.35


@dataclass(frozen=True)
class CycleObservation:
    started_at: datetime
    source_file: str
    local_cycle_index: int
    capacity_ah: float


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _cycle_observations(payload: bytes, source_file: str) -> list[CycleObservation]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.worksheets[1]
    cumulative_by_cycle: dict[int, list[float]] = defaultdict(list)
    started_by_cycle: dict[int, datetime] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cycle, observed_at, cumulative = row[5], row[2], row[9]
        if cycle is None or observed_at is None or cumulative is None:
            continue
        cycle_index = int(cycle)
        cumulative_by_cycle[cycle_index].append(float(cumulative))
        started_by_cycle.setdefault(cycle_index, observed_at)

    observations: list[CycleObservation] = []
    previous = 0.0
    for cycle_index in sorted(cumulative_by_cycle):
        cumulative = max(cumulative_by_cycle[cycle_index])
        capacity = cumulative - previous
        previous = cumulative
        if MIN_COMPLETE_CAPACITY_AH <= capacity <= MAX_PLAUSIBLE_CAPACITY_AH:
            observations.append(CycleObservation(
                started_at=started_by_cycle[cycle_index],
                source_file=source_file,
                local_cycle_index=cycle_index,
                capacity_ah=capacity,
            ))
    return observations


def _archive_path(raw_root: Path, cell_id: str) -> Path:
    path = raw_root / f"{cell_id}.zip"
    if not path.is_file():
        raise FileNotFoundError(f"CALCE archive is missing: {path}")
    return path


def require_writable_destination(path: Path) -> Path:
    """Reject every script-managed write beneath the immutable source tree."""

    resolved = path.resolve()
    if resolved == SOURCE_ROOT or resolved.is_relative_to(SOURCE_ROOT):
        raise ValueError(
            "data/source is the read-only source of truth; write the derived "
            "artifact outside it and promote it through an explicit review"
        )
    return resolved


def build(raw_root: Path, output: Path, report: Path) -> None:
    output = require_writable_destination(output)
    report = require_writable_destination(report)
    rows: list[dict[str, object]] = []
    source_archives: list[dict[str, object]] = []
    cell_summaries: list[dict[str, object]] = []
    for cell_id, discharge_rate_c in CELLS.items():
        archive_path = _archive_path(raw_root, cell_id)
        archive_bytes = archive_path.read_bytes()
        source_archives.append({
            "name": archive_path.name,
            "sha256": _sha256(archive_bytes),
            "bytes": len(archive_bytes),
        })
        observations: list[CycleObservation] = []
        seen_workbooks: set[str] = set()
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
            for name in sorted(archive.namelist()):
                if not name.lower().endswith(".xlsx"):
                    continue
                payload = archive.read(name)
                digest = _sha256(payload)
                if digest in seen_workbooks:
                    continue
                seen_workbooks.add(digest)
                observations.extend(_cycle_observations(payload, Path(name).name))
        observations.sort(key=lambda item: (item.started_at, item.source_file, item.local_cycle_index))
        reference_capacity = statistics.median(
            item.capacity_ah for item in observations[: min(5, len(observations))]
        )
        for cycle_index, item in enumerate(observations, start=1):
            rows.append({
                "cell_id": cell_id,
                "cycle_index": cycle_index,
                "discharge_rate_c": discharge_rate_c,
                "capacity_ah": round(item.capacity_ah, 8),
                "capacity_percent": round(100 * item.capacity_ah / reference_capacity, 6),
                "observed_at": item.started_at.isoformat(sep=" "),
                "source_file": item.source_file,
                "source_local_cycle": item.local_cycle_index,
            })
        cell_summaries.append({
            "cell_id": cell_id,
            "discharge_rate_c": discharge_rate_c,
            "cycles": len(observations),
            "reference_capacity_ah": round(reference_capacity, 8),
            "minimum_capacity_percent": round(
                min(100 * item.capacity_ah / reference_capacity for item in observations),
                4,
            ),
        })

    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=(
            "cell_id",
            "cycle_index",
            "discharge_rate_c",
            "capacity_ah",
            "capacity_percent",
            "observed_at",
            "source_file",
            "source_local_cycle",
        ), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)

    report.parent.mkdir(parents=True, exist_ok=True)
    report.write_text(json.dumps({
        "schema_version": "calce-cs2-derivation-report/v1",
        "source": {
            "provider": "CALCE Battery Research Group, University of Maryland",
            "landing_page": "https://calce.umd.edu/data",
            "cells": list(CELLS),
            "archives": source_archives,
        },
        "derivation": {
            "training_row": "one complete discharge cycle",
            "capacity": "increment of cumulative Discharge_Capacity(Ah) within each workbook",
            "complete_cycle_range_ah": [
                MIN_COMPLETE_CAPACITY_AH,
                MAX_PLAUSIBLE_CAPACITY_AH,
            ],
            "capacity_percent_reference": "median of the first five retained cycles per cell",
            "duplicate_workbooks": "deduplicated by SHA-256 within each cell archive",
        },
        "rows": len(rows),
        "derived_csv_sha256": hashlib.sha256(output.read_bytes()).hexdigest(),
        "cells": cell_summaries,
    }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw-root", type=Path, required=True)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "artifacts/derived-data/battery_calce_cs2_cycles.csv"
        ),
    )
    parser.add_argument(
        "--report",
        type=Path,
        default=Path(
            "artifacts/derived-data/battery-calce-cs2-derivation.json"
        ),
    )
    args = parser.parse_args()
    build(args.raw_root, args.output, args.report)
    print(f"Wrote {args.output} and {args.report}")


if __name__ == "__main__":
    main()
