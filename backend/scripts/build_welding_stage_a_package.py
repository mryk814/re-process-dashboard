"""Build the inactive Stage A deterministic Model Package from the demo workbook.

The workbook is read-only input. Scientific transform data is snapshotted into
the package; prices are deliberately emitted only as a separate catalog fixture.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from material_workbench.adapters.builtin_deterministic_linear import (
    DeterministicLinearArtifact,
    ScientificHoopRow,
    ScientificMaterialRow,
    ScientificTransformMaster,
    WeightedMeanAuxiliaryFeature,
    WholeWireCompilerContract,
    _BuiltinDeterministicLinearTransform,
)
from material_workbench.contracts.blend_contracts import (
    BlendItem,
    CommercialMaterial,
    CommercialMaterialCatalog,
    SelectionCountConstraint,
    SparseBlendDesignSpace,
)
from material_workbench.contracts.stage_a_contracts import (
    STAGE_A_COMPONENTS,
    STAGE_A_HOOP_SOURCE_COMPONENT_COLUMNS,
    STAGE_A_SOURCE_COMPONENT_COLUMNS,
    ScientificBlendInput,
)
from material_workbench.modeling.model_packages import DeterministicTransformSpec


DEFAULT_SOURCE = Path("data/source/welding_consumable_multistage_synthetic_dataset.xlsx")
DEFAULT_DESTINATION = Path("models/packages/welding-stage-a-deterministic-v1")
DEFAULT_CATALOG_DESTINATION = Path("models/catalogs/welding-stage-a-commercial-v2.json")
DEFAULT_DESIGN_SPACE_DESTINATION = Path("models/design-spaces/welding-stage-a-v1.json")
SHEETS = {
    "materials": "原料マスタ",
    "material_composition": "原料成分",
    "hoops": "フープマスタ",
    "blends": "配合",
    "blend_items": "配合明細",
}
COLUMNS = {
    "material_id": "原料_key**",
    "material_name": "原料名",
    "material_group": "原料グループ",
    "procurement": "調達区分",
    "unit_price": "単価[円/kg]",
    "d50": "粒度D50[um]",
    "hoop_id": "フープ_key**",
    "blend_id": "配合_key**",
    "fill_ratio": "充填率[%]",
    "ratio": "配合比[%]",
}


def _json_bytes(value: Any) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, allow_nan=False, indent=2, sort_keys=True)
        + "\n"
    ).encode("utf-8")


def _records(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    rows = workbook[sheet_name].iter_rows(values_only=True)
    headers = tuple(str(value) for value in next(rows))
    return [dict(zip(headers, row, strict=True)) for row in rows]


def _artifact(root: Path, path: Path) -> dict[str, Any]:
    payload = path.read_bytes()
    return {
        "path": path.relative_to(root).as_posix(),
        "sha256": hashlib.sha256(payload).hexdigest(),
        "bytes": len(payload),
        "media_type": "application/json",
    }


def build_package(
    source: Path,
    destination: Path,
    catalog_destination: Path = DEFAULT_CATALOG_DESTINATION,
    design_space_destination: Path | None = None,
) -> None:
    design_space_destination = design_space_destination or (
        DEFAULT_DESIGN_SPACE_DESTINATION
        if catalog_destination == DEFAULT_CATALOG_DESTINATION
        else catalog_destination.with_name(f"{catalog_destination.stem}-design-space.json")
    )
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        records = {
            key: _records(workbook, sheet_name) for key, sheet_name in SHEETS.items()
        }
    finally:
        workbook.close()

    material_rows = records["material_composition"]
    expected_material_headers = (
        COLUMNS["material_id"],
        "原料名",
        *STAGE_A_SOURCE_COMPONENT_COLUMNS,
    )
    actual_material_headers = tuple(material_rows[0])
    if actual_material_headers != expected_material_headers:
        raise ValueError(
            "原料成分 sheet does not match the fixed Stage A source-column contract: "
            f"expected {expected_material_headers}, got {actual_material_headers}"
        )
    expected_hoop_headers = (
        COLUMNS["hoop_id"],
        "フープ名",
        "材質記号",
        "板厚[mm]",
        "幅[mm]",
        *STAGE_A_HOOP_SOURCE_COMPONENT_COLUMNS,
    )
    actual_hoop_headers = tuple(records["hoops"][0])
    if actual_hoop_headers != expected_hoop_headers:
        raise ValueError(
            "フープマスタ sheet does not match the fixed Stage A source-column contract"
        )
    material_meta = {
        row[COLUMNS["material_id"]]: row for row in records["materials"]
    }
    materials = tuple(
        ScientificMaterialRow(
            material_id=str(row[COLUMNS["material_id"]]),
            group=str(material_meta[row[COLUMNS["material_id"]]][COLUMNS["material_group"]]),
            d50_um=float(material_meta[row[COLUMNS["material_id"]]][COLUMNS["d50"]]),
            composition={
                component: float(row[column] or 0.0)
                for component, column in zip(
                    STAGE_A_COMPONENTS,
                    STAGE_A_SOURCE_COMPONENT_COLUMNS,
                    strict=True,
                )
            },
        )
        for row in material_rows
    )
    # Source hoop rows contain only elemental columns; absent compound axes are zero.
    hoops = tuple(
        ScientificHoopRow(
            hoop_id=str(row[COLUMNS["hoop_id"]]),
            composition={
                component: float(row.get(column) or 0.0)
                for component, column in zip(
                    STAGE_A_COMPONENTS,
                    STAGE_A_SOURCE_COMPONENT_COLUMNS,
                    strict=True,
                )
            },
        )
        for row in records["hoops"]
    )
    master = ScientificTransformMaster(
        schema_version="stage-a-scientific-master/v1",
        resource_id="welding-stage-a-science",
        revision=1,
        components=STAGE_A_COMPONENTS,
        materials=materials,
        hoops=hoops,
    )
    artifact = DeterministicLinearArtifact(
        schema_version="deterministic-linear-artifact/v1",
        compiler=WholeWireCompilerContract(
            id="sparse_blend_whole_wire.v1",
            core_ratio_unit="mass_percent_core",
            fill_ratio_unit="mass_percent_whole_wire",
            absolute_mass_fraction_unit="mass_fraction_whole_wire",
            composition_unit="mass_percent",
        ),
        scientific_master=master,
        auxiliary_features=(
            WeightedMeanAuxiliaryFeature(
                name="alloy_powder_d50_um",
                source="d50_um",
                aggregation="core_ratio_weighted_mean",
                included_groups=("合金鉄", "純金属粉", "脱酸剤"),
                unit="um",
                empty_value=120.0,
            ),
        ),
    )
    catalog = CommercialMaterialCatalog(
        schema_version="commercial-material-catalog/v2",
        resource_id="welding-stage-a-commercial",
        revision=1,
        materials=tuple(
            CommercialMaterial(
                material_id=str(row[COLUMNS["material_id"]]),
                name=str(row[COLUMNS["material_name"]]),
                material_type=str(row[COLUMNS["material_name"]]).rsplit("-", 1)[0],
                group=str(row[COLUMNS["material_group"]]),
                main_components=tuple(
                    component
                    for component, value in sorted(
                        next(
                            item.composition
                            for item in materials
                            if item.material_id == str(row[COLUMNS["material_id"]])
                        ).items(),
                        key=lambda item: (-item[1], item[0]),
                    )[:3]
                    if value > 0
                ),
                procurement=str(row[COLUMNS["procurement"]]),
                unit_price_yen_per_kg_core=float(row[COLUMNS["unit_price"]]),
            )
            for row in records["materials"]
        ),
    )
    design_space = SparseBlendDesignSpace(
        schema_version="sparse-blend-design-space/v1",
        resource_id="welding-stage-a-design-space",
        revision=1,
        scientific_master=master.ref,
        commercial_catalog=catalog.ref,
        allowed_material_ids=tuple(item.material_id for item in materials),
        selection_count=SelectionCountConstraint(minimum=1, maximum=20),
        fixed_hoop_id="HP-01",
        fixed_fill_ratio=19.06,
        balance_material_id="RM-0013",
    )
    transform_spec = DeterministicTransformSpec(
        id="material-composition",
        runtime_type="builtin.deterministic_linear.v1",
        artifact="transform/stage-a.json",
        compiler_id="sparse_blend_whole_wire.v1",
        scientific_master_digest=master.ref.digest,
        output_names=STAGE_A_COMPONENTS,
        output_unit="mass_percent_whole_wire",
        auxiliary_feature_names=("alloy_powder_d50_um",),
    )
    transform = _BuiltinDeterministicLinearTransform(transform_spec, artifact)

    lines_by_blend: dict[str, list[BlendItem]] = {}
    for row in records["blend_items"]:
        lines_by_blend.setdefault(str(row[COLUMNS["blend_id"]]), []).append(
            BlendItem(
                material_id=str(row[COLUMNS["material_id"]]),
                ratio=float(row[COLUMNS["ratio"]]),
            )
        )
    blends: list[ScientificBlendInput] = []
    for row in records["blends"]:
        items = tuple(lines_by_blend[str(row[COLUMNS["blend_id"]])])
        blends.append(
            ScientificBlendInput(
                items=items,
                hoop_id=str(row[COLUMNS["hoop_id"]]),
                fill_ratio=float(row[COLUMNS["fill_ratio"]]),
                scientific_master=master.ref,
            )
        )

    if destination.exists():
        shutil.rmtree(destination)
    (destination / "transform").mkdir(parents=True)
    (destination / "smoke").mkdir()
    (destination / "reference").mkdir()
    artifact_path = destination / transform_spec.artifact
    artifact_path.write_bytes(_json_bytes(artifact.model_dump(mode="json")))

    smoke_input_path = destination / "smoke/input.json"
    smoke_expected_path = destination / "smoke/expected.json"
    smoke_input_path.write_bytes(_json_bytes(blends[0].model_dump(mode="json")))
    smoke_expected_path.write_bytes(
        _json_bytes(transform.transform(blends[0]).model_dump(mode="json"))
    )
    golden_path = destination / "reference/stage-a-golden-120.json"
    golden_path.write_bytes(
        _json_bytes(
            {
                "schema_version": "stage-a-golden/v1",
                "scientific_source_digest": master.ref.digest,
                "rows": [
                    {
                        "blend_id": str(source_row[COLUMNS["blend_id"]]),
                        "blend": blend.model_dump(mode="json"),
                        "expected": transform.transform(blend).model_dump(mode="json"),
                    }
                    for source_row, blend in zip(records["blends"], blends, strict=True)
                ],
            }
        )
    )
    catalog_destination.parent.mkdir(parents=True, exist_ok=True)
    catalog_destination.write_bytes(_json_bytes(catalog.model_dump(mode="json")))
    design_space_destination.parent.mkdir(parents=True, exist_ok=True)
    design_space_destination.write_bytes(
        _json_bytes(design_space.model_dump(mode="json"))
    )

    files = (artifact_path, smoke_input_path, smoke_expected_path, golden_path)
    manifest = {
        "schema_version": "model-package/v1",
        "package_id": "welding-stage-a-deterministic-v1",
        "package_version": "1.0.0",
        "task_id": "welding-stage-a-v1",
        "input_schema_version": "sparse-blend/v1",
        "package_kind": "deterministic_transform",
        "deterministic_transforms": [transform_spec.model_dump(mode="json")],
        "provenance": {
            "training_data_id": master.ref.digest,
            "feature_dataset_id": master.ref.digest,
            "training_code_revision": "build_welding_stage_a_package.py/v1",
        },
        "artifacts": [_artifact(destination, path) for path in files],
        "smoke_test": {
            "input": smoke_input_path.relative_to(destination).as_posix(),
            "expected": smoke_expected_path.relative_to(destination).as_posix(),
        },
        "deterministic_golden": {
            "path": golden_path.relative_to(destination).as_posix(),
            "schema_version": "stage-a-golden/v1",
            "expected_rows": 120,
        },
    }
    (destination / "manifest.json").write_bytes(_json_bytes(manifest))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--destination", type=Path, default=DEFAULT_DESTINATION)
    parser.add_argument(
        "--catalog-destination",
        type=Path,
        default=DEFAULT_CATALOG_DESTINATION,
    )
    parser.add_argument(
        "--design-space-destination",
        type=Path,
        default=DEFAULT_DESIGN_SPACE_DESTINATION,
    )
    args = parser.parse_args()
    build_package(
        args.source,
        args.destination,
        args.catalog_destination,
        args.design_space_destination,
    )


if __name__ == "__main__":
    main()
