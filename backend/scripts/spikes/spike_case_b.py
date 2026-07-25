"""反証ケースB：複数sheet・複数観測family。

溶接語彙を一切使わない3シート構成（工程条件 + 試験A + 試験B）をObservation Profileで
表現し、Training Viewまで通せるかを測定する。
併せて、Observation familyのruntime / builderが別Taskへ再利用できるかを確認する。
"""
from __future__ import annotations

import inspect
import json
import os
import random
import sys
import tempfile
import traceback
from pathlib import Path

REPO = Path(os.environ.get("SPIKE_REPO_ROOT") or Path.cwd()).resolve()
if not (REPO / "pyproject.toml").exists():
    raise SystemExit(f"run from the repository root (got {REPO})")


def _work_dir(name: str) -> Path:
    """Spike成果物はリポジトリ外へ書く。SPIKE_WORK_DIRで上書きできる。"""

    root = Path(os.environ.get("SPIKE_WORK_DIR") or Path(tempfile.gettempdir()) / "material-workbench-spikes")
    return root / name

sys.path.insert(0, str(REPO / "backend" / "src"))

SCRATCH = _work_dir("case-b")
CONDITIONS = 40
ALPHA_PER_CONDITION = 3
BETA_PER_CONDITION = 2


def build_workbook(path: Path) -> None:
    from openpyxl import Workbook

    rng = random.Random(4711)
    book = Workbook()

    conditions = book.active
    conditions.title = "conditions"
    conditions.append(["condition_key", "feed_rate_mm_s", "line_tension_n", "process_route"])

    stock = book.create_sheet("stock")
    stock.append(["stock_key", "thickness_mm", "supplier_lot"])

    relation = book.create_sheet("relation")
    relation.append(["condition_key", "stock_key", "alpha_key", "beta_key"])

    alpha = book.create_sheet("test_alpha")
    alpha.append(["alpha_key", "test_load_kgf", "strength_mpa", "operator"])

    beta = book.create_sheet("test_beta")
    beta.append(["beta_key", "indent_load_kgf", "hardness_hv", "operator"])

    for index in range(CONDITIONS):
        condition_key = f"cond-{index:03d}"
        stock_key = f"stock-{index % 12:03d}"
        feed = rng.uniform(2.0, 22.0)
        tension = rng.uniform(120.0, 900.0)
        route = ("route_a", "route_b", "route_c")[index % 3]
        conditions.append([condition_key, round(feed, 4), round(tension, 3), route])
        if index < 12:
            stock.append([stock_key, round(rng.uniform(0.8, 6.0), 3), f"lot-{index:03d}"])

        for repeat in range(ALPHA_PER_CONDITION):
            alpha_key = f"{condition_key}-a{repeat}"
            load = rng.uniform(5.0, 60.0)
            strength = 320.0 + 6.2 * feed + 0.21 * tension + rng.gauss(0.0, 9.0)
            alpha.append([alpha_key, round(load, 3), round(strength, 2), "op-1"])
            relation.append([condition_key, stock_key, alpha_key, None])

        for repeat in range(BETA_PER_CONDITION):
            beta_key = f"{condition_key}-b{repeat}"
            load = rng.uniform(1.0, 30.0)
            # 一部の測定を欠損させ、target別cohortを作る
            hardness = (
                None
                if rng.random() < 0.2
                else round(118.0 + 2.4 * feed + 0.04 * tension + rng.gauss(0.0, 4.0), 2)
            )
            beta.append([beta_key, round(load, 3), hardness, "op-2"])
            relation.append([condition_key, stock_key, None, beta_key])

    book.save(path)


def profile_document() -> dict:
    def condition_inputs() -> list[dict]:
        return [
            {
                "path": "process.feed_rate_mm_s",
                "role": "condition",
                "column": "feed_rate_mm_s",
                "kind": "numeric",
                "source_unit": "mm/s",
                "canonical_unit": "mm/s",
            },
            {
                "path": "process.line_tension_n",
                "role": "condition",
                "column": "line_tension_n",
                "kind": "numeric",
                "source_unit": "N",
                "canonical_unit": "N",
            },
            {
                "path": "categorical.process_route",
                "role": "condition",
                "column": "process_route",
                "kind": "categorical",
            },
            {
                "path": "process.thickness_mm",
                "role": "stock",
                "column": "thickness_mm",
                "kind": "numeric",
                "source_unit": "mm",
                "canonical_unit": "mm",
            },
        ]

    return {
        "schema_version": "observation-dataset-profile/v1",
        "id": "spike-two-family-observations-v1",
        "task_id": "spike-two-family-v1",
        "relation_sheet": "relation",
        "entities": [
            {
                "role": "condition",
                "sheet": "conditions",
                "key_column": "condition_key",
                "relation_column": "condition_key",
            },
            {
                "role": "stock",
                "sheet": "stock",
                "key_column": "stock_key",
                "relation_column": "stock_key",
            },
        ],
        "families": [
            {
                "id": "alpha",
                "sheet": "test_alpha",
                "relation_column": "alpha_key",
                "observation_id_column": "alpha_key",
                "split_group_role": "condition",
                "inputs": [
                    *condition_inputs(),
                    {
                        # 試験行固有入力（試験Aのシートにしかない列）
                        "path": "process.test_load_n",
                        "role": "alpha",
                        "column": "test_load_kgf",
                        "kind": "numeric",
                        "source_unit": "kgf",
                        "canonical_unit": "N",
                    },
                ],
                "outputs": [
                    {
                        "key": "strength_mpa",
                        "column": "strength_mpa",
                        "source_unit": "MPa",
                        "canonical_unit": "MPa",
                    }
                ],
                "metadata": [{"key": "operator", "column": "operator"}],
            },
            {
                "id": "beta",
                "sheet": "test_beta",
                "relation_column": "beta_key",
                "observation_id_column": "beta_key",
                "split_group_role": "condition",
                "inputs": [
                    *condition_inputs(),
                    {
                        "path": "process.indent_load_n",
                        "role": "beta",
                        "column": "indent_load_kgf",
                        "kind": "numeric",
                        "source_unit": "kgf",
                        "canonical_unit": "N",
                    },
                ],
                "outputs": [
                    {
                        "key": "hardness_hv",
                        "column": "hardness_hv",
                        "source_unit": "HV",
                        "canonical_unit": "HV",
                    }
                ],
                "metadata": [{"key": "operator", "column": "operator"}],
            },
        ],
    }


def main() -> int:
    from material_workbench.data.observation_profile import (
        ObservationDatasetProfile,
        build_observation_training_dataset,
    )

    SCRATCH.mkdir(parents=True, exist_ok=True)
    findings: list[str] = []
    source = SCRATCH / "two_family_source.xlsx"
    profile_path = SCRATCH / "observation-profile-spike-two-family-v1.json"
    document = profile_document()
    profile_path.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
    build_workbook(source)

    try:
        profile = ObservationDatasetProfile.model_validate(document)
        _check(findings, "Observation Profileを溶接語彙なしで宣言できる", True)
    except Exception as exc:
        _check(findings, "Observation Profileを溶接語彙なしで宣言できる", False, exc)
        return _report(findings)

    try:
        dataset = build_observation_training_dataset(source, profile)
        _check(findings, "Training Viewの構築（build_observation_training_dataset）", True)
    except Exception as exc:
        _check(findings, "Training Viewの構築（build_observation_training_dataset）", False, exc)
        return _report(findings)

    for family_id in ("alpha", "beta"):
        view = dataset.views[family_id]
        usable = [row for row in view.rows if row.eligible]
        print(
            f"[view] {family_id}: rows={len(view.rows)} eligible={len(usable)} "
            f"split_groups={view.summary.split_groups} features={len(view.feature_names)}"
        )
    alpha_targets = {target.target for target in dataset.views["alpha"].summary.targets}
    beta_targets = {target.target for target in dataset.views["beta"].summary.targets}
    _check(
        findings,
        "target別cohortがfamilyごとに分離される",
        alpha_targets == {"strength_mpa"} and beta_targets == {"hardness_hv"},
        f"alpha={sorted(alpha_targets)} beta={sorted(beta_targets)}",
    )
    beta_summary = next(
        item for item in dataset.views["beta"].summary.targets if item.target == "hardness_hv"
    )
    beta_view = dataset.views["beta"]
    missing_rows = sum(
        1 for row in beta_view.rows if not row.target_status["hardness_hv"].usable
    )
    print(
        f"[cohort] beta.hardness_hv usable={beta_summary.usable_rows} "
        f"入力行={beta_view.summary.usable_input_rows} target欠損={missing_rows} "
        f"exclusion={dict(beta_summary.exclusion_reasons)}"
    )
    _check(
        findings,
        "欠損targetが入力行と別に不適格として数えられる",
        missing_rows > 0 and beta_summary.usable_rows == len(beta_view.rows) - missing_rows,
        f"target欠損={missing_rows} usable={beta_summary.usable_rows} 全行={len(beta_view.rows)}",
    )

    # 試験行固有入力が canonical input として乗るか
    alpha_features = set(dataset.views["alpha"].feature_names)
    _check(
        findings,
        "試験行固有入力を canonical input として宣言できる",
        "process.test_load_n" in alpha_features,
        sorted(alpha_features),
    )

    # 宣言した単位変換が Training View で適用されるか（kgf -> N）
    sample = next(row for row in dataset.views["alpha"].rows if row.eligible)
    converted = sample.inputs["process.test_load_n"]
    _check(
        findings,
        "宣言した単位変換（kgf→N）がTraining Viewで適用される",
        converted is not None and converted > 100.0,
        f"canonical_unit=N と宣言した値が {converted}（生値のまま。9.80665倍されていない）",
    )

    # family名が分岐条件になっていないか（契約側のみで判定）
    _check(
        findings,
        "family名が契約の分岐条件になっていない",
        _families_are_data_driven(),
        "build_observation_training_dataset がfamily idで分岐している",
    )

    _probe_runtime_reuse(findings)
    return _report(findings)


def _families_are_data_driven() -> bool:
    from material_workbench.data import observation_profile

    source = inspect.getsource(observation_profile.build_observation_training_dataset)
    return not any(
        marker in source for marker in ('== "tensile"', '== "charpy"', '== "corrosion"')
    )


def _probe_runtime_reuse(findings: list[str]) -> None:
    """Observation familyのruntime / builderが2つ目のTaskへ再利用できるかを確認する。"""

    from material_workbench.modeling import stage_c_model_builder, stage_c_regression

    builder_parameters = set(inspect.signature(stage_c_model_builder.build).parameters)
    _check(
        findings,
        "Observation family builderがprofileでパラメタ化されている",
        "profile_path" in builder_parameters or "profile" in builder_parameters,
        f"stage_c_model_builder.build の引数={sorted(builder_parameters)}",
    )
    _check(
        findings,
        "Observation family runtimeがtask_idでパラメタ化されている",
        not hasattr(stage_c_regression, "TASK_ID"),
        f"stage_c_regression.TASK_ID={getattr(stage_c_regression, 'TASK_ID', None)!r} "
        f"PROFILE_PATH={Path(getattr(stage_c_regression, 'PROFILE_PATH', '')).name!r} "
        "がmodule定数として固定されている",
    )
    loader_parameters = set(inspect.signature(stage_c_regression.load_stage_c_data).parameters)
    _check(
        findings,
        "Observation family loaderがprofileを受け取れる",
        "profile" in loader_parameters,
        f"load_stage_c_data の引数={sorted(loader_parameters)}",
    )


def _check(findings: list[str], label: str, ok: bool, detail: object = "") -> None:
    print(f"[{'OK  ' if ok else 'FAIL'}] {label}" + (f"\n        -> {detail}" if not ok and detail else ""))
    if not ok:
        findings.append(f"{label} :: {detail}")


def _report(findings: list[str]) -> int:
    print("\n=== 発見事項 ===")
    if not findings:
        print("なし（Observation family全体をそのまま再利用できた）")
    for item in findings:
        print(f"- {item}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
