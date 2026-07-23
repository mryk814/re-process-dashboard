from __future__ import annotations

from collections import defaultdict
from copy import deepcopy
from io import BytesIO
from typing import Any, Callable

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .dataset_profile import DatasetInputProfile, load_dataset_profile
from .hot_rolling_feature_pipeline import PROCESS_NAMES
from .importer import WorkbookData, composition_names
from .task_modules import PredictionRuntime
from .task_registry import load_task_contracts
from .schemas import Candidate, CandidateInput, HeatPoint, ScreeningRequest
from .screening_score import GoalDirection, evaluate_screening_goal, score_contract


COMPOSITION_COLUMNS = composition_names(task_id="annealed-properties-v1")
SCREENING_SEED = 20260719


def _set_screen_value(candidate: Candidate, name: str, value: float | str) -> Candidate:
    updated = candidate.model_copy(deep=True)
    heat_parts = name.split(".")
    if len(heat_parts) == 3 and heat_parts[0] == "heat_pattern" and heat_parts[1].isdigit() and heat_parts[2] in {"time_s", "temperature_c"}:
        points = updated.inputs.heat_pattern
        index = int(heat_parts[1])
        if points is None or index >= len(points):
            raise ValueError(f"ヒートパターンに存在しない点です: {name}")
        setattr(points[index], heat_parts[2], float(value))
        return updated
    group, separator, field = name.partition(".")
    if not separator or group not in {"composition", "process", "categorical"}:
        raise ValueError(f"スクリーニング対象外の変数です: {name}")
    values = getattr(updated.inputs, group)
    values[field] = str(value) if group == "categorical" else float(value)
    return updated


def run_latin_hypercube(
    runtime: PredictionRuntime,
    base: Candidate,
    request: ScreeningRequest,
    *,
    goal_directions: dict[str, GoalDirection | None],
    probability_available: dict[str, bool],
    candidate_validator: Callable[[CandidateInput], None],
) -> dict[str, Any]:
    rng = np.random.default_rng(SCREENING_SEED)
    pool_size = request.samples * 4
    sample_values: dict[str, list[float | str]] = {}
    for name in sorted(request.variables):
        spec = request.variables[name]
        if spec.mode == "fixed":
            sample_values[name] = [spec.value] * pool_size  # type: ignore[list-item]
        elif spec.mode == "range":
            permutation = rng.permutation(pool_size)
            sample_values[name] = [float(spec.min + (permutation[index] + rng.random()) / pool_size * (spec.max - spec.min)) for index in range(pool_size)]  # type: ignore[operator]
        else:
            values = spec.values or []
            permutation = rng.permutation(pool_size)
            sample_values[name] = [values[int(np.floor((permutation[index] + rng.random()) / pool_size * len(values))) % len(values)] for index in range(pool_size)]
    points: list[dict[str, Any]] = []
    base_prediction = runtime.predict(base, detailed=False)
    for sample_index in range(pool_size):
        if len(points) >= request.samples:
            break
        candidate = base.model_copy(deep=True)
        applied = {name: sample_values[name][sample_index] for name in sorted(sample_values)}
        for name, value in applied.items():
            candidate = _set_screen_value(candidate, name, value)
        try:
            candidate_input = CandidateInput.model_validate(candidate.model_dump())
            candidate_validator(candidate_input)
        except ValueError:
            continue
        candidate = Candidate.model_validate({**candidate.model_dump(), **candidate_input.model_dump()})
        target_values = {
            key: value
            for key, value in {request.target: request.target_value, **request.secondary_targets}.items()
            if value is not None and probability_available.get(key, False)
        }
        prediction = runtime.predict(
            candidate,
            detailed=False,
            target_values=target_values,
        )
        selected = prediction["predictions"][request.target]
        support = prediction["support"]
        evaluation = evaluate_screening_goal(
            selected.value,
            target_value=request.target_value,
            direction=goal_directions.get(request.target),
            at_least_probability=selected.goal_probability if probability_available.get(request.target, False) else None,
            support_distance=support.distance,
        )
        secondary_evaluations = {
            key: evaluate_screening_goal(
                prediction["predictions"][key].value,
                target_value=value,
                direction=goal_directions.get(key),
                at_least_probability=prediction["predictions"][key].goal_probability if probability_available.get(key, False) else None,
                support_distance=support.distance,
            )
            for key, value in request.secondary_targets.items()
        }
        prediction_payload = selected.model_dump()
        prediction_payload["goal_direction"] = goal_directions.get(request.target)
        predictions_payload = {}
        for key, item in prediction["predictions"].items():
            predictions_payload[key] = item.model_dump()
            predictions_payload[key]["goal_direction"] = goal_directions.get(key)
        points.append({
            "index": len(points),
            "inputs": applied,
            "candidate": CandidateInput.model_validate(candidate.model_dump()).model_dump(mode="json"),
            "prediction": prediction_payload,
            "predictions": predictions_payload,
            "color_value": selected.value,
            "support": support.model_dump(),
            "warnings": prediction.get("warnings", []),
            "similar": [item.model_dump() if hasattr(item, "model_dump") else item for item in prediction.get("similar", [])],
            "score": None if evaluation.score is None else round(float(evaluation.score), 6),
            "goal_evaluation": evaluation.model_dump(),
            "secondary_goal_evaluations": {key: item.model_dump() for key, item in secondary_evaluations.items()},
        })
    if len(points) < request.samples:
        raise ValueError(f"指定範囲では制約を満たす点を{request.samples}件作れませんでした。範囲を見直してください")
    support_rank = {"supported": 0, "caution": 1, "extrapolated": 2}
    ranked = sorted(
        points,
        key=lambda point: (
            sum(item["achieved"] is False for item in point["secondary_goal_evaluations"].values()),
            point["score"] is None,
            point["score"] if point["score"] is not None else support_rank[point["support"]["status"]],
            support_rank[point["support"]["status"]],
            point["index"],
        ),
    )
    return {
        "schema_version": "screening-run/v2",
        "seed": SCREENING_SEED,
        "base_candidate_id": base.id,
        "base_inputs": base.inputs.model_dump(mode="json"),
        "base_canonical_input": base_prediction["canonical_input"],
        "model_provenance": base_prediction["model_meta"],
        "target": request.target,
        "target_value": request.target_value,
        "secondary_targets": request.secondary_targets,
        "score_contract": score_contract(
            goal_directions.get(request.target),
            request.target_value,
            probability_available=probability_available.get(request.target, False),
        ),
        "samples": request.samples,
        "variables": {name: spec.model_dump() for name, spec in request.variables.items()},
        "points": points,
        "representative_points": ranked[:10],
    }


def candidate_from_lineage(data: WorkbookData, entity_key: str) -> CandidateInput:
    process_role = "annealing"
    process_key = entity_key
    if process_key not in data.anneal_features and process_key not in data.hot_rolling_features:
        relations = data.lineage.get(entity_key, {})
        anneal_candidates = relations.get(data.role_to_key["annealing"], [])
        hot_candidates = relations.get(data.role_to_key["hot_rolling"], [])
        if len(anneal_candidates) == 1:
            process_key = anneal_candidates[0]
        elif len(hot_candidates) == 1:
            process_key = hot_candidates[0]
            process_role = "hot_rolling"
        else:
            raise ValueError("焼鈍または熱延条件を一意にたどれないため候補化できません")
    elif process_key in data.hot_rolling_features:
        process_role = "hot_rolling"

    relations = data.lineage.get(process_key, {})
    melt_keys = sorted(set(relations.get(data.role_to_key["melt"], [])))
    if len(melt_keys) != 1 or melt_keys[0] not in data.composition:
        raise ValueError("工程条件に一意な成分が接続されていません")
    feature = data.anneal_features.get(process_key) if process_role == "annealing" else data.hot_rolling_features.get(process_key)
    if feature is None:
        raise ValueError("候補化できる工程条件が見つかりません")
    heat_pattern = None
    if process_role == "annealing":
        heat_pattern = [HeatPoint.model_validate(point) for point in deepcopy(feature["heat_pattern"])]
        if len(heat_pattern) < 2:
            raise ValueError("候補化に必要な焼鈍履歴がありません")
        process_values = {"ls_mpm": float(feature["ls_mpm"])}
        entity_type = "annealing"
    else:
        process_values = {name: float(feature[name]) for name in PROCESS_NAMES}
        entity_type = "hot_rolling"
    return CandidateInput(
        name=f"過去条件 {process_key}",
        inputs={
            "composition": deepcopy(data.composition[melt_keys[0]]),
            "process": process_values,
            "categorical": {},
            "heat_pattern": heat_pattern,
        },
        provenance={
                "source_kind": "lineage",
                "source_ref": {
                    "entity_type": entity_type,
                    "entity_key": process_key,
                "data_source_digest": data.source_sha256,
            },
        },
    )


def _candidate_xlsx_names(
    task_id: str,
    profile_path: str | None = None,
    profile: DatasetInputProfile | None = None,
) -> dict[str, str]:
    profile = profile or load_dataset_profile(profile_path)
    task_profile = profile.tasks[task_id]
    names = {
        "schema_version": "形式バージョン",
        "id": "候補ID",
        "name": "候補名",
        "support_status": "学習範囲判定",
        "support_distance": "学習範囲からの距離",
        "segment_start": "工程境界",
        "heat_time_basis": "時間基準",
    }
    names.update({mapping.path: mapping.column for mapping in task_profile.mappings if mapping.column})
    heat_mapping = next((mapping for mapping in task_profile.mappings if mapping.kind == "ordered_heat_series"), None)
    if heat_mapping and heat_mapping.series_columns:
        names["time_s"] = heat_mapping.series_columns.time
        names["temperature_c"] = heat_mapping.series_columns.value
    technical_names = {
        item.name: item.column
        for item in profile.shared.technical
        if item.role == "anneal_history"
    }
    names["stage_name"] = technical_names.get("stage_name", "工程")
    names["stage_category"] = technical_names.get("stage_category", "標準工程カテゴリ")
    for observation in task_profile.observations:
        for target in (*observation.targets, *observation.auxiliary):
            names.setdefault(target.key, target.column)
    return names


def _candidate_template_unit(field: Any) -> str | None:
    unit = {"mpm": "m/min"}.get(field.unit, field.unit)
    return "mass%" if field.path.startswith("composition.") and unit == "%" else unit


def _candidate_template_names(definition: Any) -> dict[str, str]:
    names = {
        "name": "候補名",
        "time_s": "経過時間[s]",
        "temperature_c": "温度[℃]",
        "heat_time_basis": "時間基準",
    }
    response_labels = {
        item.path: item.label
        for item in definition.response_curve_variables
        if item.kind == "numeric_input" and item.path
    }
    for group in definition.input_groups:
        for field in group.fields:
            if field.kind not in {"number", "categorical"}:
                continue
            label = "ライン速度" if field.path == "process.ls_mpm" else response_labels.get(field.path, field.label)
            unit = _candidate_template_unit(field)
            names[field.path] = f"{label}[{unit}]" if unit else label
    return names


def _xlsx_position(positions: dict[str, int], internal_name: str, *display_names: str | None) -> int | None:
    for name in (*display_names, internal_name):
        if name and name in positions:
            return positions[name]
    return None


def import_candidates_xlsx(
    contents: bytes,
    task_id: str = "annealed-properties-v1",
    profile_path: str | None = None,
    profile: DatasetInputProfile | None = None,
    validate_candidate: Callable[[CandidateInput], Any] | None = None,
) -> tuple[list[CandidateInput], list[dict[str, Any]]]:
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        return [], [{"row": 0, "message": f"Excelを読み込めません: {exc}"}]
    definition = load_task_contracts()[task_id].task_definition
    fields = {
        field.path: field
        for group in definition.input_groups
        for field in group.fields
        if field.kind in {"number", "categorical"}
    }
    composition_fields = [field for field in fields.values() if field.path.startswith("composition.")]
    process_fields = [field for field in fields.values() if field.path.startswith("process.")]
    categorical_fields = [field for field in fields.values() if field.path.startswith("categorical.")]
    heat_enabled = any(field.kind == "heat_pattern" for group in definition.input_groups for field in group.fields)
    display_names = _candidate_xlsx_names(task_id, profile_path, profile)
    template_names = _candidate_template_names(definition)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
    duplicate_headers = sorted({header for header in headers if header and headers.count(header) > 1})
    if duplicate_headers:
        workbook.close()
        return [], [{"row": 1, "message": f"列名が重複しています: {', '.join(duplicate_headers)}"}]
    positions = {header: index for index, header in enumerate(headers) if header}
    if not headers or _xlsx_position(positions, "name", template_names["name"], display_names["name"]) is None:
        return [], [{"row": 1, "message": f"必須列 {template_names['name']} がありません"}]
    imported: list[CandidateInput] = []
    errors: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None for value in row):
            continue
        try:
            def value(header: str, *display_headers: str | None, default: Any = None) -> Any:
                position = _xlsx_position(positions, header, *display_headers)
                return row[position] if position is not None and position < len(row) else default

            composition = {
                field.path.removeprefix("composition."): float(value(field.path.removeprefix("composition."), template_names.get(field.path), display_names.get(field.path)))
                for field in composition_fields
                if value(field.path.removeprefix("composition."), template_names.get(field.path), display_names.get(field.path)) is not None
            }
            process = {
                field.path.removeprefix("process."): float(value(field.path.removeprefix("process."), template_names.get(field.path), display_names.get(field.path)))
                for field in process_fields
                if value(field.path.removeprefix("process."), template_names.get(field.path), display_names.get(field.path)) is not None
            }
            categorical = {
                field.path.removeprefix("categorical."): str(value(field.path.removeprefix("categorical."), template_names.get(field.path), display_names.get(field.path)))
                for field in categorical_fields
                if value(field.path.removeprefix("categorical."), template_names.get(field.path), display_names.get(field.path)) is not None
            }
            points: list[dict[str, Any]] = []
            if heat_enabled:
                index = 1
                while (
                    _xlsx_position(positions, f"time_s_{index}", f"{template_names['time_s']}_{index}", f"{display_names['time_s']}_{index}") is not None
                    and _xlsx_position(positions, f"temperature_c_{index}", f"{template_names['temperature_c']}_{index}", f"{display_names['temperature_c']}_{index}") is not None
                ):
                    time = value(f"time_s_{index}", f"{template_names['time_s']}_{index}", f"{display_names['time_s']}_{index}")
                    temperature = value(f"temperature_c_{index}", f"{template_names['temperature_c']}_{index}", f"{display_names['temperature_c']}_{index}")
                    if time is not None and temperature is not None:
                        segment_raw = value(f"segment_start_{index}", f"{display_names['segment_start']}_{index}", default=False)
                        segment_start = segment_raw is True or str(segment_raw).strip().lower() in {"1", "true", "yes", "あり"}
                        points.append({
                            "time_s": float(time),
                            "temperature_c": float(temperature),
                            "segment_start": segment_start,
                            "stage_name": str(value(f"stage_name_{index}", f"{display_names['stage_name']}_{index}")).strip() if value(f"stage_name_{index}", f"{display_names['stage_name']}_{index}") is not None else None,
                            "stage_category": str(value(f"stage_category_{index}", f"{display_names['stage_category']}_{index}")).strip() if value(f"stage_category_{index}", f"{display_names['stage_category']}_{index}") is not None else None,
                        })
                    index += 1
                if len(points) < 2:
                    raise ValueError(f"{template_names['time_s']}_1 / {template_names['temperature_c']}_1 から少なくとも2点が必要です")
            name = value("name", template_names["name"], display_names["name"])
            if name is None or not str(name).strip():
                raise ValueError(f"{template_names['name']}は空にできません")
            raw_time_basis = str(value(
                "heat_time_basis",
                template_names["heat_time_basis"],
                display_names["heat_time_basis"],
                default="ライン速度連動",
            )).strip()
            if raw_time_basis in {"", "line_speed", "ライン速度連動"}:
                heat_time_basis = "line_speed"
            elif raw_time_basis in {"elapsed_time", "経過時間を直接指定", "経過時間"}:
                heat_time_basis = "elapsed_time"
            else:
                raise ValueError(
                    f"{template_names['heat_time_basis']}は"
                    "「ライン速度連動」または「経過時間を直接指定」で入力してください"
                )
            payload = CandidateInput(
                name=str(name).strip(),
                inputs={
                    "composition": composition,
                    "process": process,
                    "categorical": categorical,
                    "heat_pattern": points if heat_enabled else None,
                    "heat_time_basis": heat_time_basis,
                },
            )
            if validate_candidate is not None:
                validate_candidate(payload)
            imported.append(payload)
        except (TypeError, ValueError) as exc:
            errors.append({"row": row_number, "message": str(exc)})
    workbook.close()
    return imported, errors


def _candidate_xlsx_input_schema(
    task_id: str,
    heat_point_count: int,
    profile_path: str | None = None,
    profile: DatasetInputProfile | None = None,
    *,
    for_template: bool = False,
) -> tuple[Any, list[Any], list[Any], bool, dict[str, str], list[str], list[str]]:
    definition = load_task_contracts()[task_id].task_definition
    numeric_fields = [
        field for group in definition.input_groups
        for field in group.fields
        if field.kind == "number"
    ]
    categorical_fields = [
        field for group in definition.input_groups
        for field in group.fields
        if field.kind == "categorical"
    ]
    heat_enabled = any(field.kind == "heat_pattern" for group in definition.input_groups for field in group.fields)
    display_names = _candidate_template_names(definition) if for_template else _candidate_xlsx_names(task_id, profile_path, profile)
    heat_headers = [
        item for index in range(1, heat_point_count + 1)
        for item in ((f"time_s_{index}", f"temperature_c_{index}") if for_template else (f"time_s_{index}", f"temperature_c_{index}", f"segment_start_{index}", f"stage_name_{index}", f"stage_category_{index}"))
    ] if heat_enabled else []
    input_headers = [
        display_names["name"],
        *[display_names.get(field.path, field.label) for field in numeric_fields],
        *[display_names.get(field.path, field.label) for field in categorical_fields],
        *([display_names["heat_time_basis"]] if heat_enabled else []),
        *[
            f"{display_names.get(header.rsplit('_', 1)[0], header.rsplit('_', 1)[0])}_{header.rsplit('_', 1)[1]}"
            for header in heat_headers
        ],
    ]
    return definition, numeric_fields, categorical_fields, heat_enabled, display_names, heat_headers, input_headers


def candidate_template_xlsx(runtime: PredictionRuntime, task_id: str = "annealed-properties-v1") -> bytes:
    contract = load_task_contracts()[task_id]
    canonical = contract.canonical_candidate
    heat_point_count = max(2, len(canonical.heat_pattern or ())) if canonical.heat_pattern is not None else 0
    definition, numeric_fields, categorical_fields, heat_enabled, display_names, heat_headers, input_headers = _candidate_xlsx_input_schema(
        task_id,
        heat_point_count,
        runtime.data.profile_path if not runtime.data.profile_path.startswith("catalog:") else None,
        getattr(runtime.data, "profile", None),
        for_template=True,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "候補"
    sheet.append(input_headers)
    heat_values = [
        item
        for point in (canonical.heat_pattern or ())
        for item in (point.time_s, point.temperature_c)
    ] if heat_enabled else []
    heat_values.extend([None] * (len(heat_headers) - len(heat_values)))
    example_values = [
        canonical.composition.get(field.path.removeprefix("composition."))
        if field.path.startswith("composition.")
        else canonical.process.get(field.path.removeprefix("process."))
        for field in numeric_fields
    ]
    example_values.extend([
        canonical.categorical.get(field.path.removeprefix("categorical."))
        for field in categorical_fields
    ])
    if heat_enabled:
        example_values.append("ライン速度連動")
    example_sheet = workbook.create_sheet("記入例")
    example_sheet.append(input_headers)
    example_sheet.append(["記入例（候補シートへコピーして変更）", *example_values, *heat_values])
    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F5FC4")
    example_fill = PatternFill("solid", fgColor="FFF4D6")
    for target_sheet in (sheet, example_sheet):
        for cell in target_sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        target_sheet.row_dimensions[1].height = 28
        target_sheet.freeze_panes = "B2"
        target_sheet.auto_filter.ref = target_sheet.dimensions
        for column in target_sheet.columns:
            letter = column[0].column_letter
            target_sheet.column_dimensions[letter].width = min(24, max(13, max(len(str(cell.value or "")) for cell in column) + 2))
        target_sheet.column_dimensions["A"].width = 40
        for column_index, header in enumerate(input_headers, start=1):
            if header == "ライン速度[m/min]":
                target_sheet.column_dimensions[target_sheet.cell(1, column_index).column_letter].width = 20
            elif header.startswith(("経過時間[s]_", "温度[℃]_")):
                target_sheet.column_dimensions[target_sheet.cell(1, column_index).column_letter].width = 18
    for cell in example_sheet[2]:
        cell.fill = example_fill
    for index, field in enumerate((*numeric_fields, *categorical_fields), start=2):
        column_letter = sheet.cell(1, index).column_letter
        if field.kind == "number" and field.allowed_range is not None:
            validation = DataValidation(type="decimal", operator="between", formula1=str(field.allowed_range.min), formula2=str(field.allowed_range.max), allow_blank=not field.required)
            validation.error = f"{field.allowed_range.min}～{field.allowed_range.max} の数値を入力してください"
            validation.errorTitle = "入力範囲外"
            validation.prompt = f"単位: {_candidate_template_unit(field) or 'なし'} / 許容範囲: {field.allowed_range.min}～{field.allowed_range.max}"
            validation.promptTitle = field.label
            validation.showErrorMessage = True
            validation.showInputMessage = True
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}501")
        elif field.kind == "categorical":
            validation = DataValidation(type="list", formula1='"' + ','.join(field.choices) + '"', allow_blank=not field.required)
            validation.showErrorMessage = True
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}501")
    if heat_enabled:
        basis_column = sheet.cell(1, 2 + len(numeric_fields) + len(categorical_fields)).column_letter
        validation = DataValidation(
            type="list",
            formula1='"ライン速度連動,経過時間を直接指定"',
            allow_blank=True,
        )
        validation.showErrorMessage = True
        sheet.add_data_validation(validation)
        validation.add(f"{basis_column}2:{basis_column}501")

    guide = workbook.create_sheet("入力ルール")
    guide.append([f"{definition.label} 候補XLSX", "「記入例」を参考に「候補」シートへ入力し、アプリへ読み込みます。読み込まれるのは「候補」シートだけです。"])
    guide.append(["基本ルール", "1行＝1候補です。1行目の列名と単位は変更しないでください。空行は無視されます。"])
    guide.append(["候補名", f"{display_names['name']} は必須です。「記入例」の2行目を「候補」へコピーし、名前と値を変更して使えます。"])
    guide.append(["数値", "数値だけを入力します。単位は列名に記載されているため、セル内には書きません。"])
    if heat_enabled:
        guide.append(["時間基準", "「ライン速度連動」はLS変更時に全時刻を再計算します。「経過時間を直接指定」は入力した時刻をそのまま使います。空欄はライン速度連動です。"])
        guide.append(["履歴点", f"{display_names['time_s']}_N と {display_names['temperature_c']}_N を同じ番号で最低2点入力します。時間は候補全体の開始からの連続した経過時間で、工程が変わっても0へ戻さず昇順にします。点を増やす場合は右へ同じ列セットを追加します。"])
    guide.append([])
    rules_header_row = guide.max_row + 1
    for row_number in range(1, rules_header_row - 1):
        guide.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=5)
    guide.append(["列名", "区分", "必須", "入力ルール", "学習データ範囲（参照）"])
    guide.append([display_names["name"], "識別", "必須", "候補ごとに異なる名前", "—"])
    for field in (*numeric_fields, *categorical_fields):
        if field.kind == "number":
            assert field.allowed_range is not None and field.training_range is not None
            display_unit = _candidate_template_unit(field) or ""
            rule = f"{field.allowed_range.min}～{field.allowed_range.max} {display_unit}".strip()
            training = f"{field.training_range.min}～{field.training_range.max} {display_unit}".strip()
        else:
            rule = " / ".join(field.choices)
            training = "—"
        guide.append([display_names.get(field.path, field.label), field.label, "必須" if field.required else "任意", rule, training])
    if heat_enabled:
        guide.append([display_names["heat_time_basis"], "時間軸", "任意", "ライン速度連動 / 経過時間を直接指定", "—"])
        guide.append([f"{display_names['time_s']}_N / {display_names['temperature_c']}_N", "履歴点", "最低2点", "同じNの時刻と温度を対で入力", "—"])
    guide.freeze_panes = f"A{rules_header_row + 1}"
    guide.column_dimensions["A"].width = 34
    guide.column_dimensions["B"].width = 22
    guide.column_dimensions["C"].width = 12
    guide.column_dimensions["D"].width = 42
    guide.column_dimensions["E"].width = 34
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in guide[rules_header_row]:
        cell.fill = PatternFill("solid", fgColor="DCEAFB")
        cell.font = Font(bold=True, color="173F75")
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in range(1, rules_header_row - 1):
        guide.cell(row_number, 2).alignment = Alignment(vertical="center", wrap_text=False)
        guide.row_dimensions[row_number].height = 24
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def candidates_xlsx(candidates: list[Candidate], runtime: PredictionRuntime, task_id: str = "annealed-properties-v1") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "候補"
    max_points = max((len(candidate.inputs.heat_pattern or []) for candidate in candidates), default=2)
    definition, numeric_fields, categorical_fields, heat_enabled, display_names, heat_headers, input_headers = _candidate_xlsx_input_schema(
        task_id,
        max_points,
        runtime.data.profile_path if not runtime.data.profile_path.startswith("catalog:") else None,
        getattr(runtime.data, "profile", None),
    )
    headers = [
        display_names["schema_version"], display_names["id"], *input_headers,
        *[display_names.get(output.key, output.label) for output in definition.outputs],
        display_names["support_status"], display_names["support_distance"],
    ]
    if len(headers) != len(set(headers)):
        duplicates = sorted({header for header in headers if headers.count(header) > 1})
        raise ValueError(f"候補XLSXの見出しが重複しています: {', '.join(duplicates)}")
    sheet.append(headers)
    for candidate in candidates:
        result = runtime.predict(candidate, detailed=False)
        heat_values = [
            item
            for point in (candidate.inputs.heat_pattern or [])
            for item in (point.time_s, point.temperature_c, point.segment_start, point.stage_name, point.stage_category)
        ] if heat_enabled else []
        heat_values.extend([None] * (len(heat_headers) - len(heat_values)))
        input_values = [
            candidate.inputs.composition.get(field.path.removeprefix("composition."))
            if field.path.startswith("composition.")
            else candidate.inputs.process.get(field.path.removeprefix("process."))
            for field in numeric_fields
        ]
        input_values.extend([
            candidate.inputs.categorical.get(field.path.removeprefix("categorical."))
            for field in categorical_fields
        ])
        if heat_enabled:
            input_values.append(
                "経過時間を直接指定"
                if candidate.inputs.heat_time_basis == "elapsed_time"
                else "ライン速度連動"
            )
        sheet.append([
            "material-workbench-candidate-v2", candidate.id, candidate.name, *input_values, *heat_values,
            *(result["predictions"][output.key].value for output in definition.outputs),
            result["support"].status, result["support"].distance,
        ])
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(22, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    guide = workbook.create_sheet("説明")
    guide.append([display_names["schema_version"], "material-workbench-candidate-v2"])
    guide.append(["用途", "候補入力と軽量プレビュー予測の出力。候補シートはそのまま候補importに使用できます。"])
    guide.append(["入力項目", " / ".join(f"{field.label}[{field.unit}]" for field in (*numeric_fields, *categorical_fields))])
    if heat_enabled:
        guide.append(["時間基準", "ライン速度連動ではLS変更時に全時刻を再計算します。経過時間を直接指定では入力時刻を保持します。"])
        guide.append(["焼鈍履歴", f"{display_names['time_s']}_N / {display_names['temperature_c']}_N を履歴点として指定し、{display_names['stage_name']}_N を工程名として表示します。工程の先頭は {display_names['segment_start']}_N で表します。"])
    guide.append(["予測", "出力単位はTaskDefinitionの定義に従います。"])
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 110
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
