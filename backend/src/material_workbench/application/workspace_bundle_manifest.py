from __future__ import annotations

import json
import shutil
import sqlite3
import tempfile
from functools import lru_cache
from hashlib import sha256
from pathlib import Path, PurePosixPath
from openpyxl import load_workbook
from material_workbench.data.evidence_images import (
    EvidenceImageError,
    resolve_evidence_image,
)
from material_workbench.contracts.workspace_bundle_contracts import (
    WorkspaceBundleDiagnostic,
    WorkspaceBundleFile,
    WorkspaceBundleManifest,
    WorkspaceBundleMigration,
    WorkspaceBundlePackageReference,
    WorkspaceBundleResource,
    WorkspaceTableEvidence,
)
from material_workbench.modeling.model_package_verification import ModelPackageLoader
from material_workbench.persistence.sqlite_connection import (
    connect_sqlite,
)
from material_workbench.persistence.store import Store
from material_workbench.persistence.row_payload_store import (
    RowPayloadError,
    RowPayloadReference,
    RowPayloadStore,
)
from material_workbench.persistence.data_lifecycle_payload_storage import (
    QuarantinedPayloadReference,
    StoredLifecycleRowResource,
)
from material_workbench.application.workspace_bundle_shared import (
    RESOURCE_ARCHIVE_ROOT,
    ROW_PAYLOAD_ARCHIVE_ROOT,
    WorkspaceBundleError,
    _canonical_digest,
    _file_digest,
    _json_value,
)


def _normalized_row(
    table: str,
    columns: tuple[str, ...],
    row: sqlite3.Row,
) -> list[object]:
    values: list[object] = []
    for column in columns:
        value = row[column]
        if column == "locator" and table == "data_assets":
            value = f"data-asset:{row['id']}"
        elif column == "locator_kind" and table == "data_assets":
            value = "workspace-managed"
        elif column == "locator" and table == "model_package_refs":
            value = f"model-package:{row['id']}"
        values.append(_json_value(value))
    return values


def _table_evidence(
    connection: sqlite3.Connection,
    *,
    expected: tuple[WorkspaceTableEvidence, ...] | None = None,
) -> tuple[WorkspaceTableEvidence, ...]:
    if expected is None:
        table_columns = []
        tables = sorted(
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name NOT LIKE 'sqlite_%' AND name <> 'schema_migrations'"
            )
        )
        for table in tables:
            columns = tuple(
                str(row[1])
                for row in connection.execute(
                    f'PRAGMA table_info("{table.replace(chr(34), chr(34) * 2)}")'
                )
            )
            table_columns.append((table, columns))
    else:
        table_columns = [(item.table, item.columns) for item in expected]
    evidence: list[WorkspaceTableEvidence] = []
    for table, columns in table_columns:
        actual_columns = {
            str(row[1])
            for row in connection.execute(
                f'PRAGMA table_info("{table.replace(chr(34), chr(34) * 2)}")'
            )
        }
        if not set(columns) <= actual_columns:
            raise WorkspaceBundleError(
                f"Migration removed evidence columns from {table}"
            )
        select_columns = ",".join(
            f'"{column.replace(chr(34), chr(34) * 2)}"' for column in columns
        )
        rows = [
            _normalized_row(table, columns, row)
            for row in connection.execute(
                f'SELECT {select_columns} FROM "{table.replace(chr(34), chr(34) * 2)}"'
            )
        ]
        rows.sort(key=lambda row: json.dumps(row, ensure_ascii=False, sort_keys=True))
        evidence.append(
            WorkspaceTableEvidence(
                table=table,
                columns=columns,
                row_count=len(rows),
                digest=_canonical_digest(rows),
            )
        )
    return tuple(evidence)


def _database_evidence(
    database: Path,
    *,
    expected_tables: tuple[WorkspaceTableEvidence, ...] | None = None,
) -> tuple[
    tuple[WorkspaceBundleMigration, ...],
    tuple[WorkspaceBundlePackageReference, ...],
    tuple[WorkspaceTableEvidence, ...],
    tuple[WorkspaceBundleDiagnostic, ...],
]:
    connection = connect_sqlite(database)
    try:
        integrity_rows = connection.execute("PRAGMA integrity_check").fetchall()
        foreign_keys = connection.execute("PRAGMA foreign_key_check").fetchall()
        tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )
        }
        migrations = (
            tuple(
                WorkspaceBundleMigration(
                    id=str(row["id"]),
                    checksum=str(row["checksum"]),
                    applied_at=str(row["applied_at"]),
                )
                for row in connection.execute(
                    "SELECT id,checksum,applied_at FROM schema_migrations ORDER BY id"
                )
            )
            if "schema_migrations" in tables
            else ()
        )
        packages = (
            tuple(
                WorkspaceBundlePackageReference(
                    reference_id=str(row["id"]),
                    package_id=str(row["package_id"]),
                    task_id=str(row["task_id"]),
                    manifest_digest=str(row["manifest_digest"]),
                    archived=row["archived_at"] is not None,
                )
                for row in connection.execute(
                    "SELECT id,package_id,task_id,manifest_digest,locator,archived_at "
                    "FROM model_package_refs ORDER BY id"
                )
            )
            if "model_package_refs" in tables
            else ()
        )
        evidence = _table_evidence(connection, expected=expected_tables)
    finally:
        connection.close()
    integrity_failures = [str(row[0]) for row in integrity_rows if row[0] != "ok"]
    if integrity_failures:
        raise WorkspaceBundleError(
            "Workspace DB integrity_check failed: " + "; ".join(integrity_failures[:10])
        )
    if foreign_keys:
        raise WorkspaceBundleError(
            f"Workspace DB has {len(foreign_keys)} foreign key violations"
        )
    diagnostics = (
        WorkspaceBundleDiagnostic(
            id="sqlite-integrity-check", status="ok", detail="ok"
        ),
        WorkspaceBundleDiagnostic(
            id="sqlite-foreign-keys", status="ok", detail="0 violations"
        ),
    )
    return migrations, packages, evidence, diagnostics


def _logical_database_backup(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    source_connection = connect_sqlite(source)
    target_connection = sqlite3.connect(destination)
    try:
        source_connection.backup(target_connection)
        target_connection.commit()
        result = [
            str(row[0])
            for row in target_connection.execute("PRAGMA integrity_check")
            if row[0] != "ok"
        ]
        if result:
            raise WorkspaceBundleError(
                "SQLite Backup API output failed integrity_check: "
                + "; ".join(result[:10])
            )
    finally:
        target_connection.close()
        source_connection.close()


def _migration_inventory(database: Path) -> dict[str, str]:
    connection = connect_sqlite(database)
    try:
        tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )
        }
        if "schema_migrations" not in tables:
            return {}
        return {
            str(row["id"]): str(row["checksum"])
            for row in connection.execute(
                "SELECT id,checksum FROM schema_migrations ORDER BY id"
            )
        }
    finally:
        connection.close()


@lru_cache(maxsize=1)
def _current_migration_inventory() -> dict[str, str]:
    with tempfile.TemporaryDirectory(prefix="material-workbench-schema-") as temporary:
        database = Path(temporary) / "current.db"
        Store(database)
        return _migration_inventory(database)


def _validate_migration_inventory(
    database: Path,
    manifest: WorkspaceBundleManifest,
) -> None:
    actual = _migration_inventory(database)
    declared = {item.id: item.checksum for item in manifest.schema_migrations}
    if actual != declared:
        raise WorkspaceBundleError(
            "Workspace bundle migration inventory does not match its database"
        )
    supported = _current_migration_inventory()
    unknown = sorted(set(actual) - set(supported))
    if unknown:
        raise WorkspaceBundleError(
            "Workspace bundle was created by a newer or unsupported application; "
            f"unknown migrations: {', '.join(unknown)}"
        )
    mismatched = sorted(
        migration_id
        for migration_id, checksum in actual.items()
        if supported[migration_id] != checksum
    )
    if mismatched:
        raise WorkspaceBundleError(
            "Workspace bundle migration checksum is not supported: "
            + ", ".join(mismatched)
        )


def _file_record(path: Path, archive_path: str) -> WorkspaceBundleFile:
    return WorkspaceBundleFile(
        path=archive_path,
        sha256=_file_digest(path),
        size_bytes=path.stat().st_size,
    )


def _copy_file_verified(source: Path, destination: Path, digest: str) -> None:
    if not source.is_file() or source.is_symlink():
        raise WorkspaceBundleError(
            f"Workspace resource is not a regular file: {source}"
        )
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, destination)
    if _file_digest(destination) != digest.removeprefix("sha256:"):
        destination.unlink(missing_ok=True)
        raise WorkspaceBundleError(f"Workspace resource digest mismatch: {source}")


def _copy_package_verified(
    source: Path,
    destination: Path,
    manifest_digest: str,
) -> None:
    if not source.is_dir() or source.is_symlink():
        raise WorkspaceBundleError(f"Model Package directory is missing: {source}")
    linked = next(
        (path for path in source.rglob("*") if path.is_symlink()),
        None,
    )
    if linked is not None:
        raise WorkspaceBundleError(
            f"Model Package symlink is not allowed in a Workspace bundle: {linked}"
        )
    verified = ModelPackageLoader().load(source)
    expected = manifest_digest.removeprefix("sha256:")
    if verified.manifest_sha256 != expected:
        raise WorkspaceBundleError(f"Model Package manifest digest mismatch: {source}")
    shutil.copytree(
        source,
        destination,
        symlinks=False,
        ignore_dangling_symlinks=False,
    )
    copied = ModelPackageLoader().load(destination)
    if copied.manifest_sha256 != expected:
        raise WorkspaceBundleError(
            f"Copied Model Package manifest digest mismatch: {source}"
        )


def _declared_evidence_files(
    source: Path,
    profile_payloads: tuple[dict[str, object], ...],
) -> tuple[tuple[tuple[PurePosixPath, Path], ...], tuple[str, ...]]:
    declarations: set[tuple[str, str]] = set()
    for payload in profile_payloads:
        shared = payload.get("shared")
        if not isinstance(shared, dict):
            continue
        sheets = shared.get("sheets")
        technical = shared.get("technical")
        if not isinstance(sheets, dict) or not isinstance(technical, list):
            continue
        for item in technical:
            if not isinstance(item, dict) or item.get("name") != "evidence_image":
                continue
            role = item.get("role")
            column = item.get("column")
            sheet = sheets.get(role) if isinstance(role, str) else None
            if isinstance(sheet, str) and isinstance(column, str):
                declarations.add((sheet, column))
    if not declarations:
        return (), ()
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        return (), (f"{source.name}: evidence image columns require an XLSX workbook",)

    found: dict[str, Path] = {}
    warnings: set[str] = set()
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        for sheet_name, column_name in sorted(declarations):
            if sheet_name not in workbook.sheetnames:
                warnings.add(
                    f"{source.name}: evidence image sheet is missing: {sheet_name}"
                )
                continue
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, ())
            try:
                column_index = next(
                    index
                    for index, value in enumerate(headers)
                    if str(value).strip() == column_name
                )
            except StopIteration:
                warnings.add(
                    f"{source.name}: evidence image column is missing: "
                    f"{sheet_name}/{column_name}"
                )
                continue
            for row in rows:
                if column_index >= len(row) or row[column_index] is None:
                    continue
                declared = str(row[column_index]).strip()
                if not declared:
                    continue
                try:
                    resolved = resolve_evidence_image(declared, source)
                except EvidenceImageError as exc:
                    warnings.add(f"{source.name}: {declared}: {exc}")
                    continue
                if resolved.available and resolved.resolved is not None:
                    relative = resolved.resolved.relative_to(source.resolve().parent)
                    found[relative.as_posix()] = resolved.resolved
                else:
                    warnings.add(
                        f"{source.name}: {declared}: "
                        f"{resolved.reason or 'evidence image is unavailable'}"
                    )
    finally:
        workbook.close()
    return (
        tuple((PurePosixPath(path), found[path]) for path in sorted(found)),
        tuple(sorted(warnings)),
    )


def _resource_bundle_digest(
    records: tuple[WorkspaceBundleFile, ...],
) -> str:
    return _canonical_digest(
        [
            {
                "path": record.path,
                "sha256": record.sha256,
                "size_bytes": record.size_bytes,
            }
            for record in sorted(records, key=lambda item: item.path)
        ]
    )


def _snapshot_resources(
    database: Path,
    staging: Path,
) -> tuple[
    tuple[WorkspaceBundleResource, ...],
    tuple[WorkspaceBundleFile, ...],
    tuple[str, ...],
]:
    connection = connect_sqlite(database)
    try:
        assets = connection.execute(
            "SELECT id,sha256,locator FROM data_assets ORDER BY id"
        ).fetchall()
        packages = connection.execute(
            "SELECT id,manifest_digest,locator FROM model_package_refs ORDER BY id"
        ).fetchall()
        table_names = {
            str(row["name"])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }
        if {"dataset_revisions", "dataset_profile_revisions"} <= table_names:
            profile_rows = connection.execute(
                "SELECT dr.data_asset_id,pr.effective_profile_json "
                "FROM dataset_revisions dr "
                "JOIN dataset_profile_revisions pr ON pr.id=dr.profile_revision_id "
                "ORDER BY dr.data_asset_id,pr.id"
            ).fetchall()
        else:
            profile_rows = []
    finally:
        connection.close()
    profiles_by_asset: dict[str, list[dict[str, object]]] = {}
    for row in profile_rows:
        try:
            payload = json.loads(str(row["effective_profile_json"]))
        except ValueError:
            continue
        if isinstance(payload, dict):
            profiles_by_asset.setdefault(str(row["data_asset_id"]), []).append(payload)
    resources: list[WorkspaceBundleResource] = []
    file_records: dict[str, WorkspaceBundleFile] = {}
    warnings: list[str] = []
    copied_roots: dict[
        tuple[str, str],
        tuple[str, str, str | None, tuple[str, ...]],
    ] = {}

    for row in assets:
        digest = str(row["sha256"])
        reference_id = str(row["id"])
        key = ("data_asset", f"{digest}:{reference_id}")
        copied = copied_roots.get(key)
        if copied is None:
            source = Path(str(row["locator"]))
            suffix = Path(str(row["locator"])).suffix.lower()
            reference_digest = sha256(reference_id.encode("utf-8")).hexdigest()[:12]
            relative_root = f"data-assets/sha256/{digest}/{reference_digest}"
            relative_file = f"{relative_root}/asset{suffix}"
            destination = staging / relative_file
            _copy_file_verified(source, destination, digest)
            archive_path = f"{RESOURCE_ARCHIVE_ROOT}/{relative_file}"
            records = [_file_record(destination, archive_path)]
            file_records[archive_path] = records[0]
            evidence_files, evidence_warnings = _declared_evidence_files(
                source,
                tuple(profiles_by_asset.get(reference_id, ())),
            )
            warnings.extend(evidence_warnings)
            for relative, evidence_source in evidence_files:
                auxiliary_relative = f"{relative_root}/{relative.as_posix()}"
                auxiliary_destination = staging / Path(auxiliary_relative)
                auxiliary_destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.copyfile(evidence_source, auxiliary_destination)
                auxiliary_archive = f"{RESOURCE_ARCHIVE_ROOT}/{auxiliary_relative}"
                record = _file_record(auxiliary_destination, auxiliary_archive)
                records.append(record)
                file_records[auxiliary_archive] = record
            resource_files = tuple(record.path for record in records)
            copied = (
                f"{RESOURCE_ARCHIVE_ROOT}/{relative_root}",
                _resource_bundle_digest(tuple(records)),
                archive_path,
                resource_files,
            )
            copied_roots[key] = copied
        resources.append(
            WorkspaceBundleResource(
                kind="data_asset",
                reference_id=reference_id,
                content_digest=digest,
                bundle_digest=copied[1],
                bundle_root=copied[0],
                primary_file=copied[2],
                files=copied[3],
            )
        )

    for row in packages:
        digest = str(row["manifest_digest"])
        key = ("model_package", digest)
        copied = copied_roots.get(key)
        if copied is None:
            safe_digest = digest.removeprefix("sha256:")
            relative_root = f"model-packages/sha256/{safe_digest}"
            destination = staging / relative_root
            _copy_package_verified(Path(str(row["locator"])), destination, digest)
            paths: list[str] = []
            for path in sorted(
                item for item in destination.rglob("*") if item.is_file()
            ):
                relative = path.relative_to(staging).as_posix()
                archive_path = f"{RESOURCE_ARCHIVE_ROOT}/{relative}"
                file_records[archive_path] = _file_record(path, archive_path)
                paths.append(archive_path)
            records = tuple(file_records[path] for path in paths)
            copied = (
                f"{RESOURCE_ARCHIVE_ROOT}/{relative_root}",
                _resource_bundle_digest(records),
                None,
                tuple(paths),
            )
            copied_roots[key] = copied
        resources.append(
            WorkspaceBundleResource(
                kind="model_package",
                reference_id=str(row["id"]),
                content_digest=digest,
                bundle_digest=copied[1],
                bundle_root=copied[0],
                primary_file=None,
                files=copied[3],
            )
        )
    return (
        tuple(resources),
        tuple(file_records[path] for path in sorted(file_records)),
        tuple(sorted(set(warnings))),
    )


def _row_payload_references(
    database: Path,
) -> tuple[RowPayloadReference, ...]:
    connection = connect_sqlite(database)
    references: dict[str, RowPayloadReference] = {}
    try:
        for table, kind in (
            ("raw_source_snapshots", "raw-json-record/v1"),
            ("source_curation_runs", "curated-row/v1"),
        ):
            columns = {
                str(row["name"])
                for row in connection.execute(f'PRAGMA table_info("{table}")')
            }
            if "row_payload_sha256" not in columns:
                continue
            rows = connection.execute(
                f"SELECT id,payload,row_payload_sha256,row_payload_bytes,row_count "
                f"FROM {table} ORDER BY id"
            ).fetchall()
            for row in rows:
                wrapper = StoredLifecycleRowResource.model_validate_json(row["payload"])
                expected_resource_kind = (
                    "raw_source_snapshot"
                    if table == "raw_source_snapshots"
                    else "curation_run"
                )
                indexed_values = (
                    row["row_payload_sha256"],
                    row["row_payload_bytes"],
                    row["row_count"],
                )
                if all(value is None for value in indexed_values):
                    if (
                        wrapper.resource_kind == expected_resource_kind
                        and wrapper.resource_id == str(row["id"])
                        and wrapper.row_payload is None
                        and wrapper.unavailable_reason
                    ):
                        continue
                    raise WorkspaceBundleError(
                        f"Lifecycle row payload is unavailable without quarantine: "
                        f"{table}/{row['id']}"
                    )
                if any(value is None for value in indexed_values):
                    raise WorkspaceBundleError(
                        f"Lifecycle row payload reference is incomplete: "
                        f"{table}/{row['id']}"
                    )
                reference = RowPayloadReference(
                    record_kind=kind,
                    sha256=str(indexed_values[0]),
                    size_bytes=int(indexed_values[1]),
                    row_count=int(indexed_values[2]),
                )
                if (
                    wrapper.resource_kind != expected_resource_kind
                    or wrapper.resource_id != str(row["id"])
                    or wrapper.row_payload != reference
                ):
                    raise WorkspaceBundleError(
                        f"Lifecycle row payload reference is inconsistent: "
                        f"{table}/{row['id']}"
                    )
                previous = references.get(reference.sha256)
                if previous is not None and (
                    previous.size_bytes != reference.size_bytes
                    or previous.row_count != reference.row_count
                ):
                    raise WorkspaceBundleError(
                        "Lifecycle row payload reference is inconsistent"
                    )
                references[reference.sha256] = reference
    finally:
        connection.close()
    return tuple(references[digest] for digest in sorted(references))


def _quarantined_payload_references(
    database: Path,
) -> tuple[QuarantinedPayloadReference, ...]:
    connection = connect_sqlite(database)
    references: dict[str, QuarantinedPayloadReference] = {}
    try:
        for table in ("raw_source_snapshots", "source_curation_runs"):
            columns = {
                str(row["name"])
                for row in connection.execute(f'PRAGMA table_info("{table}")')
            }
            if "row_payload_sha256" not in columns:
                continue
            for row in connection.execute(
                f'SELECT id,payload FROM "{table}" ORDER BY id'
            ):
                wrapper = StoredLifecycleRowResource.model_validate_json(row["payload"])
                if wrapper.unavailable_reason is None:
                    continue
                reference = wrapper.quarantined_payload
                if reference is None:
                    raise WorkspaceBundleError(
                        f"Lifecycle quarantine reference is missing: "
                        f"{table}/{row['id']}"
                    )
                previous = references.get(reference.path)
                if previous is not None and previous != reference:
                    raise WorkspaceBundleError(
                        "Lifecycle quarantine reference is inconsistent"
                    )
                references[reference.path] = reference
    finally:
        connection.close()
    return tuple(references[path] for path in sorted(references))


def _snapshot_row_payloads(
    *,
    database: Path,
    staged_database: Path,
    staging: Path,
) -> tuple[WorkspaceBundleFile, ...]:
    source_store = RowPayloadStore(database)
    records: list[WorkspaceBundleFile] = []
    for reference in _row_payload_references(staged_database):
        try:
            source_store.verify(reference)
        except RowPayloadError as exc:
            raise WorkspaceBundleError(
                f"Lifecycle row payload cannot be backed up: {reference.sha256}"
            ) from exc
        source = source_store.path_for(reference)
        archive_path = (
            f"{ROW_PAYLOAD_ARCHIVE_ROOT}/sha256/"
            f"{reference.sha256[:2]}/{reference.sha256}.jsonl"
        )
        destination = staging / Path(archive_path)
        _copy_file_verified(source, destination, reference.sha256)
        records.append(_file_record(destination, archive_path))
    expected_quarantine_paths: set[Path] = set()
    for reference in _quarantined_payload_references(staged_database):
        source = source_store.database.parent / Path(reference.path)
        expected_quarantine_paths.add(source.resolve())
        if (
            not source.is_file()
            or source.is_symlink()
            or source.stat().st_size != reference.size_bytes
            or _file_digest(source) != reference.sha256
        ):
            raise WorkspaceBundleError(
                f"Lifecycle payload quarantine cannot be backed up: {reference.path}"
            )
        archive_path = f"workspace/{reference.path}"
        destination = staging / Path(archive_path)
        _copy_file_verified(source, destination, reference.sha256)
        records.append(_file_record(destination, archive_path))
    quarantine_root = source_store.database.parent / "row-payloads" / "quarantine"
    actual_quarantine_paths = (
        {path.resolve() for path in quarantine_root.rglob("*.json") if path.is_file()}
        if quarantine_root.exists()
        else set()
    )
    if actual_quarantine_paths != expected_quarantine_paths:
        raise WorkspaceBundleError(
            "Lifecycle payload quarantine inventory is inconsistent"
        )
    return tuple(records)
