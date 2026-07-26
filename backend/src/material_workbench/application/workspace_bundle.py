from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import stat
import tempfile
import zipfile
from datetime import UTC, datetime, timedelta
from functools import lru_cache
from hashlib import sha256
from pathlib import Path, PurePosixPath
from uuid import uuid4

from openpyxl import load_workbook

from material_workbench.data.evidence_images import (
    EvidenceImageError,
    resolve_evidence_image,
)
from material_workbench.contracts.chain_contracts import (
    semantic_digest,
    task_contract_surface,
    validate_chain_revision,
)
from material_workbench.contracts.workspace_bundle_contracts import (
    WorkspaceBackupResult,
    WorkspaceBundleDiagnostic,
    WorkspaceBundleFile,
    WorkspaceBundleManifest,
    WorkspaceBundleMigration,
    WorkspaceBundlePackageReference,
    WorkspaceBundleResource,
    WorkspaceRestoreCommitResult,
    WorkspaceRestorePrepared,
    WorkspaceRestoreResolution,
    WorkspaceTableEvidence,
)
from material_workbench.modeling.model_packages import ModelPackageLoader
from material_workbench.modeling.transform_catalog import (
    DeterministicTransformCatalog,
)
from material_workbench.persistence.sqlite_connection import (
    connect_sqlite,
    validate_sqlite_foreign_keys,
)
from material_workbench.persistence.store import Store
from material_workbench.persistence.welding_chain_bootstrap import (
    welding_stage_a_surface,
)
from material_workbench.persistence.workspace_catalog import WorkspaceCatalog
from material_workbench.tasks.project_runtime_resolver import ProjectRuntimeResolver
from material_workbench.tasks.task_registry import TaskRegistry


DATABASE_ARCHIVE_PATH = "workspace/workbench.db"
MANIFEST_ARCHIVE_PATH = "manifest.json"
RESOURCE_ARCHIVE_ROOT = "workspace/resources"
MAX_BUNDLE_ENTRIES = 20_000
MAX_BUNDLE_BYTES = 2 * 1024 * 1024 * 1024
MAX_ENTRY_BYTES = 1024 * 1024 * 1024
MAX_MANIFEST_BYTES = 8 * 1024 * 1024
MAX_COMPRESSION_RATIO = 200
MIN_FREE_SPACE_RESERVE = 64 * 1024 * 1024
RESTORE_EXPIRY_HOURS = 24
WINDOWS_RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


class WorkspaceBundleError(RuntimeError):
    pass


def _file_digest(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_digest(value: object) -> str:
    return sha256(
        json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()


def _json_value(value: object) -> object:
    if isinstance(value, bytes):
        return {"$bytes": value.hex()}
    return value


def _normalized_row(
    table: str,
    columns: tuple[str, ...],
    row: sqlite3.Row,
) -> list[object]:
    values: list[object] = []
    for column in columns:
        value = row[column]
        if column == "locator" and table == "data_assets":
            value = f"data-asset:{row['id']}"
        elif column == "locator_kind" and table == "data_assets":
            value = "workspace-managed"
        elif column == "locator" and table == "model_package_refs":
            value = f"model-package:{row['id']}"
        values.append(_json_value(value))
    return values


def _table_evidence(
    connection: sqlite3.Connection,
    *,
    expected: tuple[WorkspaceTableEvidence, ...] | None = None,
) -> tuple[WorkspaceTableEvidence, ...]:
    if expected is None:
        table_columns = []
        tables = sorted(
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name NOT LIKE 'sqlite_%' AND name <> 'schema_migrations'"
            )
        )
        for table in tables:
            columns = tuple(
                str(row[1])
                for row in connection.execute(
                    f'PRAGMA table_info("{table.replace(chr(34), chr(34) * 2)}")'
                )
            )
            table_columns.append((table, columns))
    else:
        table_columns = [(item.table, item.columns) for item in expected]
    evidence: list[WorkspaceTableEvidence] = []
    for table, columns in table_columns:
        actual_columns = {
            str(row[1])
            for row in connection.execute(
                f'PRAGMA table_info("{table.replace(chr(34), chr(34) * 2)}")'
            )
        }
        if not set(columns) <= actual_columns:
            raise WorkspaceBundleError(
                f"Migration removed evidence columns from {table}"
            )
        select_columns = ",".join(
            f'"{column.replace(chr(34), chr(34) * 2)}"' for column in columns
        )
        rows = [
            _normalized_row(table, columns, row)
            for row in connection.execute(
                f'SELECT {select_columns} FROM '
                f'"{table.replace(chr(34), chr(34) * 2)}"'
            )
        ]
        rows.sort(key=lambda row: json.dumps(row, ensure_ascii=False, sort_keys=True))
        evidence.append(
            WorkspaceTableEvidence(
                table=table,
                columns=columns,
                row_count=len(rows),
                digest=_canonical_digest(rows),
            )
        )
    return tuple(evidence)


def _database_evidence(
    database: Path,
    *,
    expected_tables: tuple[WorkspaceTableEvidence, ...] | None = None,
) -> tuple[
    tuple[WorkspaceBundleMigration, ...],
    tuple[WorkspaceBundlePackageReference, ...],
    tuple[WorkspaceTableEvidence, ...],
    tuple[WorkspaceBundleDiagnostic, ...],
]:
    connection = connect_sqlite(database)
    try:
        integrity_rows = connection.execute("PRAGMA integrity_check").fetchall()
        foreign_keys = connection.execute("PRAGMA foreign_key_check").fetchall()
        tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )
        }
        migrations = (
            tuple(
                WorkspaceBundleMigration(
                    id=str(row["id"]),
                    checksum=str(row["checksum"]),
                    applied_at=str(row["applied_at"]),
                )
                for row in connection.execute(
                    "SELECT id,checksum,applied_at FROM schema_migrations ORDER BY id"
                )
            )
            if "schema_migrations" in tables
            else ()
        )
        packages = (
            tuple(
                WorkspaceBundlePackageReference(
                    reference_id=str(row["id"]),
                    package_id=str(row["package_id"]),
                    task_id=str(row["task_id"]),
                    manifest_digest=str(row["manifest_digest"]),
                    archived=row["archived_at"] is not None,
                )
                for row in connection.execute(
                    "SELECT id,package_id,task_id,manifest_digest,locator,archived_at "
                    "FROM model_package_refs ORDER BY id"
                )
            )
            if "model_package_refs" in tables
            else ()
        )
        evidence = _table_evidence(connection, expected=expected_tables)
    finally:
        connection.close()
    integrity_failures = [str(row[0]) for row in integrity_rows if row[0] != "ok"]
    if integrity_failures:
        raise WorkspaceBundleError(
            "Workspace DB integrity_check failed: "
            + "; ".join(integrity_failures[:10])
        )
    if foreign_keys:
        raise WorkspaceBundleError(
            f"Workspace DB has {len(foreign_keys)} foreign key violations"
        )
    diagnostics = (
        WorkspaceBundleDiagnostic(
            id="sqlite-integrity-check", status="ok", detail="ok"
        ),
        WorkspaceBundleDiagnostic(
            id="sqlite-foreign-keys", status="ok", detail="0 violations"
        ),
    )
    return migrations, packages, evidence, diagnostics


def _logical_database_backup(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    source_connection = connect_sqlite(source)
    target_connection = sqlite3.connect(destination)
    try:
        source_connection.backup(target_connection)
        target_connection.commit()
        result = [
            str(row[0])
            for row in target_connection.execute("PRAGMA integrity_check")
            if row[0] != "ok"
        ]
        if result:
            raise WorkspaceBundleError(
                "SQLite Backup API output failed integrity_check: "
                + "; ".join(result[:10])
            )
    finally:
        target_connection.close()
        source_connection.close()


def _migration_inventory(database: Path) -> dict[str, str]:
    connection = connect_sqlite(database)
    try:
        tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )
        }
        if "schema_migrations" not in tables:
            return {}
        return {
            str(row["id"]): str(row["checksum"])
            for row in connection.execute(
                "SELECT id,checksum FROM schema_migrations ORDER BY id"
            )
        }
    finally:
        connection.close()


@lru_cache(maxsize=1)
def _current_migration_inventory() -> dict[str, str]:
    with tempfile.TemporaryDirectory(prefix="material-workbench-schema-") as temporary:
        database = Path(temporary) / "current.db"
        Store(database)
        return _migration_inventory(database)


def _validate_migration_inventory(
    database: Path,
    manifest: WorkspaceBundleManifest,
) -> None:
    actual = _migration_inventory(database)
    declared = {item.id: item.checksum for item in manifest.schema_migrations}
    if actual != declared:
        raise WorkspaceBundleError(
            "Workspace bundle migration inventory does not match its database"
        )
    supported = _current_migration_inventory()
    unknown = sorted(set(actual) - set(supported))
    if unknown:
        raise WorkspaceBundleError(
            "Workspace bundle was created by a newer or unsupported application; "
            f"unknown migrations: {', '.join(unknown)}"
        )
    mismatched = sorted(
        migration_id
        for migration_id, checksum in actual.items()
        if supported[migration_id] != checksum
    )
    if mismatched:
        raise WorkspaceBundleError(
            "Workspace bundle migration checksum is not supported: "
            + ", ".join(mismatched)
        )


def _file_record(path: Path, archive_path: str) -> WorkspaceBundleFile:
    return WorkspaceBundleFile(
        path=archive_path,
        sha256=_file_digest(path),
        size_bytes=path.stat().st_size,
    )


def _copy_file_verified(source: Path, destination: Path, digest: str) -> None:
    if not source.is_file() or source.is_symlink():
        raise WorkspaceBundleError(f"Workspace resource is not a regular file: {source}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, destination)
    if _file_digest(destination) != digest.removeprefix("sha256:"):
        destination.unlink(missing_ok=True)
        raise WorkspaceBundleError(f"Workspace resource digest mismatch: {source}")


def _copy_package_verified(
    source: Path,
    destination: Path,
    manifest_digest: str,
) -> None:
    if not source.is_dir() or source.is_symlink():
        raise WorkspaceBundleError(f"Model Package directory is missing: {source}")
    linked = next(
        (path for path in source.rglob("*") if path.is_symlink()),
        None,
    )
    if linked is not None:
        raise WorkspaceBundleError(
            f"Model Package symlink is not allowed in a Workspace bundle: {linked}"
        )
    verified = ModelPackageLoader().load(source)
    expected = manifest_digest.removeprefix("sha256:")
    if verified.manifest_sha256 != expected:
        raise WorkspaceBundleError(
            f"Model Package manifest digest mismatch: {source}"
        )
    shutil.copytree(
        source,
        destination,
        symlinks=False,
        ignore_dangling_symlinks=False,
    )
    copied = ModelPackageLoader().load(destination)
    if copied.manifest_sha256 != expected:
        raise WorkspaceBundleError(
            f"Copied Model Package manifest digest mismatch: {source}"
        )


def _declared_evidence_files(
    source: Path,
    profile_payloads: tuple[dict[str, object], ...],
) -> tuple[tuple[tuple[PurePosixPath, Path], ...], tuple[str, ...]]:
    declarations: set[tuple[str, str]] = set()
    for payload in profile_payloads:
        shared = payload.get("shared")
        if not isinstance(shared, dict):
            continue
        sheets = shared.get("sheets")
        technical = shared.get("technical")
        if not isinstance(sheets, dict) or not isinstance(technical, list):
            continue
        for item in technical:
            if not isinstance(item, dict) or item.get("name") != "evidence_image":
                continue
            role = item.get("role")
            column = item.get("column")
            sheet = sheets.get(role) if isinstance(role, str) else None
            if isinstance(sheet, str) and isinstance(column, str):
                declarations.add((sheet, column))
    if not declarations:
        return (), ()
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        return (), (f"{source.name}: evidence image columns require an XLSX workbook",)

    found: dict[str, Path] = {}
    warnings: set[str] = set()
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        for sheet_name, column_name in sorted(declarations):
            if sheet_name not in workbook.sheetnames:
                warnings.add(f"{source.name}: evidence image sheet is missing: {sheet_name}")
                continue
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, ())
            try:
                column_index = next(
                    index
                    for index, value in enumerate(headers)
                    if str(value).strip() == column_name
                )
            except StopIteration:
                warnings.add(
                    f"{source.name}: evidence image column is missing: "
                    f"{sheet_name}/{column_name}"
                )
                continue
            for row in rows:
                if column_index >= len(row) or row[column_index] is None:
                    continue
                declared = str(row[column_index]).strip()
                if not declared:
                    continue
                try:
                    resolved = resolve_evidence_image(declared, source)
                except EvidenceImageError as exc:
                    warnings.add(f"{source.name}: {declared}: {exc}")
                    continue
                if resolved.available and resolved.resolved is not None:
                    relative = resolved.resolved.relative_to(source.resolve().parent)
                    found[relative.as_posix()] = resolved.resolved
                else:
                    warnings.add(
                        f"{source.name}: {declared}: "
                        f"{resolved.reason or 'evidence image is unavailable'}"
                    )
    finally:
        workbook.close()
    return (
        tuple((PurePosixPath(path), found[path]) for path in sorted(found)),
        tuple(sorted(warnings)),
    )


def _resource_bundle_digest(
    records: tuple[WorkspaceBundleFile, ...],
) -> str:
    return _canonical_digest(
        [
            {"path": record.path, "sha256": record.sha256, "size_bytes": record.size_bytes}
            for record in sorted(records, key=lambda item: item.path)
        ]
    )


def _snapshot_resources(
    database: Path,
    staging: Path,
) -> tuple[
    tuple[WorkspaceBundleResource, ...],
    tuple[WorkspaceBundleFile, ...],
    tuple[str, ...],
]:
    connection = connect_sqlite(database)
    try:
        assets = connection.execute(
            "SELECT id,sha256,locator FROM data_assets ORDER BY id"
        ).fetchall()
        packages = connection.execute(
            "SELECT id,manifest_digest,locator FROM model_package_refs ORDER BY id"
        ).fetchall()
        table_names = {
            str(row["name"])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }
        if {"dataset_revisions", "dataset_profile_revisions"} <= table_names:
            profile_rows = connection.execute(
                "SELECT dr.data_asset_id,pr.effective_profile_json "
                "FROM dataset_revisions dr "
                "JOIN dataset_profile_revisions pr ON pr.id=dr.profile_revision_id "
                "ORDER BY dr.data_asset_id,pr.id"
            ).fetchall()
        else:
            profile_rows = []
    finally:
        connection.close()
    profiles_by_asset: dict[str, list[dict[str, object]]] = {}
    for row in profile_rows:
        try:
            payload = json.loads(str(row["effective_profile_json"]))
        except ValueError:
            continue
        if isinstance(payload, dict):
            profiles_by_asset.setdefault(str(row["data_asset_id"]), []).append(payload)
    resources: list[WorkspaceBundleResource] = []
    file_records: dict[str, WorkspaceBundleFile] = {}
    warnings: list[str] = []
    copied_roots: dict[
        tuple[str, str],
        tuple[str, str, str | None, tuple[str, ...]],
    ] = {}

    for row in assets:
        digest = str(row["sha256"])
        reference_id = str(row["id"])
        key = ("data_asset", f"{digest}:{reference_id}")
        copied = copied_roots.get(key)
        if copied is None:
            source = Path(str(row["locator"]))
            suffix = Path(str(row["locator"])).suffix.lower()
            reference_digest = sha256(reference_id.encode("utf-8")).hexdigest()[:12]
            relative_root = f"data-assets/sha256/{digest}/{reference_digest}"
            relative_file = f"{relative_root}/asset{suffix}"
            destination = staging / relative_file
            _copy_file_verified(source, destination, digest)
            archive_path = f"{RESOURCE_ARCHIVE_ROOT}/{relative_file}"
            records = [_file_record(destination, archive_path)]
            file_records[archive_path] = records[0]
            evidence_files, evidence_warnings = _declared_evidence_files(
                source,
                tuple(profiles_by_asset.get(reference_id, ())),
            )
            warnings.extend(evidence_warnings)
            for relative, evidence_source in evidence_files:
                auxiliary_relative = f"{relative_root}/{relative.as_posix()}"
                auxiliary_destination = staging / Path(auxiliary_relative)
                auxiliary_destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.copyfile(evidence_source, auxiliary_destination)
                auxiliary_archive = f"{RESOURCE_ARCHIVE_ROOT}/{auxiliary_relative}"
                record = _file_record(auxiliary_destination, auxiliary_archive)
                records.append(record)
                file_records[auxiliary_archive] = record
            resource_files = tuple(record.path for record in records)
            copied = (
                f"{RESOURCE_ARCHIVE_ROOT}/{relative_root}",
                _resource_bundle_digest(tuple(records)),
                archive_path,
                resource_files,
            )
            copied_roots[key] = copied
        resources.append(
            WorkspaceBundleResource(
                kind="data_asset",
                reference_id=reference_id,
                content_digest=digest,
                bundle_digest=copied[1],
                bundle_root=copied[0],
                primary_file=copied[2],
                files=copied[3],
            )
        )

    for row in packages:
        digest = str(row["manifest_digest"])
        key = ("model_package", digest)
        copied = copied_roots.get(key)
        if copied is None:
            safe_digest = digest.removeprefix("sha256:")
            relative_root = f"model-packages/sha256/{safe_digest}"
            destination = staging / relative_root
            _copy_package_verified(
                Path(str(row["locator"])), destination, digest
            )
            paths: list[str] = []
            for path in sorted(item for item in destination.rglob("*") if item.is_file()):
                relative = path.relative_to(staging).as_posix()
                archive_path = f"{RESOURCE_ARCHIVE_ROOT}/{relative}"
                file_records[archive_path] = _file_record(path, archive_path)
                paths.append(archive_path)
            records = tuple(file_records[path] for path in paths)
            copied = (
                f"{RESOURCE_ARCHIVE_ROOT}/{relative_root}",
                _resource_bundle_digest(records),
                None,
                tuple(paths),
            )
            copied_roots[key] = copied
        resources.append(
            WorkspaceBundleResource(
                kind="model_package",
                reference_id=str(row["id"]),
                content_digest=digest,
                bundle_digest=copied[1],
                bundle_root=copied[0],
                primary_file=None,
                files=copied[3],
            )
        )
    return (
        tuple(resources),
        tuple(file_records[path] for path in sorted(file_records)),
        tuple(sorted(set(warnings))),
    )


def _safe_backup_destination(
    raw: str | Path,
    *,
    database: Path,
    data_library_root: Path,
) -> Path:
    destination = Path(raw).expanduser().resolve()
    if destination.suffix.lower() != ".mdwb":
        destination = destination.with_suffix(".mdwb")
    for forbidden in (
        database.resolve(),
        data_library_root.resolve(),
        (database.parent / ".workspace-restore").resolve(),
    ):
        if destination == forbidden or forbidden in destination.parents:
            raise WorkspaceBundleError(
                "Backup destination cannot be inside the active Workspace"
            )
    destination.parent.mkdir(parents=True, exist_ok=True)
    return destination


def create_workspace_backup(
    *,
    database: Path,
    data_library_root: Path,
    destination: str | Path,
    app_version: str,
) -> WorkspaceBackupResult:
    target = _safe_backup_destination(
        destination,
        database=database,
        data_library_root=data_library_root,
    )
    temporary_target = target.with_name(f".{target.name}.{uuid4().hex}.partial")
    with tempfile.TemporaryDirectory(
        prefix="material-workbench-backup-", dir=target.parent
    ) as temporary:
        staging = Path(temporary)
        staged_database = staging / "workbench.db"
        _logical_database_backup(database, staged_database)
        migrations, packages, evidence, diagnostics = _database_evidence(
            staged_database
        )
        resources, resource_files, resource_warnings = _snapshot_resources(
            staged_database, staging / "resources"
        )
        database_record = _file_record(
            staged_database, DATABASE_ARCHIVE_PATH
        )
        manifest = WorkspaceBundleManifest(
            bundle_id=f"workspace-bundle-{uuid4()}",
            created_at=datetime.now(UTC),
            app_version=app_version,
            database=database_record,
            data_library_files=resource_files,
            schema_migrations=migrations,
            model_package_references=packages,
            bundled_resources=resources,
            table_counts={item.table: item.row_count for item in evidence},
            table_evidence=evidence,
            diagnostics=(
                *diagnostics,
                WorkspaceBundleDiagnostic(
                    id="cataloged-resources",
                    status="ok",
                    detail=f"{len(resources)} Data Asset/Package references included",
                ),
                WorkspaceBundleDiagnostic(
                    id="auxiliary-relative-files",
                    status="warning" if resource_warnings else "ok",
                    detail=(
                        "; ".join(resource_warnings[:20])
                        if resource_warnings
                        else "Declared relative evidence images are included"
                    ),
                ),
            ),
        )
        try:
            with zipfile.ZipFile(
                temporary_target,
                mode="w",
                compression=zipfile.ZIP_DEFLATED,
                compresslevel=6,
            ) as bundle:
                bundle.writestr(
                    MANIFEST_ARCHIVE_PATH,
                    manifest.model_dump_json(indent=2),
                )
                bundle.write(staged_database, DATABASE_ARCHIVE_PATH)
                for record in resource_files:
                    relative = PurePosixPath(record.path).relative_to(
                        RESOURCE_ARCHIVE_ROOT
                    )
                    bundle.write(
                        staging / "resources" / Path(relative.as_posix()),
                        record.path,
                    )
            inspected, _ = _inspect_bundle(temporary_target)
            inspected.close()
            os.replace(temporary_target, target)
        finally:
            temporary_target.unlink(missing_ok=True)
    return WorkspaceBackupResult(
        destination=str(target),
        manifest=manifest,
        bundle_sha256=_file_digest(target),
        size_bytes=target.stat().st_size,
    )


def _validate_archive_name(name: str) -> None:
    path = PurePosixPath(name)
    if (
        path.is_absolute()
        or ".." in path.parts
        or "\\" in name
        or re.match(r"^[A-Za-z]:", name)
        or ":" in name
    ):
        raise WorkspaceBundleError(f"Unsafe Workspace bundle entry: {name}")
    for part in path.parts:
        stem = part.rstrip(" .").split(".", 1)[0].upper()
        if stem in WINDOWS_RESERVED_NAMES:
            raise WorkspaceBundleError(
                f"Reserved Windows path in Workspace bundle: {name}"
            )


def _inspect_bundle(source: Path) -> tuple[zipfile.ZipFile, dict[str, zipfile.ZipInfo]]:
    try:
        bundle = zipfile.ZipFile(source, mode="r")
    except (OSError, zipfile.BadZipFile) as exc:
        raise WorkspaceBundleError(f"Workspace bundle cannot be opened: {exc}") from exc
    entries = bundle.infolist()
    names = [entry.filename for entry in entries]
    try:
        if len(entries) > MAX_BUNDLE_ENTRIES:
            raise WorkspaceBundleError("Workspace bundle has too many entries")
        if len(names) != len(set(names)):
            raise WorkspaceBundleError("Workspace bundle contains duplicate entries")
        if sum(entry.file_size for entry in entries) > MAX_BUNDLE_BYTES:
            raise WorkspaceBundleError("Workspace bundle is too large")
        for entry in entries:
            _validate_archive_name(entry.filename)
            if entry.file_size > MAX_ENTRY_BYTES:
                raise WorkspaceBundleError(
                    f"Workspace bundle entry is too large: {entry.filename}"
                )
            unix_mode = entry.external_attr >> 16
            if stat.S_ISLNK(unix_mode):
                raise WorkspaceBundleError(
                    f"Workspace bundle symlink is not allowed: {entry.filename}"
                )
            if (
                entry.file_size
                and entry.compress_size
                and entry.file_size / entry.compress_size > MAX_COMPRESSION_RATIO
            ):
                raise WorkspaceBundleError(
                    f"Workspace bundle compression ratio is unsafe: {entry.filename}"
                )
        manifest = next(
            (entry for entry in entries if entry.filename == MANIFEST_ARCHIVE_PATH),
            None,
        )
        if manifest is None or manifest.file_size > MAX_MANIFEST_BYTES:
            raise WorkspaceBundleError(
                "Workspace bundle manifest is missing or too large"
            )
        return bundle, {entry.filename: entry for entry in entries}
    except Exception:
        bundle.close()
        raise


def _stream_extract(
    bundle: zipfile.ZipFile,
    entries: dict[str, zipfile.ZipInfo],
    manifest: WorkspaceBundleManifest,
    destination: Path,
) -> None:
    records = (manifest.database, *manifest.data_library_files)
    expanded_bytes = sum(record.size_bytes for record in records)
    required_free = expanded_bytes + max(
        MIN_FREE_SPACE_RESERVE,
        expanded_bytes // 10,
    )
    available_free = shutil.disk_usage(destination).free
    if available_free < required_free:
        raise WorkspaceBundleError(
            "Workspace bundleの展開に必要な空き容量がありません"
        )
    expected = {MANIFEST_ARCHIVE_PATH, *(record.path for record in records)}
    if set(entries) != expected:
        raise WorkspaceBundleError(
            "Workspace bundle entries do not match the manifest"
        )
    for record in records:
        entry = entries.get(record.path)
        if entry is None or entry.file_size != record.size_bytes:
            raise WorkspaceBundleError(
                f"Workspace bundle size mismatch: {record.path}"
            )
        target = destination / Path(record.path)
        target.parent.mkdir(parents=True, exist_ok=True)
        digest = sha256()
        written = 0
        with bundle.open(entry, mode="r") as source, target.open("xb") as output:
            while chunk := source.read(1024 * 1024):
                written += len(chunk)
                if written > record.size_bytes:
                    raise WorkspaceBundleError(
                        f"Workspace bundle entry expanded beyond manifest: {record.path}"
                    )
                digest.update(chunk)
                output.write(chunk)
        if written != record.size_bytes or digest.hexdigest() != record.sha256:
            raise WorkspaceBundleError(
                f"Workspace bundle digest mismatch: {record.path}"
            )


def _validate_resource_manifest(manifest: WorkspaceBundleManifest) -> None:
    records = {record.path: record for record in manifest.data_library_files}
    covered: set[str] = set()
    references: set[tuple[str, str]] = set()
    for resource in manifest.bundled_resources:
        reference = (resource.kind, resource.reference_id)
        if reference in references:
            raise WorkspaceBundleError(
                f"Workspace bundle has duplicate resource reference: {reference}"
            )
        references.add(reference)
        if not resource.files:
            raise WorkspaceBundleError(
                f"Workspace bundle resource has no files: {resource.reference_id}"
            )
        root = PurePosixPath(resource.bundle_root)
        selected: list[WorkspaceBundleFile] = []
        for path in resource.files:
            candidate = PurePosixPath(path)
            try:
                candidate.relative_to(root)
            except ValueError as exc:
                raise WorkspaceBundleError(
                    f"Workspace resource file escapes its root: {path}"
                ) from exc
            record = records.get(path)
            if record is None:
                raise WorkspaceBundleError(
                    f"Workspace resource file is not indexed: {path}"
                )
            selected.append(record)
            covered.add(path)
        if resource.kind == "data_asset":
            if resource.primary_file is None or resource.primary_file not in resource.files:
                raise WorkspaceBundleError(
                    f"Data Asset primary file is invalid: {resource.reference_id}"
                )
        elif resource.primary_file is not None:
            raise WorkspaceBundleError(
                f"Model Package cannot declare a primary file: {resource.reference_id}"
            )
        if _resource_bundle_digest(tuple(selected)) != resource.bundle_digest:
            raise WorkspaceBundleError(
                f"Workspace resource inventory digest mismatch: {resource.reference_id}"
            )
    if covered != set(records):
        raise WorkspaceBundleError(
            "Workspace bundle contains resource files with no resource owner"
        )


def _extract_verified_bundle(
    source: Path,
    destination: Path,
) -> tuple[WorkspaceBundleManifest, str]:
    bundle, entries = _inspect_bundle(source)
    try:
        try:
            raw_manifest = bundle.read(MANIFEST_ARCHIVE_PATH)
            manifest = WorkspaceBundleManifest.model_validate_json(raw_manifest)
        except (KeyError, ValueError) as exc:
            raise WorkspaceBundleError(
                f"Unsupported or invalid Workspace bundle manifest: {exc}"
            ) from exc
        _validate_resource_manifest(manifest)
        _stream_extract(bundle, entries, manifest, destination)
        (destination / MANIFEST_ARCHIVE_PATH).write_bytes(raw_manifest)
        return manifest, sha256(raw_manifest).hexdigest()
    finally:
        bundle.close()


def _resource_by_reference(
    manifest: WorkspaceBundleManifest,
) -> dict[tuple[str, str], WorkspaceBundleResource]:
    return {
        (resource.kind, resource.reference_id): resource
        for resource in manifest.bundled_resources
    }


def _staged_resource_root(
    extracted: Path,
    resource: WorkspaceBundleResource,
) -> Path:
    return extracted / Path(resource.bundle_root)


def _rebind_staged_locators(
    database: Path,
    extracted: Path,
    manifest: WorkspaceBundleManifest,
) -> None:
    resources = _resource_by_reference(manifest)
    connection = connect_sqlite(database)
    try:
        connection.execute("BEGIN IMMEDIATE")
        for row in connection.execute("SELECT id FROM data_assets").fetchall():
            resource = resources.get(("data_asset", str(row["id"])))
            if resource is None or resource.primary_file is None:
                raise WorkspaceBundleError(
                    f"Managed Data Asset is missing from bundle: {row['id']}"
                )
            locator = extracted / Path(resource.primary_file)
            connection.execute(
                "UPDATE data_assets SET locator_kind='managed',locator=? WHERE id=?",
                (str(locator.resolve()), row["id"]),
            )
        for row in connection.execute("SELECT id FROM model_package_refs").fetchall():
            resource = resources.get(("model_package", str(row["id"])))
            if resource is None:
                raise WorkspaceBundleError(
                    f"Model Package is missing from bundle: {row['id']}"
                )
            connection.execute(
                "UPDATE model_package_refs SET locator=? WHERE id=?",
                (str(_staged_resource_root(extracted, resource).resolve()), row["id"]),
            )
        connection.commit()
    finally:
        connection.close()


def _validate_restored_references(
    database: Path,
    task_registry: TaskRegistry,
    transform_catalog: DeterministicTransformCatalog | None = None,
) -> tuple[WorkspaceBundleDiagnostic, ...]:
    catalog = WorkspaceCatalog(database)
    store = Store(database)
    validate_sqlite_foreign_keys(database)
    resolver = ProjectRuntimeResolver(catalog, task_registry)
    unresolved: list[str] = []
    package_refs = catalog.list_model_package_refs(include_archived=True)
    for project in store.list_projects(include_archived=True):
        identity = project.scientific_identity
        if identity.identity_kind == "chain":
            revision = store.get_chain_revision(identity.chain_revision_id)
            if revision is None:
                unresolved.append(
                    f"{project.id}: 固定されたChain Revisionがありません"
                )
                continue
            revision_payload = revision.model_dump(
                mode="json", exclude={"revision_digest"}
            )
            if (
                revision.revision_digest != identity.chain_revision_digest
                or revision.revision_digest != semantic_digest(revision_payload)
            ):
                unresolved.append(
                    f"{project.id}: Chain Revision digestが一致しません"
                )
                continue
            definition = store.get_chain_definition(
                revision.chain_id, revision.chain_definition_digest
            )
            if definition is None:
                unresolved.append(
                    f"{project.id}: 固定されたChain Definitionがありません"
                )
                continue
            surfaces = {}
            for stage in revision.stages:
                if stage.stage_kind == "deterministic_transform":
                    if transform_catalog is None:
                        unresolved.append(
                            f"{project.id}/{stage.stage_id}: "
                            "決定論的Transformの実行資源を検証できません"
                        )
                        continue
                    try:
                        transform_entry = transform_catalog.entry(stage.contract_id)
                        surface = welding_stage_a_surface(transform_entry.package)
                    except Exception as exc:
                        unresolved.append(
                            f"{project.id}/{stage.stage_id}: {exc}"
                        )
                        continue
                    surfaces[(stage.stage_kind, stage.contract_id)] = surface
                    actual_package = (
                        f"sha256:{transform_entry.package.manifest_sha256}"
                    )
                    if surface.contract_digest != stage.contract_digest:
                        unresolved.append(
                            f"{project.id}/{stage.stage_id}: "
                            "Transform contract digestが一致しません"
                        )
                    if actual_package != stage.package_manifest_digest:
                        unresolved.append(
                            f"{project.id}/{stage.stage_id}: "
                            "Transform Package digestが一致しません"
                        )
                    continue
                try:
                    contract = task_registry.contract_for(stage.contract_id)
                    entry = task_registry.entry_for(stage.contract_id)
                except Exception as exc:
                    unresolved.append(
                        f"{project.id}/{stage.stage_id}: {exc}"
                    )
                    continue
                actual_contract = semantic_digest(
                    contract.task_definition.model_dump(mode="json")
                )
                surfaces[(stage.stage_kind, stage.contract_id)] = (
                    task_contract_surface(
                        contract.task_definition,
                        contract_digest=actual_contract,
                    )
                )
                if actual_contract != stage.contract_digest:
                    unresolved.append(
                        f"{project.id}/{stage.stage_id}: Task contract digestが一致しません"
                    )
                if entry.package_digest != stage.package_manifest_digest:
                    unresolved.append(
                        f"{project.id}/{stage.stage_id}: Model Package digestが一致しません"
                    )
                matching_refs = [
                    ref
                    for ref in package_refs
                    if ref.task_id == stage.contract_id
                    and ref.task_contract_digest == stage.contract_digest
                    and f"sha256:{ref.manifest_digest}"
                    == stage.package_manifest_digest
                ]
                if len(matching_refs) != 1:
                    unresolved.append(
                        f"{project.id}/{stage.stage_id}: "
                        "固定されたModel Package参照を一意に解決できません"
                    )
                else:
                    ref = matching_refs[0]
                    try:
                        package = ModelPackageLoader().load(Path(ref.locator))
                    except Exception as exc:
                        unresolved.append(
                            f"{project.id}/{stage.stage_id}: "
                            f"Model Package本体を検証できません: {exc}"
                        )
                    else:
                        if (
                            package.manifest_sha256 != ref.manifest_digest
                            or package.manifest.task_id != ref.task_id
                            or package.manifest.package_id != ref.package_id
                        ):
                            unresolved.append(
                                f"{project.id}/{stage.stage_id}: "
                                "Model Package本体と参照が一致しません"
                            )
                view = catalog.get_dataset_view_revision(
                    stage.dataset_view_revision_id or "",
                    include_archived=True,
                )
                if view is None:
                    unresolved.append(
                        f"{project.id}/{stage.stage_id}: Dataset Viewがありません"
                    )
                    continue
                profile_digests: set[str] = set()
                for member in view.members:
                    dataset = catalog.get_dataset_revision(
                        member.dataset_revision_id,
                        include_archived=True,
                    )
                    profile = (
                        catalog.get_profile_revision(
                            dataset.profile_revision_id,
                            include_archived=True,
                        )
                        if dataset is not None
                        else None
                    )
                    if profile is not None:
                        profile_digests.add(profile.profile_digest)
                if profile_digests != {stage.dataset_profile_digest}:
                    unresolved.append(
                        f"{project.id}/{stage.stage_id}: Dataset Profile digestが一致しません"
                    )
            if len(surfaces) == len(revision.stages):
                try:
                    validate_chain_revision(
                        definition,
                        revision,
                        contracts=surfaces,
                    )
                except (KeyError, ValueError) as exc:
                    unresolved.append(
                        f"{project.id}: Chain Revision内部契約が不正です: {exc}"
                    )
            continue
        try:
            resolver.resolve(project)
        except Exception as exc:
            unresolved.append(f"{project.id}: {exc}")
    return (
        WorkspaceBundleDiagnostic(
            id="restored-fixed-references",
            status="warning" if unresolved else "ok",
            detail=(
                "; ".join(unresolved[:10])
                if unresolved
                else "All Project fixed references resolved"
            ),
        ),
    )


def _state_path(root: Path) -> Path:
    return root / "state.json"


def _write_state(root: Path, state: dict[str, object]) -> None:
    temporary = root / ".state.json.tmp"
    temporary.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.replace(temporary, _state_path(root))


def prepare_workspace_restore(
    *,
    database: Path,
    data_library_root: Path,
    source: str | Path,
    task_registry: TaskRegistry,
    transform_catalog: DeterministicTransformCatalog | None = None,
) -> WorkspaceRestorePrepared:
    token = uuid4().hex
    restore_root = database.parent / ".workspace-restore" / token
    next_root = restore_root / "next"
    next_root.mkdir(parents=True, exist_ok=False)
    source_path = Path(source).expanduser().resolve()
    try:
        manifest, manifest_digest = _extract_verified_bundle(
            source_path, next_root
        )
        staged_database = next_root / Path(manifest.database.path)
        _validate_migration_inventory(staged_database, manifest)
        # Store applies only the application's allow-listed migrations.
        Store(staged_database)
        migrated_inventory = _migration_inventory(staged_database)
        supported_inventory = _current_migration_inventory()
        if migrated_inventory != supported_inventory:
            missing = sorted(set(supported_inventory) - set(migrated_inventory))
            raise WorkspaceBundleError(
                "Workspace migration did not reach the current schema"
                + (f": missing {', '.join(missing)}" if missing else "")
            )
        _rebind_staged_locators(staged_database, next_root, manifest)
        reference_diagnostics = _validate_restored_references(
            staged_database,
            task_registry,
            transform_catalog,
        )
        _, _, evidence, database_diagnostics = _database_evidence(
            staged_database,
            expected_tables=manifest.table_evidence,
        )
        expected = {item.table: item for item in manifest.table_evidence}
        for item in evidence:
            before = expected[item.table]
            if item.row_count != before.row_count or item.digest != before.digest:
                raise WorkspaceBundleError(
                    f"Migration changed Workspace evidence in {item.table}"
                )
        staged_digest = _file_digest(staged_database)
        state: dict[str, object] = {
            "schema_version": "workspace-restore-state/v1",
            "token": token,
            "status": "prepared",
            "bundle_sha256": _file_digest(source_path),
            "manifest_sha256": manifest_digest,
            "staged_database_sha256": staged_digest,
            "database_archive_path": manifest.database.path,
            "expires_at": (
                datetime.now(UTC) + timedelta(hours=RESTORE_EXPIRY_HOURS)
            ).isoformat(),
        }
        _write_state(restore_root, state)
        return WorkspaceRestorePrepared(
            restore_token=token,
            manifest=manifest,
            migrated_database_sha256=staged_digest,
            diagnostics=(*database_diagnostics, *reference_diagnostics),
        )
    except Exception:
        shutil.rmtree(restore_root, ignore_errors=True)
        raise


def _restore_root(database: Path, token: str) -> Path:
    if len(token) != 32 or any(character not in "0123456789abcdef" for character in token):
        raise WorkspaceBundleError("Invalid restore token")
    root = database.parent / ".workspace-restore" / token
    if not root.is_dir():
        raise WorkspaceBundleError("Prepared restore was not found")
    return root


def _read_state(root: Path) -> dict[str, object]:
    try:
        state = json.loads(_state_path(root).read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        raise WorkspaceBundleError("Restore state is invalid") from exc
    if state.get("schema_version") != "workspace-restore-state/v1":
        raise WorkspaceBundleError("Restore state version is unsupported")
    return state


def _final_resource_path(
    data_library_root: Path,
    resource: WorkspaceBundleResource,
) -> Path:
    digest = resource.bundle_digest
    kind = "data-assets" if resource.kind == "data_asset" else "model-packages"
    return data_library_root / "by-digest" / kind / digest


def _install_resources(
    *,
    database: Path,
    data_library_root: Path,
    root: Path,
    manifest: WorkspaceBundleManifest,
) -> tuple[str, ...]:
    staged_database = root / "next" / Path(manifest.database.path)
    connection = connect_sqlite(staged_database)
    installed: list[str] = []
    try:
        connection.execute("BEGIN IMMEDIATE")
        for resource in manifest.bundled_resources:
            source_root = _staged_resource_root(root / "next", resource)
            final_root = _final_resource_path(data_library_root, resource)
            if not final_root.exists():
                final_root.parent.mkdir(parents=True, exist_ok=True)
                temporary = final_root.with_name(f".{final_root.name}.{uuid4().hex}.tmp")
                shutil.copytree(source_root, temporary)
                try:
                    os.replace(temporary, final_root)
                    installed.append(
                        final_root.relative_to(data_library_root).as_posix()
                    )
                finally:
                    shutil.rmtree(temporary, ignore_errors=True)
            if resource.kind == "data_asset":
                if resource.primary_file is None:
                    raise WorkspaceBundleError(
                        f"Data Asset primary file is missing: {resource.reference_id}"
                    )
                relative_file = PurePosixPath(resource.primary_file).relative_to(
                    PurePosixPath(resource.bundle_root)
                )
                locator = final_root / Path(relative_file.as_posix())
                if _file_digest(locator) != resource.content_digest.removeprefix("sha256:"):
                    raise WorkspaceBundleError(
                        f"Installed Data Asset digest mismatch: {resource.reference_id}"
                    )
                connection.execute(
                    "UPDATE data_assets SET locator_kind='managed',locator=? WHERE id=?",
                    (str(locator.resolve()), resource.reference_id),
                )
            else:
                verified = ModelPackageLoader().load(final_root)
                if verified.manifest_sha256 != resource.content_digest.removeprefix("sha256:"):
                    raise WorkspaceBundleError(
                        f"Installed Model Package digest mismatch: {resource.reference_id}"
                    )
                connection.execute(
                    "UPDATE model_package_refs SET locator=? WHERE id=?",
                    (str(final_root.resolve()), resource.reference_id),
                )
        connection.commit()
    finally:
        connection.close()
    validate_sqlite_foreign_keys(staged_database)
    return tuple(installed)


def _cleanup_installed_resources(
    data_library_root: Path,
    relative_roots: object,
) -> None:
    if not isinstance(relative_roots, list | tuple):
        return
    root = data_library_root.resolve()
    for raw in relative_roots:
        if not isinstance(raw, str):
            continue
        relative = PurePosixPath(raw)
        if relative.is_absolute() or ".." in relative.parts:
            continue
        target = (root / Path(relative.as_posix())).resolve()
        expected_parent = root / "by-digest"
        if expected_parent not in target.parents:
            continue
        shutil.rmtree(target, ignore_errors=True)
        parent = target.parent
        while parent != root and parent.exists() and not any(parent.iterdir()):
            parent.rmdir()
            parent = parent.parent


def commit_workspace_restore(
    *,
    database: Path,
    data_library_root: Path,
    restore_token: str,
) -> WorkspaceRestoreCommitResult:
    root = _restore_root(database, restore_token)
    state = _read_state(root)
    if state.get("status") != "prepared":
        raise WorkspaceBundleError("Restore is not in prepared state")
    expires_at = datetime.fromisoformat(str(state["expires_at"]))
    if datetime.now(UTC) >= expires_at:
        raise WorkspaceBundleError("Prepared restore has expired")
    manifest = WorkspaceBundleManifest.model_validate_json(
        (root / "next" / MANIFEST_ARCHIVE_PATH).read_text(encoding="utf-8")
    )
    staged_database = root / "next" / Path(str(state["database_archive_path"]))
    if _file_digest(staged_database) != state["staged_database_sha256"]:
        raise WorkspaceBundleError("Prepared restore database changed before commit")

    installed_resources: tuple[str, ...] = ()
    try:
        installed_resources = _install_resources(
            database=database,
            data_library_root=data_library_root,
            root=root,
            manifest=manifest,
        )
        # Locator rebinding intentionally changes only normalized operational paths.
        _, _, evidence, _ = _database_evidence(
            staged_database, expected_tables=manifest.table_evidence
        )
        expected = {item.table: item for item in manifest.table_evidence}
        for item in evidence:
            before = expected[item.table]
            if item.row_count != before.row_count or item.digest != before.digest:
                raise WorkspaceBundleError(
                    f"Commit changed Workspace evidence in {item.table}"
                )
    except Exception:
        _cleanup_installed_resources(data_library_root, installed_resources)
        raise

    rollback_database = root / "rollback-workbench.db"
    state["status"] = "committing"
    state["commit_database_sha256"] = _file_digest(staged_database)
    state["previous_database_sha256"] = (
        _file_digest(database) if database.exists() else None
    )
    state["installed_resource_roots"] = list(installed_resources)
    _write_state(root, state)
    moved_current = False
    installed_next = False
    try:
        if database.exists():
            os.replace(database, rollback_database)
            moved_current = True
        os.replace(staged_database, database)
        installed_next = True
        state["status"] = "committed"
        state["committed_at"] = datetime.now(UTC).isoformat()
        _write_state(root, state)
    except Exception as exc:
        if installed_next and database.exists():
            failed = root / "failed-workbench.db"
            os.replace(database, failed)
        if moved_current and rollback_database.exists():
            os.replace(rollback_database, database)
        _cleanup_installed_resources(data_library_root, installed_resources)
        state["status"] = "commit_failed"
        state["failure"] = str(exc)
        _write_state(root, state)
        raise WorkspaceBundleError(
            f"Workspace restore commit failed; current Workspace was preserved: {exc}"
        ) from exc
    return WorkspaceRestoreCommitResult(
        restore_token=restore_token,
        rollback_available=moved_current,
    )


def rollback_workspace_restore(
    *,
    database: Path,
    data_library_root: Path,
    restore_token: str,
) -> WorkspaceRestoreResolution:
    root = _restore_root(database, restore_token)
    state = _read_state(root)
    rollback_database = root / "rollback-workbench.db"
    if not rollback_database.exists():
        raise WorkspaceBundleError("Rollback database is unavailable")
    failed_database = root / "failed-workbench.db"
    if database.exists():
        os.replace(database, failed_database)
    os.replace(rollback_database, database)
    _cleanup_installed_resources(
        data_library_root,
        state.get("installed_resource_roots"),
    )
    shutil.rmtree(root, ignore_errors=True)
    return WorkspaceRestoreResolution(
        status="rolled_back", restore_token=restore_token
    )


def finalize_workspace_restore(
    *,
    database: Path,
    restore_token: str,
) -> WorkspaceRestoreResolution:
    root = _restore_root(database, restore_token)
    state = _read_state(root)
    if state.get("status") != "committed":
        raise WorkspaceBundleError("Only a committed restore can be finalized")
    shutil.rmtree(root)
    parent = root.parent
    if parent.exists() and not any(parent.iterdir()):
        parent.rmdir()
    return WorkspaceRestoreResolution(
        status="finalized", restore_token=restore_token
    )


def cancel_workspace_restore(
    *,
    database: Path,
    restore_token: str,
) -> WorkspaceRestoreResolution:
    root = _restore_root(database, restore_token)
    state = _read_state(root)
    if state.get("status") not in {"prepared", "commit_failed"}:
        raise WorkspaceBundleError(
            "Only a prepared or preserved failed restore can be cancelled"
        )
    shutil.rmtree(root)
    parent = root.parent
    if parent.exists() and not any(parent.iterdir()):
        parent.rmdir()
    return WorkspaceRestoreResolution(
        status="cancelled", restore_token=restore_token
    )


def recover_incomplete_workspace_restores(
    database: Path,
    data_library_root: Path | None = None,
) -> list[str]:
    restore_parent = database.parent / ".workspace-restore"
    recovered: list[str] = []
    if not restore_parent.exists():
        return recovered
    library_root = data_library_root or database.parent / "data-library"
    for root in sorted(path for path in restore_parent.iterdir() if path.is_dir()):
        try:
            state = _read_state(root)
        except WorkspaceBundleError:
            continue
        if state.get("status") in {"committing", "committed"}:
            rollback_database = root / "rollback-workbench.db"
            if rollback_database.exists():
                failed_database = root / "failed-workbench.db"
                if database.exists():
                    os.replace(database, failed_database)
                os.replace(rollback_database, database)
                _cleanup_installed_resources(
                    library_root,
                    state.get("installed_resource_roots"),
                )
                recovered.append(str(state.get("token", root.name)))
                shutil.rmtree(root, ignore_errors=True)
        elif state.get("status") == "prepared":
            try:
                expires_at = datetime.fromisoformat(str(state["expires_at"]))
            except (KeyError, ValueError):
                continue
            if datetime.now(UTC) >= expires_at:
                recovered.append(str(state.get("token", root.name)))
                shutil.rmtree(root, ignore_errors=True)
        elif state.get("status") == "commit_failed":
            _cleanup_installed_resources(
                library_root,
                state.get("installed_resource_roots"),
            )
            recovered.append(str(state.get("token", root.name)))
            shutil.rmtree(root, ignore_errors=True)
    if restore_parent.exists() and not any(restore_parent.iterdir()):
        restore_parent.rmdir()
    return recovered
