from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from material_workbench.data.profile_workbench import validate_workbook_profile
from material_workbench.data.profiles.canonicalization import canonicalize_workbook
from material_workbench.data.profiles.loading import load_dataset_profile
from material_workbench.data.profiles.requirements import task_data_requirements
from material_workbench.data.profiles.schema import DatasetInputProfile
from material_workbench.developer_experience.commands import (
    developer_command as command,
)
from material_workbench.developer_experience.schemas import (
    DataPurposeGuidance,
    InspectionDecision,
    ProfileCandidate,
    SourceInspection,
)

PROFILE_GLOB = "dataset-input-profile-*.json"
_UNIT_SUFFIX = re.compile(r"\s*\[([^\]]+)\]\s*$")


def _sha256(path: Path) -> str:
    with path.open("rb") as stream:
        return hashlib.file_digest(stream, "sha256").hexdigest()


def _workbook_headers(source: Path) -> dict[str, set[str]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        return {
            sheet.title: {
                str(value).strip()
                for value in next(sheet.iter_rows(values_only=True), ())
                if value is not None and str(value).strip()
            }
            for sheet in workbook.worksheets
        }
    finally:
        workbook.close()


def _profile_columns(profile: DatasetInputProfile) -> dict[str, set[str]]:
    columns: dict[str, set[str]] = defaultdict(set)
    optional_roles = set(profile.shared.optional_roles)
    optional_technical = set(profile.shared.optional_technical_fields)
    task_requirements = task_data_requirements(profile)

    def add(role: str, values: Iterable[str | None]) -> None:
        if role in optional_roles:
            return
        sheet = profile.sheet_for_role(role)
        columns[sheet].update(value for value in values if value)

    for entity in profile.shared.entities:
        add(entity.role, (entity.key,))
    for join in profile.shared.relation.joins:
        if task_requirements.requires_relation(join):
            add(profile.shared.relation.role, join.source_columns)
    for policy in profile.shared.eligibility:
        add(policy.role, (policy.column,))
    for technical in profile.shared.technical:
        if technical.name in optional_technical or f"{technical.role}.{technical.name}" in optional_technical:
            continue
        add(technical.role, (technical.column,))
    for task in profile.tasks.values():
        for mapping in task.mappings:
            add(mapping.role, (mapping.column,))
            if mapping.series_columns:
                add(
                    mapping.role,
                    (
                        mapping.series_columns.parent,
                        mapping.series_columns.order,
                        mapping.series_columns.time,
                        mapping.series_columns.value,
                    ),
                )
            for source in mapping.observation_sources:
                add(source.role, (source.column,))
            # Measurement-point fallback is an alternative route, not a
            # requirement when a complete history series is already present.
        for observation in task.observations:
            required_metadata = (
                column
                for key, column in observation.metadata_columns.items()
                if key not in observation.optional_metadata_keys
            )
            required_measurements = (
                column
                for target in (*observation.targets, *observation.auxiliary)
                if target.key not in observation.optional_auxiliary_keys
                for column in target.source_columns
            )
            add(
                observation.role,
                (
                    observation.id_column,
                    observation.parent_column,
                    *required_metadata,
                    *required_measurements,
                ),
            )
    return dict(columns)


def _base_header(header: str) -> str:
    return _UNIT_SUFFIX.sub("", header).replace(" ", "").lower()


def _decision(
    decision: str,
    reason: str,
    *,
    evidence: Iterable[str] = (),
    confidence: str = "medium",
) -> InspectionDecision:
    return InspectionDecision.model_validate({
        "decision": decision,
        "reason": reason,
        "confidence": confidence,
        "evidence": list(evidence),
    })


def inspect_source_against_profiles(
    source: Path,
    *,
    profile_path: Path | None = None,
    profile_directory: Path | None = None,
) -> SourceInspection:
    source = source.resolve()
    if not source.exists():
        raise FileNotFoundError(source)
    headers = _workbook_headers(source)
    profile_root = profile_directory or Path(__file__).resolve().parents[1] / "data"
    profile_paths = [
        path for path in sorted(profile_root.glob(PROFILE_GLOB))
        if not path.stem.endswith("-base")
    ]
    if profile_path is not None:
        explicit = profile_path.resolve()
        profile_paths = [explicit, *(path for path in profile_paths if path.resolve() != explicit)]

    candidates: list[ProfileCandidate] = []
    for path in profile_paths:
        profile = load_dataset_profile(path)
        expected = _profile_columns(profile)
        expected_sheets = {
            sheet
            for role, sheet in profile.shared.sheets.items()
            if role not in profile.shared.optional_roles
        }
        actual_sheets = set(headers)
        missing_sheets = sorted(expected_sheets - actual_sheets)
        extra_sheets = sorted(actual_sheets - expected_sheets)
        missing_columns: dict[str, list[str]] = {}
        extra_columns: dict[str, list[str]] = {}
        possible_units: list[str] = []
        matched_columns = 0
        total_columns = sum(len(values) for values in expected.values())
        for sheet, wanted in expected.items():
            actual = headers.get(sheet, set())
            missing = sorted(wanted - actual)
            if missing:
                missing_columns[sheet] = missing
                by_base = {_base_header(column): column for column in actual}
                for column in missing:
                    alternate = by_base.get(_base_header(column))
                    if alternate and alternate != column:
                        possible_units.append(f"{sheet}: {column} ↔ {alternate}")
            matched_columns += len(wanted & actual)
            unused = sorted(actual - wanted)
            if unused:
                extra_columns[sheet] = unused
        sheet_score = len(expected_sheets & actual_sheets) / max(len(expected_sheets), 1)
        column_score = matched_columns / max(total_columns, 1)
        score = round((sheet_score * 0.4 + column_score * 0.6) * 100)
        validation_error = None
        if not missing_sheets and not missing_columns:
            try:
                validate_workbook_profile(source, path)
            except Exception as exc:  # diagnostic boundary: preserve the exact preflight failure
                validation_error = str(exc)
        candidates.append(ProfileCandidate(
            profile_id=profile.profile_id,
            profile_path=str(path.resolve()),
            score=score,
            task_ids=sorted(profile.tasks),
            missing_sheets=missing_sheets,
            extra_sheets=extra_sheets,
            missing_columns=missing_columns,
            extra_columns=extra_columns,
            possible_unit_differences=possible_units,
            validation_error=validation_error,
        ))
    candidates.sort(key=lambda item: (-item.score, item.profile_id))
    selected = candidates[0] if candidates else None
    if profile_path is not None:
        selected = next(item for item in candidates if Path(item.profile_path) == profile_path.resolve())
    ambiguous = (
        profile_path is None
        and len(candidates) > 1
        and candidates[0].score == candidates[1].score
    )
    counts: dict[str, object] = {}
    learning_counts: dict[str, int] = {}
    output_counts: dict[str, int] = {}
    structural_differences: dict[str, list[str]] = {
        "relation": [],
        "keys": [],
        "cardinality": [],
    }
    if selected and not selected.missing_sheets and not selected.missing_columns:
        try:
            counts = validate_workbook_profile(source, Path(selected.profile_path))
            selected_profile = load_dataset_profile(Path(selected.profile_path))
            workbook = load_workbook(source, read_only=True, data_only=True)
            try:
                canonical = canonicalize_workbook(workbook, selected_profile)
            finally:
                workbook.close()
            for observation in canonical.observations:
                if all(observation.policy_results.values()):
                    learning_counts[observation.task_id] = learning_counts.get(observation.task_id, 0) + 1
                for output in observation.canonical_measurements:
                    output_counts[output] = output_counts.get(output, 0) + 1
        except Exception as exc:
            message = str(exc)
            if "relation" in message.lower():
                structural_differences["relation"].append(message)
            if "key" in message.lower() or "参照" in message:
                structural_differences["keys"].append(message)
            if "cardinal" in message.lower() or "一意" in message:
                structural_differences["cardinality"].append(message)
    if selected:
        relation_sheet = next(
            (
                sheet for sheet in selected.missing_sheets
                if "relation" in sheet.lower() or "関係" in sheet
            ),
            None,
        )
        if relation_sheet:
            structural_differences["relation"].append(f"relation sheet missing: {relation_sheet}")
        for sheet, columns in selected.missing_columns.items():
            key_like = [column for column in columns if "key" in column.lower() or "id" in column.lower() or "キー" in column]
            if key_like:
                structural_differences["keys"].append(f"{sheet}: {', '.join(key_like)}")
    profile_difference = bool(
        selected is None
        or selected.missing_sheets
        or selected.missing_columns
        or selected.validation_error
    )
    difference_evidence = []
    if selected:
        difference_evidence.extend(f"不足シート: {item}" for item in selected.missing_sheets)
        difference_evidence.extend(
            f"不足列 {sheet}: {', '.join(columns)}"
            for sheet, columns in selected.missing_columns.items()
        )
        difference_evidence.extend(f"単位差候補: {item}" for item in selected.possible_unit_differences)
        difference_evidence.extend(
            f"未知の追加列 {sheet}: {', '.join(columns)}"
            for sheet, columns in selected.extra_columns.items()
        )
        if selected.validation_error:
            difference_evidence.append(f"Profile検証: {selected.validation_error}")
    if ambiguous:
        profile_decision = _decision(
            "review_required",
            "Profile候補が同点のため自動選択できません。",
            evidence=[candidate.profile_id for candidate in candidates[:2]],
            confidence="high",
        )
    elif profile_difference:
        profile_decision = _decision(
            "yes",
            "既存Profileの必須構造では、このExcelをそのまま正規化できません。",
            evidence=difference_evidence,
            confidence="high",
        )
    else:
        profile_decision = _decision(
            "no",
            "選択した既存Profileの必須シート・列と検証契約に一致します。",
            evidence=[selected.profile_id] if selected else [],
            confidence="high",
        )
    semantic_differences = bool(
        ambiguous
        or profile_difference
        or (selected and selected.possible_unit_differences)
    )
    task_decision = _decision(
        "review_required" if selected and selected.task_ids else "no",
        (
            "対応Task候補はありますが、入力・目的変数・学習単位の意味が同じかはExcel構造だけでは確定できません。"
            if selected and selected.task_ids
            else "対応するTask候補がありません。"
        ),
        evidence=selected.task_ids if selected else [],
        confidence="medium" if selected and selected.task_ids else "high",
    )
    feature_decision = _decision(
        "review_required" if semantic_differences else "no",
        (
            "列・単位・Profile差分があり、canonical入力の意味が変わるか確認が必要です。"
            if semantic_differences
            else "既存Profileの必須構造に一致しており、行追加だけならFeature Pipeline変更は不要です。"
        ),
        evidence=difference_evidence,
        confidence="medium" if semantic_differences else "high",
    )
    package_decision = _decision(
        "review_required",
        "Application Datasetとして参照するだけなら再学習不要です。学習データを変更する場合だけ新しいPackageが必要です。",
        evidence=[
            f"学習可能な観測行: {sum(learning_counts.values())}",
            f"出力値を持つ観測: {sum(output_counts.values())}",
        ],
        confidence="high",
    )
    new_task_decision = _decision(
        "review_required" if semantic_differences else "no",
        (
            "未知列や構造差分が新しい目的変数・学習単位を表す可能性があります。"
            if semantic_differences
            else "既存Task候補の構造に一致しています。新Taskが必要という証拠はありません。"
        ),
        evidence=difference_evidence,
        confidence="medium",
    )
    recommendations = []
    if ambiguous:
        recommendations.append("候補Profileが同点です。--profile で明示してください。")
    if profile_difference:
        recommendations.append("既存Profileとの差分を確認し、元ExcelではなくProfile側で吸収できるか判断してください。")
    else:
        recommendations.append("既存Profileを再利用できます。行追加だけならTaskDefinitionや特徴量コードの変更は不要です。")
    recommendations.append("このデータを「参照・探索」「モデル学習」「候補入力」のどれに使うか先に決めてください。")
    return SourceInspection(
        source=str(source),
        source_sha256=_sha256(source),
        selected_profile=selected.profile_path if selected else None,
        ambiguous=ambiguous,
        candidates=candidates,
        canonical_counts=counts,
        learning_counts=learning_counts,
        output_counts=output_counts,
        structural_differences=structural_differences,
        decisions={
            "new_profile_required": profile_decision,
            "reuse_task_definition": task_decision,
            "feature_pipeline_change_required": feature_decision,
            "model_package_rebuild_required": package_decision,
            "new_task_may_be_required": new_task_decision,
        },
        data_purposes=[
            DataPurposeGuidance(
                id="application",
                label="Projectで参照・探索するデータ",
                description="Data Libraryへ登録し、Projectの固定参照や類似実績の確認に使います。",
                package_rebuild="no",
            ),
            DataPurposeGuidance(
                id="training",
                label="モデルを学習するデータ",
                description="学習行や出力値を変えるため、新しいModel Packageと検証が必要です。",
                package_rebuild="yes",
            ),
            DataPurposeGuidance(
                id="candidate_input",
                label="候補入力として使うデータ",
                description="TaskDefinitionとFeature Pipelineのcanonical入力へ変換できるか確認します。",
                package_rebuild="review_required",
            ),
        ],
        recommendations=recommendations,
        commands=[
            command(
                "uv",
                [
                    "run", "python", "backend/scripts/operations/profile_workbench.py", "inspect", str(source),
                    *([] if not selected else ["--profile", selected.profile_path]),
                ],
            ),
            *(
                [
                    command(
                        "uv",
                        ["run", "python", "backend/scripts/operations/profile_workbench.py", "validate", str(source), "--profile", selected.profile_path],
                    ),
                    command(
                        "uv",
                        [
                            "run", "python", "backend/scripts/operations/profile_workbench.py", "register", str(source),
                            "--profile", selected.profile_path,
                            "--database", "data/workbench.db",
                            "--library", "data/data-library",
                        ],
                    ),
                ]
                if selected else []
            ),
            command(
                "npm",
                ["run", "model:build", "--", "--task", "<task-id>", "--source", "<source>", "--output", "models/packages/<new-id>"],
            ),
            command(
                "npm",
                ["run", "model:verify", "--", "--task", "<task-id>", "--source", "<source>", "--package", "models/packages/<new-id>"],
            ),
        ],
    )
