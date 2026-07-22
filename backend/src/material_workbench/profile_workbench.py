"""Read-only developer tooling for inspecting workbook Profiles."""
from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .dataset_profile import canonicalize_workbook, load_dataset_profile
from .dataset_registration import file_sha256
from .importer import detect_dataset_profile_path
from .model_lifecycle import dataset_profile_digest


def inspect_workbook(source: Path, profile_path: Path | None = None) -> dict[str, Any]:
    """Return a compact workbook inventory and canonicalization preflight."""

    source = source.resolve()
    selected = profile_path.resolve() if profile_path else None
    profile_error: str | None = None
    if selected is None:
        try:
            selected = detect_dataset_profile_path(source).resolve()
        except (OSError, ValueError) as exc:
            profile_error = str(exc)

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value) if value is not None else "" for value in next(rows, ())]
            nonempty_rows = sum(1 for row in rows if any(value is not None for value in row))
            sheets.append({"name": sheet.title, "headers": headers, "rows": nonempty_rows})

        result: dict[str, Any] = {
            "source": str(source),
            "source_sha256": file_sha256(source),
            "sheets": sheets,
            "profile": str(selected) if selected else None,
            "profile_error": profile_error,
            "canonicalization": None,
        }
        if selected is not None:
            result["canonicalization"] = validate_workbook_profile(source, selected)
        return result
    finally:
        workbook.close()


def validate_workbook_profile(source: Path, profile_path: Path) -> dict[str, Any]:
    """Validate one explicit effective Profile without runtime/model assumptions."""

    source = source.resolve()
    profile_path = profile_path.resolve()
    profile = load_dataset_profile(profile_path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        canonical = canonicalize_workbook(workbook, profile)
    finally:
        workbook.close()
    observations_by_task = Counter(item.task_id for item in canonical.observations)
    return {
        "ok": True,
        "registration_ready": True,
        "source": str(source),
        "source_sha256": file_sha256(source),
        "profile": str(profile_path),
        "profile_id": profile.profile_id,
        "profile_digest": dataset_profile_digest(profile_path),
        "task_ids": sorted(profile.tasks),
        "entities": len(canonical.entities),
        "relations": len(canonical.relations),
        "observations": len(canonical.observations),
        "observations_by_task": dict(sorted(observations_by_task.items())),
        "heat_series_parents": len(canonical.heat_series),
    }
