"""Profile-driven Stage B training rows for the welding multistage example.

The weld-metal analysis is the observation grain. ``relationEx`` is only an
index: repeated test relations must never multiply the 300 weld-run rows.
"""
from __future__ import annotations

import hashlib
import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, model_validator

from material_workbench.modeling.tabular_regression import (
    TabularData,
    TabularDatasetProfile,
)


def _digest(value: Any) -> str:
    encoded = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return "sha256:" + hashlib.sha256(encoded).hexdigest()


class StageBProfileModel(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)


class StageBField(StageBProfileModel):
    path: str
    role: Literal["welding_condition"]
    column: str
    kind: Literal["number", "categorical"]
    source_unit: str = ""
    canonical_unit: str = ""
    choices: tuple[str, ...] = ()

    @model_validator(mode="after")
    def choices_match_kind(self) -> "StageBField":
        if (self.kind == "categorical") != bool(self.choices):
            raise ValueError("categorical Stage B fields require choices")
        return self


class StageBWorkbookProfile(StageBProfileModel):
    schema_version: Literal["welding-stage-b-profile/v1"]
    id: str
    task_id: str
    transform_id: str
    transform_version: str
    folds: int = Field(ge=2, le=10)
    sheets: dict[
        Literal[
            "relation", "raw_material", "raw_composition", "hoop", "blend",
            "blend_line", "welding_condition", "weld_metal",
        ],
        str,
    ]
    keys: dict[str, str]
    raw_component_columns: dict[str, str]
    hoop_component_columns: dict[str, str]
    weld_output_columns: dict[str, str]
    welding_context: tuple[StageBField, ...]
    fill_ratio_column: str
    blend_ratio_column: str
    raw_group_column: str
    raw_d50_column: str
    d50_groups: tuple[str, ...]
    d50_default_um: float

    @model_validator(mode="after")
    def scientific_axes_are_complete(self) -> "StageBWorkbookProfile":
        if tuple(self.raw_component_columns) != tuple(self.hoop_component_columns):
            raise ValueError("raw-material and hoop component axes must match exactly")
        if len(self.weld_output_columns) != 16:
            raise ValueError("Stage B must declare the 16 weld-metal output axes")
        return self

    @property
    def profile_digest(self) -> str:
        return _digest(self.model_dump(mode="json"))

    @property
    def transform_digest(self) -> str:
        return _digest({
            "id": self.transform_id,
            "version": self.transform_version,
            "raw_components": self.raw_component_columns,
            "hoop_components": self.hoop_component_columns,
            "fill_ratio_column": self.fill_ratio_column,
            "blend_ratio_column": self.blend_ratio_column,
            "d50_groups": self.d50_groups,
            "d50_default_um": self.d50_default_um,
        })


@dataclass(frozen=True)
class StageBTrainingData:
    data: TabularData
    profile_digest: str
    transform_digest: str
    cohort_digests: dict[str, str]
    fold_digests: dict[str, str]
    missing_by_target: dict[str, int]


def load_stage_b_profile(path: str | Path) -> StageBWorkbookProfile:
    return StageBWorkbookProfile.model_validate_json(
        Path(path).read_text(encoding="utf-8")
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _table(workbook: Any, name: str) -> tuple[tuple[str, ...], list[dict[str, Any]]]:
    worksheet = workbook[name]
    values = worksheet.iter_rows(values_only=True)
    headers = tuple(str(value).strip() for value in next(values))
    if any(not value for value in headers) or len(headers) != len(set(headers)):
        raise ValueError(f"{name}: headers must be non-empty and unique")
    rows = [
        dict(zip(headers, row, strict=False))
        for row in values
        if any(value is not None and value != "" for value in row)
    ]
    return headers, rows


def _key(value: Any) -> str | None:
    text = "" if value is None else str(value).strip()
    return text or None


def _number(value: Any, *, label: str) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label}: finite number is required") from exc
    if not math.isfinite(result):
        raise ValueError(f"{label}: finite number is required")
    return result


def _index(rows: list[dict[str, Any]], column: str, *, label: str) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = _key(row.get(column))
        if key is None:
            continue
        if key in result:
            raise ValueError(f"{label}: duplicate key {key}")
        result[key] = row
    return result


def _runtime_profile(profile: StageBWorkbookProfile) -> TabularDatasetProfile:
    inputs = [
        {
            "path": f"composition.{component}",
            "column": column,
            "kind": "number",
            "unit": "mass% whole wire",
            "transform": "linear",
        }
        for component, column in profile.raw_component_columns.items()
    ]
    inputs.extend({
        "path": field.path,
        "column": field.column,
        "kind": field.kind,
        "unit": field.canonical_unit,
        "choices": field.choices,
        "transform": "linear" if field.kind == "number" else "quadratic",
    } for field in profile.welding_context)
    inputs.append({
        "path": "process.alloy_powder_d50_um",
        "column": profile.raw_d50_column,
        "kind": "number",
        "unit": "µm",
        "transform": "linear",
    })
    return TabularDatasetProfile.model_validate({
        "schema_version": "tabular-dataset-profile/v1",
        "profile_id": profile.id,
        "name": "溶接材料 Stage B：溶着金属成分",
        "task_id": profile.task_id,
        "package_id": "welding-consumable-stage-b-ridge-v1",
        "id_column": profile.keys["weld_metal"],
        "group_column": profile.keys["weld_run"],
        "model_family": "ridge",
        "ridge_alpha": 2.0,
        "inputs": inputs,
        "outputs": [
            {
                "key": key,
                "column": column,
                "unit": "mass% deposited metal",
                "lower_bound": 0,
            }
            for key, column in profile.weld_output_columns.items()
        ],
    })


def build_stage_b_training_data(
    source: str | Path,
    profile: StageBWorkbookProfile,
) -> StageBTrainingData:
    source_path = Path(source)
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        tables = {
            role: _table(workbook, sheet)
            for role, sheet in profile.sheets.items()
        }
    finally:
        workbook.close()

    def rows(role: str) -> list[dict[str, Any]]:
        return tables[role][1]

    raw = _index(rows("raw_material"), profile.keys["raw_material"], label="raw material")
    raw_composition = _index(
        rows("raw_composition"), profile.keys["raw_material"], label="raw composition"
    )
    hoops = _index(rows("hoop"), profile.keys["hoop"], label="hoop")
    blends = _index(rows("blend"), profile.keys["blend"], label="blend")
    conditions = _index(
        rows("welding_condition"), profile.keys["welding_condition"], label="welding condition"
    )
    weld_metals = _index(
        rows("weld_metal"), profile.keys["weld_metal"], label="weld metal"
    )
    lines_by_blend: dict[str, list[dict[str, Any]]] = {}
    for row in rows("blend_line"):
        blend_key = _key(row.get(profile.keys["blend"]))
        if blend_key:
            lines_by_blend.setdefault(blend_key, []).append(row)

    # relationEx may repeat once per downstream specimen. Collapse only when the
    # scientific upstream identity is exactly the same.
    relation_by_weld: dict[str, set[tuple[str, str, str, str]]] = {}
    for relation in rows("relation"):
        weld_key = _key(relation.get(profile.keys["weld_metal"]))
        identity = tuple(
            _key(relation.get(profile.keys[key])) or ""
            for key in ("blend", "hoop", "welding_condition", "weld_run")
        )
        if weld_key:
            relation_by_weld.setdefault(weld_key, set()).add(identity)

    runtime_profile = _runtime_profile(profile)
    observations: list[dict[str, Any]] = []
    for weld_key, weld_row in weld_metals.items():
        identities = relation_by_weld.get(weld_key, set())
        reasons: list[str] = []
        if len(identities) != 1:
            reasons.append(
                "upstream relation missing" if not identities else "conflicting upstream relations"
            )
            blend_key = hoop_key = condition_key = run_key = ""
        else:
            blend_key, hoop_key, condition_key, run_key = next(iter(identities))
        blend = blends.get(blend_key)
        hoop = hoops.get(hoop_key)
        condition = conditions.get(condition_key)
        blend_lines = lines_by_blend.get(blend_key, [])
        if blend is None:
            reasons.append("blend entity missing")
        if hoop is None:
            reasons.append("hoop entity missing")
        if condition is None:
            reasons.append("welding condition entity missing")
        if not blend_lines:
            reasons.append("blend line missing")

        composition: dict[str, float] = {}
        process: dict[str, float] = {}
        categorical: dict[str, str] = {}
        if not reasons:
            assert blend is not None and hoop is not None and condition is not None
            fill = _number(blend[profile.fill_ratio_column], label="fill ratio") / 100
            core = {component: 0.0 for component in profile.raw_component_columns}
            ratio_total = 0.0
            d50_numerator = d50_denominator = 0.0
            for line in blend_lines:
                material_key = _key(line.get(profile.keys["raw_material"]))
                if material_key not in raw or material_key not in raw_composition:
                    raise ValueError(f"unknown raw material in blend {blend_key}: {material_key}")
                ratio = _number(line[profile.blend_ratio_column], label="blend ratio")
                ratio_total += ratio
                for component, column in profile.raw_component_columns.items():
                    core[component] += ratio / 100 * _number(
                        raw_composition[material_key][column],
                        label=f"{material_key}.{column}",
                    )
                if str(raw[material_key][profile.raw_group_column]) in profile.d50_groups:
                    d50_numerator += ratio * _number(
                        raw[material_key][profile.raw_d50_column],
                        label=f"{material_key}.D50",
                    )
                    d50_denominator += ratio
            if not math.isclose(ratio_total, 100.0, abs_tol=1e-6):
                reasons.append("blend ratio total is not 100%")
            for component, column in profile.hoop_component_columns.items():
                composition[component] = (
                    fill * core[component]
                    + (1 - fill) * _number(hoop.get(column, 0), label=f"hoop.{column}")
                )
            for field in profile.welding_context:
                raw_value = condition.get(field.column)
                if field.kind == "number":
                    process[field.path.split(".", 1)[1]] = _number(
                        raw_value, label=field.path
                    )
                else:
                    value = str(raw_value).strip()
                    if value not in field.choices:
                        reasons.append(f"unknown category: {field.path}={value}")
                    categorical[field.path.split(".", 1)[1]] = value
            process["alloy_powder_d50_um"] = (
                d50_numerator / d50_denominator
                if d50_denominator
                else profile.d50_default_um
            )

        outputs: dict[str, float] = {}
        for target, column in profile.weld_output_columns.items():
            value = weld_row.get(column)
            if value not in (None, ""):
                outputs[target] = _number(value, label=f"{weld_key}.{column}")
        observations.append({
            "id": weld_key,
            "parent_key": run_key,
            "task_id": profile.task_id,
            "composition": composition,
            "features": process,
            "categorical": categorical,
            "outputs": outputs,
            "eligible": not reasons,
            "exclusion_reasons": reasons,
            "run_context": {
                "curation": {"status": "accepted" if not reasons else "blocked"},
                "entity_keys": {
                    "blend": blend_key,
                    "hoop": hoop_key,
                    "welding_condition": condition_key,
                    "weld_run": run_key,
                    "weld_metal": weld_key,
                },
            },
        })

    eligible = [row for row in observations if row["eligible"]]
    numeric_keys = [
        item.path.split(".", 1)[1]
        for item in runtime_profile.inputs
        if item.kind == "number"
    ]
    medians = {
        key: float(sorted(
            (
                row["composition"].get(key)
                if key in row["composition"]
                else row["features"].get(key)
            )
            for row in eligible
        )[len(eligible) // 2])
        for key in numeric_keys
    }
    cohort_digests: dict[str, str] = {}
    fold_digests: dict[str, str] = {}
    missing_by_target: dict[str, int] = {}
    for target in profile.weld_output_columns:
        cohort = [
            (row["id"], row["parent_key"])
            for row in eligible
            if target in row["outputs"]
        ]
        assignment = {
            group: index % profile.folds
            for index, group in enumerate(sorted({group for _, group in cohort}))
        }
        cohort_digests[target] = _digest(cohort)
        fold_digests[target] = _digest(assignment)
        missing_by_target[target] = len(eligible) - len(cohort)

    data = TabularData(
        source_path=str(source_path),
        source_mtime_ns=source_path.stat().st_mtime_ns,
        source_sha256=_sha256(source_path),
        profile_path=f"catalog:{profile.id}",
        profile=runtime_profile,
        profile_id=profile.id,
        observations=observations,
        medians=medians,
        measurement_labels={
            key: key for key in profile.weld_output_columns
        },
        row_count=len(observations),
        quality=[],
        detected_quality=[],
        technical_columns={},
    )
    return StageBTrainingData(
        data=data,
        profile_digest=profile.profile_digest,
        transform_digest=profile.transform_digest,
        cohort_digests=cohort_digests,
        fold_digests=fold_digests,
        missing_by_target=missing_by_target,
    )
