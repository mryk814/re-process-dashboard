from __future__ import annotations

import hashlib
import math
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from statistics import fmean, median, pstdev
from typing import Any, Mapping

from openpyxl import load_workbook

from material_workbench.data.dataset_profile import CanonicalDataset, DatasetInputProfile, DatasetProfileError, canonicalize_workbook, load_dataset_profile
from material_workbench.contracts.task_contracts import TaskDefinition

def composition_names(
    profile: DatasetInputProfile | None = None,
    task_id: str | None = None,
) -> tuple[str, ...]:
    selected = profile or load_dataset_profile()
    definitions = (
        (selected.task_definitions[task_id],)
        if task_id is not None
        else tuple(selected.task_definitions.values())
    )
    names: list[str] = []
    for task in definitions:
        fields = (
            field for group in sorted(task.input_groups, key=lambda item: item.order)
            for field in sorted(group.fields, key=lambda item: item.order)
        )
        for field in fields:
            if field.path.startswith("composition."):
                name = field.path.removeprefix("composition.")
                if name not in names:
                    names.append(name)
    return tuple(names)


def _as_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)) and value > 30_000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    return str(value) if value is not None else None


def _records(sheet: Any) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    keys = [str(header) if header is not None else f"column_{i}" for i, header in enumerate(headers)]
    result: list[dict[str, Any]] = []
    for row in rows:
        record = {keys[i]: value for i, value in enumerate(row) if i < len(keys)}
        if any(value is not None for value in record.values()):
            result.append(record)
    return result


@dataclass(frozen=True)
class WorkbookData:
    source_path: str
    source_mtime_ns: int
    source_sha256: str
    profile_path: str
    profile: DatasetInputProfile
    sheets: dict[str, list[dict[str, Any]]]
    composition: dict[str, dict[str, float]]
    hot_rolling_features: dict[str, dict[str, Any]]
    anneal_features: dict[str, dict[str, Any]]
    lineage: dict[str, dict[str, list[str]]]
    observations: list[dict[str, Any]]
    quality: list[dict[str, Any]]
    detected_quality: list[dict[str, Any]]
    entities: dict[str, dict[str, dict[str, Any]]]
    medians: dict[str, float]
    entity_sheets: dict[str, str]
    key_to_sheet: dict[str, str]
    profile_id: str
    metadata_columns: tuple[str, ...]
    role_to_key: dict[str, str]
    role_to_sheet: dict[str, str]
    technical_columns: dict[tuple[str, str], str]
    policy_columns: dict[tuple[str, str], str]
    relation_sheet: str
    relation_join_columns: dict[str, tuple[str, ...]]
    lineage_stage_order: dict[str, int]
    lineage_adjacencies: tuple[tuple[str, str], ...]

    @property
    def source_summary(self) -> dict[str, Any]:
        return {
            "source": self.source_path,
            "source_sha256": self.source_sha256,
            "profile": self.profile_path,
            "profile_id": self.profile_id,
            "sheets": {name: len(rows) for name, rows in self.sheets.items()},
            "observations": len(self.observations),
            "quality_issues": len(self.quality),
            "detected_quality_issues": len(self.detected_quality),
        }


def lineage_reference_keys(data: WorkbookData, parent_key: str, process_role: str) -> dict[str, str | None]:
    relations = data.lineage.get(parent_key, {})

    def first_key(role: str) -> str | None:
        key_column = data.role_to_key.get(role)
        if not key_column:
            return None
        values = sorted(set(str(value) for value in relations.get(key_column, []) if value))
        return values[0] if len(values) == 1 else None

    return {
        "melt_key": first_key("melt"),
        "process_key": first_key(process_role),
        "process_label": "焼鈍履歴" if process_role == "annealing" else "熱延履歴" if process_role == "hot_rolling" else "工程履歴",
    }


def _scalar(value: Any) -> str | float | int | bool | None:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (str, float, int, bool)) or value is None:
        return value
    return str(value)


def _detect_data_quality(
    sheets: dict[str, list[dict[str, Any]]],
    entities: dict[str, dict[str, dict[str, Any]]],
    entity_sheets: Mapping[str, str] | None = None,
    key_to_sheet: Mapping[str, str] | None = None,
    observation_parent_columns: Mapping[str, str] | None = None,
    relation_sheet: str | None = None,
    relation_columns: tuple[str, ...] | None = None,
    relation_entity_keys: Mapping[str, str] | None = None,
) -> list[dict[str, str]]:
    """Detect structural data problems without trusting the workbook's scenarios.

    The workbook's 想定異常 sheet remains useful as a curated verification
    fixture, but it is not evidence that this copy of the source was checked.
    """
    profile = None
    if entity_sheets is None or observation_parent_columns is None or relation_sheet is None or relation_columns is None:
        profile = load_dataset_profile()
    if entity_sheets is None:
        assert profile is not None
        entity_sheets = {
            profile.sheet_for_role(entity.role): entity.key
            for entity in profile.shared.entities
        }
    if key_to_sheet is None:
        key_to_sheet = {key: sheet for sheet, key in entity_sheets.items()}
    if observation_parent_columns is None:
        assert profile is not None
        observation_parent_columns = {
            profile.sheet_for_role(observation.role): observation.parent_column
            for task in profile.tasks.values() for observation in task.observations
        }
    if relation_sheet is None or relation_columns is None:
        assert profile is not None
        relation_sheet = profile.sheet_for_role(profile.shared.relation.role)
        relation_columns = tuple(
            column
            for join in profile.shared.relation.joins
            for column in join.source_columns
        )
    relation_entity_keys = relation_entity_keys or {column: column for column in relation_columns}
    issues: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str]] = set()

    def add(issue_type: str, source_sheet: str, entity_key: str, detail: str) -> None:
        item = (issue_type, source_sheet, entity_key, detail)
        if item in seen:
            return
        seen.add(item)
        issues.append({
            "issue_id": f"{issue_type}:{source_sheet}:{entity_key or 'row'}:{len(issues) + 1}",
            "issue_type": issue_type,
            "source_sheet": source_sheet,
            "entity_key": entity_key,
            "detail": detail,
        })

    for sheet_name, key_column in entity_sheets.items():
        rows = sheets[sheet_name]
        counts: dict[str, int] = defaultdict(int)
        for row_index, row in enumerate(rows, start=2):
            value = row.get(key_column)
            if value is None or not str(value).strip():
                add("missing_key", sheet_name, "", f"{row_index}行目に{key_column}がありません")
                continue
            counts[str(value)] += 1
        for key, count in counts.items():
            if count > 1:
                add("duplicate_key", sheet_name, key, f"{key_column}が{count}回出現します")

    # A test row without its parent condition is unusable even when it is not
    # represented in relation, so detect that source-level missing reference.
    for sheet_name, parent_column in observation_parent_columns.items():
        for row_index, row in enumerate(sheets[sheet_name], start=2):
            if row.get(parent_column) is None or not str(row.get(parent_column)).strip():
                key_column = entity_sheets[sheet_name]
                add("missing_key", sheet_name, str(row.get(key_column) or ""), f"{row_index}行目に{parent_column}がありません")

    relation_keys: set[str] = set()
    for row_index, edge in enumerate(sheets[relation_sheet], start=2):
        for key_column, value in edge.items():
            if key_column not in relation_columns or value is None or not str(value).strip():
                continue
            key = str(value)
            relation_keys.add(key)
            entity_key_column = relation_entity_keys.get(key_column, key_column)
            source_sheet = key_to_sheet.get(entity_key_column)
            if source_sheet is None:
                continue
            if key not in entities.get(entity_key_column, {}):
                add("invalid_reference", relation_sheet, key, f"{row_index}行目の{key_column}に対応する{source_sheet}レコードがありません")

    for key_column, records in entities.items():
        source_sheet = key_to_sheet[key_column]
        for key in records:
            if key not in relation_keys:
                add("orphan_entity", source_sheet, key, "どのrelation行からも参照されていません")
    return issues


def _attach_quality_navigation(
    issues: list[dict[str, str]],
    lineage: Mapping[str, Mapping[str, list[str]]],
) -> list[dict[str, Any]]:
    """Describe where each structural finding can be investigated."""
    enriched: list[dict[str, Any]] = []
    for issue in issues:
        entity_key = issue["entity_key"]
        focus_entity_key = entity_key if entity_key and entity_key in lineage else None
        related_entity_keys = sorted({
            key
            for values in lineage.get(focus_entity_key or "", {}).values()
            for key in values
            if key != focus_entity_key
        })
        enriched.append({
            **issue,
            "focus_entity_key": focus_entity_key,
            "related_entity_keys": related_entity_keys,
            "missing_reference_key": entity_key if issue["issue_type"] == "invalid_reference" else None,
            "suggested_view": "lineage" if focus_entity_key else "source_sheet",
        })
    return enriched


def lineage_node_detail(data: WorkbookData, entity_key: str) -> dict[str, Any]:
    """Return inspectable source facts for a lineage node, not only its edges."""
    entity_type = next((key_column for key_column, records in data.entities.items() if entity_key in records), None)
    if entity_type is None:
        relations = data.lineage.get(entity_key)
        if relations is None:
            raise KeyError(entity_key)
        entity_type = next((column for column, values in relations.items() if entity_key in values), None)
        if entity_type not in data.key_to_sheet:
            raise KeyError(entity_key)
        source_sheet = data.key_to_sheet[entity_type]
        source_row: dict[str, Any] = {}
        missing_source = True
    else:
        source_sheet = data.key_to_sheet[entity_type]
        source_row = data.entities[entity_type][entity_key]
        missing_source = False
    relations = data.lineage.get(entity_key, {})
    melt_key_type = data.role_to_key["melt"]
    melt_keys = [entity_key] if entity_type == melt_key_type else relations.get(melt_key_type, [])
    composition = data.composition.get(melt_keys[0], {}) if len(melt_keys) == 1 else {}
    anneal_key_type = data.role_to_key["annealing"]
    anneal_keys = [entity_key] if entity_type == anneal_key_type else relations.get(anneal_key_type, [])
    heat_pattern: list[dict[str, Any]] = []
    if len(anneal_keys) == 1:
        heat_pattern = [dict(point) for point in data.anneal_features.get(anneal_keys[0], {}).get("heat_pattern", [])]
    connected_keys = {entity_key, *(key for values in relations.values() for key in values)}
    aggregate: dict[str, list[float]] = defaultdict(list)
    grouped_values: dict[tuple[str, str], list[float]] = defaultdict(list)
    grouped_observations: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    observation_count = 0
    connected_observations: list[dict[str, Any]] = []
    for observation in data.observations:
        if observation["id"] not in connected_keys and observation["parent_key"] not in connected_keys:
            continue
        observation_count += 1
        connected_observations.append({
            "id": observation["id"],
            "source": observation["source"],
            "parent_key": observation["parent_key"],
            "outputs": observation["outputs"],
        })
        for property_name, value in observation["outputs"].items():
            aggregate[property_name].append(float(value))
            group = (str(observation["source"]), property_name)
            grouped_values[group].append(float(value))
            grouped_observations[group].append(connected_observations[-1])
    property_summary = {
        property_name: {
            "count": len(values), "min": min(values), "mean": float(fmean(values)),
            "std": float(pstdev(values)), "median": float(median(values)), "max": max(values),
        }
        for property_name, values in sorted(aggregate.items())
    }
    observation_groups = [
        {
            "stage": "熱延後" if source.startswith("熱延") else "焼鈍後",
            "test_type": source,
            "property": property_name,
            "count": len(values),
            "min": min(values),
            "mean": float(fmean(values)),
            "std": float(pstdev(values)),
            "median": float(median(values)),
            "max": max(values),
            "observations": grouped_observations[(source, property_name)],
        }
        for (source, property_name), values in sorted(grouped_values.items())
    ]
    primary_conditions = {
        column: _scalar(value)
        for column, value in source_row.items()
        if column != entity_type and column not in data.key_to_sheet and column not in data.metadata_columns and value is not None
    }
    canonical_entity_type = data.key_to_sheet.get(entity_type, entity_type)
    return {
        "key": entity_key,
        "entity_type": canonical_entity_type,
        "source_sheet": source_sheet,
        "source_row": {column: _scalar(value) for column, value in source_row.items()},
        "primary_conditions": primary_conditions,
        "composition": composition,
        "heat_pattern": heat_pattern,
        "connected_observation_count": observation_count,
        "connected_observations": connected_observations,
        "observation_groups": observation_groups,
        "property_summary": property_summary,
        "related_entities": relations,
        "missing_source": missing_source,
    }


def lineage_neighborhood(data: WorkbookData, entity_key: str, max_nodes: int = 40) -> dict[str, Any]:
    """Build a bounded route graph while preserving evidence from relation rows."""
    if entity_key not in data.lineage:
        raise KeyError(entity_key)
    route_rows: list[tuple[int, dict[str, Any]]] = []
    relation_columns = {
        column
        for columns in data.relation_join_columns.values()
        for column in columns
    }
    for row_number, row in enumerate(data.sheets[data.relation_sheet], start=2):
        keys = {str(value) for column, value in row.items() if column in relation_columns and value}
        if entity_key in keys:
            route_rows.append((row_number, row))

    candidate_nodes: dict[str, str] = {}
    if not route_rows:
        own_type = next((column for column, records in data.entities.items() if entity_key in records), "")
        if own_type:
            candidate_nodes[entity_key] = own_type
    for _, row in route_rows:
        for column, value in row.items():
            if column in data.lineage_stage_order and value:
                candidate_nodes[str(value)] = column

    ordered = sorted(
        candidate_nodes.items(),
        key=lambda item: (item[0] != entity_key, data.lineage_stage_order.get(item[1], 99), item[0]),
    )
    node_limit = max(1, max_nodes)
    visible = dict(ordered[:node_limit])
    issue_by_key: dict[str, list[str]] = defaultdict(list)
    for issue in data.detected_quality:
        if issue["entity_key"] in visible and issue["issue_type"] not in issue_by_key[issue["entity_key"]]:
            issue_by_key[issue["entity_key"]].append(issue["issue_type"])
    nodes = [
        {
            "key": key,
            "entity_type": data.key_to_sheet.get(column, column),
            "source_sheet": data.key_to_sheet.get(column, column),
            "exists": key in data.entities.get(column, {}),
            "selected": key == entity_key,
            "issue_types": sorted(issue_by_key.get(key, [])),
        }
        for key, column in visible.items()
    ]
    edge_rows: dict[tuple[str, str], list[int]] = defaultdict(list)
    for row_number, row in route_rows:
        for source_column, target_column in data.lineage_adjacencies:
            source, target = row.get(source_column), row.get(target_column)
            if source and target and str(source) in visible and str(target) in visible:
                pair = (str(source), str(target))
                if row_number not in edge_rows[pair]:
                    edge_rows[pair].append(row_number)
        # A direct hot-roll → anneal edge is real only for relation rows where
        # cold rolling is absent. Never draw it as a parallel bypass around an
        # existing cold-roll condition.
        hot_key = data.role_to_key["hot_rolling"]
        anneal_key = data.role_to_key["annealing"]
        cold_key = data.role_to_key["cold_rolling"]
        if row.get(hot_key) and row.get(anneal_key) and not row.get(cold_key):
            pair = (str(row[hot_key]), str(row[anneal_key]))
            if pair[0] in visible and pair[1] in visible and row_number not in edge_rows[pair]:
                edge_rows[pair].append(row_number)
    return {
        "nodes": nodes,
        "edges": [
            {"source": source, "target": target, "route_rows": rows}
            for (source, target), rows in sorted(edge_rows.items())
        ],
        "relation_row_count": len(route_rows),
        "visible_node_count": len(visible),
        "total_node_count": len(ordered),
        "node_limit": node_limit,
        "has_more": len(visible) < len(ordered),
        "omitted_node_count": max(0, len(ordered) - len(visible)),
    }


def _normalize_stage_local_times(points: list[dict[str, Any]]) -> None:
    """Stitch stage-local clock resets without adding the new stage's initial timestamp."""
    stage_raw_origin = 0.0
    stage_normalized_origin = 0.0
    previous_raw = 0.0
    previous_normalized = -1.0
    for point in points:
        raw = point["time_s"]
        if previous_normalized >= 0 and raw < previous_raw:
            stage_raw_origin = raw
            stage_normalized_origin = previous_normalized + 1e-6
            point["segment_start"] = True
        normalized = stage_normalized_origin + (raw - stage_raw_origin)
        point["time_s"] = max(normalized, previous_normalized + 1e-6)
        previous_raw = raw
        previous_normalized = point["time_s"]


def _time_at_or_above(points: list[dict[str, Any]], threshold: float) -> float:
    total = 0.0
    for left, right in zip(points, points[1:]):
        if right.get("segment_start"):
            continue
        t0, t1 = float(left["time_s"]), float(right["time_s"])
        y0, y1 = float(left["temperature_c"]), float(right["temperature_c"])
        if y0 >= threshold and y1 >= threshold:
            total += t1 - t0
        elif (y0 >= threshold) != (y1 >= threshold) and y1 != y0:
            crossing = t0 + (threshold - y0) * (t1 - t0) / (y1 - y0)
            total += (t1 - crossing) if y1 >= threshold else (crossing - t0)
    return total


def _derived_anneal_feature_row(
    parent_key: str,
    mapped_process: dict[str, Any],
    points: list[dict[str, Any]],
) -> dict[str, Any]:
    finite_points = [
        point for point in points
        if isinstance(point.get("time_s"), (int, float))
        and isinstance(point.get("temperature_c"), (int, float))
    ]
    eligible = len(finite_points) >= 2 and mapped_process.get("ls_mpm") is not None
    peak = max((float(point["temperature_c"]) for point in finite_points), default=0.0)
    stage_names = [str(point.get("stage_name") or "") for point in finite_points]
    stage_categories = [str(point.get("stage_category") or "") for point in finite_points]
    return {
        **mapped_process,
        "max_temperature_c": peak,
        "hold_time_s": _time_at_or_above(finite_points, peak * 0.95) if eligible and peak else 0.0,
        "input_points": float(len(finite_points)),
        "reheat": 1.0 if "REHEAT" in stage_categories else 0.0,
        "alloying": 1.0 if any(category in {"ZINC_BATH", "ALLOYING"} for category in stage_categories) else 0.0,
        "feature_status": "特徴量化可" if eligible else "特徴量化不可",
        "feature_eligible": eligible,
        "standard_route": " > ".join(dict.fromkeys(category for category in stage_categories if category)),
        "process_signature": " > ".join(stage_names),
        "unmapped_stage_count": sum(not category for category in stage_categories),
        "heat_pattern": points,
        "parent_key": parent_key,
    }


def _line_speed_from_history(
    points: list[dict[str, Any]],
    master_rows: list[dict[str, Any]],
    *,
    equipment_column: str,
    equipment_value: str,
    stage_column: str,
    position_column: str,
) -> float | None:
    """Infer line speed from elapsed history and fixed measurement positions."""
    stage_positions = {
        str(row.get(stage_column)): float(position)
        for row in master_rows
        if str(row.get(equipment_column) or "") == equipment_value
        and isinstance((position := row.get(position_column)), (int, float))
        and math.isfinite(float(position))
    }
    positioned_points = sorted(
        (
            (float(point["time_s"]), stage_positions[str(point.get("stage_name"))])
            for point in points
            if isinstance(point.get("time_s"), (int, float))
            and math.isfinite(float(point["time_s"]))
            and str(point.get("stage_name")) in stage_positions
        ),
        key=lambda item: item[0],
    )
    speed_estimates = [
        60.0 * (right_position - left_position) / (right_time - left_time)
        for (left_time, left_position), (right_time, right_position)
        in zip(positioned_points, positioned_points[1:])
        if right_time > left_time and right_position > left_position
    ]
    if not speed_estimates:
        return None
    inferred = float(median(speed_estimates))
    return inferred if math.isfinite(inferred) and inferred > 0 else None


def detect_dataset_profile_path(
    source_path: str | Path,
    profile_directory: str | Path | None = None,
    task_definitions: Mapping[str, TaskDefinition] | None = None,
) -> Path:
    """Choose the most specific profile whose required sheets are present."""
    source_path = Path(source_path)
    profile_root = Path(profile_directory) if profile_directory else Path(__file__).parent
    candidates = sorted(profile_root.glob("dataset-input-profile-*.json"))
    if not candidates:
        raise DatasetProfileError([f"no dataset input profiles found in {profile_root}"])
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet_names = set(workbook.sheetnames)
        matches: list[tuple[tuple[int, int], Path]] = []
        diagnostics: list[str] = []
        for candidate in candidates:
            try:
                profile = load_dataset_profile(candidate, task_definitions)
            except DatasetProfileError as exc:
                diagnostics.append(f"{candidate.name}: profile invalid ({exc})")
                continue
            required_sheets = {
                sheet for role, sheet in profile.shared.sheets.items()
                if role not in profile.shared.optional_roles
            }
            missing = sorted(required_sheets - sheet_names)
            if missing:
                diagnostics.append(f"{candidate.name}: missing sheets {', '.join(missing)}")
                continue
            marker_failure = _source_marker_failure(workbook, profile)
            if marker_failure:
                diagnostics.append(f"{candidate.name}: {marker_failure}")
                continue
            matches.append(((len(required_sheets), len(profile.source_markers)), candidate))
    finally:
        workbook.close()
    if not matches:
        raise DatasetProfileError([
            f"no dataset input profile matches workbook {source_path}",
            *diagnostics,
        ])
    best_score = max(score for score, _ in matches)
    best = [candidate for score, candidate in matches if score == best_score]
    if len(best) > 1:
        raise DatasetProfileError([
            f"multiple dataset input profiles match workbook {source_path}; pass an explicit profile",
            *[candidate.name for candidate in sorted(best)],
        ])
    return best[0]


def _source_marker_failure(workbook: Any, profile: DatasetInputProfile) -> str | None:
    for marker in profile.source_markers:
        if marker.sheet not in workbook.sheetnames:
            return f"source marker sheet {marker.sheet!r} is missing"
        rows = workbook[marker.sheet].iter_rows(values_only=True)
        headers = tuple(str(value) if value is not None else "" for value in next(rows, ()))
        if marker.column not in headers:
            return f"source marker column {marker.column!r} is missing from {marker.sheet!r}"
        index = headers.index(marker.column)
        observed = {
            str(row[index])
            for row in rows
            if index < len(row) and row[index] is not None and str(row[index]).strip()
        }
        missing = sorted(set(marker.values) - observed)
        if missing:
            return f"source marker values missing from {marker.sheet!r}.{marker.column}: {', '.join(missing)}"
    return None


def load_workbook_data(
    path: str | Path,
    profile_path: str | Path | None = None,
    task_definitions: Mapping[str, TaskDefinition] | None = None,
    *,
    profile: DatasetInputProfile | None = None,
) -> WorkbookData:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel source not found: {path}")
    if profile is not None and profile_path is not None:
        raise ValueError("profileとprofile_pathは同時に指定できません")
    if profile is None:
        selected_profile_path = Path(profile_path) if profile_path else detect_dataset_profile_path(path, task_definitions=task_definitions)
        profile = load_dataset_profile(selected_profile_path, task_definitions)
        profile_locator = str(selected_profile_path)
    else:
        profile_locator = f"catalog:{profile.profile_id}"
    anneal_task_id = next(
        task_id for task_id, task in profile.tasks.items()
        if any(mapping.kind == "ordered_heat_series" for mapping in task.mappings)
    )
    hot_task = next(
        task_id for task_id, task in profile.tasks.items()
        if any(mapping.role == "hot_rolling" and mapping.path.startswith("process.") for mapping in task.mappings)
    )
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        canonical = canonicalize_workbook(wb, profile)
    finally:
        wb.close()
    sheets = dict(canonical.source_rows)
    entity_sheets = {
        profile.sheet_for_role(entity.role): entity.key
        for entity in profile.shared.entities
    }
    role_to_key = {entity.role: entity.key for entity in profile.shared.entities}
    role_to_sheet = {role: profile.sheet_for_role(role) for role in profile.shared.sheets}
    technical_columns = {(item.role, item.name): item.column for item in profile.shared.technical}
    policy_columns = {(item.role, item.policy): item.column for item in profile.shared.eligibility}
    key_to_sheet = {key: sheet for sheet, key in entity_sheets.items()}
    entities: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for sheet_name, key_column in entity_sheets.items():
        for row in sheets[sheet_name]:
            value = row.get(key_column)
            if value is not None and str(value).strip():
                # Keep the first row for node inspection. Duplicate rows are
                # independently surfaced by the quality detector below.
                entities[key_column].setdefault(str(value), row)
    melt_rows = canonical.rows("melt")
    composition_mappings: dict[str, tuple[str, str]] = {}
    for task_id, task in profile.tasks.items():
        for mapping in task.mappings:
            if mapping.path.startswith("composition."):
                composition_mappings.setdefault(mapping.path, (task_id, mapping.path))
    composition_paths = tuple(composition_mappings)
    melt_key = next(entity.key for entity in profile.shared.entities if entity.role == "melt")
    composition = {
        str(row[melt_key]): {
            path.removeprefix("composition."): float(value)
            for path in composition_paths
            if (value := canonical.value(row, *composition_mappings[path])) is not None
        }
        for row in melt_rows
    }
    hot_key = next(entity.key for entity in profile.shared.entities if entity.role == "hot_rolling")
    hot_rolling_features = {
        str(row[hot_key]): {
            **canonical.mapped_values(row, hot_task, ("process.", "categorical.")),
            "equipment": str(canonical.technical_value(row, "hot_rolling", "equipment") or ""),
        }
        for row in canonical.rows("hot_rolling")
    }
    anneal_key = next(entity.key for entity in profile.shared.entities if entity.role == "annealing")
    anneal_history_by_key: dict[str, list[dict[str, float]]] = defaultdict(list)
    anneal_history_by_key.update({identity[1]: [dict(point) for point in points] for identity, points in canonical.heat_series.items()})
    anneal_feature_sheet = profile.shared.sheets.get("anneal_features")
    if anneal_feature_sheet and anneal_feature_sheet in sheets:
        anneal_features = {
            str(canonical.technical_value(row, "anneal_features", "parent_key")): {
                **canonical.mapped_values(row, anneal_task_id, ("process.", "categorical.")),
                "max_temperature_c": float(canonical.technical_value(row, "anneal_features", "max_temperature_c")),
                "hold_time_s": float(canonical.technical_value(row, "anneal_features", "hold_time_s")),
                "input_points": float(canonical.technical_value(row, "anneal_features", "input_points")),
                "reheat": 1.0 if canonical.technical_value(row, "anneal_features", "reheat") == "あり" else 0.0,
                "alloying": 1.0 if canonical.technical_value(row, "anneal_features", "alloying") == "通過" else 0.0,
                "feature_status": str(canonical.technical_value(row, "anneal_features", "feature_status")),
                "feature_eligible": canonical.policy_allows(row, "anneal_features", "anneal_feature_status/v1"),
                "standard_route": str(canonical.technical_value(row, "anneal_features", "standard_route") or ""),
                "process_signature": str(canonical.technical_value(row, "anneal_features", "process_signature") or ""),
                "unmapped_stage_count": int(canonical.technical_value(row, "anneal_features", "unmapped_stage_count") or 0),
                "heat_pattern": anneal_history_by_key.get(str(canonical.technical_value(row, "anneal_features", "parent_key")), []),
            }
            for row in canonical.rows("anneal_features")
        }
    else:
        anneal_rows_by_key = {
            str(row[anneal_key]): row
            for row in canonical.rows("annealing")
            if row.get(anneal_key) is not None and str(row[anneal_key]).strip()
        }
        anneal_keys = set(anneal_rows_by_key) | set(anneal_history_by_key)
        history_mapping = next(
            mapping
            for mapping in profile.tasks[anneal_task_id].mappings
            if mapping.kind == "ordered_heat_series"
        )
        history_parent_column = (
            history_mapping.series_columns.parent
            if history_mapping.series_columns is not None else None
        )
        source_history_keys = {
            str(row[history_parent_column])
            for row in sheets.get(profile.sheet_for_role(history_mapping.role), [])
            if history_parent_column
            and row.get(history_parent_column) is not None
            and str(row[history_parent_column]).strip()
        }
        anneal_keys |= source_history_keys
        fallback = history_mapping.measurement_point_fallback
        anneal_features = {}
        for parent in sorted(anneal_keys):
            row = anneal_rows_by_key.get(parent)
            mapped_process = (
                canonical.mapped_values(row, anneal_task_id, ("process.", "categorical."))
                if row is not None else {}
            )
            points = anneal_history_by_key.get(parent, [])
            if mapped_process.get("ls_mpm") is None and fallback is not None:
                inferred_speed = _line_speed_from_history(
                    points,
                    sheets[profile.sheet_for_role(fallback.master_role)],
                    equipment_column=fallback.equipment_column,
                    equipment_value=fallback.equipment_value,
                    stage_column=fallback.stage_column,
                    position_column=fallback.position_column,
                )
                if inferred_speed is not None:
                    mapped_process["ls_mpm"] = inferred_speed
            anneal_features[parent] = _derived_anneal_feature_row(
                parent,
                mapped_process,
                points,
            )
    lineage: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    key_by_type = {entity.type: entity.key for entity in profile.shared.entities}
    for relation in canonical.relations:
        edge = {key_by_type[entity_type]: identity[1] for entity_type, identity in relation.items()}
        for column, value in edge.items():
            if value:
                lineage[str(value)][column].append(str(value))
        anchor = {key: str(value) for key, value in edge.items() if value}
        for current_key in anchor.values():
            bucket = lineage[current_key]
            for related_column, related_key in anchor.items():
                if related_key not in bucket[related_column]:
                    bucket[related_column].append(related_key)
    # Ensure every source entity can be inspected, including isolated tests and
    # process conditions. Relation references without source rows already exist
    # in lineage and are exposed as missing-source nodes.
    for records in entities.values():
        for key in records:
            lineage.setdefault(key, defaultdict(list))

    def upstream_composition(parent_key: str) -> dict[str, float] | None:
        melt_keys = sorted(set(lineage.get(parent_key, {}).get(melt_key, [])))
        return composition.get(melt_keys[0]) if len(melt_keys) == 1 else None

    anneal_status = {
        str(row[anneal_key]): canonical.policy_allows(row, "annealing", "learning_flag/v1")
        for row in canonical.rows("annealing")
    }
    anneal_condition_keys = set(anneal_status)
    hot_status = {
        str(row[hot_key]): canonical.policy_allows(row, "hot_rolling", "learning_flag/v1")
        for row in canonical.rows("hot_rolling")
    }
    observations: list[dict[str, Any]] = []
    for canonical_observation in canonical.observations:
            parent = canonical_observation.parent_key
            is_anneal = canonical_observation.task_id == anneal_task_id
            process = anneal_features.get(parent) if is_anneal else hot_rolling_features.get(parent)
            comp = upstream_composition(parent)
            canonical_output_labels = {
                output.key: output.measurement_keys[0]
                for output in profile.task_definitions[canonical_observation.task_id].outputs
            }
            measurement_labels = {
                target.key: canonical_output_labels.get(target.key, target.source_columns[0])
                for observation in profile.tasks[canonical_observation.task_id].observations
                for target in (*observation.targets, *observation.auxiliary)
            }
            outputs = {
                measurement_labels[key]: value
                for key, value in canonical_observation.canonical_measurements.items()
            }
            if not outputs:
                continue
            eligibility_reasons: list[str] = []
            if not process:
                eligibility_reasons.append("工程条件が見つかりません")
            if not comp:
                eligibility_reasons.append("上流の成分が一意に決まりません")
            if is_anneal:
                if process and not process["feature_eligible"]:
                    eligibility_reasons.append("焼鈍履歴を特徴量化できません")
                if parent in anneal_condition_keys and not anneal_status[parent]:
                    eligibility_reasons.append("焼鈍条件が学習対象外です")
                if not canonical_observation.policy_results.get("valid_observation/v1", False):
                    eligibility_reasons.append("試験判定が有効ではありません")
            else:
                if not hot_status.get(parent, False):
                    eligibility_reasons.append("熱延条件が学習対象外です")
                if not canonical_observation.policy_results.get("valid_observation/v1", False):
                    eligibility_reasons.append("試験判定が有効ではありません")
            physical_ranges = profile.shared.physical_ranges.get(canonical_observation.task_id, {})
            for property_name, value in canonical_observation.canonical_measurements.items():
                bounds = physical_ranges.get(property_name)
                if bounds and not bounds[0] <= value <= bounds[1]:
                    label = measurement_labels[property_name]
                    eligibility_reasons.append(f"{label}が物理範囲外です")
            observations.append({
                "id": canonical_observation.id, "task_id": canonical_observation.task_id, "source": profile.sheet_for_role(canonical_observation.source_role), "parent_key": parent,
                "features": process, "composition": comp, "outputs": outputs,
                "eligible": not eligibility_reasons,
                "eligibility_reasons": eligibility_reasons,
                "date": _as_date(canonical_observation.metadata.get("date")),
                "test_direction": str(direction) if not is_anneal and (direction := canonical_observation.metadata.get("direction")) else None,
            })
    values = {
        path: sorted(float(value) for row in melt_rows if isinstance((value := canonical.value(row, *composition_mappings[path])), (int, float)))
        for path in composition_paths
    }
    medians = {path.removeprefix("composition."): series[len(series) // 2] for path, series in values.items() if series}
    with path.open("rb") as source_file:
        source_sha256 = hashlib.file_digest(source_file, "sha256").hexdigest()
    normalized_lineage = {k: {inner: sorted(set(values)) for inner, values in v.items()} for k, v in lineage.items()}
    observation_parent_columns = {
        profile.sheet_for_role(observation.role): observation.parent_column
        for task in profile.tasks.values() for observation in task.observations
        if observation.parent_column is not None
    }
    relation_sheet = profile.sheet_for_role(profile.shared.relation.role)
    relation_join_columns = {
        join.entity_type: join.source_columns
        for join in profile.shared.relation.joins
    }
    relation_entity_keys = {
        column: next(entity.key for entity in profile.shared.entities if entity.type == join.entity_type)
        for join in profile.shared.relation.joins
        for column in join.source_columns
    }
    lineage_stage_order = {
        column: join.stage
        for join in profile.shared.relation.joins
        for column in join.source_columns
    }
    lineage_adjacencies = tuple(
        (parent_column, child_column)
        for join in profile.shared.relation.joins
        for parent_type in (
            join.edge_parent_entity_types
            if join.edge_parent_entity_types is not None
            else join.parent_entity_types
        )
        for parent_column in relation_join_columns[parent_type]
        for child_column in join.source_columns
    )
    detected_quality = _attach_quality_navigation(
        _detect_data_quality(
            sheets,
            entities,
            entity_sheets,
            key_to_sheet,
            observation_parent_columns,
            relation_sheet,
            tuple(
                column
                for join in profile.shared.relation.joins
                for column in join.source_columns
            ),
            relation_entity_keys,
        ),
        normalized_lineage,
    )
    return WorkbookData(
        source_path=str(path),
        source_mtime_ns=path.stat().st_mtime_ns,
        source_sha256=source_sha256,
        profile_path=profile_locator,
        profile=profile,
        sheets=sheets,
        composition=composition,
        hot_rolling_features=hot_rolling_features,
        anneal_features=anneal_features,
        lineage=normalized_lineage,
        observations=observations,
        quality=canonical.source_rows.get(profile.shared.sheets.get("quality", ""), []),
        detected_quality=detected_quality,
        entities=dict(entities),
        medians=medians,
        entity_sheets=entity_sheets,
        key_to_sheet=key_to_sheet,
        profile_id=profile.profile_id,
        metadata_columns=profile.shared.metadata_columns,
        role_to_key=role_to_key,
        role_to_sheet=role_to_sheet,
        technical_columns=technical_columns,
        policy_columns=policy_columns,
        relation_sheet=relation_sheet,
        relation_join_columns=relation_join_columns,
        lineage_stage_order=lineage_stage_order,
        lineage_adjacencies=lineage_adjacencies,
    )
