from __future__ import annotations

from collections import defaultdict
from copy import deepcopy
from io import BytesIO
from statistics import fmean, pstdev
from typing import Any, Callable

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from material_workbench.data.dataset_profile import DatasetInputProfile, load_dataset_profile
from material_workbench.modeling.hot_rolling_feature_pipeline import PROCESS_NAMES
from material_workbench.data.importer import WorkbookData, composition_names
from material_workbench.task_modules import BatchPredictionRuntime, PredictionRuntime
from material_workbench.tasks.task_registry import load_task_contracts
from material_workbench.contracts.schemas import (
    Candidate,
    CandidateInput,
    CandidateOriginEvidence,
    DEFAULT_SCREENING_SEED,
    HeatPoint,
    LineageCandidateOption,
    ScreeningGoal,
    ScreeningRequest,
)
from material_workbench.contracts.task_contracts import LineageReference, TaskDefinition
from material_workbench.contracts.design_space_contracts import DesignSpaceDefinition
from material_workbench.contracts.proposal_contracts import ProposalStrategyDefinition
from material_workbench.contracts.batch_proposal_contracts import BatchProposalDefinition
from material_workbench.domain.screening_score import evaluate_screening_goal, score_contract, screening_goal_runtime_value
from material_workbench.domain.proposal_acquisition import acquisition_value
from material_workbench.domain.proposal_generation import generate_candidates
from material_workbench.domain.batch_selector import (
    candidate_design_values,
    select_experiment_batch,
)


COMPOSITION_COLUMNS = composition_names(task_id="annealed-properties-v1")
SCREENING_SEED = DEFAULT_SCREENING_SEED


def generate_from_design_space(
    base: Candidate,
    design_space: DesignSpaceDefinition,
    *,
    count: int,
    seed: int = SCREENING_SEED,
) -> list[tuple[Candidate, dict[str, float | str]]]:
    """Generate candidates from the immutable design-space contract.

    Sampling is deliberately kept as the first allow-listed strategy.  The
    important boundary is that it consumes the DesignSpace, rather than a UI
    request, so later simplex, Bayesian, or program decoders share the same
    fixed values, domains and conditional rules.
    """
    return generate_candidates(
        "latin_hypercube",
        base,
        design_space,
        count=count,
        seed=seed,
    )


def _validate_screening_pool(
    generated: list[tuple[Candidate, dict[str, float | str]]],
    candidate_validator: Callable[[CandidateInput], None],
) -> tuple[
    list[tuple[int, Candidate, CandidateInput, dict[str, float | str]]],
    dict[str, int],
    list[dict[str, Any]],
]:
    """Validate the complete proposal pool so rejection counts have a fixed denominator."""
    valid_candidates: list[
        tuple[int, Candidate, CandidateInput, dict[str, float | str]]
    ] = []
    rejected_by_reason: defaultdict[str, int] = defaultdict(int)
    rejections: list[dict[str, Any]] = []
    for pool_index, (candidate, applied) in enumerate(generated):
        try:
            candidate_input = CandidateInput.model_validate(candidate.model_dump())
            candidate_validator(candidate_input)
        except ValueError as exc:
            reason = str(exc) or "candidate_constraint"
            rejected_by_reason[reason] += 1
            rejections.append(
                {
                    "pool_index": pool_index,
                    "inputs": applied,
                    "reason": reason,
                }
            )
            continue
        candidate = Candidate.model_validate({**candidate.model_dump(), **candidate_input.model_dump()})
        valid_candidates.append((pool_index, candidate, candidate_input, applied))
    return valid_candidates, dict(sorted(rejected_by_reason.items())), rejections


def _apply_screening_goal_metadata(
    prediction: dict[str, Any],
    goal: ScreeningGoal | None,
) -> dict[str, Any]:
    prediction["goal_direction"] = goal.direction if goal else None
    prediction["goal_value"] = (
        goal.lower if goal and goal.direction == "at_least"
        else goal.upper if goal and goal.direction == "at_most"
        else None
    )
    prediction["goal_lower"] = goal.lower if goal and goal.direction == "between" else None
    prediction["goal_upper"] = goal.upper if goal and goal.direction == "between" else None
    return prediction


def _proposal_coverage(
    generated: list[tuple[Candidate, dict[str, float | str]]],
    design_space: DesignSpaceDefinition,
) -> dict[str, dict[str, float]]:
    evidence: dict[str, dict[str, float]] = {}
    domains = (*design_space.numeric_domains, *design_space.heat_pattern_domains)
    for domain in domains:
        values = [
            float(applied[domain.path])
            for _, applied in generated
            if domain.path in applied
        ]
        if not values:
            continue
        if domain.range is not None:
            span = domain.range.max - domain.range.min
        else:
            span = max(domain.values) - min(domain.values)
        observed_min = min(values)
        observed_max = max(values)
        evidence[domain.path] = {
            "observed_min": observed_min,
            "observed_max": observed_max,
            "observed_mean": sum(values) / len(values),
            "normalized_span": min(
                1.0,
                max(0.0, (observed_max - observed_min) / max(span, 1e-12)),
            ),
        }
    return evidence


def _evaluate_proposal_pool(
    runtime: PredictionRuntime,
    candidates: list[Candidate],
    *,
    target_values: dict[str, Any],
) -> list[dict[str, Any]]:
    """Evaluate a generated pool without building discarded similarity rows."""

    evaluation_candidates = [
        candidate.model_copy(
            update={
                "id": f"{candidate.id}:proposal-pool:{index}",
            }
        )
        for index, candidate in enumerate(candidates)
    ]
    predictions = (
        runtime.predict_batch(
            evaluation_candidates,
            detailed=False,
            target_values=target_values,
        )
        if (
            isinstance(runtime, BatchPredictionRuntime)
            and runtime.supports_batch_prediction
        )
        else [
            runtime.predict_core(
                candidate,
                detailed=False,
                target_values=target_values,
            )
            for candidate in evaluation_candidates
        ]
    )
    if len(predictions) != len(candidates):
        raise ValueError("batch prediction did not preserve candidate cardinality")
    for candidate, evaluation_candidate, prediction in zip(
        candidates,
        evaluation_candidates,
        predictions,
        strict=True,
    ):
        if prediction.get("candidate_id") != evaluation_candidate.id:
            raise ValueError("batch prediction did not preserve candidate order")
        prediction["candidate_id"] = candidate.id
        if "support" in prediction:
            continue
        support = runtime.support_summary(evaluation_candidate)
        prediction["support"] = support
        prediction["similar"] = []
        if (
            support.status != "supported"
            and support.message not in prediction["warnings"]
        ):
            prediction["warnings"].append(support.message)
    return predictions


def run_proposal(
    runtime: PredictionRuntime,
    base: Candidate,
    request: ScreeningRequest,
    *,
    probability_available: dict[str, bool],
    candidate_validator: Callable[[CandidateInput], None],
    design_space: DesignSpaceDefinition | None = None,
    strategy: ProposalStrategyDefinition,
    batch_reference_candidates: dict[str, Candidate] | None = None,
) -> dict[str, Any]:
    pool_size = request.samples * request.proposal.pool_multiplier
    if design_space is None:
        raise ValueError("Design Spaceなしでは候補を生成できません")
    generated = generate_candidates(
        strategy.generator_id,
        base,
        design_space,
        count=pool_size,
        seed=request.seed,
        generator_version=strategy.generator_version,
        parameters=strategy.generator_parameters,
    )
    points: list[dict[str, Any]] = []
    base_prediction = runtime.predict(base, detailed=False)
    valid_candidates, rejected_by_reason, proposal_rejections = (
        _validate_screening_pool(generated, candidate_validator)
    )

    if len(valid_candidates) < request.samples:
        raise ValueError(
            f"生成{pool_size}件中、制約を満たす点は{len(valid_candidates)}件でした。"
            f"{request.samples}件を作れるよう範囲を見直してください"
        )

    configured_goals: dict[str, ScreeningGoal] = dict(request.secondary_goals)
    if request.target_goal is not None:
        configured_goals[request.target] = request.target_goal
    target_values = {
        key: screening_goal_runtime_value(goal)
        for key, goal in configured_goals.items()
        if probability_available.get(key, False)
    }
    proposal_candidates = [
        candidate for _, candidate, _, _ in valid_candidates
    ]
    prediction_rows = _evaluate_proposal_pool(
        runtime,
        proposal_candidates,
        target_values=target_values,
    )
    for (
        pool_index,
        candidate,
        candidate_input,
        applied,
    ), prediction in zip(valid_candidates, prediction_rows, strict=True):
        selected = prediction["predictions"][request.target]
        support = prediction["support"]
        evaluation = evaluate_screening_goal(
            selected.value,
            goal=request.target_goal,
            achievement_probability=selected.goal_probability if probability_available.get(request.target, False) else None,
            support_distance=support.distance,
        )
        secondary_evaluations = {
            key: evaluate_screening_goal(
                prediction["predictions"][key].value,
                goal=goal,
                achievement_probability=prediction["predictions"][key].goal_probability if probability_available.get(key, False) else None,
                support_distance=support.distance,
            )
            for key, goal in request.secondary_goals.items()
        }
        acquisition_score, acquisition_components = acquisition_value(
            strategy.acquisition_id,
            prediction=selected,
            goal=request.target_goal,
            support_distance=support.distance,
            exploration_parameter=request.proposal.exploration_parameter,
            incumbent_value=request.proposal.incumbent_value,
        )
        prediction_payload = _apply_screening_goal_metadata(
            selected.model_dump(),
            request.target_goal,
        )
        predictions_payload = {}
        for key, item in prediction["predictions"].items():
            predictions_payload[key] = _apply_screening_goal_metadata(
                item.model_dump(),
                configured_goals.get(key),
            )
        points.append({
            "index": pool_index,
            "pool_index": pool_index,
            "inputs": applied,
            "candidate": candidate_input.model_dump(mode="json"),
            "prediction": prediction_payload,
            "predictions": predictions_payload,
            "color_value": selected.value,
            "support": support.model_dump(),
            "warnings": prediction.get("warnings", []),
            "similar": [item.model_dump() if hasattr(item, "model_dump") else item for item in prediction.get("similar", [])],
            "score": round(float(acquisition_score), 6),
            "acquisition_components": acquisition_components,
            "goal_evaluation": evaluation.model_dump(),
            "secondary_goal_evaluations": {key: item.model_dump() for key, item in secondary_evaluations.items()},
        })
    support_rank = {"supported": 0, "caution": 1, "extrapolated": 2}
    eligible = [
        point
        for point in points
        if not (
            request.proposal.support_policy == "exclude_extrapolated"
            and point["support"]["status"] == "extrapolated"
        )
    ]
    if len(eligible) < request.samples:
        raise ValueError(
            f"support方針を満たす点は{len(eligible)}件でした。"
            f"{request.samples}件を選べるよう範囲または方針を見直してください"
        )
    ranked = (
        sorted(eligible, key=lambda point: point["pool_index"])
        if request.purpose == "design_space_map"
        else sorted(
            eligible,
            key=lambda point: (
                support_rank[point["support"]["status"]]
                if request.proposal.support_policy == "supported_first"
                else 0,
                sum(item["achieved"] is False for item in point["secondary_goal_evaluations"].values()),
                point["score"],
                point["pool_index"],
            ),
        )
    )
    selected = ranked[:request.samples]
    batch_definition = request.batch_definition
    if (
        batch_definition is not None
        and "candidate_pool_size" not in batch_definition.model_fields_set
        and batch_definition.candidate_pool_size > request.samples
    ):
        batch_definition = BatchProposalDefinition.model_validate(
            {
                **batch_definition.model_dump(mode="json"),
                "candidate_pool_size": request.samples,
            }
        )
    exact_control_points: list[dict[str, Any]] = []
    if batch_definition is not None:
        references = batch_reference_candidates or {}
        for control_index, requirement in enumerate(
            batch_definition.controls
        ):
            candidate = references[requirement.candidate_id]
            exact_control_points.append(
                {
                    "pool_index": pool_size + control_index,
                    "inputs": candidate_design_values(candidate, design_space),
                    "candidate": CandidateInput.model_validate(
                        candidate.model_dump()
                    ).model_dump(mode="json"),
                    "score": 0.0,
                    "secondary_goal_evaluations": {},
                    "_batch_source": "exact_control",
                    "_candidate_id": candidate.id,
                    "_candidate_revision": candidate.revision,
                }
            )
    batch_proposal = (
        select_experiment_batch(
            [
                *ranked[: batch_definition.candidate_pool_size],
                *exact_control_points,
            ],
            batch_definition,
            design_space,
            seed=request.seed,
            reference_candidates=batch_reference_candidates or {},
            distance_id=strategy.distance_id,
            distance_version=strategy.distance_version,
            distance_parameters=strategy.distance_parameters,
        )
        if batch_definition is not None
        else None
    )
    candidate_by_pool_index = {
        pool_index: candidate
        for pool_index, candidate, _, _ in valid_candidates
    }
    for point in selected:
        point["similar"] = [
            item.model_dump() if hasattr(item, "model_dump") else item
            for item in runtime.similarity(
                candidate_by_pool_index[point["pool_index"]]
            )
        ]
    selected_pool_indices = {point["pool_index"] for point in selected}
    selected_points = []
    selected_rank = {
        point["pool_index"]: rank for rank, point in enumerate(selected, start=1)
    }
    for index, point in enumerate(selected):
        selected_points.append({**point, "index": index})
    if batch_proposal is not None:
        point_index_by_pool = {
            point["pool_index"]: point["index"] for point in selected_points
        }
        batch_proposal["selected"] = [
            {
                "point_index": point_index_by_pool.get(
                    item["point"]["pool_index"]
                ),
                "pool_index": item["point"]["pool_index"],
                "order": order,
                "role": item["role"],
                "reason": item["reason"],
                "acquisition_component": item["acquisition_component"],
                "diversity_component": item["diversity_component"],
                "pending_penalty": item["pending_penalty"],
                "resource_penalty": item["resource_penalty"],
                "combined_score": item["combined_score"],
                "estimated_cost": item["estimated_cost"],
                "setup_group": item["setup_group"],
                "source": item["source"],
                "candidate_id": item["candidate_id"],
                "candidate_revision": item["candidate_revision"],
                "canonical_identity_digest": item[
                    "canonical_identity_digest"
                ],
            }
            for order, item in enumerate(batch_proposal["selected"], start=1)
        ]
    proposal_pool = [
        {
            "pool_index": point["pool_index"],
            "inputs": point["inputs"],
            "acquisition_score": point["score"],
            "acquisition_components": point["acquisition_components"],
            "support_status": point["support"]["status"],
            "selected_rank": selected_rank.get(point["pool_index"]),
            "exclusion_reason": (
                None
                if point["pool_index"] in selected_pool_indices
                else "support_policy_extrapolated"
                if request.proposal.support_policy == "exclude_extrapolated"
                and point["support"]["status"] == "extrapolated"
                else "ranked_below_selection_cutoff"
            ),
        }
        for point in points
    ]
    return {
        "schema_version": "screening-run/v4",
        "purpose": request.purpose,
        "source_run_id": request.source_run_id,
        "seed": request.seed,
        "base_candidate_id": base.id,
        "base_inputs": base.inputs.model_dump(mode="json"),
        "base_canonical_input": base_prediction["canonical_input"],
        "model_provenance": base_prediction["model_meta"],
        "target": request.target,
        "target_goal": request.target_goal.model_dump(mode="json") if request.target_goal else None,
        "secondary_goals": {
            key: goal.model_dump(mode="json")
            for key, goal in request.secondary_goals.items()
        },
        "score_contract": score_contract(
            request.target_goal,
            probability_available=probability_available.get(request.target, False),
        ),
        "samples": request.samples,
        "variables": {name: spec.model_dump() for name, spec in request.variables.items()},
        "points": selected_points,
        "representative_points": selected_points[:10],
        "proposal_pool": proposal_pool,
        "proposal_rejections": proposal_rejections,
        "batch_proposal": batch_proposal,
        "_proposal_diagnostics": {
            "generated_count": pool_size,
            "valid_count": len(valid_candidates),
            "evaluated_count": len(points),
            "selected_count": len(selected_points),
            "rejected_count": sum(rejected_by_reason.values()),
            "rejection_rate": sum(rejected_by_reason.values()) / pool_size,
            "rejected_by_reason": dict(sorted(rejected_by_reason.items())),
            "coverage_by_path": _proposal_coverage(generated, design_space),
        },
    }


def lineage_candidate_options(data: WorkbookData, entity_key: str) -> list[LineageCandidateOption]:
    routes = tuple(
        route for route in data.relation_routes
        if entity_key in route.members.values()
    )
    process_roles = (
        ("annealing",)
        if entity_key in data.anneal_features
        else ("hot_rolling",)
        if entity_key in data.hot_rolling_features
        else ("annealing", "hot_rolling")
    )
    options: list[LineageCandidateOption] = []
    seen: set[tuple[str, str, str]] = set()
    for route in routes:
        for process_role in process_roles:
            features = (
                data.anneal_features
                if process_role == "annealing"
                else data.hot_rolling_features
            )
            process_key = route.members.get(process_role)
            if not process_key or process_key not in features:
                continue
            melt_key = route.members.get("melt")
            if not melt_key:
                process_melt_keys = {
                    candidate_route.members["melt"]
                    for candidate_route in data.relation_routes
                    if candidate_route.members.get(process_role) == process_key
                    and candidate_route.members.get("melt") in data.composition
                }
                melt_key = next(iter(process_melt_keys)) if len(process_melt_keys) == 1 else None
            if not melt_key or melt_key not in data.composition:
                continue
            identity = (process_role, process_key, melt_key)
            if identity in seen:
                continue
            seen.add(identity)
            feature = features[process_key]
            if process_role == "annealing" and len(feature.get("heat_pattern", [])) < 2:
                continue
            options.append(LineageCandidateOption(
                process_key=process_key,
                process_role=process_role,
                process_label="焼鈍条件" if process_role == "annealing" else "熱延条件",
                melt_key=melt_key,
            ))
    return sorted(
        options,
        key=lambda option: (option.process_role, option.process_key, option.melt_key),
    )


def candidate_from_lineage(
    data: WorkbookData,
    entity_key: str,
    *,
    process_key: str | None = None,
    melt_key: str | None = None,
) -> CandidateInput:
    options = lineage_candidate_options(data, entity_key)
    if process_key is not None or melt_key is not None:
        selected = next((
            option for option in options
            if option.process_key == process_key and option.melt_key == melt_key
        ), None)
        if selected is None:
            raise ValueError("選択した工程条件と成分の組み合わせを候補化できません")
    elif len(options) == 1:
        selected = options[0]
    elif options:
        raise ValueError("候補化できる上流条件が複数あります。工程条件と成分を選択してください")
    else:
        raise ValueError("候補化できる工程条件と成分の組み合わせが見つかりません")
    process_role = selected.process_role
    process_key = selected.process_key
    melt_key = selected.melt_key
    feature = data.anneal_features.get(process_key) if process_role == "annealing" else data.hot_rolling_features.get(process_key)
    if feature is None:
        raise ValueError("候補化できる工程条件が見つかりません")
    heat_pattern = None
    if process_role == "annealing":
        heat_pattern = [HeatPoint.model_validate(point) for point in deepcopy(feature["heat_pattern"])]
        if len(heat_pattern) < 2:
            raise ValueError("候補化に必要な焼鈍履歴がありません")
        process_values = (
            {"ls_mpm": float(feature["ls_mpm"])}
            if isinstance(feature.get("ls_mpm"), (int, float))
            else {}
        )
        entity_type = "annealing"
    else:
        process_values = {name: float(feature[name]) for name in PROCESS_NAMES}
        entity_type = "hot_rolling"
    direct_context_routes = [
        route
        for route in data.relation_routes
        if entity_key in route.members.values()
        and route.members.get(process_role) == process_key
        and route.members.get("melt") in {None, melt_key}
    ]
    relation_context_ids = {route.id for route in direct_context_routes}
    if not any(route.members.get("melt") == melt_key for route in direct_context_routes):
        bridge_routes = [
            route
            for route in data.relation_routes
            if route.members.get(process_role) == process_key
            and route.members.get("melt") == melt_key
        ]
        if bridge_routes:
            minimum_members = min(len(route.members) for route in bridge_routes)
            relation_context_ids.update(
                route.id
                for route in bridge_routes
                if len(route.members) == minimum_members
            )
    return CandidateInput(
        name=f"実績 {process_key} / {melt_key}",
        inputs={
            "composition": deepcopy(data.composition[melt_key]),
            "process": process_values,
            "categorical": {},
            "heat_pattern": heat_pattern,
            "heat_time_basis": (
                "elapsed_time"
                if process_role == "annealing" and "ls_mpm" not in process_values
                else "line_speed"
            ),
        },
        provenance={
            "source_kind": "lineage",
            "source_ref": {
                "entity_type": entity_type,
                "entity_key": process_key,
                "composition_entity_key": melt_key,
                "relation_context_ids": sorted(relation_context_ids),
                "data_source_digest": data.source_sha256,
            },
        },
    )


def lineage_candidate_origin_evidence(
    data: WorkbookData,
    *,
    candidate_id: str,
    task_id: str,
    reference: LineageReference,
    task_definition: TaskDefinition,
) -> CandidateOriginEvidence:
    """Aggregate only observations that created this lineage candidate."""
    if reference.data_source_digest and reference.data_source_digest != data.source_sha256:
        raise ValueError("候補化した時点の参照データと現在のデータが一致しません")
    route_ids = set(reference.relation_context_ids)
    if reference.composition_entity_key is None and not route_ids:
        raise ValueError("作成元実測を特定するための成分またはrelation経路がありません")

    observations = [
        observation
        for observation in data.observations
        if observation["task_id"] == task_id
        and observation["parent_key"] == reference.entity_key
        and (
            reference.composition_entity_key is None
            or observation.get("composition_key") == reference.composition_entity_key
        )
        and (
            not route_ids
            or bool(route_ids.intersection(observation.get("relation_context_ids", ())))
        )
    ]
    declared_measurements = {
        measurement_key
        for output in task_definition.outputs
        for measurement_key in (*output.measurement_keys, output.key, output.label)
    }
    values_by_measurement: defaultdict[str, list[float]] = defaultdict(list)
    for observation in observations:
        for measurement_key, value in observation.get("outputs", {}).items():
            if measurement_key in declared_measurements and isinstance(value, (int, float)):
                values_by_measurement[measurement_key].append(float(value))

    return CandidateOriginEvidence(
        candidate_id=candidate_id,
        task_id=task_id,
        process_key=reference.entity_key,
        composition_key=reference.composition_entity_key,
        relation_context_ids=sorted(route_ids),
        observation_ids=sorted(observation["id"] for observation in observations),
        repeat_summary={
            key: {
                "mean": fmean(values),
                "std": pstdev(values),
                "n": len(values),
            }
            for key, values in sorted(values_by_measurement.items())
        },
    )


def _candidate_xlsx_names(
    task_id: str,
    profile_path: str | None = None,
    profile: DatasetInputProfile | None = None,
) -> dict[str, str]:
    profile = profile or load_dataset_profile(profile_path)
    task_profile = profile.tasks[task_id]
    names = {
        "schema_version": "形式バージョン",
        "id": "候補ID",
        "name": "候補名",
        "support_status": "学習範囲判定",
        "support_distance": "学習範囲からの距離",
        "segment_start": "工程境界",
        "heat_time_basis": "時間基準",
    }
    names.update({mapping.path: mapping.column for mapping in task_profile.mappings if mapping.column})
    heat_mapping = next((mapping for mapping in task_profile.mappings if mapping.kind == "ordered_heat_series"), None)
    if heat_mapping and heat_mapping.series_columns:
        names["time_s"] = heat_mapping.series_columns.time
        names["temperature_c"] = heat_mapping.series_columns.value
    technical_names = {
        item.name: item.column
        for item in profile.shared.technical
        if item.role == "anneal_history"
    }
    names["stage_name"] = technical_names.get("stage_name", "工程")
    names["stage_category"] = technical_names.get("stage_category", "標準工程カテゴリ")
    for observation in task_profile.observations:
        for target in (*observation.targets, *observation.auxiliary):
            names.setdefault(target.key, target.column)
    return names


def _candidate_template_unit(field: Any) -> str | None:
    unit = {"mpm": "m/min"}.get(field.unit, field.unit)
    return "mass%" if field.path.startswith("composition.") and unit == "%" else unit


def _candidate_template_names(definition: Any) -> dict[str, str]:
    names = {
        "name": "候補名",
        "time_s": "経過時間[s]",
        "temperature_c": "温度[℃]",
        "heat_time_basis": "時間基準",
    }
    response_labels = {
        item.path: item.label
        for item in definition.response_curve_variables
        if item.kind == "numeric_input" and item.path
    }
    for group in definition.input_groups:
        for field in group.fields:
            if field.kind not in {"number", "categorical"}:
                continue
            label = "ライン速度" if field.path == "process.ls_mpm" else response_labels.get(field.path, field.label)
            unit = _candidate_template_unit(field)
            names[field.path] = f"{label}[{unit}]" if unit else label
    return names


def _xlsx_position(positions: dict[str, int], internal_name: str, *display_names: str | None) -> int | None:
    for name in (*display_names, internal_name):
        if name and name in positions:
            return positions[name]
    return None


def import_candidates_xlsx(
    contents: bytes,
    task_id: str = "annealed-properties-v1",
    profile_path: str | None = None,
    profile: DatasetInputProfile | None = None,
    validate_candidate: Callable[[CandidateInput], Any] | None = None,
) -> tuple[list[CandidateInput], list[dict[str, Any]]]:
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        return [], [{"row": 0, "message": f"Excelを読み込めません: {exc}"}]
    definition = load_task_contracts()[task_id].task_definition
    fields = {
        field.path: field
        for group in definition.input_groups
        for field in group.fields
        if field.kind in {"number", "categorical"}
    }
    composition_fields = [field for field in fields.values() if field.path.startswith("composition.")]
    process_fields = [field for field in fields.values() if field.path.startswith("process.")]
    categorical_fields = [field for field in fields.values() if field.path.startswith("categorical.")]
    heat_enabled = any(field.kind == "heat_pattern" for group in definition.input_groups for field in group.fields)
    display_names = _candidate_xlsx_names(task_id, profile_path, profile)
    template_names = _candidate_template_names(definition)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
    duplicate_headers = sorted({header for header in headers if header and headers.count(header) > 1})
    if duplicate_headers:
        workbook.close()
        return [], [{"row": 1, "message": f"列名が重複しています: {', '.join(duplicate_headers)}"}]
    positions = {header: index for index, header in enumerate(headers) if header}
    if not headers or _xlsx_position(positions, "name", template_names["name"], display_names["name"]) is None:
        return [], [{"row": 1, "message": f"必須列 {template_names['name']} がありません"}]
    imported: list[CandidateInput] = []
    errors: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None for value in row):
            continue
        try:
            def value(header: str, *display_headers: str | None, default: Any = None) -> Any:
                position = _xlsx_position(positions, header, *display_headers)
                return row[position] if position is not None and position < len(row) else default

            composition = {
                field.path.removeprefix("composition."): float(value(field.path.removeprefix("composition."), template_names.get(field.path), display_names.get(field.path)))
                for field in composition_fields
                if value(field.path.removeprefix("composition."), template_names.get(field.path), display_names.get(field.path)) is not None
            }
            process = {
                field.path.removeprefix("process."): float(value(field.path.removeprefix("process."), template_names.get(field.path), display_names.get(field.path)))
                for field in process_fields
                if value(field.path.removeprefix("process."), template_names.get(field.path), display_names.get(field.path)) is not None
            }
            categorical = {
                field.path.removeprefix("categorical."): str(value(field.path.removeprefix("categorical."), template_names.get(field.path), display_names.get(field.path)))
                for field in categorical_fields
                if value(field.path.removeprefix("categorical."), template_names.get(field.path), display_names.get(field.path)) is not None
            }
            points: list[dict[str, Any]] = []
            if heat_enabled:
                index = 1
                while (
                    _xlsx_position(positions, f"time_s_{index}", f"{template_names['time_s']}_{index}", f"{display_names['time_s']}_{index}") is not None
                    and _xlsx_position(positions, f"temperature_c_{index}", f"{template_names['temperature_c']}_{index}", f"{display_names['temperature_c']}_{index}") is not None
                ):
                    time = value(f"time_s_{index}", f"{template_names['time_s']}_{index}", f"{display_names['time_s']}_{index}")
                    temperature = value(f"temperature_c_{index}", f"{template_names['temperature_c']}_{index}", f"{display_names['temperature_c']}_{index}")
                    if time is not None and temperature is not None:
                        segment_raw = value(f"segment_start_{index}", f"{display_names['segment_start']}_{index}", default=False)
                        segment_start = segment_raw is True or str(segment_raw).strip().lower() in {"1", "true", "yes", "あり"}
                        points.append({
                            "time_s": float(time),
                            "temperature_c": float(temperature),
                            "segment_start": segment_start,
                            "stage_name": str(value(f"stage_name_{index}", f"{display_names['stage_name']}_{index}")).strip() if value(f"stage_name_{index}", f"{display_names['stage_name']}_{index}") is not None else None,
                            "stage_category": str(value(f"stage_category_{index}", f"{display_names['stage_category']}_{index}")).strip() if value(f"stage_category_{index}", f"{display_names['stage_category']}_{index}") is not None else None,
                        })
                    index += 1
                if len(points) < 2:
                    raise ValueError(f"{template_names['time_s']}_1 / {template_names['temperature_c']}_1 から少なくとも2点が必要です")
            name = value("name", template_names["name"], display_names["name"])
            if name is None or not str(name).strip():
                raise ValueError(f"{template_names['name']}は空にできません")
            raw_time_basis = str(value(
                "heat_time_basis",
                template_names["heat_time_basis"],
                display_names["heat_time_basis"],
                default="ライン速度連動",
            )).strip()
            if raw_time_basis in {"", "line_speed", "ライン速度連動"}:
                heat_time_basis = "line_speed"
            elif raw_time_basis in {"elapsed_time", "経過時間を直接指定", "経過時間"}:
                heat_time_basis = "elapsed_time"
            else:
                raise ValueError(
                    f"{template_names['heat_time_basis']}は"
                    "「ライン速度連動」または「経過時間を直接指定」で入力してください"
                )
            payload = CandidateInput(
                name=str(name).strip(),
                inputs={
                    "composition": composition,
                    "process": process,
                    "categorical": categorical,
                    "heat_pattern": points if heat_enabled else None,
                    "heat_time_basis": heat_time_basis,
                },
            )
            if validate_candidate is not None:
                validate_candidate(payload)
            imported.append(payload)
        except (TypeError, ValueError) as exc:
            errors.append({"row": row_number, "message": str(exc)})
    workbook.close()
    return imported, errors


def _candidate_xlsx_input_schema(
    task_id: str,
    heat_point_count: int,
    profile_path: str | None = None,
    profile: DatasetInputProfile | None = None,
    *,
    for_template: bool = False,
) -> tuple[Any, list[Any], list[Any], bool, dict[str, str], list[str], list[str]]:
    definition = load_task_contracts()[task_id].task_definition
    numeric_fields = [
        field for group in definition.input_groups
        for field in group.fields
        if field.kind == "number"
    ]
    categorical_fields = [
        field for group in definition.input_groups
        for field in group.fields
        if field.kind == "categorical"
    ]
    heat_enabled = any(field.kind == "heat_pattern" for group in definition.input_groups for field in group.fields)
    display_names = _candidate_template_names(definition) if for_template else _candidate_xlsx_names(task_id, profile_path, profile)
    heat_headers = [
        item for index in range(1, heat_point_count + 1)
        for item in ((f"time_s_{index}", f"temperature_c_{index}") if for_template else (f"time_s_{index}", f"temperature_c_{index}", f"segment_start_{index}", f"stage_name_{index}", f"stage_category_{index}"))
    ] if heat_enabled else []
    input_headers = [
        display_names["name"],
        *[display_names.get(field.path, field.label) for field in numeric_fields],
        *[display_names.get(field.path, field.label) for field in categorical_fields],
        *([display_names["heat_time_basis"]] if heat_enabled else []),
        *[
            f"{display_names.get(header.rsplit('_', 1)[0], header.rsplit('_', 1)[0])}_{header.rsplit('_', 1)[1]}"
            for header in heat_headers
        ],
    ]
    return definition, numeric_fields, categorical_fields, heat_enabled, display_names, heat_headers, input_headers


def candidate_template_xlsx(runtime: PredictionRuntime, task_id: str = "annealed-properties-v1") -> bytes:
    contract = load_task_contracts()[task_id]
    canonical = contract.canonical_candidate
    heat_point_count = max(2, len(canonical.heat_pattern or ())) if canonical.heat_pattern is not None else 0
    definition, numeric_fields, categorical_fields, heat_enabled, display_names, heat_headers, input_headers = _candidate_xlsx_input_schema(
        task_id,
        heat_point_count,
        runtime.data.profile_path if not runtime.data.profile_path.startswith("catalog:") else None,
        getattr(runtime.data, "profile", None),
        for_template=True,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "候補"
    sheet.append(input_headers)
    heat_values = [
        item
        for point in (canonical.heat_pattern or ())
        for item in (point.time_s, point.temperature_c)
    ] if heat_enabled else []
    heat_values.extend([None] * (len(heat_headers) - len(heat_values)))
    example_values = [
        canonical.composition.get(field.path.removeprefix("composition."))
        if field.path.startswith("composition.")
        else canonical.process.get(field.path.removeprefix("process."))
        for field in numeric_fields
    ]
    example_values.extend([
        canonical.categorical.get(field.path.removeprefix("categorical."))
        for field in categorical_fields
    ])
    if heat_enabled:
        example_values.append("ライン速度連動")
    example_sheet = workbook.create_sheet("記入例")
    example_sheet.append(input_headers)
    example_sheet.append(["記入例（候補シートへコピーして変更）", *example_values, *heat_values])
    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F5FC4")
    example_fill = PatternFill("solid", fgColor="FFF4D6")
    for target_sheet in (sheet, example_sheet):
        for cell in target_sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        target_sheet.row_dimensions[1].height = 28
        target_sheet.freeze_panes = "B2"
        target_sheet.auto_filter.ref = target_sheet.dimensions
        for column in target_sheet.columns:
            letter = column[0].column_letter
            target_sheet.column_dimensions[letter].width = min(24, max(13, max(len(str(cell.value or "")) for cell in column) + 2))
        target_sheet.column_dimensions["A"].width = 40
        for column_index, header in enumerate(input_headers, start=1):
            if header == "ライン速度[m/min]":
                target_sheet.column_dimensions[target_sheet.cell(1, column_index).column_letter].width = 20
            elif header.startswith(("経過時間[s]_", "温度[℃]_")):
                target_sheet.column_dimensions[target_sheet.cell(1, column_index).column_letter].width = 18
    for cell in example_sheet[2]:
        cell.fill = example_fill
    for index, field in enumerate((*numeric_fields, *categorical_fields), start=2):
        column_letter = sheet.cell(1, index).column_letter
        if field.kind == "number" and field.allowed_range is not None:
            validation = DataValidation(type="decimal", operator="between", formula1=str(field.allowed_range.min), formula2=str(field.allowed_range.max), allow_blank=not field.required)
            validation.error = f"{field.allowed_range.min}～{field.allowed_range.max} の数値を入力してください"
            validation.errorTitle = "入力範囲外"
            validation.prompt = f"単位: {_candidate_template_unit(field) or 'なし'} / 許容範囲: {field.allowed_range.min}～{field.allowed_range.max}"
            validation.promptTitle = field.label
            validation.showErrorMessage = True
            validation.showInputMessage = True
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}501")
        elif field.kind == "categorical":
            validation = DataValidation(type="list", formula1='"' + ','.join(field.choices) + '"', allow_blank=not field.required)
            validation.showErrorMessage = True
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}501")
    if heat_enabled:
        basis_column = sheet.cell(1, 2 + len(numeric_fields) + len(categorical_fields)).column_letter
        validation = DataValidation(
            type="list",
            formula1='"ライン速度連動,経過時間を直接指定"',
            allow_blank=True,
        )
        validation.showErrorMessage = True
        sheet.add_data_validation(validation)
        validation.add(f"{basis_column}2:{basis_column}501")

    guide = workbook.create_sheet("入力ルール")
    guide.append([f"{definition.label} 候補XLSX", "「記入例」を参考に「候補」シートへ入力し、アプリへ読み込みます。読み込まれるのは「候補」シートだけです。"])
    guide.append(["基本ルール", "1行＝1候補です。1行目の列名と単位は変更しないでください。空行は無視されます。"])
    guide.append(["候補名", f"{display_names['name']} は必須です。「記入例」の2行目を「候補」へコピーし、名前と値を変更して使えます。"])
    guide.append(["数値", "数値だけを入力します。単位は列名に記載されているため、セル内には書きません。"])
    if heat_enabled:
        guide.append(["時間基準", "「ライン速度連動」はLS変更時に全時刻を再計算します。「経過時間を直接指定」は入力した時刻をそのまま使います。空欄はライン速度連動です。"])
        guide.append(["履歴点", f"{display_names['time_s']}_N と {display_names['temperature_c']}_N を同じ番号で最低2点入力します。時間は候補全体の開始からの連続した経過時間で、工程が変わっても0へ戻さず昇順にします。点を増やす場合は右へ同じ列セットを追加します。"])
    guide.append([])
    rules_header_row = guide.max_row + 1
    for row_number in range(1, rules_header_row - 1):
        guide.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=5)
    guide.append(["列名", "区分", "必須", "入力ルール", "学習データ範囲（参照）"])
    guide.append([display_names["name"], "識別", "必須", "候補ごとに異なる名前", "—"])
    for field in (*numeric_fields, *categorical_fields):
        if field.kind == "number":
            assert field.allowed_range is not None and field.training_range is not None
            display_unit = _candidate_template_unit(field) or ""
            rule = f"{field.allowed_range.min}～{field.allowed_range.max} {display_unit}".strip()
            training = f"{field.training_range.min}～{field.training_range.max} {display_unit}".strip()
        else:
            rule = " / ".join(field.choices)
            training = "—"
        guide.append([display_names.get(field.path, field.label), field.label, "必須" if field.required else "任意", rule, training])
    if heat_enabled:
        guide.append([display_names["heat_time_basis"], "時間軸", "任意", "ライン速度連動 / 経過時間を直接指定", "—"])
        guide.append([f"{display_names['time_s']}_N / {display_names['temperature_c']}_N", "履歴点", "最低2点", "同じNの時刻と温度を対で入力", "—"])
    guide.freeze_panes = f"A{rules_header_row + 1}"
    guide.column_dimensions["A"].width = 34
    guide.column_dimensions["B"].width = 22
    guide.column_dimensions["C"].width = 12
    guide.column_dimensions["D"].width = 42
    guide.column_dimensions["E"].width = 34
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in guide[rules_header_row]:
        cell.fill = PatternFill("solid", fgColor="DCEAFB")
        cell.font = Font(bold=True, color="173F75")
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in range(1, rules_header_row - 1):
        guide.cell(row_number, 2).alignment = Alignment(vertical="center", wrap_text=False)
        guide.row_dimensions[row_number].height = 24
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def candidates_xlsx(candidates: list[Candidate], runtime: PredictionRuntime, task_id: str = "annealed-properties-v1") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "候補"
    max_points = max((len(candidate.inputs.heat_pattern or []) for candidate in candidates), default=2)
    definition, numeric_fields, categorical_fields, heat_enabled, display_names, heat_headers, input_headers = _candidate_xlsx_input_schema(
        task_id,
        max_points,
        runtime.data.profile_path if not runtime.data.profile_path.startswith("catalog:") else None,
        getattr(runtime.data, "profile", None),
    )
    headers = [
        display_names["schema_version"], display_names["id"], *input_headers,
        *[display_names.get(output.key, output.label) for output in definition.outputs],
        display_names["support_status"], display_names["support_distance"],
    ]
    if len(headers) != len(set(headers)):
        duplicates = sorted({header for header in headers if headers.count(header) > 1})
        raise ValueError(f"候補XLSXの見出しが重複しています: {', '.join(duplicates)}")
    sheet.append(headers)
    for candidate in candidates:
        result = (
            None
            if candidate.blend_validation.status == "invalid"
            else runtime.predict(candidate, detailed=False)
        )
        heat_values = [
            item
            for point in (candidate.inputs.heat_pattern or [])
            for item in (point.time_s, point.temperature_c, point.segment_start, point.stage_name, point.stage_category)
        ] if heat_enabled else []
        heat_values.extend([None] * (len(heat_headers) - len(heat_values)))
        input_values = [
            candidate.inputs.composition.get(field.path.removeprefix("composition."))
            if field.path.startswith("composition.")
            else candidate.inputs.process.get(field.path.removeprefix("process."))
            for field in numeric_fields
        ]
        input_values.extend([
            candidate.inputs.categorical.get(field.path.removeprefix("categorical."))
            for field in categorical_fields
        ])
        if heat_enabled:
            input_values.append(
                "経過時間を直接指定"
                if candidate.inputs.heat_time_basis == "elapsed_time"
                else "ライン速度連動"
            )
        sheet.append([
            "material-workbench-candidate-v2", candidate.id, candidate.name, *input_values, *heat_values,
            *(
                result["predictions"][output.key].value if result is not None else None
                for output in definition.outputs
            ),
            result["support"].status if result is not None else None,
            result["support"].distance if result is not None else None,
        ])
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(22, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    guide = workbook.create_sheet("説明")
    guide.append([display_names["schema_version"], "material-workbench-candidate-v2"])
    guide.append(["用途", "候補入力と軽量プレビュー予測の出力。候補シートはそのまま候補importに使用できます。"])
    guide.append(["入力項目", " / ".join(f"{field.label}[{field.unit}]" for field in (*numeric_fields, *categorical_fields))])
    if heat_enabled:
        guide.append(["時間基準", "ライン速度連動ではLS変更時に全時刻を再計算します。経過時間を直接指定では入力時刻を保持します。"])
        guide.append(["焼鈍履歴", f"{display_names['time_s']}_N / {display_names['temperature_c']}_N を履歴点として指定し、{display_names['stage_name']}_N を工程名として表示します。工程の先頭は {display_names['segment_start']}_N で表します。"])
    guide.append(["予測", "出力単位はTaskDefinitionの定義に従います。"])
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 110
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
