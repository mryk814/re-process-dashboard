"""UI-only onboarding for a new tabular CSV or single-sheet XLSX task.

The browser never receives a path or a CLI command. The original upload remains
read-only; a canonical CSV snapshot is stored in the user-owned Task store and
registered with the original source provenance in the managed Data Library.
"""
from __future__ import annotations

from dataclasses import dataclass
import json
import logging
from pathlib import Path
import shutil
from tempfile import TemporaryDirectory
from typing import Annotated, Any, Literal

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from starlette.concurrency import run_in_threadpool

from decision_workbench.application.dataset_registration import (
    DatasetRegistrationResult,
    ManagedDatasetRegistrationCheckpoint,
    managed_dataset_registration_checkpoint,
    register_managed_dataset,
    rollback_managed_dataset_registration,
)
from decision_workbench.application.personal_task_packages import (
    build_standard_package,
    promote_personal_package,
    rollback_personal_package_attempt,
)
from decision_workbench.developer_experience.task_scaffolding import (
    ScaffoldField,
    TASK_BUNDLE_SCHEMA_VERSION,
    TASK_SCAFFOLD_SCHEMA_VERSION,
    TASK_ID_MIN_LENGTH,
    TASK_ID_PATTERN,
    TASK_ID_EXAMPLE,
    TaskScaffoldResult,
    create_task_scaffold,
    inspect_task_source,
    validate_personal_task_store_path,
)
from decision_workbench.data.file_integrity import file_sha256
from decision_workbench.modeling.model_lifecycle import (
    load_available_packages,
    validate_personal_model_store_path,
)
from decision_workbench.modeling.model_package_verify import verify_model_package
from decision_workbench.modeling.training.recipe import (
    CSV_ONBOARDING_ESTIMATOR_IDS,
    CsvOnboardingEstimatorOption,
    csv_onboarding_estimator_options,
    csv_onboarding_estimator_recipe,
    estimator_recipe,
)
from decision_workbench.contracts.evidence_contracts import ApiError


router = APIRouter(prefix="/api/data-library/csv-onboarding", tags=["csv-onboarding"])
MAX_TABULAR_BYTES = 100 * 1024 * 1024
logger = logging.getLogger(__name__)

_FAILURE_MESSAGES = {
    "inspect": "表を確認できませんでした。ファイル、sheet、ヘッダー、行数を確認してもう一度試してください。",
    "fields": "列の設定を確認してください。",
    "storage": "個人TaskまたはModelの保存先を確認してください。",
    "prepare": "Taskとモデルを準備できませんでした。列の役割と範囲を確認してもう一度試してください。",
    "build": "Estimatorの学習またはPackage buildに失敗しました。選択したEstimatorと入力列を確認してください。",
    "verify": "作成したModel Packageが安全契約を満たしませんでした。Estimatorの説明と入力設定を確認してください。",
    "promote": "検証済みModel Packageを利用可能状態へ昇格できませんでした。個人Model保存先を確認してください。",
    "dataset": "DatasetをData Libraryへ登録できませんでした。準備した内容は取り消しました。もう一度試してください。",
    "refresh": "新しいTaskを再読込できませんでした。準備した内容は取り消しました。もう一度試してください。",
}


class OnboardingStorageError(ValueError):
    def __init__(self, code: str, message: str, next_action: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.next_action = next_action

    def public_detail(self) -> dict[str, str]:
        return {
            "code": self.code,
            "message": self.message,
            "next_action": self.next_action,
        }


class CsvInspectionColumn(BaseModel):
    name: str
    kind: Literal["number", "categorical"]
    non_empty: int
    observed_min: float | None
    observed_max: float | None
    choices: list[str]


class OnboardingWorksheet(BaseModel):
    name: str
    state: Literal["visible", "hidden", "veryHidden"]


class CsvTaskIdContract(BaseModel):
    pattern: str
    min_length: int
    example: str


class CsvInspectionResponse(BaseModel):
    source_filename: str
    source_sha256: str
    source_kind: Literal["csv", "xlsx"]
    reader_policy: str
    worksheets: list[OnboardingWorksheet]
    selected_sheet: str | None
    requires_sheet_selection: bool
    rows: int
    relations: Literal[0]
    grain: Literal["one-row-one-observation"]
    columns: list[CsvInspectionColumn]
    task_id_contract: CsvTaskIdContract
    estimators: list[CsvOnboardingEstimatorOption]
    default_estimator_id: str
    notice: str


class CsvPrepareResponse(BaseModel):
    state: Literal["ready"]
    task_id: str
    dataset_view_revision_id: str
    dataset_revision_id: str
    source_sha256: str
    model_package_ref_id: str
    reused_existing: bool


@dataclass(frozen=True)
class OnboardingReservation:
    task_root: Path
    model_store: Path
    package_id: str
    reusable_task: TaskScaffoldResult | None
    reusable_package: bool


def _normalized_json(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _normalized_json(item) for key, item in value.items()}
    if isinstance(value, (tuple, list)):
        return [_normalized_json(item) for item in value]
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        return float(value)
    return value


def _task_file(root: Path, relative: str, *, label: str) -> Path:
    candidate = (root / relative).resolve()
    if root not in candidate.parents or not candidate.is_file():
        raise ValueError(f"existing Task is missing {label}")
    return candidate


def _load_reusable_task(
    *,
    root: Path,
    task_id: str,
    label: str,
    source: Path,
    fields: tuple[ScaffoldField, ...],
    estimator_id: str,
    grain_confirmation: str,
    relation_confirmation: str,
    selected_sheet: str | None,
) -> TaskScaffoldResult:
    scaffold_path = _task_file(root, "scaffold.json", label="scaffold")
    bundle_path = _task_file(root, "bundle.json", label="bundle")
    scaffold = json.loads(scaffold_path.read_text(encoding="utf-8"))
    bundle = json.loads(bundle_path.read_text(encoding="utf-8"))
    source_meta = scaffold["source"]
    safety = scaffold["safety"]
    expected_fields = [field.__dict__ for field in fields]
    scaffold_estimator = scaffold.get("estimator", {})
    bundle_estimator = str(bundle.get("estimator_id", ""))
    standard_estimator_ids = tuple(
        str(item)
        for item in bundle.get("standard_estimator_ids", (bundle_estimator,))
    )
    if (
        scaffold.get("schema_version") != TASK_SCAFFOLD_SCHEMA_VERSION
        or scaffold.get("task_id") != task_id
        or scaffold.get("label") != label.strip()
        or scaffold.get("state") != "ready"
        or source_meta.get("original_sha256") != file_sha256(source)
        or source_meta.get("selected_sheet") != selected_sheet
        or _normalized_json(scaffold.get("fields"))
        != _normalized_json(expected_fields)
        or _normalized_json(scaffold_estimator)
        != _normalized_json(estimator_recipe(bundle_estimator).model_dump(mode="json"))
        or tuple(scaffold.get("standard_estimator_ids", standard_estimator_ids))
        != standard_estimator_ids
        or not standard_estimator_ids
        or len(set(standard_estimator_ids)) != len(standard_estimator_ids)
        or bundle_estimator not in standard_estimator_ids
        or estimator_id not in standard_estimator_ids
        or not set(standard_estimator_ids).issubset(CSV_ONBOARDING_ESTIMATOR_IDS)
        or safety.get("grain_confirmation") != grain_confirmation
        or safety.get("relation_confirmation") != relation_confirmation
        or bundle.get("schema_version") != TASK_BUNDLE_SCHEMA_VERSION
        or bundle.get("task_id") != task_id
        or bundle.get("state") != "ready"
    ):
        raise ValueError("existing Task identity does not match this CSV setup")
    source_path = _task_file(
        root,
        str(source_meta["materialized_csv"]),
        label="source",
    )
    profile_path = _task_file(root, "dataset-profile.json", label="profile")
    task_definition_path = _task_file(
        root,
        "task-definition.json",
        label="Task definition",
    )
    training_recipe_path = _task_file(
        root,
        "training-recipe.json",
        label="training recipe",
    )
    return TaskScaffoldResult(
        task_id=task_id,
        state="ready",
        root=root,
        source_path=source_path,
        profile_path=profile_path,
        task_definition_path=task_definition_path,
        training_recipe_path=training_recipe_path,
        unresolved=(),
    )


async def _uploaded_tabular(file: UploadFile) -> tuple[TemporaryDirectory[str], Path]:
    filename = Path(file.filename or "").name
    if not filename or Path(filename).suffix.lower() not in {".csv", ".xlsx"}:
        raise HTTPException(422, "CSVまたはExcel .xlsxファイルを選択してください")
    temporary = TemporaryDirectory(prefix="material-workbench-tabular-")
    target = Path(temporary.name) / filename
    size = 0
    try:
        with target.open("wb") as stream:
            while chunk := await file.read(1024 * 1024):
                size += len(chunk)
                if size > MAX_TABULAR_BYTES:
                    raise HTTPException(413, "CSVまたはExcelファイルは100 MB以下にしてください")
                stream.write(chunk)
        return temporary, target
    except Exception:
        temporary.cleanup()
        raise
    finally:
        await file.close()


def _workbook_sheets(source: Path) -> list[OnboardingWorksheet]:
    workbook = load_workbook(source, read_only=False, data_only=False)
    try:
        return [
            OnboardingWorksheet(name=sheet.title, state=sheet.sheet_state)
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _validate_selected_worksheet(source: Path, sheet_name: str) -> None:
    workbook = load_workbook(source, read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError("選択したsheetがworkbookにありません")
        sheet = workbook[sheet_name]
        if sheet.sheet_state != "visible":
            raise ValueError("hidden sheetは単一表onboardingで選択できません")
        if sheet.merged_cells.ranges:
            raise ValueError("merged cellを含むsheetは単一の矩形表として扱えません")
        if any(
            cell.data_type == "f"
            for row in sheet.iter_rows()
            for cell in row
        ):
            raise ValueError("formula cellを含むsheetは保存値の由来を固定できないため扱えません")
        populated_rows = [
            row
            for row in sheet.iter_rows(values_only=True)
            if any(value not in (None, "") for value in row)
        ]
        if not populated_rows:
            raise ValueError("選択したsheetに表がありません")
        first = next(
            index
            for index, row in enumerate(sheet.iter_rows(values_only=True), start=1)
            if any(value not in (None, "") for value in row)
        )
        last = max(
            index
            for index, row in enumerate(sheet.iter_rows(values_only=True), start=1)
            if any(value not in (None, "") for value in row)
        )
        if first != 1:
            raise ValueError("一行目がheaderではありません")
        if any(
            not any(value not in (None, "") for value in row)
            for row in sheet.iter_rows(min_row=first, max_row=last, values_only=True)
        ):
            raise ValueError("表の途中に空行があり、単一の矩形表として扱えません")
    finally:
        workbook.close()


def _inspection_payload(
    source: Path,
    *,
    sheet_name: str | None = None,
) -> CsvInspectionResponse:
    source_kind: Literal["csv", "xlsx"] = (
        "xlsx" if source.suffix.lower() == ".xlsx" else "csv"
    )
    worksheets = _workbook_sheets(source) if source_kind == "xlsx" else []
    if source_kind == "xlsx" and sheet_name is None:
        return CsvInspectionResponse(
            source_filename=source.name,
            source_sha256=file_sha256(source),
            source_kind="xlsx",
            reader_policy="xlsx-stored-values-no-formulas/v1",
            worksheets=worksheets,
            selected_sheet=None,
            requires_sheet_selection=True,
            rows=0,
            relations=0,
            grain="one-row-one-observation",
            columns=[],
            task_id_contract=CsvTaskIdContract(
                pattern=TASK_ID_PATTERN,
                min_length=TASK_ID_MIN_LENGTH,
                example=TASK_ID_EXAMPLE,
            ),
            estimators=list(csv_onboarding_estimator_options()),
            default_estimator_id="ridge.v1",
            notice="visible sheetを明示選択してください。hidden sheetは選択できません。",
        )
    if source_kind == "xlsx":
        assert sheet_name is not None
        _validate_selected_worksheet(source, sheet_name)
    inspected = inspect_task_source(source, sheet=sheet_name)
    return CsvInspectionResponse(
        source_filename=inspected.source.name,
        source_sha256=inspected.source_sha256,
        source_kind=source_kind,
        reader_policy=(
            "xlsx-stored-values-no-formulas/v1"
            if source_kind == "xlsx"
            else "csv-utf8-header/v1"
        ),
        worksheets=worksheets,
        selected_sheet=inspected.selected_sheet,
        requires_sheet_selection=False,
        rows=inspected.row_count,
        relations=0,
        grain="one-row-one-observation",
        columns=[
            {
                "name": item.name,
                "kind": item.kind,
                "non_empty": item.non_empty,
                "observed_min": item.minimum,
                "observed_max": item.maximum,
                "choices": list(item.choices),
            }
            for item in inspected.columns
        ],
        task_id_contract=CsvTaskIdContract(
            pattern=TASK_ID_PATTERN,
            min_length=TASK_ID_MIN_LENGTH,
            example=TASK_ID_EXAMPLE,
        ),
        estimators=list(csv_onboarding_estimator_options()),
        default_estimator_id="ridge.v1",
        notice="観測最小値・最大値は要約です。物理的な許容範囲や目標値には自動で使いません。",
    )


@router.post(
    "/inspect",
    response_model=CsvInspectionResponse,
    responses={422: {"model": ApiError, "description": "Validation Error"}},
)
async def inspect_csv(
    file: UploadFile = File(...),
    sheet_name: str | None = Form(None),
) -> CsvInspectionResponse:
    temporary, source = await _uploaded_tabular(file)
    try:
        return await run_in_threadpool(
            _inspection_payload,
            source,
            sheet_name=sheet_name,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.info("CSV_ONBOARDING_FAILED stage=inspect error=%s", exc)
        raise HTTPException(
            422,
            {
                "code": "validation_error",
                "message": str(exc) or _FAILURE_MESSAGES["inspect"],
                "next_action": (
                    "複数表、relation、formula、merged cellを含むExcelは"
                    "Profile Workbenchで構造を確認してください。"
                ),
            },
        ) from exc
    finally:
        temporary.cleanup()


def _fields_from_json(fields_json: str) -> tuple[ScaffoldField, ...]:
    raw_fields = json.loads(fields_json)
    if not isinstance(raw_fields, list):
        raise ValueError("fields must be a list")
    return tuple(
        ScaffoldField(
            **{
                **item,
                **{
                    key: tuple(item[key])
                    for key in (
                        "allowed_range",
                        "default_range",
                        "training_range",
                        "plausible_range",
                        "display_range",
                    )
                    if item.get(key) is not None
                },
            }
        )
        for item in raw_fields
    )


def _new_onboarding_paths(
    *,
    task_id: str,
    task_store_path: Path | None,
    model_store_path: Path | None,
    label: str,
    source: Path,
    fields: tuple[ScaffoldField, ...],
    estimator_id: str,
    grain_confirmation: str,
    relation_confirmation: str,
    selected_sheet: str | None,
) -> OnboardingReservation:
    """Reserve new paths or recover one fully matching immutable identity."""

    csv_onboarding_estimator_recipe(estimator_id)
    if task_store_path is None:
        raise OnboardingStorageError(
            "task-store-unconfigured",
            "個人Taskの保存先がこのWorkspaceに設定されていません。",
            "ワークスペース → 保存場所を管理で個人Taskの保存先を確認し、APIを再起動してください。",
        )
    try:
        task_store = validate_personal_task_store_path(task_store_path)
    except (OSError, ValueError) as exc:
        raise OnboardingStorageError(
            "task-store-unavailable",
            "個人Taskの保存先を利用できません。保存先フォルダが作成できないか、安全境界外です。",
            "ワークスペース → 保存場所を管理で場所を確認し、書き込み可能なユーザー領域へ直してから再確認してください。",
        ) from exc
    task_root = (task_store / task_id).resolve()
    if task_root.parent != task_store.resolve():
        raise OnboardingStorageError(
            "task-id-invalid",
            "Task IDが保存先の安全境界を越えるため使用できません。",
            "Data Library → 新しい予測問題で、英数字・ハイフン中心のTask IDを入力してください。",
        )
    if model_store_path is None:
        raise OnboardingStorageError(
            "model-store-unconfigured",
            "個人Model / Packageの保存先がこのWorkspaceに設定されていません。",
            "ワークスペース → 保存場所を管理で個人Model / Packageの保存先を確認し、APIを再起動してください。",
        )
    try:
        model_store = validate_personal_model_store_path(model_store_path)
        model_store.mkdir(parents=True, exist_ok=True)
    except (OSError, ValueError) as exc:
        raise OnboardingStorageError(
            "model-store-unavailable",
            "個人Model / Packageの保存先を利用できません。保存先フォルダを作成できないか、安全境界外です。",
            "ワークスペース → 保存場所を管理で場所を確認し、書き込み可能なユーザー領域へ直してから再確認してください。",
        ) from exc
    package_id = (
        f"{task_id}-personal-1"
        if estimator_id == "ridge.v1"
        else f"{task_id}-{estimator_id.replace('.', '-')}-personal-1"
    )
    packages_root = (model_store / "packages").resolve()
    package_root = (packages_root / package_id).resolve()
    if package_root.parent != packages_root:
        raise OnboardingStorageError(
            "package-id-invalid",
            "Model Package IDが保存先の安全境界を越えるため使用できません。",
            "Data Library → 新しい予測問題で別のTask IDを入力してください。",
        )
    if package_root.exists() and not task_root.exists():
        raise OnboardingStorageError(
            "task-id-conflict",
            f"Task ID「{task_id}」には準備途中または異なる個人Task / Modelが残っています。既存内容は上書きしません。",
            "Data Library → 新しい予測問題で別のTask IDを入力し、準備を再試行してください。",
        )
    if task_root.exists():
        try:
            reusable_task = _load_reusable_task(
                root=task_root,
                task_id=task_id,
                label=label,
                source=source,
                fields=fields,
                estimator_id=estimator_id,
                grain_confirmation=grain_confirmation,
                relation_confirmation=relation_confirmation,
                selected_sheet=selected_sheet,
            )
            config_path = model_store / "available-packages.json"
            relative_package = package_root.relative_to(model_store).as_posix()
            package_is_available = (
                package_root.is_dir()
                and config_path.is_file()
                and relative_package in load_available_packages(config_path).packages
            )
            if package_root.exists() and not package_is_available:
                raise ValueError("existing Model Package is not available")
            if package_is_available:
                verify_model_package(
                    package_root,
                    task_id=task_id,
                    source=reusable_task.source_path,
                    profile=reusable_task.profile_path,
                )
        except (KeyError, OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
            raise OnboardingStorageError(
                "task-id-conflict",
                f"Task ID「{task_id}」は今回とは異なる個人Task / Modelとして保存されています。既存内容は上書きしません。",
                "Data Library → 新しい予測問題で別のTask IDを入力し、準備を再試行してください。",
            ) from exc
        return OnboardingReservation(
            task_root=task_root,
            model_store=model_store,
            package_id=package_id,
            reusable_task=reusable_task,
            reusable_package=package_is_available,
        )
    return OnboardingReservation(
        task_root=task_root,
        model_store=model_store,
        package_id=package_id,
        reusable_task=None,
        reusable_package=False,
    )


async def _rollback_prepare_attempt(
    *,
    request: Request,
    task_root: Path | None,
    model_store: Path | None,
    package_id: str | None,
    registration: DatasetRegistrationResult | None,
    registration_checkpoint: ManagedDatasetRegistrationCheckpoint | None,
    refresh: Any,
    runtime_published: bool,
) -> None:
    """Compensate only paths and catalog records reserved by this request."""

    if registration is not None and registration_checkpoint is not None:
        try:
            await run_in_threadpool(
                rollback_managed_dataset_registration,
                database=Path(request.app.state.workspace_database),
                registration=registration,
                checkpoint=registration_checkpoint,
            )
        except Exception:
            logger.exception("CSV_ONBOARDING_ROLLBACK_FAILED stage=dataset")
    if model_store is not None and package_id is not None:
        try:
            await run_in_threadpool(
                rollback_personal_package_attempt,
                package_id,
                store=model_store,
            )
        except Exception:
            logger.exception("CSV_ONBOARDING_ROLLBACK_FAILED stage=package")
    if task_root is not None and task_root.exists():
        try:
            await run_in_threadpool(shutil.rmtree, task_root)
        except Exception:
            logger.exception("CSV_ONBOARDING_ROLLBACK_FAILED stage=task")
    if runtime_published and refresh is not None:
        try:
            await refresh()
        except Exception:
            logger.exception("CSV_ONBOARDING_ROLLBACK_FAILED stage=runtime")


@router.post(
    "/prepare",
    response_model=CsvPrepareResponse,
    responses={422: {"model": ApiError, "description": "Validation Error"}},
)
async def prepare_csv_task(
    request: Request,
    task_id: Annotated[str, Form(min_length=TASK_ID_MIN_LENGTH, pattern=TASK_ID_PATTERN)],
    file: UploadFile = File(...),
    label: str = Form(...),
    estimator_id: str = Form("ridge.v1"),
    fields_json: str = Form(...),
    grain_confirmation: str = Form(...),
    relation_confirmation: str = Form(...),
    sheet_name: str | None = Form(None),
) -> CsvPrepareResponse:
    """Create, verify, promote, register, and reload one reviewed tabular Task."""

    try:
        csv_onboarding_estimator_recipe(estimator_id)
    except ValueError as exc:
        raise HTTPException(
            422,
            {
                "code": "validation_error",
                "message": "選択したEstimatorは、このCSV Taskで利用できません。",
                "next_action": str(exc),
                "field_errors": [
                    {"path": "estimator_id", "message": str(exc)}
                ],
            },
        ) from exc
    try:
        fields = _fields_from_json(fields_json)
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        logger.info("CSV_ONBOARDING_FAILED stage=fields error_type=%s", type(exc).__name__)
        raise HTTPException(422, _FAILURE_MESSAGES["fields"]) from exc
    try:
        temporary, source = await _uploaded_tabular(file)
        if source.suffix.lower() == ".xlsx" and not sheet_name:
            raise HTTPException(422, "Excelはsheetを明示確認してから準備してください")
        if source.suffix.lower() == ".csv" and sheet_name:
            raise HTTPException(422, "CSVにsheetは指定できません")
        await run_in_threadpool(
            _inspection_payload,
            source,
            sheet_name=sheet_name,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("CSV_ONBOARDING_FAILED stage=inspect")
        raise HTTPException(422, _FAILURE_MESSAGES["inspect"]) from exc
    try:
        lock = request.app.state.csv_onboarding_lock
        async with lock:
            task_root: Path | None = None
            model_store: Path | None = None
            package_id: str | None = None
            task_attempt_owned = False
            package_attempt_owned = False
            registration: DatasetRegistrationResult | None = None
            registration_checkpoint: ManagedDatasetRegistrationCheckpoint | None = None
            refresh = getattr(request.app.state, "refresh_task_resources", None)
            runtime_published = False
            stage = "storage"
            try:
                reservation = await run_in_threadpool(
                    _new_onboarding_paths,
                    task_id=task_id,
                    task_store_path=getattr(request.app.state, "task_store_path", None),
                    model_store_path=getattr(request.app.state, "model_store_path", None),
                    label=label,
                    source=source,
                    fields=fields,
                    estimator_id=estimator_id,
                    grain_confirmation=grain_confirmation,
                    relation_confirmation=relation_confirmation,
                    selected_sheet=sheet_name,
                )
                task_root = reservation.task_root
                model_store = reservation.model_store
                package_id = reservation.package_id
                reused_existing = (
                    reservation.reusable_task is not None
                    and reservation.reusable_package
                )
                task_attempt_owned = reservation.reusable_task is None
                package_attempt_owned = not reservation.reusable_package
                stage = "prepare"
                result = reservation.reusable_task
                if result is None:
                    result = await run_in_threadpool(
                        create_task_scaffold,
                        source=source,
                        task_id=task_id,
                        label=label,
                        fields=fields,
                        grain_confirmation=grain_confirmation,
                        relation_confirmation=relation_confirmation,
                        estimator_id=estimator_id,
                        sheet=sheet_name,
                    )
                if result.state != "ready" or result.profile_path is None:
                    raise ValueError("CSV onboarding requires a complete Task definition")
                if not reservation.reusable_package:
                    estimator_slug = estimator_id.replace(".", "-")
                    candidate = result.root / f"candidate-package-{estimator_slug}"
                    feature_dataset = result.root / f"feature-dataset-{estimator_slug}.json"
                    stage = "build"
                    await run_in_threadpool(
                        build_standard_package,
                        task_id,
                        result.source_path,
                        candidate,
                        feature_dataset,
                        package_id=package_id,
                        package_version="1.0.0",
                        replace=False,
                        estimator=estimator_id,
                        profile=result.profile_path,
                    )
                    stage = "verify"
                    await run_in_threadpool(
                        verify_model_package,
                        candidate,
                        task_id=task_id,
                        source=result.source_path,
                        profile=result.profile_path,
                    )
                    stage = "promote"
                    await run_in_threadpool(
                        promote_personal_package,
                        task_id,
                        candidate,
                        result.source_path,
                        model_store,
                        profile=result.profile_path,
                    )
                stage = "dataset"
                registration_checkpoint = await run_in_threadpool(
                    managed_dataset_registration_checkpoint,
                    Path(request.app.state.workspace_database),
                )
                registration = await run_in_threadpool(
                    register_managed_dataset,
                    database=Path(request.app.state.workspace_database),
                    source=result.source_path,
                    library_root=Path(request.app.state.data_library_root),
                    profile_path=result.profile_path,
                    name=label,
                    member_provenance={
                        "lineage_kind": "new_task_tabular_onboarding",
                        "original_filename": source.name,
                        "original_source_sha256": file_sha256(source),
                        "selected_sheet": sheet_name,
                        "reader_policy": (
                            "xlsx-stored-values-no-formulas/v1"
                            if source.suffix.lower() == ".xlsx"
                            else "csv-utf8-header/v1"
                        ),
                        "header_policy": "first-row-non-empty-unique/v1",
                        "canonical_snapshot_sha256": file_sha256(result.source_path),
                    },
                    promote_existing_bundled=False,
                )
                if refresh is None:
                    raise RuntimeError("task resource refresh is not available")
                stage = "refresh"
                await refresh()
                runtime_published = True
                runtime_catalog = request.app.state.runtime_context.workspace_catalog
                dataset_view = next(
                    (
                        item
                        for item in runtime_catalog.list_dataset_view_revisions()
                        if item.id == registration.dataset_view_revision_id
                    ),
                    None,
                )
                package = next(
                    (
                        item
                        for item in runtime_catalog.list_model_package_refs(
                            include_archived=False,
                        )
                        if item.task_id == task_id and item.package_id == package_id
                    ),
                    None,
                )
                if dataset_view is None or package is None:
                    raise RuntimeError("refreshed resources do not match the registered Dataset")
                return CsvPrepareResponse(
                    state="ready",
                    task_id=task_id,
                    dataset_view_revision_id=dataset_view.id,
                    dataset_revision_id=registration.dataset_revision_id,
                    source_sha256=registration.source_sha256,
                    model_package_ref_id=package.id,
                    reused_existing=reused_existing,
                )
            except OnboardingStorageError as exc:
                logger.info(
                    "CSV_ONBOARDING_FAILED stage=storage task_id=%s code=%s",
                    task_id,
                    exc.code,
                )
                await _rollback_prepare_attempt(
                    request=request,
                    task_root=task_root if task_attempt_owned else None,
                    model_store=model_store if package_attempt_owned else None,
                    package_id=package_id if package_attempt_owned else None,
                    registration=registration,
                    registration_checkpoint=registration_checkpoint,
                    refresh=refresh,
                    runtime_published=runtime_published,
                )
                raise HTTPException(422, exc.public_detail()) from exc
            except Exception as exc:
                logger.exception(
                    "CSV_ONBOARDING_FAILED stage=%s task_id=%s",
                    stage,
                    task_id,
                )
                await _rollback_prepare_attempt(
                    request=request,
                    task_root=task_root if task_attempt_owned else None,
                    model_store=model_store if package_attempt_owned else None,
                    package_id=package_id if package_attempt_owned else None,
                    registration=registration,
                    registration_checkpoint=registration_checkpoint,
                    refresh=refresh,
                    runtime_published=runtime_published,
                )
                raise HTTPException(422, _FAILURE_MESSAGES[stage]) from exc
    finally:
        temporary.cleanup()
