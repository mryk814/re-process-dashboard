"""Read-only intake diagnosis for the supported data shapes.

The catalog deliberately describes the shipped profile and Task seams; it does
not infer scientific meaning, create a Dataset, or select a Model Package.
"""
from __future__ import annotations

import csv
from pathlib import Path
import re
from typing import Any, Literal

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict

from decision_workbench.data.file_integrity import file_sha256
from decision_workbench.task_composition.builtin.catalog import BUILTIN_TASK_MODULES
from decision_workbench.tasks.task_registry import load_task_contracts


READINESS_SCHEMA_VERSION = "data-readiness/v1"

SourceShape = Literal[
    "independent_rows", "repeated_measurements", "longitudinal_curve",
    "wide_multi_target", "relational_workbook", "variable_length_series",
]
ReadinessState = Literal["ready", "profile_needed", "task_slice_needed", "unsupported"]
ColumnRole = Literal[
    "row_id", "candidate_input", "fixed_context", "grouping", "time_axis",
    "technical_metadata", "target", "unknown",
]


class _ReadinessModel(BaseModel):
    model_config = ConfigDict(frozen=True)


class CatalogEntry(_ReadinessModel):
    source_shape: SourceShape
    state: ReadinessState
    profile_families: tuple[str, ...]
    standard_onboarding: bool
    reason: str


class ReadinessCatalog(_ReadinessModel):
    schema_version: Literal[READINESS_SCHEMA_VERSION] = READINESS_SCHEMA_VERSION
    entries: tuple[CatalogEntry, ...]


class ReadinessColumn(_ReadinessModel):
    name: str
    suggested_role: ColumnRole
    non_empty: int
    missing: int


class TaskReadinessEntry(_ReadinessModel):
    task_id: str
    label: str
    source_shape: SourceShape
    profile_family: str
    split_policy: str
    input_kinds: tuple[str, ...]
    target_kind: Literal["single_target", "multi_target"]
    standard_authoring: bool


class ReadinessInventory(_ReadinessModel):
    schema_version: Literal[READINESS_SCHEMA_VERSION] = READINESS_SCHEMA_VERSION
    generated_from: tuple[str, ...]
    tasks: tuple[TaskReadinessEntry, ...]


class ReadinessPreflight(_ReadinessModel):
    schema_version: Literal[READINESS_SCHEMA_VERSION] = READINESS_SCHEMA_VERSION
    source_filename: str
    source_sha256: str
    source_shape: SourceShape | None
    state: ReadinessState
    available_profile_families: tuple[str, ...]
    standard_onboarding: bool
    route: Literal["standard_onboarding", "profile_workbench", "task_slice", "unsupported"]
    reasons: tuple[str, ...]
    ambiguities: tuple[str, ...]
    columns: tuple[ReadinessColumn, ...]
    target_availability: Literal["not_declared", "complete", "partial_by_target"]
    workbook_sheets: tuple[str, ...] = ()


_CATALOG = ReadinessCatalog(entries=(
    CatalogEntry(source_shape="independent_rows", state="ready", profile_families=("tabular-dataset-profile/v1",), standard_onboarding=True, reason="1行=1独立観測を人が確認できれば標準Tabular onboardingへ進めます。"),
    CatalogEntry(source_shape="repeated_measurements", state="profile_needed", profile_families=("observation-dataset-profile/v1", "tabular-dataset-profile/v1"), standard_onboarding=False, reason="conditionごとの反復はgroup splitと観測粒度をProfileで固定します。"),
    CatalogEntry(source_shape="longitudinal_curve", state="profile_needed", profile_families=("tabular-dataset-profile/v1", "observation-dataset-profile/v1"), standard_onboarding=False, reason="entity/run/cell内の軸順序とgroup splitをProfileで固定します。"),
    CatalogEntry(source_shape="wide_multi_target", state="ready", profile_families=("tabular-dataset-profile/v1",), standard_onboarding=True, reason="全targetが同じ観測行にある場合だけ、targetを明示して標準Tabular onboardingへ進めます。"),
    CatalogEntry(source_shape="relational_workbook", state="profile_needed", profile_families=("dataset-input-profile/v2",), standard_onboarding=False, reason="複数表とrelationは標準onboardingで自動結合せず、Profile Workbenchでkeyとjoinを確認します。"),
    CatalogEntry(source_shape="variable_length_series", state="task_slice_needed", profile_families=(), standard_onboarding=False, reason="可変長Seriesは通常Candidateとは別assetです。Task contract、feature表現、runtimeを含む縦スライスが必要です。"),
))

_SOURCE_KIND = {
    "welding_multistage": ("relational_workbook", "profile_defined"),
    "primary": ("relational_workbook", "profile_defined"),
    "flank_wear": ("longitudinal_curve", "grouped_tool_condition"),
    "external_wear_curve": ("longitudinal_curve", "grouped_tool_condition"),
    "external_battery_degradation": ("longitudinal_curve", "grouped_cell"),
    "welding_stage_c": ("repeated_measurements", "grouped_weld_run"),
    "welding_graph_synthetic_demonstration": (
        "repeated_measurements",
        "grouped_weld_run",
    ),
}
_PROFILE_FAMILY_BY_SOURCE_KIND = {
    "welding_multistage": "welding-stage-b-profile/v1",
    "primary": "dataset-input-profile/v2",
    "flank_wear": "dataset-input-profile/v2",
    "welding_stage_c": "observation-dataset-profile/v1",
    "welding_graph_synthetic_demonstration": "observation-dataset-profile/v1",
}
_GROUP = re.compile(r"(?:^|[_\-])(group|batch|run|cell|specimen|entity|condition)(?:[_\-]|$)", re.I)
_ROW_ID = re.compile(r"(?:^|[_\-])(id|uuid|record)(?:[_\-]|$)", re.I)
_AXIS = re.compile(r"(?:^|[_\-])(time|timestamp|cycle|sequence|step|index|age)(?:[_\-]|$)", re.I)
_TECHNICAL = re.compile(r"(?:operator|instrument|file|source|sheet|note|comment|metadata|version)", re.I)
_SERIES = re.compile(r"(?:series|curve|trace|waveform|spectrum|signal).*(?:path|file|ref|reference)|(?:path|file|ref|reference).*(?:series|curve|trace|waveform|spectrum|signal)", re.I)
_SHIPPED_TECHNICAL_COLUMNS = {
    "test_date",
    "specimen_number",
    "specimen_position",
    "notch_position",
    "試験日",
    "試験者",
    "試験片番号",
    "試験片位置",
    "ノッチ位置",
}


def readiness_catalog() -> ReadinessCatalog:
    return _CATALOG


def _profile_schema(module: Any) -> str:
    # These source kinds are the stable adapter seams used by the shipped loaders.
    # Unlisted built-in tabular tasks all use TabularProfileFamilyAdapter.
    return _PROFILE_FAMILY_BY_SOURCE_KIND.get(
        module.source_kind, "tabular-dataset-profile/v1"
    )


def build_readiness_inventory() -> ReadinessInventory:
    contracts = load_task_contracts()
    entries: list[TaskReadinessEntry] = []
    for task_id, module in sorted(BUILTIN_TASK_MODULES.items()):
        shape, split_policy = _SOURCE_KIND.get(module.source_kind, ("independent_rows", "row_independent"))
        contract = contracts[task_id].task_definition
        entries.append(TaskReadinessEntry(
            task_id=task_id,
            label=contract.label,
            source_shape=shape,
            profile_family=_profile_schema(module),
            split_policy=split_policy,
            input_kinds=tuple(group.key for group in contract.input_groups),
            target_kind="single_target" if len(contract.outputs) == 1 else "multi_target",
            standard_authoring=(shape == "independent_rows" and module.standard_model_authoring is not None),
        ))
    return ReadinessInventory(
        generated_from=(
            "backend/src/decision_workbench/developer_experience/readiness.py",
            "backend/src/decision_workbench/task_composition/builtin/catalog.py",
            "backend/src/decision_workbench/tasks/task_definitions",
        ),
        tasks=tuple(entries),
    )


def _load_rows(source: Path) -> tuple[list[dict[str, Any]], tuple[str, ...]]:
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            if not reader.fieldnames:
                raise ValueError("CSV header is required for readiness preflight")
            names = tuple(str(name).strip() for name in reader.fieldnames)
            if not all(names) or len(set(names)) != len(names):
                raise ValueError("CSV headers must be non-empty and unique")
            return [dict(row) for row in reader], ()
    if source.suffix.lower() != ".xlsx":
        raise ValueError("readiness preflight accepts .csv or .xlsx")
    workbook = load_workbook(source, read_only=True, data_only=False)
    try:
        visible = tuple(sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible")
        if not visible:
            raise ValueError("workbook has no visible sheet")
        if len(visible) != 1:
            return [], visible
        values = workbook[visible[0]].iter_rows(values_only=True)
        headers = next(values, None)
        if headers is None:
            raise ValueError("worksheet header is required")
        names = tuple(str(value).strip() if value is not None else "" for value in headers)
        if not all(names) or len(set(names)) != len(names):
            raise ValueError("worksheet headers must be non-empty and unique")
        rows = [
            {name: value for name, value in zip(names, row, strict=True)}
            for row in values
            if any(value not in (None, "") for value in row)
        ]
        return rows, visible
    finally:
        workbook.close()


def _role(name: str, targets: set[str]) -> ColumnRole:
    if name in targets:
        return "target"
    if _SERIES.search(name):
        return "fixed_context"
    if name.strip() in _SHIPPED_TECHNICAL_COLUMNS:
        return "technical_metadata"
    if _AXIS.search(name):
        return "time_axis"
    if _TECHNICAL.search(name):
        return "technical_metadata"
    if _GROUP.search(name):
        return "grouping"
    if _ROW_ID.search(name):
        return "row_id"
    return "candidate_input"


def _catalog_entry(shape: SourceShape) -> CatalogEntry:
    return next(entry for entry in _CATALOG.entries if entry.source_shape == shape)


def preflight_source(
    source: Path,
    *,
    target_columns: tuple[str, ...] = (),
) -> ReadinessPreflight:
    rows, sheets = _load_rows(source)
    digest = file_sha256(source)
    if len(sheets) > 1:
        entry = _catalog_entry("relational_workbook")
        return ReadinessPreflight(
            source_filename=source.name, source_sha256=digest, source_shape="relational_workbook",
            state=entry.state, available_profile_families=entry.profile_families,
            standard_onboarding=False, route="profile_workbench", reasons=(entry.reason,),
            ambiguities=("table間のkey、relationの向き、学習行を確認してください。",),
            columns=(), target_availability="not_declared", workbook_sheets=sheets,
        )
    if not rows:
        return ReadinessPreflight(
            source_filename=source.name, source_sha256=digest, source_shape=None, state="unsupported",
            available_profile_families=(), standard_onboarding=False, route="unsupported",
            reasons=("headerを持つデータ行を確認できません。未知のTabularとして続行しません。",),
            ambiguities=(), columns=(), target_availability="not_declared", workbook_sheets=sheets,
        )
    headers = tuple(rows[0])
    unknown_targets = tuple(name for name in target_columns if name not in headers)
    if unknown_targets:
        raise ValueError(f"target columns were not found: {', '.join(unknown_targets)}")
    targets = set(target_columns)
    columns = tuple(ReadinessColumn(
        name=name,
        suggested_role=_role(name, targets),
        non_empty=sum(row.get(name) not in (None, "") for row in rows),
        missing=sum(row.get(name) in (None, "") for row in rows),
    ) for name in headers)
    role_by_name = {column.name: column.suggested_role for column in columns}
    group_columns = [name for name, role in role_by_name.items() if role == "grouping"]
    axis_columns = [name for name, role in role_by_name.items() if role == "time_axis"]
    has_series_reference = any(_SERIES.search(name) for name in headers)
    repeated = any(len({str(row.get(name)) for row in rows}) < len(rows) for name in group_columns)
    target_missing_masks = {
        name: tuple(index for index, row in enumerate(rows) if row.get(name) in (None, ""))
        for name in targets
    }
    any_target_missing = any(target_missing_masks.values())
    target_masks_differ = len(set(target_missing_masks.values())) > 1
    if has_series_reference:
        shape: SourceShape = "variable_length_series"
    elif group_columns and axis_columns and repeated:
        shape = "longitudinal_curve"
    elif group_columns and repeated:
        shape = "repeated_measurements"
    elif any_target_missing:
        shape = "wide_multi_target"
    elif len(targets) > 1:
        shape = "wide_multi_target"
    else:
        shape = "independent_rows"
    entry = _catalog_entry(shape)
    state = entry.state
    reasons = [entry.reason]
    ambiguities: list[str] = []
    if not targets:
        ambiguities.append("目的変数は未宣言です。候補入力へ自動採用せず、Data Libraryで明示してください。")
    if shape == "independent_rows":
        ambiguities.append("1行=1独立観測であり、同一conditionの反復ではないことを確認してください。")
    if any_target_missing:
        state = "profile_needed"
        reasons.append("目的変数に欠損があります。標準onboardingで暗黙に行を落とさず、Profileで欠損方針を明示します。")
    if target_masks_differ:
        reasons.append("targetごとに欠損行が異なります。全targetの共通行だけへ縮めず、target別cohortをProfileで保持します。")
    missing_inputs = [column.name for column in columns if column.suggested_role == "candidate_input" and column.missing]
    if missing_inputs:
        state = "profile_needed"
        reasons.append("候補入力に欠損があります。補完・除外・fixed contextの扱いをProfileで明示します。")
    route: Literal["standard_onboarding", "profile_workbench", "task_slice", "unsupported"] = {
        "ready": "standard_onboarding", "profile_needed": "profile_workbench",
        "task_slice_needed": "task_slice", "unsupported": "unsupported",
    }[state]
    return ReadinessPreflight(
        source_filename=source.name, source_sha256=digest, source_shape=shape, state=state,
        available_profile_families=entry.profile_families,
        standard_onboarding=state == "ready" and entry.standard_onboarding,
        route=route, reasons=tuple(reasons), ambiguities=tuple(ambiguities),
        columns=columns,
        target_availability=("not_declared" if not targets else "partial_by_target" if any_target_missing else "complete"),
        workbook_sheets=sheets,
    )
