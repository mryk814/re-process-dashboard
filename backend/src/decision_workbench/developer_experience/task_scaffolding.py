"""Generate reviewed, data-only tabular Task bundles outside the repository."""
from __future__ import annotations

import csv
from contextlib import contextmanager
from contextvars import ContextVar
from dataclasses import dataclass
import json
import math
import os
from pathlib import Path
import re
import shutil
from statistics import median
from tempfile import NamedTemporaryFile
import time
from typing import Any, Literal, Sequence
from uuid import uuid4

from openpyxl import load_workbook

from decision_workbench.contracts.task_contracts import TaskContractFixture
from decision_workbench.data.file_integrity import file_sha256
from decision_workbench.modeling.tabular.profile import TabularDatasetProfile
from decision_workbench.modeling.training.recipe import (
    CSV_ONBOARDING_ESTIMATOR_IDS,
    estimator_recipe,
)
from decision_workbench.bootstrap.dev_workspace_storage import (
    validate_personal_or_dev_store,
)


TASK_BUNDLE_SCHEMA_VERSION = "external-task-bundle/v1"
TASK_SCAFFOLD_SCHEMA_VERSION = "task-scaffold/v1"
SUPPORTED_ESTIMATORS = CSV_ONBOARDING_ESTIMATOR_IDS
TASK_ID_PATTERN = r"^[a-z][a-z0-9-]{2,79}-v[1-9][0-9]*$"
TASK_ID_MIN_LENGTH = 6
TASK_ID_EXAMPLE = "concrete-slump-v1"
_TASK_ID = re.compile(TASK_ID_PATTERN)
_PERSONAL_TASK_STORE_OVERRIDE: ContextVar[Path | None] = ContextVar(
    "personal_task_store_override",
    default=None,
)
_KEY = re.compile(r"^[A-Za-z][A-Za-z0-9_]*$")
_WINDOWS_REPLACE_RETRY_ENABLED = os.name == "nt"
_WINDOWS_REPLACE_MAX_RETRIES = 5
_WINDOWS_REPLACE_RETRY_SECONDS = 0.05


@dataclass(frozen=True)
class SourceColumn:
    name: str
    kind: Literal["number", "categorical"]
    non_empty: int
    minimum: float | None
    maximum: float | None
    median: float | str | None
    choices: tuple[str, ...]


@dataclass(frozen=True)
class SourceInspection:
    source: Path
    source_sha256: str
    selected_sheet: str | None
    row_count: int
    columns: tuple[SourceColumn, ...]
    rows: tuple[dict[str, Any], ...]


@dataclass(frozen=True)
class ScaffoldField:
    column: str
    role: Literal["composition", "process", "categorical", "output"]
    key: str
    label: str
    unit: str | None
    goal_direction: Literal["at_least", "at_most", "target"] = "target"
    allowed_range: tuple[float, float] | None = None
    default_range: tuple[float, float] | None = None
    training_range: tuple[float, float] | None = None
    plausible_range: tuple[float, float] | None = None
    display_range: tuple[float, float] | None = None


@dataclass(frozen=True)
class TaskScaffoldResult:
    task_id: str
    state: Literal["draft", "ready"]
    root: Path
    source_path: Path
    profile_path: Path | None
    task_definition_path: Path | None
    training_recipe_path: Path | None
    unresolved: tuple[str, ...]


def personal_task_store_path() -> Path:
    overridden = _PERSONAL_TASK_STORE_OVERRIDE.get()
    if overridden is not None:
        return overridden
    configured = os.getenv("WORKBENCH_TASK_STORE_PATH", "").strip()
    if configured:
        return Path(configured).expanduser()
    local = os.getenv("LOCALAPPDATA", "").strip()
    if local:
        return Path(local) / "Material Decision Workbench" / "tasks"
    return Path.home() / ".material-decision-workbench" / "tasks"


@contextmanager
def use_personal_task_store(path: Path):
    """Bind one app's configured Task store across async request work."""

    token = _PERSONAL_TASK_STORE_OVERRIDE.set(path.expanduser().resolve())
    try:
        yield
    finally:
        _PERSONAL_TASK_STORE_OVERRIDE.reset(token)


def validate_personal_task_store_path(path: Path | None = None) -> Path:
    store = validate_personal_or_dev_store(
        path or personal_task_store_path(),
        resource_kind="task",
    )
    store.mkdir(parents=True, exist_ok=True)
    return store


def _tabular_rows(source: Path, sheet: str | None) -> tuple[str | None, list[dict[str, Any]]]:
    suffix = source.suffix.lower()
    if suffix == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            raw_names = reader.fieldnames
            if raw_names is None:
                raise ValueError("CSV has no header row")
            names = [str(value).strip() if value is not None else "" for value in raw_names]
            if any(not name for name in names) or len(names) != len(set(names)):
                raise ValueError("header names must be non-empty and unique")
            reader.fieldnames = names
            return None, [dict(row) for row in reader]
    if suffix != ".xlsx":
        raise ValueError("Task scaffold accepts .csv or .xlsx")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        selected = sheet or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            raise ValueError(f"Excel sheet was not found: {selected}")
        values = workbook[selected].iter_rows(values_only=True)
        headers = next(values, None)
        if headers is None:
            raise ValueError("selected sheet has no header row")
        names = [str(value).strip() if value is not None else "" for value in headers]
        if any(not name for name in names) or len(names) != len(set(names)):
            raise ValueError("header names must be non-empty and unique")
        rows = [
            {name: value for name, value in zip(names, row, strict=True)}
            for row in values
            if any(value not in (None, "") for value in row)
        ]
        return selected, rows
    finally:
        workbook.close()


def _number(value: Any, *, preserve_stored_type: bool = False) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if preserve_stored_type and isinstance(value, str):
        return None
    if (
        isinstance(value, str)
        and re.fullmatch(r"[+-]?0\d+", value.strip()) is not None
    ):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def inspect_task_source(source: Path, *, sheet: str | None = None) -> SourceInspection:
    source = source.expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    selected_sheet, rows = _tabular_rows(source, sheet)
    if len(rows) < 3:
        raise ValueError("Task scaffold requires at least 3 data rows")
    headers = tuple(rows[0])
    columns: list[SourceColumn] = []
    for name in headers:
        raw = [row.get(name) for row in rows if row.get(name) not in (None, "")]
        numbers = [
            _number(value, preserve_stored_type=selected_sheet is not None)
            for value in raw
        ]
        numeric = bool(raw) and all(value is not None for value in numbers)
        if numeric:
            values = [float(value) for value in numbers if value is not None]
            columns.append(SourceColumn(
                name=name,
                kind="number",
                non_empty=len(values),
                minimum=min(values),
                maximum=max(values),
                median=median(values),
                choices=(),
            ))
        else:
            choices = tuple(sorted({str(value).strip() for value in raw}))
            columns.append(SourceColumn(
                name=name,
                kind="categorical",
                non_empty=len(raw),
                minimum=None,
                maximum=None,
                median=choices[0] if choices else None,
                choices=choices,
            ))
    return SourceInspection(
        source=source,
        source_sha256=file_sha256(source),
        selected_sheet=selected_sheet,
        row_count=len(rows),
        columns=tuple(columns),
        rows=tuple(rows),
    )


def _range_payload(bounds: tuple[float, float]) -> dict[str, float]:
    return {"min": bounds[0], "max": bounds[1]}


def _valid_range(bounds: tuple[float, float] | None) -> bool:
    return (
        bounds is not None
        and all(math.isfinite(value) for value in bounds)
        and bounds[0] < bounds[1]
    )


def _contains(
    outer: tuple[float, float],
    inner: tuple[float, float],
) -> bool:
    return outer[0] <= inner[0] and inner[1] <= outer[1]


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _write_canonical_csv(
    path: Path,
    rows: Sequence[dict[str, Any]],
    columns: Sequence[str],
) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(columns), lineterminator="\n")
        writer.writeheader()
        writer.writerows({column: row.get(column) for column in columns} for row in rows)


def create_task_scaffold(
    *,
    source: Path,
    task_id: str,
    label: str,
    fields: Sequence[ScaffoldField],
    grain_confirmation: Literal["one-row-one-observation"],
    relation_confirmation: Literal["no-relations"],
    estimator_id: str = "ridge.v1",
    sheet: str | None = None,
    store: Path | None = None,
) -> TaskScaffoldResult:
    if _TASK_ID.fullmatch(task_id) is None:
        raise ValueError("task id must look like material-property-v1")
    from decision_workbench.task_composition.builtin.catalog import (
        BUILTIN_TASK_MODULES,
    )

    if task_id in BUILTIN_TASK_MODULES:
        raise ValueError(
            f"個人Taskは同梱Task IDを置き換えられません: {task_id}"
        )
    if not label.strip():
        raise ValueError("task label is required")
    if estimator_id not in SUPPORTED_ESTIMATORS:
        raise ValueError(
            f"unsupported scaffold estimator: {estimator_id}; "
            f"choose one of {', '.join(SUPPORTED_ESTIMATORS)}"
        )
    if grain_confirmation != "one-row-one-observation":
        raise ValueError(
            "学習一行を確認してください。標準scaffoldはone-row-one-observationだけを扱います"
        )
    if relation_confirmation != "no-relations":
        raise ValueError(
            "relationを確認してください。relationがあるデータは専用Task設計が必要です"
        )
    inspection = inspect_task_source(source, sheet=sheet)
    by_name = {column.name: column for column in inspection.columns}
    unresolved: list[str] = []
    seen_columns: set[str] = set()
    seen_paths: set[str] = set()
    outputs = [field for field in fields if field.role == "output"]
    inputs = [field for field in fields if field.role != "output"]
    if not inputs:
        unresolved.append("入力列を1件以上指定してください")
    if not outputs:
        unresolved.append("予測対象を1件以上指定してください")
    for field in fields:
        if field.role not in {"composition", "process", "categorical", "output"}:
            unresolved.append(f"列の役割を確認してください: {field.column}")
            continue
        if (
            field.role == "output"
            and field.goal_direction not in {"at_least", "at_most", "target"}
        ):
            unresolved.append(f"目標方向を確認してください: {field.column}")
        if field.column not in by_name:
            unresolved.append(f"列が見つかりません: {field.column}")
        if field.column in seen_columns:
            unresolved.append(f"列の役割が重複しています: {field.column}")
        seen_columns.add(field.column)
        if _KEY.fullmatch(field.key) is None:
            unresolved.append(f"canonical keyを確認してください: {field.key}")
        path = field.key if field.role == "output" else f"{field.role}.{field.key}"
        if path in seen_paths:
            unresolved.append(f"canonical pathが重複しています: {path}")
        seen_paths.add(path)
        column = by_name.get(field.column)
        if column is None:
            continue
        if field.role != "categorical" and column.kind != "number":
            unresolved.append(f"数値列として読めません: {field.column}")
        if (
            field.role != "categorical"
            and column.kind == "number"
            and column.minimum == column.maximum
        ):
            unresolved.append(
                f"数値範囲を確認できません（全行が同じ値です）: {field.column}"
            )
        if field.role == "categorical" and not column.choices:
            unresolved.append(f"カテゴリ値を確認できません: {field.column}")
        if field.role != "categorical" and not (field.unit or "").strip():
            unresolved.append(f"単位を明示してください: {field.column}")
        if not field.label.strip():
            unresolved.append(f"表示名を明示してください: {field.column}")
        if field.role in {"composition", "process"}:
            if not _valid_range(field.allowed_range):
                unresolved.append(f"物理的な許容範囲を明示してください: {field.column}")
            if not _valid_range(field.default_range):
                unresolved.append(f"通常使う範囲を明示してください: {field.column}")
            if not _valid_range(field.training_range):
                unresolved.append(f"学習範囲を確認してください: {field.column}")
            if (
                _valid_range(field.allowed_range)
                and _valid_range(field.default_range)
                and not _contains(field.allowed_range, field.default_range)
            ):
                unresolved.append(f"通常使う範囲が許容範囲外です: {field.column}")
            if (
                _valid_range(field.allowed_range)
                and _valid_range(field.training_range)
                and not _contains(field.allowed_range, field.training_range)
            ):
                unresolved.append(f"学習範囲が許容範囲外です: {field.column}")
            observed = (column.minimum, column.maximum)
            if (
                _valid_range(field.training_range)
                and observed[0] is not None
                and observed[1] is not None
                and field.training_range != observed
            ):
                unresolved.append(
                    f"学習範囲はinspect結果と一致させて確認してください: {field.column}"
                )
        if field.role == "output":
            if not _valid_range(field.plausible_range):
                unresolved.append(f"出力の妥当範囲を明示してください: {field.column}")
            if not _valid_range(field.display_range):
                unresolved.append(f"出力の表示範囲を明示してください: {field.column}")
            if (
                _valid_range(field.plausible_range)
                and _valid_range(field.display_range)
                and not _contains(field.plausible_range, field.display_range)
            ):
                unresolved.append(f"出力の表示範囲が妥当範囲外です: {field.column}")

    root = validate_personal_task_store_path(store) / task_id
    replaced_draft: Path | None = None
    if root.exists():
        scaffold_path = root / "scaffold.json"
        existing = (
            json.loads(scaffold_path.read_text(encoding="utf-8"))
            if scaffold_path.is_file()
            else {}
        )
        if (
            existing.get("state") != "draft"
            or existing.get("source", {}).get("original_sha256")
            != inspection.source_sha256
        ):
            raise FileExistsError(f"Task scaffold already exists: {root}")
        replaced_draft = root.with_name(f".{root.name}.draft-{uuid4().hex}")
        root.replace(replaced_draft)
    root.mkdir(parents=True)
    try:
        original = root / f"source-original{inspection.source.suffix.lower()}"
        shutil.copyfile(inspection.source, original)
        source_csv = root / "source.csv"
        _write_canonical_csv(
            source_csv,
            inspection.rows,
            [column.name for column in inspection.columns],
        )
        scaffold_payload = {
            "schema_version": TASK_SCAFFOLD_SCHEMA_VERSION,
            "task_id": task_id,
            "label": label.strip(),
            "state": "draft" if unresolved else "ready",
            "source": {
                "original_filename": inspection.source.name,
                "original_sha256": inspection.source_sha256,
                "selected_sheet": inspection.selected_sheet,
                "reader_policy": (
                    "xlsx-stored-values-no-formulas/v1"
                    if inspection.source.suffix.lower() == ".xlsx"
                    else "csv-utf8-header/v1"
                ),
                "header_policy": "first-row-non-empty-unique/v1",
                "materialized_csv": source_csv.name,
                "materialized_sha256": file_sha256(source_csv),
                "rows": inspection.row_count,
            },
            "fields": [field.__dict__ for field in fields],
            "estimator": estimator_recipe(estimator_id).model_dump(mode="json"),
            "standard_estimator_ids": list(SUPPORTED_ESTIMATORS),
            "unresolved": unresolved,
            "safety": {
                "meaning_and_units_confirmed": not unresolved,
                "grain_confirmation": grain_confirmation,
                "relation_confirmation": relation_confirmation,
                "ranges_explicitly_confirmed": not unresolved,
                "loads_python_code": False,
                "adapter_family": "tabular-regression",
                "store_scope": "personal",
            },
        }
        _write_json(root / "scaffold.json", scaffold_payload)
        if unresolved:
            if replaced_draft is not None:
                shutil.rmtree(replaced_draft)
            return TaskScaffoldResult(
                task_id=task_id,
                state="draft",
                root=root,
                source_path=source_csv,
                profile_path=None,
                task_definition_path=None,
                training_recipe_path=None,
                unresolved=tuple(unresolved),
            )

        input_groups: list[dict[str, Any]] = []
        candidate: dict[str, Any] = {
            "schema_version": "canonical-candidate/v1",
            "task_id": task_id,
            "composition": {},
            "process": {},
            "categorical": {},
            "heat_pattern": None,
            "provenance": {"source_kind": "direct", "source_ref": None},
        }
        profile_inputs: list[dict[str, Any]] = []
        display_decimals: dict[str, int] = {}
        for group_order, role in enumerate(("composition", "process", "categorical")):
            group_fields: list[dict[str, Any]] = []
            for field in [item for item in inputs if item.role == role]:
                column = by_name[field.column]
                path = f"{role}.{field.key}"
                if role == "categorical":
                    choices = list(column.choices)
                    definition = {
                        "path": path,
                        "kind": "categorical",
                        "order": len(group_fields),
                        "label": field.label.strip(),
                        "required": True,
                        "editable": True,
                        "choices": choices,
                    }
                    candidate[role][field.key] = str(column.median)
                    profile_inputs.append({
                        "path": path,
                        "column": field.column,
                        "kind": "categorical",
                        "choices": choices,
                    })
                else:
                    assert column.minimum is not None and column.maximum is not None
                    assert field.training_range is not None
                    assert field.allowed_range is not None
                    assert field.default_range is not None
                    definition = {
                        "path": path,
                        "kind": "number",
                        "order": len(group_fields),
                        "label": field.label.strip(),
                        "unit": field.unit,
                        "required": True,
                        "editable": True,
                        "default_range": _range_payload(field.default_range),
                        "allowed_range": _range_payload(field.allowed_range),
                        "training_range": _range_payload(field.training_range),
                    }
                    candidate[role][field.key] = float(column.median)
                    profile_inputs.append({
                        "path": path,
                        "column": field.column,
                        "kind": "number",
                        "unit": field.unit,
                    })
                    display_decimals[path] = 4
                group_fields.append(definition)
            if group_fields:
                input_groups.append({
                    "key": role,
                    "order": len(input_groups),
                    "label": {
                        "composition": "組成",
                        "process": "工程条件",
                        "categorical": "カテゴリ",
                    }[role],
                    "fields": group_fields,
                })

        output_definitions: list[dict[str, Any]] = []
        profile_outputs: list[dict[str, Any]] = []
        runtime_targets: list[dict[str, Any]] = []
        for field in outputs:
            column = by_name[field.column]
            assert column.minimum is not None and column.maximum is not None
            assert field.plausible_range is not None
            assert field.display_range is not None
            output_definitions.append({
                "key": field.key,
                "label": field.label.strip(),
                "unit": field.unit,
                "goal_direction": field.goal_direction,
                "measurement_keys": [field.key],
                "plausibility_range": _range_payload(field.plausible_range),
                "preferred_display_range": _range_payload(field.display_range),
            })
            profile_outputs.append({
                "key": field.key,
                "column": field.column,
                "unit": field.unit,
            })
            display_decimals[f"output.{field.key}"] = 4
            runtime_targets.append({
                "target": field.key,
                "point_statistics": ["mean"],
                "standard_deviation": False,
                "quantiles": True,
                "samples": False,
                "parametric_distribution": False,
                "uncertainty_components": False,
                "support": True,
                "warnings": True,
                "goal_probability": "unavailable",
            })

        fixture_payload = {
            "task_definition": {
                "schema_version": "task-definition/v1",
                "id": task_id,
                "label": label.strip(),
                "canonical_candidate_schema_version": "canonical-candidate/v1",
                "input_groups": input_groups,
                "outputs": output_definitions,
                "display_decimals": display_decimals,
                "fixed_context": [{
                    "path": "context.source",
                    "order": 0,
                    "label": "データ",
                    "value": inspection.source.name,
                }],
                "constraints": [],
                "response_curve_variables": [
                    {
                        "kind": "numeric_input",
                        "order": index,
                        "label": field.label.strip(),
                        "path": f"{field.role}.{field.key}",
                    }
                    for index, field in enumerate(
                        item for item in inputs if item.role != "categorical"
                    )
                ],
            },
            "canonical_candidate": candidate,
            "runtime_capability": {
                "schema_version": "runtime-capability/v1",
                "task_id": task_id,
                "model_package_schema_version": "model-package/v1",
                "targets": runtime_targets,
                "joint_samples": False,
                "operations": {
                    "preview": True,
                    "detailed_prediction": True,
                    "response_curve": True,
                    "similarity": True,
                    "snapshot": True,
                    "actual_measurement": True,
                },
            },
        }
        fixture = TaskContractFixture.model_validate(fixture_payload)
        profile = TabularDatasetProfile.model_validate({
            "schema_version": "tabular-dataset-profile/v1",
            "profile_id": f"{task_id}-profile",
            "name": f"{label.strip()} Dataset Profile",
            "task_id": task_id,
            "package_id": f"{task_id}-personal",
            "id_column": None,
            "group_column": None,
            "inputs": profile_inputs,
            "outputs": profile_outputs,
        })
        task_path = root / "task-definition.json"
        profile_path = root / "dataset-profile.json"
        recipe_path = root / "training-recipe.json"
        _write_json(task_path, fixture.model_dump(mode="json"))
        _write_json(profile_path, profile.model_dump(mode="json", exclude_none=True))
        _write_json(recipe_path, {
            "schema_version": "model-training-recipe/v1",
            "estimator": estimator_recipe(estimator_id).model_dump(mode="json"),
        })
        _write_json(root / "bundle.json", {
            "schema_version": TASK_BUNDLE_SCHEMA_VERSION,
            "task_id": task_id,
            "label": label.strip(),
            "state": "ready",
            "source_path": source_csv.name,
            "profile_path": profile_path.name,
            "task_definition_path": task_path.name,
            "training_recipe_path": recipe_path.name,
            "estimator_id": estimator_id,
            "standard_estimator_ids": list(SUPPORTED_ESTIMATORS),
            "package_path": None,
            "loads_python_code": False,
            "grain_confirmation": grain_confirmation,
            "relation_confirmation": relation_confirmation,
        })
        if replaced_draft is not None:
            shutil.rmtree(replaced_draft)
        return TaskScaffoldResult(
            task_id=task_id,
            state="ready",
            root=root,
            source_path=source_csv,
            profile_path=profile_path,
            task_definition_path=task_path,
            training_recipe_path=recipe_path,
            unresolved=(),
        )
    except Exception:
        shutil.rmtree(root, ignore_errors=True)
        if replaced_draft is not None and replaced_draft.exists():
            replaced_draft.replace(root)
        raise


def link_promoted_package(
    task_id: str,
    package_path: Path,
    *,
    store: Path | None = None,
) -> bool:
    root = validate_personal_task_store_path(store)
    if not root.exists():
        return False
    bundle_path = root / task_id / "bundle.json"
    if not bundle_path.is_file():
        return False
    payload = json.loads(bundle_path.read_text(encoding="utf-8"))
    if payload.get("schema_version") != TASK_BUNDLE_SCHEMA_VERSION:
        raise ValueError(f"unsupported external Task bundle: {bundle_path}")
    payload["package_path"] = str(package_path.expanduser().resolve())
    with NamedTemporaryFile(
        "w",
        encoding="utf-8",
        dir=bundle_path.parent,
        prefix=".bundle.",
        suffix=".json",
        delete=False,
    ) as stream:
        json.dump(payload, stream, ensure_ascii=False, indent=2)
        stream.write("\n")
        temporary = Path(stream.name)
    try:
        retries = 0
        while True:
            try:
                temporary.replace(bundle_path)
                break
            except PermissionError:
                if (
                    not _WINDOWS_REPLACE_RETRY_ENABLED
                    or retries >= _WINDOWS_REPLACE_MAX_RETRIES
                ):
                    raise
                retries += 1
                time.sleep(_WINDOWS_REPLACE_RETRY_SECONDS)
    finally:
        temporary.unlink(missing_ok=True)
    return True
