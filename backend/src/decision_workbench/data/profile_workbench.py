"""Read-only developer tooling for inspecting workbook Profiles."""
from __future__ import annotations

import json
import os
import re
from collections import Counter
from copy import deepcopy
from difflib import SequenceMatcher
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import load_workbook

from decision_workbench.contracts.dataset_disposition_contracts import (
    DATASET_CANONICALIZATION_CONTRACT_DIGEST,
    build_count_disposition,
    build_dataset_disposition,
)
from decision_workbench.data.evidence_images import technical_field_is_optional
from decision_workbench.data.file_integrity import file_sha256
from decision_workbench.data.importer import detect_dataset_profile_path
from decision_workbench.data.profiles.canonicalization import canonicalize_workbook
from decision_workbench.data.profiles.loading import (
    load_dataset_profile,
    materialize_dataset_profile_document,
)
from decision_workbench.data.profiles.requirements import task_data_requirements
from decision_workbench.data.profiles.schema import (
    DatasetProfileError,
    source_units_for,
    unit_conversion,
)
from decision_workbench.modeling.model_lifecycle import dataset_profile_digest
from decision_workbench.bootstrap.dev_workspace_storage import (
    validate_personal_or_dev_store,
)

_REPOSITORY_ROOT = Path(__file__).resolve().parents[4]
_PROFILE_SCHEMA_VERSION = "dataset-input-profile/v2"
_HEADER_UNIT = re.compile(r"\[([^\[\]]+)\]\s*$")
_SEMANTIC_KIND_ORDER = {
    "entity_key": 0,
    "relation_join": 1,
    "input": 2,
    "output": 3,
    "series": 4,
    "policy": 5,
    "technical": 6,
}


def _header_unit(column: str) -> str | None:
    match = _HEADER_UNIT.search(column)
    if match:
        return match.group(1)
    if column.endswith("%") and len(column) > 1:
        return "%"
    return None


def personal_profile_store_path() -> Path:
    """Return the external, user-owned Profile store."""

    configured = os.getenv("WORKBENCH_PROFILE_STORE_PATH", "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    local_app_data = os.getenv("LOCALAPPDATA", "").strip()
    if local_app_data:
        return (Path(local_app_data) / "Material Decision Workbench" / "profiles").resolve()
    return (Path.home() / ".local" / "share" / "Material Decision Workbench" / "profiles").resolve()


def validate_personal_profile_store_path(path: Path | None = None) -> Path:
    """Reject a personal store inside the source checkout."""

    return validate_personal_or_dev_store(
        path or personal_profile_store_path(),
        resource_kind="profile",
    )


def personal_profile_paths(path: Path | None = None) -> tuple[Path, ...]:
    """List valid persisted effective Profiles without trusting sidecar files."""

    from decision_workbench.data.profile_family_registry import load_profile_document

    store = validate_personal_profile_store_path(path)
    if not store.is_dir():
        return ()
    result: list[Path] = []
    for candidate in sorted(store.glob("*.json")):
        if not re.fullmatch(r"[0-9a-f]{64}\.json", candidate.name):
            continue
        try:
            raw = json.loads(candidate.read_text(encoding="utf-8"))
            if "extends" in raw:
                continue
            load_profile_document(candidate)
            if dataset_profile_digest(candidate) != f"sha256:{candidate.stem}":
                continue
        except (DatasetProfileError, json.JSONDecodeError, OSError, TypeError, ValueError):
            continue
        result.append(candidate.resolve())
    return tuple(result)


def profile_locator_for_digest(profile_digest: str) -> Path | None:
    """Resolve a persisted Profile without reconstructing it from catalog JSON."""

    bundled = [
        *Path(__file__).resolve().parent.glob("dataset-input-profile-*.json"),
        Path(__file__).resolve().parent / "welding-stage-b-profile-v1.json",
    ]
    for candidate in [*sorted(bundled), *personal_profile_paths()]:
        if candidate.is_file() and dataset_profile_digest(candidate) == profile_digest:
            return candidate.resolve()
    return None


def _normalized_name(value: str) -> str:
    return "".join(character.casefold() for character in value if character.isalnum())


def _rank_candidates(expected_names: list[str], available: list[str]) -> list[dict[str, Any]]:
    expected_normalized = [_normalized_name(item) for item in expected_names]
    ranked: list[tuple[float, str]] = []
    for source_name in available:
        normalized = _normalized_name(source_name)
        score = max(
            (
                SequenceMatcher(None, expected, normalized).ratio()
                for expected in expected_normalized
                if expected and normalized
            ),
            default=0.0,
        )
        ranked.append((score, source_name))
    ranked.sort(key=lambda item: (-item[0], item[1]))
    return [
        {"source_name": source_name, "score": round(score, 3)}
        for score, source_name in ranked
    ]


def _initial_binding(
    *,
    expected_names: list[str],
    available: list[str],
) -> tuple[str, str | None, list[dict[str, Any]]]:
    exact = next((name for name in expected_names if name in available), None)
    candidates = _rank_candidates(expected_names, available)
    if exact is not None:
        return "confirmed", exact, candidates
    if candidates:
        top = float(candidates[0]["score"])
        second = float(candidates[1]["score"]) if len(candidates) > 1 else 0.0
        if top >= 0.55 and top - second >= 0.08:
            return "suggested", str(candidates[0]["source_name"]), candidates
    return "unresolved", None, candidates


def _add_column_requirement(
    requirements: dict[tuple[str, str], dict[str, Any]],
    *,
    role: str,
    canonical_names: list[str],
    expected_names: list[str],
    semantic_kind: str,
    required: bool,
    source_unit: str | None = None,
    canonical_unit: str | None = None,
) -> None:
    canonical_names = [str(item) for item in canonical_names if str(item).strip()]
    expected_names = [str(item) for item in expected_names if str(item).strip()]
    if not role or not canonical_names or not expected_names:
        return
    key = (role, canonical_names[0])
    current = requirements.get(key)
    if current is None:
        requirements[key] = {
            "role": role,
            "canonical_names": canonical_names,
            "expected_names": expected_names,
            "semantic_kind": semantic_kind,
            "required": required,
            "source_unit": source_unit,
            "canonical_unit": canonical_unit,
        }
        return
    current["required"] = bool(current["required"] or required)
    if _SEMANTIC_KIND_ORDER[semantic_kind] < _SEMANTIC_KIND_ORDER[current["semantic_kind"]]:
        current["semantic_kind"] = semantic_kind
    current["canonical_names"] = list(dict.fromkeys([*current["canonical_names"], *canonical_names]))
    current["expected_names"] = list(dict.fromkeys([*current["expected_names"], *expected_names]))
    current["source_unit"] = current["source_unit"] or source_unit
    current["canonical_unit"] = current["canonical_unit"] or canonical_unit


def _profile_column_requirements(
    document: dict[str, Any],
    *,
    required_relation_entity_types: frozenset[str],
) -> dict[tuple[str, str], dict[str, Any]]:
    shared = dict(document.get("shared", {}))
    aliases = {
        str(role): {str(key): str(value) for key, value in dict(values).items()}
        for role, values in dict(shared.get("column_aliases", {})).items()
    }
    optional_roles = {str(item) for item in shared.get("optional_roles", [])}
    optional_technical = {str(item) for item in shared.get("optional_technical_fields", [])}
    requirements: dict[tuple[str, str], dict[str, Any]] = {}

    def add(
        role: str,
        canonical_names: list[str],
        semantic_kind: str,
        *,
        required: bool = True,
        source_unit: str | None = None,
        canonical_unit: str | None = None,
    ) -> None:
        expected = [
            aliases.get(role, {}).get(name, name)
            for name in canonical_names
        ]
        _add_column_requirement(
            requirements,
            role=role,
            canonical_names=canonical_names,
            expected_names=expected,
            semantic_kind=semantic_kind,
            required=required and role not in optional_roles,
            source_unit=source_unit,
            canonical_unit=canonical_unit,
        )

    for entity in shared.get("entities", []):
        add(str(entity.get("role", "")), [str(entity.get("key", ""))], "entity_key")
    relation = dict(shared.get("relation", {}))
    relation_role = str(relation.get("role", ""))
    for join in relation.get("joins", []):
        names = [str(join.get("column", "")), *[str(item) for item in join.get("alternate_columns", [])]]
        add(
            relation_role,
            names,
            "relation_join",
            required=str(join.get("entity_type", "")) in required_relation_entity_types,
        )
    for policy in shared.get("eligibility", []):
        role = str(policy.get("role", ""))
        name = str(policy.get("column", ""))
        add(role, [name], "policy")
    for technical in shared.get("technical", []):
        role = str(technical.get("role", ""))
        name = str(technical.get("name", ""))
        add(
            role,
            [str(technical.get("column", ""))],
            "technical",
            required=not technical_field_is_optional(
                role,
                name,
                optional_technical,
            ),
        )

    for task in dict(document.get("tasks", {})).values():
        for mapping in task.get("mappings", []):
            role = str(mapping.get("role", ""))
            if mapping.get("kind") == "ordered_heat_series":
                series = dict(mapping.get("series_columns", {}))
                for key in ("parent", "order", "time", "value"):
                    add(
                        role,
                        [str(series.get(key, ""))],
                        "series",
                        source_unit=str(series.get(f"{key}_source_unit", "") or "") or None,
                        canonical_unit=str(series.get(f"{key}_canonical_unit", "") or "") or None,
                    )
                fallback = mapping.get("measurement_point_fallback")
                if isinstance(fallback, dict):
                    fallback_role = str(fallback.get("master_role", ""))
                    for key in ("equipment_column", "order_column", "stage_column", "position_column"):
                        add(fallback_role, [str(fallback.get(key, ""))], "series", required=False)
                continue
            add(
                role,
                [str(mapping.get("column", ""))],
                "input",
                source_unit=str(mapping.get("source_unit", "") or "") or None,
                canonical_unit=str(mapping.get("canonical_unit", "") or "") or None,
            )
            for observation_source in mapping.get("observation_sources", []):
                add(
                    str(observation_source.get("role", "")),
                    [str(observation_source.get("column", ""))],
                    "input",
                    source_unit=str(mapping.get("source_unit", "") or "") or None,
                    canonical_unit=str(mapping.get("canonical_unit", "") or "") or None,
                )
        for observation in task.get("observations", []):
            role = str(observation.get("role", ""))
            add(role, [str(observation.get("id_column", ""))], "entity_key")
            if observation.get("parent_column"):
                add(role, [str(observation["parent_column"])], "relation_join")
            optional_metadata = {str(item) for item in observation.get("optional_metadata_keys", [])}
            for key, column in dict(observation.get("metadata_columns", {})).items():
                add(role, [str(column)], "technical", required=str(key) not in optional_metadata)
            for target in observation.get("targets", []):
                names = (
                    [str(item) for item in target.get("columns", [])]
                    or [str(target.get("column", ""))]
                )
                add(
                    role,
                    names,
                    "output",
                    source_unit=str(target.get("unit", "") or "") or None,
                    canonical_unit=str(target.get("unit", "") or "") or None,
                )
            optional_auxiliary = {str(item) for item in observation.get("optional_auxiliary_keys", [])}
            for auxiliary in observation.get("auxiliary", []):
                names = (
                    [str(item) for item in auxiliary.get("columns", [])]
                    or [str(auxiliary.get("column", ""))]
                )
                add(
                    role,
                    names,
                    "output",
                    required=str(auxiliary.get("key", "")) not in optional_auxiliary,
                    source_unit=str(auxiliary.get("unit", "") or "") or None,
                    canonical_unit=str(auxiliary.get("unit", "") or "") or None,
                )
    return requirements


def create_source_binding_draft(
    source: Path,
    base_profile_path: Path,
) -> dict[str, Any] | None:
    """Describe source-only bindings over an immutable v2 semantic Profile."""

    document = materialize_dataset_profile_document(base_profile_path)
    if document.get("schema_version") != _PROFILE_SCHEMA_VERSION:
        return None
    requirements = task_data_requirements(load_dataset_profile(base_profile_path))
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        inventory: dict[str, list[str]] = {}
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            inventory[sheet.title] = [
                str(value) if value is not None else ""
                for value in next(rows, ())
            ]
    finally:
        workbook.close()

    shared = dict(document["shared"])
    slots: list[dict[str, Any]] = []
    selected_sheets: dict[str, str | None] = {}
    optional_roles = {str(item) for item in shared.get("optional_roles", [])}
    for role, expected_sheet in dict(shared.get("sheets", {})).items():
        state, selected, candidates = _initial_binding(
            expected_names=[str(expected_sheet)],
            available=list(inventory),
        )
        selected_sheets[str(role)] = selected
        slots.append({
            "slot_id": f"sheet:{role}",
            "binding_type": "sheet",
            "role": str(role),
            "semantic_kind": "technical",
            "canonical_name": str(role),
            "expected_source_name": str(expected_sheet),
            "required": str(role) not in optional_roles,
            "state": state,
            "selected_source_name": selected,
            "candidates": candidates,
        })

    for requirement in _profile_column_requirements(
        document,
        required_relation_entity_types=requirements.relation_entity_types,
    ).values():
        role = str(requirement["role"])
        sheet_name = selected_sheets.get(role)
        available = inventory.get(sheet_name or "", [])
        all_headers = list(dict.fromkeys(
            header
            for headers in inventory.values()
            for header in headers
            if header
        ))
        state, selected, candidates = _initial_binding(
            expected_names=list(requirement["expected_names"]),
            available=available,
        )
        ranked_all = _rank_candidates(list(requirement["expected_names"]), all_headers)
        candidates = list({
            str(item["source_name"]): item
            for item in [*candidates, *ranked_all]
        }.values())
        candidates.sort(key=lambda item: (-float(item["score"]), str(item["source_name"])))
        if sheet_name is None:
            state = "unresolved"
            selected = None
        canonical_name = str(requirement["canonical_names"][0])
        slot = {
            "slot_id": f"column:{role}:{canonical_name}",
            "binding_type": "column",
            "role": role,
            "semantic_kind": str(requirement["semantic_kind"]),
            "canonical_name": canonical_name,
            "expected_source_name": str(requirement["expected_names"][0]),
            "required": bool(requirement["required"]),
            "state": state,
            "selected_source_name": selected,
            "candidates": candidates,
        }
        if requirement.get("source_unit"):
            slot["source_unit"] = str(requirement["source_unit"])
        if requirement.get("canonical_unit"):
            canonical_unit = str(requirement["canonical_unit"])
            slot["canonical_unit"] = canonical_unit
            slot["source_unit_candidates"] = list(source_units_for(canonical_unit))
        slots.append(slot)

    return {
        "base_profile_digest": dataset_profile_digest(base_profile_path),
        "source_sha256": file_sha256(source),
        "complete": all(
            slot["state"] == "confirmed"
            for slot in slots
            if slot["required"]
        ),
        "slots": slots,
    }


def _apply_confirmed_bindings(
    document: dict[str, Any],
    draft: dict[str, Any],
    bindings: list[dict[str, Any]],
) -> dict[str, Any]:
    slots = {str(item["slot_id"]): dict(item) for item in draft["slots"]}
    supplied_ids: set[str] = set()
    for binding in bindings:
        slot_id = str(binding.get("slot_id", ""))
        if slot_id in supplied_ids:
            raise ValueError(f"binding slot is duplicated: {slot_id}")
        supplied_ids.add(slot_id)
        slot = slots.get(slot_id)
        if slot is None:
            raise ValueError(f"unknown binding slot: {slot_id}")
        if binding.get("state") != "confirmed":
            raise ValueError(f"binding must be explicitly confirmed: {slot_id}")
        source_name = str(binding.get("source_name", "")).strip()
        if not source_name:
            raise ValueError(f"confirmed binding has no source name: {slot_id}")
        available = {str(item["source_name"]) for item in slot.get("candidates", [])}
        if source_name not in available and source_name != slot.get("selected_source_name"):
            raise ValueError(f"binding source is not present in the inspected workbook: {slot_id}")
        canonical_unit = str(slot.get("canonical_unit", "")).strip()
        if canonical_unit:
            raw_source_unit = binding.get("source_unit")
            source_unit = raw_source_unit.strip() if isinstance(raw_source_unit, str) else ""
            if not source_unit:
                raise ValueError(f"binding source unit must be explicitly confirmed: {slot_id}")
            header_unit = _header_unit(source_name)
            if header_unit is not None and header_unit != source_unit:
                raise ValueError(
                    f"binding source unit {source_unit!r} does not match "
                    f"the workbook header unit {header_unit!r}: {slot_id}"
                )
            if unit_conversion(source_unit, canonical_unit) is None:
                raise ValueError(
                    f"unsupported source unit conversion {source_unit!r} -> "
                    f"{canonical_unit!r}: {slot_id}"
                )
            slot["selected_source_unit"] = source_unit
        slot["state"] = "confirmed"
        slot["selected_source_name"] = source_name

    incomplete = [
        slot_id
        for slot_id, slot in slots.items()
        if slot.get("required") and slot.get("state") != "confirmed"
    ]
    if incomplete:
        raise ValueError("required bindings remain unresolved: " + ", ".join(incomplete[:10]))

    result = deepcopy(document)
    shared = result["shared"]
    original_sheets = dict(shared["sheets"])
    for slot in slots.values():
        if slot["binding_type"] == "sheet" and slot["state"] == "confirmed":
            shared["sheets"][slot["role"]] = slot["selected_source_name"]

    aliases = {
        str(role): {str(key): str(value) for key, value in dict(values).items()}
        for role, values in dict(shared.get("column_aliases", {})).items()
    }
    for slot in slots.values():
        if slot["binding_type"] != "column" or slot["state"] != "confirmed":
            continue
        role = str(slot["role"])
        canonical = str(slot["canonical_name"])
        source_name = str(slot["selected_source_name"])
        role_aliases = aliases.setdefault(role, {})
        if canonical == source_name:
            role_aliases.pop(canonical, None)
        else:
            role_aliases[canonical] = source_name
    shared["column_aliases"] = {
        role: values
        for role, values in aliases.items()
        if values
    }

    for slot in slots.values():
        source_unit = slot.get("selected_source_unit")
        if (
            slot["binding_type"] != "column"
            or slot["state"] != "confirmed"
            or not source_unit
        ):
            continue
        role = str(slot["role"])
        canonical = str(slot["canonical_name"])
        for task in result.get("tasks", {}).values():
            for mapping in task.get("mappings", []):
                if mapping.get("kind") == "ordered_heat_series":
                    if str(mapping.get("role", "")) != role:
                        continue
                    series = mapping.get("series_columns", {})
                    for key in ("parent", "order", "time", "value"):
                        if series.get(key) == canonical and f"{key}_source_unit" in series:
                            series[f"{key}_source_unit"] = source_unit
                    continue
                mapping_owns_slot = (
                    str(mapping.get("role", "")) == role
                    and mapping.get("column") == canonical
                ) or any(
                    str(observation_source.get("role", "")) == role
                    and observation_source.get("column") == canonical
                    for observation_source in mapping.get("observation_sources", [])
                )
                if mapping_owns_slot and "source_unit" in mapping:
                    mapping["source_unit"] = source_unit

    sheet_role_by_original = {name: role for role, name in original_sheets.items()}
    for marker in result.get("source_markers", []):
        role = sheet_role_by_original.get(marker.get("sheet"))
        if role is None:
            continue
        marker["sheet"] = shared["sheets"][role]
        marker["column"] = shared.get("column_aliases", {}).get(role, {}).get(
            marker.get("column"),
            marker.get("column"),
        )
    result.pop("extends", None)
    return result


def save_source_binding_profile(
    *,
    source: Path,
    base_profile_path: Path,
    expected_source_sha256: str,
    bindings: list[dict[str, Any]],
    store_path: Path | None = None,
) -> dict[str, Any]:
    """Validate and atomically persist a standalone source-bound Profile."""

    actual_source_sha256 = file_sha256(source)
    if actual_source_sha256 != expected_source_sha256:
        raise RuntimeError("source_changed")
    document = materialize_dataset_profile_document(base_profile_path)
    if document.get("schema_version") != _PROFILE_SCHEMA_VERSION:
        raise ValueError("Only dataset-input-profile/v2 can be used as a binding base")
    draft = create_source_binding_draft(source, base_profile_path)
    if draft is None:
        raise ValueError("The selected Profile cannot produce a source-binding draft")
    effective = _apply_confirmed_bindings(document, draft, bindings)

    store = validate_personal_profile_store_path(store_path)
    store.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    metadata_temporary_path: Path | None = None
    try:
        with NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=store,
            prefix=".profile-",
            suffix=".tmp",
            delete=False,
        ) as stream:
            json.dump(effective, stream, ensure_ascii=False, indent=2)
            stream.write("\n")
            temporary_path = Path(stream.name)
        load_dataset_profile(temporary_path)
        validation = validate_workbook_profile(source, temporary_path)
        profile_digest = dataset_profile_digest(temporary_path)
        digest_value = profile_digest.removeprefix("sha256:")
        final_path = store / f"{digest_value}.json"
        metadata_path = store / f"{digest_value}.meta.json"
        with NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=store,
            prefix=".profile-meta-",
            suffix=".tmp",
            delete=False,
        ) as metadata_stream:
            json.dump(
                {
                    "source_sha256": actual_source_sha256,
                    "base_profile_digest": dataset_profile_digest(base_profile_path),
                },
                metadata_stream,
                ensure_ascii=False,
                indent=2,
            )
            metadata_stream.write("\n")
            metadata_temporary_path = Path(metadata_stream.name)
        # The Profile JSON is the registry's commit marker. Publish metadata first
        # and the validated Profile last so a failed save never becomes visible.
        os.replace(metadata_temporary_path, metadata_path)
        metadata_temporary_path = None
        os.replace(temporary_path, final_path)
        temporary_path = None
        return {
            "profile_id": str(validation["profile_id"]),
            "profile_digest": profile_digest,
            "profile_locator": str(final_path.resolve()),
            "source_sha256": actual_source_sha256,
            "base_profile_digest": dataset_profile_digest(base_profile_path),
            "validation": validation,
        }
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
        if metadata_temporary_path is not None:
            metadata_temporary_path.unlink(missing_ok=True)


def inspect_workbook(source: Path, profile_path: Path | None = None) -> dict[str, Any]:
    """Return a compact workbook inventory and canonicalization preflight."""

    source = source.resolve()
    selected = profile_path.resolve() if profile_path else None
    profile_error: str | None = None
    if selected is None:
        try:
            selected = detect_dataset_profile_path(source).resolve()
        except (OSError, ValueError) as exc:
            profile_error = str(exc)

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value) if value is not None else "" for value in next(rows, ())]
            nonempty_rows = sum(1 for row in rows if any(value is not None for value in row))
            sheets.append({"name": sheet.title, "headers": headers, "rows": nonempty_rows})

        result: dict[str, Any] = {
            "source": str(source),
            "source_sha256": file_sha256(source),
            "sheets": sheets,
            "profile": str(selected) if selected else None,
            "profile_error": profile_error,
            "canonicalization": None,
        }
        if selected is not None:
            try:
                result["canonicalization"] = validate_source_profile(source, selected)
            except (DatasetProfileError, ValueError) as exc:
                result["profile_error"] = "\n".join(
                    getattr(exc, "errors", (str(exc),))
                )
        return result
    finally:
        workbook.close()


def validate_source_profile(source: Path, profile_path: Path) -> dict[str, Any]:
    """Validate one explicit Profile through its allow-listed family adapter."""

    from decision_workbench.data.profile_family_registry import validate_profile_source

    return validate_profile_source(source.resolve(), profile_path.resolve())


def _stage_b_validation(source: Path, profile: Any, profile_path: Path) -> dict[str, Any]:
    """Return the stable validation payload for the Stage B Profile family."""

    source = source.resolve()
    profile_path = profile_path.resolve()
    from decision_workbench.data.stage_b_training import build_stage_b_training_data

    training = build_stage_b_training_data(
        source,
        profile,
        profile_locator=profile_path,
    )
    observations = training.data.observations
    usable = [row for row in observations if row["eligible"]]
    source_sha256 = file_sha256(source)
    rejected_by_policy = {
        reason: sum(reason in row["exclusion_reasons"] for row in observations)
        for reason in sorted({
            reason for row in observations for reason in row["exclusion_reasons"]
        })
    }
    disposition = build_count_disposition(
        source_sha256=source_sha256,
        profile_digest=training.profile_digest,
        canonicalization_contract_digest=DATASET_CANONICALIZATION_CONTRACT_DIGEST,
        task_ids=[profile.task_id],
        entities=len(observations),
        observations_by_task={profile.task_id: len(observations)},
        usable_observations_by_task={profile.task_id: len(usable)},
        rejected_by_policy=rejected_by_policy,
    )
    return {
        "ok": True,
        "registration_ready": bool(usable),
        "source": str(source),
        "source_sha256": source_sha256,
        "profile": str(profile_path),
        "profile_id": profile.id,
        "profile_digest": training.profile_digest,
        "task_ids": [profile.task_id],
        "entities": len(observations),
        "relations": 0,
        "observations": len(observations),
        "observations_by_task": {profile.task_id: len(usable)},
        "heat_series_parents": 0,
        "unresolved_heat_series_by_task": {},
        "rejected_by_policy": rejected_by_policy,
        "entity_preview": [
            {
                "entity_type": "weld_metal",
                "field_count": len(row),
                "eligible": bool(row["eligible"]),
            }
            for row in observations[:5]
        ],
        "disposition": disposition.model_dump(mode="json"),
    }


def validate_workbook_profile(source: Path, profile_path: Path) -> dict[str, Any]:
    """Compatibility entry point for validation through the Profile family registry."""

    return validate_source_profile(source, profile_path)


def _validate_dataset_input_workbook(source: Path, profile_path: Path) -> dict[str, Any]:
    """Validate one Dataset Input Profile without runtime/model assumptions."""

    source = source.resolve()
    profile_path = profile_path.resolve()
    profile = load_dataset_profile(profile_path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        canonical = canonicalize_workbook(workbook, profile)
    finally:
        workbook.close()
    observations_by_task = Counter(item.task_id for item in canonical.observations)
    rejected_by_policy = Counter(
        f"{item.source_role}.{policy}"
        for item in canonical.observations
        for policy, accepted in item.policy_results.items()
        if not accepted
    )
    source_sha256 = file_sha256(source)
    profile_digest = dataset_profile_digest(profile_path)
    disposition = build_dataset_disposition(
        canonical,
        source_sha256=source_sha256,
        profile_digest=profile_digest,
        canonicalization_contract_digest=DATASET_CANONICALIZATION_CONTRACT_DIGEST,
    )
    unresolved_heat_series_by_task = {
        task_id: item.unresolved_heat_series_parent_count
        for task_id, item in disposition.task_dispositions.items()
        if item.unresolved_heat_series_parent_count
    }
    entity_preview = [
        {
            "entity_type": identity[0],
            "task_count": len(entity.values),
            "field_count": sum(
                len(fields)
                for fields in entity.values.values()
                if isinstance(fields, dict)
            ),
        }
        for identity, entity in list(canonical.entities.items())[:5]
    ]
    return {
        "ok": True,
        "registration_ready": any(
            item.usable_observation_count > 0
            for item in disposition.task_dispositions.values()
        ),
        "source": str(source),
        "source_sha256": source_sha256,
        "profile": str(profile_path),
        "profile_id": profile.profile_id,
        "profile_digest": profile_digest,
        "task_ids": sorted(profile.tasks),
        "entities": len(canonical.entities),
        "relations": len(canonical.relations),
        "observations": len(canonical.observations),
        "observations_by_task": dict(sorted(observations_by_task.items())),
        "heat_series_parents": len(canonical.heat_series),
        "unresolved_heat_series_by_task": unresolved_heat_series_by_task,
        "rejected_by_policy": dict(sorted(rejected_by_policy.items())),
        "entity_preview": entity_preview,
        "disposition": disposition.model_dump(mode="json"),
    }
