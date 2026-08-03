from __future__ import annotations

import hashlib
import json
import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Literal, Mapping

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, ValidationError, field_validator, model_validator


OBSERVATION_PROFILE_SCHEMA_VERSION = "observation-dataset-profile/v1"
Scalar = str | float | int | bool | None


class ObservationProfileError(ValueError):
    def __init__(self, errors: Iterable[str]):
        self.errors = tuple(dict.fromkeys(str(error) for error in errors))
        super().__init__("Observation dataset profile failed:\n- " + "\n- ".join(self.errors))


class ProfileModel(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)


class EntitySource(ProfileModel):
    role: str
    sheet: str
    key_column: str
    relation_column: str


class CanonicalInput(ProfileModel):
    path: str
    role: str
    column: str
    kind: Literal["numeric", "categorical"]
    source_unit: str | None = None
    canonical_unit: str | None = None

    @model_validator(mode="after")
    def numeric_units_are_explicit(self) -> "CanonicalInput":
        if self.kind == "numeric" and (not self.source_unit or not self.canonical_unit):
            raise ValueError("numeric input requires source_unit and canonical_unit")
        if self.kind == "categorical" and (self.source_unit or self.canonical_unit):
            raise ValueError("categorical input cannot declare units")
        return self


class CanonicalOutput(ProfileModel):
    key: str
    column: str
    source_unit: str
    canonical_unit: str


class MetadataSource(ProfileModel):
    key: str
    column: str


class FixedContext(ProfileModel):
    path: str
    column: str
    expected: str | float | int | bool
    unit: str | None = None


class ObservationFamily(ProfileModel):
    id: str
    sheet: str
    relation_column: str
    observation_id_column: str
    split_group_role: str
    inputs: tuple[CanonicalInput, ...]
    outputs: tuple[CanonicalOutput, ...]
    metadata: tuple[MetadataSource, ...] = ()
    fixed_context: tuple[FixedContext, ...] = ()

    @field_validator("inputs")
    @classmethod
    def input_paths_are_unique(cls, values: tuple[CanonicalInput, ...]) -> tuple[CanonicalInput, ...]:
        paths = [value.path for value in values]
        if len(paths) != len(set(paths)):
            raise ValueError("family input paths must be unique")
        return values

    @field_validator("outputs")
    @classmethod
    def output_keys_are_unique(cls, values: tuple[CanonicalOutput, ...]) -> tuple[CanonicalOutput, ...]:
        keys = [value.key for value in values]
        if not keys or len(keys) != len(set(keys)):
            raise ValueError("family outputs must be non-empty and unique")
        return values


class SingleTableObservationSource(ProfileModel):
    """Typed layout for one table containing repeated measurement rows."""

    group_column: str
    source_sheet: str = "observations"


class ObservationValidationPlan(ProfileModel):
    strategy: Literal["grouped-k-fold"] = "grouped-k-fold"
    folds: int = Field(default=5, ge=2, le=10)


class ObservationFeatureRecipe(ProfileModel):
    id: Literal["observation-identity-v1"] = "observation-identity-v1"
    version: Literal["1.0.0"] = "1.0.0"


class ObservationEstimatorChoice(ProfileModel):
    id: Literal["ridge"] = "ridge"
    alpha: float = Field(default=1.0, gt=0)


class ObservationAuthoringContract(ProfileModel):
    source_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    observation_grain: str
    technical_metadata: tuple[str, ...] = ()
    target_eligibility: Literal["non-missing-per-target"] = "non-missing-per-target"
    validation_plan: ObservationValidationPlan = ObservationValidationPlan()
    feature_recipe: ObservationFeatureRecipe = ObservationFeatureRecipe()
    estimator: ObservationEstimatorChoice = ObservationEstimatorChoice()


# This family reads source values as-is; it never rescales. A declared
# source -> canonical pair may therefore only be a relabel, and every allowed
# relabel is listed here. A pair that would need a numeric factor is rejected at
# load time instead of silently passing raw values under the canonical label.
ALLOWED_UNIT_RELABELS = frozenset({
    ("%", "mass% deposited metal"),
    ("%", "mass% whole wire"),
    ("℃", "°C"),
    ("degC", "°C"),
    ("Mpa", "MPa"),
    ("秒", "s"),
})


def _relabel_is_allowed(source: str | None, canonical: str | None) -> bool:
    if source is None and canonical is None:
        return True
    if source is None or canonical is None:
        return False
    return source == canonical or (source, canonical) in ALLOWED_UNIT_RELABELS


class ObservationDatasetProfile(ProfileModel):
    schema_version: Literal["observation-dataset-profile/v1"]
    id: str
    task_id: str
    relation_sheet: str
    entities: tuple[EntitySource, ...]
    families: tuple[ObservationFamily, ...]
    single_table: SingleTableObservationSource | None = Field(
        default=None, exclude_if=lambda value: value is None
    )
    authoring: ObservationAuthoringContract | None = Field(
        default=None, exclude_if=lambda value: value is None
    )

    @property
    def profile_id(self) -> str:
        return self.id

    @model_validator(mode="after")
    def declared_units_are_relabels_only(self) -> "ObservationDatasetProfile":
        rescales = [
            f"{family.id}.{path}: {source!r} -> {canonical!r}"
            for family in self.families
            for path, source, canonical in (
                *(
                    (item.path, item.source_unit, item.canonical_unit)
                    for item in family.inputs
                ),
                *(
                    (item.key, item.source_unit, item.canonical_unit)
                    for item in family.outputs
                ),
            )
            if not _relabel_is_allowed(source, canonical)
        ]
        if rescales:
            raise ValueError(
                "この family は元の値をそのまま読むため、単位の数値変換を宣言できません。"
                "元データ側で換算するか、係数1の読み替えとしてallow-listへ追加してください: "
                + "; ".join(rescales)
            )
        return self

    @model_validator(mode="after")
    def references_are_consistent(self) -> "ObservationDatasetProfile":
        roles = [entity.role for entity in self.entities]
        family_ids = [family.id for family in self.families]
        if len(roles) != len(set(roles)):
            raise ValueError("entity roles must be unique")
        if len(family_ids) != len(set(family_ids)):
            raise ValueError("family ids must be unique")
        known_roles = set(roles)
        errors = []
        for family in self.families:
            if self.single_table is None and family.split_group_role not in known_roles:
                errors.append(f"{family.id}: unknown split_group_role {family.split_group_role!r}")
            for mapping in family.inputs:
                if self.single_table is not None and mapping.role != family.id:
                    errors.append(
                        f"{family.id}: single-table input must use its family role: {mapping.path}"
                    )
                elif mapping.role != family.id and mapping.role not in known_roles:
                    errors.append(f"{family.id}: input references unknown role {mapping.role!r}")
        if self.single_table is not None:
            if self.entities:
                errors.append("single-table profile cannot declare entity tables")
            if len(self.families) != 1:
                errors.append("single-table profile requires exactly one observation family")
            if self.authoring is None:
                errors.append("single-table profile requires authoring contract")
        if errors:
            raise ValueError("; ".join(errors))
        return self

    def family(self, family_id: str) -> ObservationFamily:
        match = next((family for family in self.families if family.id == family_id), None)
        if match is None:
            raise KeyError(f"unknown observation family: {family_id}")
        return match


class TargetCurationState(ProfileModel):
    usable: bool
    reasons: tuple[str, ...] = ()


class ObservationRowProvenance(ProfileModel):
    source_sheet: str
    source_row: int
    relation_sheet: str
    relation_rows: tuple[int, ...]
    entity_keys: Mapping[str, str]


class ObservationTrainingRow(ProfileModel):
    family: str
    observation_id: str
    split_group_key: str | None
    inputs: Mapping[str, Scalar]
    outputs: Mapping[str, float]
    metadata: Mapping[str, Scalar]
    fixed_context: Mapping[str, Scalar]
    eligible: bool
    exclusion_reasons: tuple[str, ...]
    target_status: Mapping[str, TargetCurationState]
    provenance: ObservationRowProvenance


class TargetTrainingSummary(ProfileModel):
    target: str
    usable_rows: int
    split_groups: int
    exclusion_reasons: Mapping[str, int]


class FamilyTrainingSummary(ProfileModel):
    family: str
    source_rows: int
    usable_input_rows: int
    excluded_input_rows: int
    split_groups: int
    exclusion_reasons: Mapping[str, int]
    targets: tuple[TargetTrainingSummary, ...]


class ObservationTrainingView(ProfileModel):
    schema_version: Literal["observation-training-view/v1"] = "observation-training-view/v1"
    profile_id: str
    profile_digest: str
    source_sha256: str
    family: str
    feature_names: tuple[str, ...]
    rows: tuple[ObservationTrainingRow, ...]
    summary: FamilyTrainingSummary


class ObservationTrainingDataset(ProfileModel):
    schema_version: Literal["observation-training-dataset/v1"] = "observation-training-dataset/v1"
    profile_id: str
    profile_digest: str
    source_sha256: str
    views: Mapping[str, ObservationTrainingView]


class ObservationTrainingInspectionPage(ProfileModel):
    profile_id: str
    profile_digest: str
    source_sha256: str
    family: str
    target: str
    source_rows: int
    usable_rows: int
    split_groups: int
    exclusion_reasons: Mapping[str, int]
    offset: int
    limit: int
    rows: tuple[ObservationTrainingRow, ...]


class ObservationTrainingProfileSummary(ProfileModel):
    profile_id: str
    profile_digest: str
    source_filename: str
    source_sha256: str
    families: tuple[FamilyTrainingSummary, ...]


@dataclass(frozen=True)
class _SheetTable:
    headers: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]


def load_observation_profile(path: str | Path) -> ObservationDatasetProfile:
    profile_path = Path(path)
    try:
        raw = json.loads(profile_path.read_text(encoding="utf-8"))
        return ObservationDatasetProfile.model_validate(raw)
    except (OSError, json.JSONDecodeError, ValidationError) as exc:
        raise ObservationProfileError([f"invalid observation profile {profile_path}: {exc}"]) from exc


def observation_profile_digest(profile: ObservationDatasetProfile) -> str:
    document = profile.model_dump(mode="json")
    if document.get("single_table") is None:
        document.pop("single_table", None)
    if document.get("authoring") is None:
        document.pop("authoring", None)
    payload = json.dumps(
        document,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return f"sha256:{hashlib.sha256(payload).hexdigest()}"


def _source_digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_table(workbook: Any, sheet_name: str) -> _SheetTable:
    if sheet_name not in workbook.sheetnames:
        raise ObservationProfileError([f"missing worksheet: {sheet_name}"])
    worksheet = workbook[sheet_name]
    values = worksheet.iter_rows(values_only=True)
    header_row = next(values, None)
    if not header_row:
        raise ObservationProfileError([f"{sheet_name}: worksheet is empty"])
    headers = tuple(str(value).strip() if value is not None else "" for value in header_row)
    if any(not header for header in headers) or len(headers) != len(set(headers)):
        raise ObservationProfileError([f"{sheet_name}: headers must be non-empty and unique"])
    rows = tuple(
        {header: value for header, value in zip(headers, row)}
        for row in values
        if any(value is not None and value != "" for value in row)
    )
    return _SheetTable(headers=headers, rows=rows)


def _read_single_table(path: Path, source_sheet: str) -> _SheetTable:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            headers = tuple(str(value).strip() for value in (reader.fieldnames or ()))
            if not headers:
                raise ObservationProfileError(["CSV header is required"])
            if any(not header for header in headers) or len(headers) != len(set(headers)):
                raise ObservationProfileError(["CSV headers must be non-empty and unique"])
            rows = tuple(
                dict(row)
                for row in reader
                if any(value not in (None, "") for value in row.values())
            )
            return _SheetTable(headers=headers, rows=rows)
    if path.suffix.lower() != ".xlsx":
        raise ObservationProfileError(["single-table observation source must be .csv or .xlsx"])
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        visible = tuple(sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible")
        if len(visible) != 1:
            raise ObservationProfileError([
                "single-table observation authoring requires exactly one visible worksheet"
            ])
        return _read_table(workbook, visible[0])
    finally:
        workbook.close()


def _required_columns(
    profile: ObservationDatasetProfile,
) -> dict[str, set[str]]:
    required: dict[str, set[str]] = defaultdict(set)
    for entity in profile.entities:
        required[profile.relation_sheet].add(entity.relation_column)
        required[entity.sheet].add(entity.key_column)
    for family in profile.families:
        required[profile.relation_sheet].add(family.relation_column)
        required[family.sheet].update((family.observation_id_column, family.relation_column))
        for mapping in family.inputs:
            source_sheet = family.sheet if mapping.role == family.id else next(
                entity.sheet for entity in profile.entities if entity.role == mapping.role
            )
            required[source_sheet].add(mapping.column)
        required[family.sheet].update(output.column for output in family.outputs)
        required[family.sheet].update(item.column for item in family.metadata)
        required[family.sheet].update(item.column for item in family.fixed_context)
    return required


def _as_key(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _numeric(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if result == result and result not in {float("inf"), float("-inf")} else None


def _scalar(value: Any) -> Scalar:
    if value is None or isinstance(value, (str, float, int, bool)):
        return value
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        return str(isoformat())
    return str(value)


def _same_fixed_value(actual: Any, expected: Any) -> bool:
    if isinstance(expected, (int, float)) and not isinstance(expected, bool):
        parsed = _numeric(actual)
        return parsed is not None and abs(parsed - float(expected)) <= 1e-9
    return str(actual).strip() == str(expected).strip()


def _summarize(
    family: ObservationFamily,
    rows: list[ObservationTrainingRow],
) -> FamilyTrainingSummary:
    family_reasons = Counter(reason for row in rows for reason in row.exclusion_reasons)
    usable_rows = [row for row in rows if row.eligible]
    targets = []
    for output in family.outputs:
        target_rows = [row for row in rows if row.target_status[output.key].usable]
        target_reasons = Counter(
            reason
            for row in rows
            for reason in row.target_status[output.key].reasons
        )
        targets.append(TargetTrainingSummary(
            target=output.key,
            usable_rows=len(target_rows),
            split_groups=len({row.split_group_key for row in target_rows if row.split_group_key}),
            exclusion_reasons=dict(sorted(target_reasons.items(), key=lambda item: (-item[1], item[0]))),
        ))
    return FamilyTrainingSummary(
        family=family.id,
        source_rows=len(rows),
        usable_input_rows=len(usable_rows),
        excluded_input_rows=len(rows) - len(usable_rows),
        split_groups=len({row.split_group_key for row in usable_rows if row.split_group_key}),
        exclusion_reasons=dict(sorted(family_reasons.items(), key=lambda item: (-item[1], item[0]))),
        targets=tuple(targets),
    )


def _build_single_table_training_dataset(
    source_path: Path,
    profile: ObservationDatasetProfile,
) -> ObservationTrainingDataset:
    assert profile.single_table is not None
    assert profile.authoring is not None
    actual_sha256 = _source_digest(source_path)
    if actual_sha256 != profile.authoring.source_sha256:
        raise ObservationProfileError(["source identity differs from the authored Profile"])
    family = profile.families[0]
    table = _read_single_table(source_path, profile.single_table.source_sheet)
    required = {
        family.observation_id_column,
        profile.single_table.group_column,
        *(mapping.column for mapping in family.inputs),
        *(output.column for output in family.outputs),
        *(item.column for item in family.metadata),
        *(item.column for item in family.fixed_context),
    }
    missing = sorted(required - set(table.headers))
    if missing:
        raise ObservationProfileError([f"missing columns: {', '.join(missing)}"])

    rows: list[ObservationTrainingRow] = []
    seen_observations: set[str] = set()
    for source_row_number, observation in enumerate(table.rows, start=2):
        observation_id = _as_key(observation.get(family.observation_id_column))
        if observation_id is None:
            observation_id = f"{family.id}:row-{source_row_number}"
        reasons: list[str] = []
        if observation_id in seen_observations:
            reasons.append("観測IDが重複")
        seen_observations.add(observation_id)
        split_group_key = _as_key(observation.get(profile.single_table.group_column))
        if split_group_key is None:
            reasons.append("分割groupなし")

        inputs: dict[str, Scalar] = {}
        for mapping in family.inputs:
            raw = observation.get(mapping.column)
            if mapping.kind == "numeric":
                value = _numeric(raw)
                if value is None:
                    reasons.append(f"入力値なし: {mapping.path}")
                inputs[mapping.path] = value
            else:
                value = _as_key(raw)
                if value is None:
                    reasons.append(f"入力値なし: {mapping.path}")
                inputs[mapping.path] = value

        fixed_context: dict[str, Scalar] = {}
        for item in family.fixed_context:
            actual = observation.get(item.column)
            fixed_context[item.path] = _scalar(actual)
            if not _same_fixed_value(actual, item.expected):
                reasons.append(f"固定context不一致: {item.path}")
        metadata = {item.key: _scalar(observation.get(item.column)) for item in family.metadata}

        outputs: dict[str, float] = {}
        target_status: dict[str, TargetCurationState] = {}
        for output in family.outputs:
            value = _numeric(observation.get(output.column))
            if value is not None:
                outputs[output.key] = value
            output_reasons = tuple(dict.fromkeys([
                *reasons,
                *(("値なし",) if value is None else ()),
            ]))
            target_status[output.key] = TargetCurationState(
                usable=not output_reasons,
                reasons=output_reasons,
            )

        rows.append(ObservationTrainingRow(
            family=family.id,
            observation_id=observation_id,
            split_group_key=split_group_key,
            inputs=inputs,
            outputs=outputs,
            metadata=metadata,
            fixed_context=fixed_context,
            eligible=not reasons,
            exclusion_reasons=tuple(dict.fromkeys(reasons)),
            target_status=target_status,
            provenance=ObservationRowProvenance(
                source_sheet=profile.single_table.source_sheet,
                source_row=source_row_number,
                relation_sheet=profile.single_table.source_sheet,
                relation_rows=(source_row_number,),
                entity_keys={"group": split_group_key} if split_group_key else {},
            ),
        ))

    digest = observation_profile_digest(profile)
    view = ObservationTrainingView(
        profile_id=profile.id,
        profile_digest=digest,
        source_sha256=actual_sha256,
        family=family.id,
        feature_names=tuple(mapping.path for mapping in family.inputs),
        rows=tuple(rows),
        summary=_summarize(family, rows),
    )
    return ObservationTrainingDataset(
        profile_id=profile.id,
        profile_digest=digest,
        source_sha256=actual_sha256,
        views={family.id: view},
    )


def build_observation_training_dataset(
    source: str | Path,
    profile: ObservationDatasetProfile,
) -> ObservationTrainingDataset:
    source_path = Path(source)
    if profile.single_table is not None:
        return _build_single_table_training_dataset(source_path, profile)
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet_names = {
            profile.relation_sheet,
            *(entity.sheet for entity in profile.entities),
            *(family.sheet for family in profile.families),
        }
        tables = {name: _read_table(workbook, name) for name in sheet_names}
    finally:
        workbook.close()

    errors = []
    for sheet, columns in _required_columns(profile).items():
        missing = sorted(columns - set(tables[sheet].headers))
        if missing:
            errors.append(f"{sheet}: missing columns: {', '.join(missing)}")
    if errors:
        raise ObservationProfileError(errors)

    entity_by_role = {entity.role: entity for entity in profile.entities}
    entity_indexes: dict[str, dict[str, dict[str, Any]]] = {}
    for entity in profile.entities:
        index: dict[str, dict[str, Any]] = {}
        duplicates = set()
        for row in tables[entity.sheet].rows:
            key = _as_key(row.get(entity.key_column))
            if key is None:
                continue
            if key in index:
                duplicates.add(key)
            else:
                index[key] = row
        if duplicates:
            raise ObservationProfileError([
                f"{entity.sheet}: duplicate {entity.key_column}: {', '.join(sorted(duplicates)[:5])}"
            ])
        entity_indexes[entity.role] = index

    relation_rows = tables[profile.relation_sheet].rows
    relation_indexes: dict[str, dict[str, list[tuple[int, dict[str, Any]]]]] = {}
    for family in profile.families:
        index: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
        for row_number, relation in enumerate(relation_rows, start=2):
            key = _as_key(relation.get(family.relation_column))
            if key:
                index[key].append((row_number, relation))
        relation_indexes[family.id] = index

    digest = observation_profile_digest(profile)
    source_sha256 = _source_digest(source_path)
    views: dict[str, ObservationTrainingView] = {}
    for family in profile.families:
        rows: list[ObservationTrainingRow] = []
        seen_observations: set[str] = set()
        for source_row_number, observation in enumerate(tables[family.sheet].rows, start=2):
            observation_id = _as_key(observation.get(family.observation_id_column))
            if observation_id is None:
                observation_id = f"{family.id}:row-{source_row_number}"
            duplicate_observation = observation_id in seen_observations
            seen_observations.add(observation_id)
            matches = relation_indexes[family.id].get(observation_id, [])
            relation = matches[0][1] if len(matches) == 1 else {}
            reasons: list[str] = []
            if duplicate_observation:
                reasons.append("観測IDが重複")
            if not matches:
                reasons.append("relationに対応行なし")
            elif len(matches) > 1:
                reasons.append("relationが複数行に対応")

            entity_rows: dict[str, dict[str, Any]] = {}
            entity_keys: dict[str, str] = {}
            if relation:
                for entity in profile.entities:
                    key = _as_key(relation.get(entity.relation_column))
                    if key is None:
                        reasons.append(f"{entity.role}キーなし")
                        continue
                    entity_keys[entity.role] = key
                    entity_row = entity_indexes[entity.role].get(key)
                    if entity_row is None:
                        reasons.append(f"{entity.role}実体なし")
                    else:
                        entity_rows[entity.role] = entity_row

            inputs: dict[str, Scalar] = {}
            for mapping in family.inputs:
                source_row = observation if mapping.role == family.id else entity_rows.get(mapping.role)
                raw = source_row.get(mapping.column) if source_row else None
                if mapping.kind == "numeric":
                    value = _numeric(raw)
                    if value is None and source_row is not None:
                        reasons.append(f"入力値なし: {mapping.path}")
                    inputs[mapping.path] = value
                else:
                    value = _as_key(raw)
                    if value is None and source_row is not None:
                        reasons.append(f"入力値なし: {mapping.path}")
                    inputs[mapping.path] = value

            fixed_context: dict[str, Scalar] = {}
            for item in family.fixed_context:
                actual = observation.get(item.column)
                fixed_context[item.path] = _scalar(actual)
                if not _same_fixed_value(actual, item.expected):
                    reasons.append(f"固定context不一致: {item.path}")

            metadata = {
                item.key: _scalar(observation.get(item.column))
                for item in family.metadata
            }
            split_entity = entity_by_role[family.split_group_role]
            split_group_key = _as_key(relation.get(split_entity.relation_column)) if relation else None
            if relation and split_group_key is None:
                reasons.append("分割groupなし")

            eligible = not reasons
            outputs: dict[str, float] = {}
            target_status: dict[str, TargetCurationState] = {}
            for output in family.outputs:
                value = _numeric(observation.get(output.column))
                if value is not None:
                    outputs[output.key] = value
                output_reasons = tuple(dict.fromkeys([
                    *reasons,
                    *(("値なし",) if value is None else ()),
                ]))
                target_status[output.key] = TargetCurationState(
                    usable=not output_reasons,
                    reasons=output_reasons,
                )

            rows.append(ObservationTrainingRow(
                family=family.id,
                observation_id=observation_id,
                split_group_key=split_group_key,
                inputs=inputs,
                outputs=outputs,
                metadata=metadata,
                fixed_context=fixed_context,
                eligible=eligible,
                exclusion_reasons=tuple(dict.fromkeys(reasons)),
                target_status=target_status,
                provenance=ObservationRowProvenance(
                    source_sheet=family.sheet,
                    source_row=source_row_number,
                    relation_sheet=profile.relation_sheet,
                    relation_rows=tuple(row_number for row_number, _ in matches),
                    entity_keys=entity_keys,
                ),
            ))

        feature_names = tuple(mapping.path for mapping in family.inputs)
        views[family.id] = ObservationTrainingView(
            profile_id=profile.id,
            profile_digest=digest,
            source_sha256=source_sha256,
            family=family.id,
            feature_names=feature_names,
            rows=tuple(rows),
            summary=_summarize(family, rows),
        )
    return ObservationTrainingDataset(
        profile_id=profile.id,
        profile_digest=digest,
        source_sha256=source_sha256,
        views=views,
    )


def materialize_observation_training_dataset(
    dataset: ObservationTrainingDataset,
    destination: str | Path,
) -> Path:
    destination_path = Path(destination)
    destination_path.mkdir(parents=True, exist_ok=True)
    summary = {
        "schema_version": dataset.schema_version,
        "profile_id": dataset.profile_id,
        "profile_digest": dataset.profile_digest,
        "source_sha256": dataset.source_sha256,
        "families": {
            family: view.summary.model_dump(mode="json")
            for family, view in dataset.views.items()
        },
    }
    (destination_path / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    for family, view in dataset.views.items():
        with (destination_path / f"{family}.jsonl").open("w", encoding="utf-8", newline="\n") as stream:
            for row in view.rows:
                stream.write(json.dumps(row.model_dump(mode="json"), ensure_ascii=False, separators=(",", ":")) + "\n")
    return destination_path


def inspect_observation_training_view(
    dataset: ObservationTrainingDataset,
    *,
    family: str,
    target: str,
    offset: int = 0,
    limit: int = 25,
) -> ObservationTrainingInspectionPage:
    if offset < 0:
        raise ValueError("offset must be non-negative")
    if limit < 1 or limit > 100:
        raise ValueError("limit must be between 1 and 100")
    try:
        view = dataset.views[family]
    except KeyError as exc:
        raise KeyError(f"unknown observation family: {family}") from exc
    summary = next((item for item in view.summary.targets if item.target == target), None)
    if summary is None:
        raise KeyError(f"{family} does not define target: {target}")
    return ObservationTrainingInspectionPage(
        profile_id=dataset.profile_id,
        profile_digest=dataset.profile_digest,
        source_sha256=dataset.source_sha256,
        family=family,
        target=target,
        source_rows=view.summary.source_rows,
        usable_rows=summary.usable_rows,
        split_groups=summary.split_groups,
        exclusion_reasons=summary.exclusion_reasons,
        offset=offset,
        limit=limit,
        rows=view.rows[offset:offset + limit],
    )
